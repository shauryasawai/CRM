from email.message import EmailMessage
import os
import traceback
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.urls import reverse
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives, send_mail
from django.utils import timezone
from django.db.models import Sum, Count, Q, Avg
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.core.exceptions import PermissionDenied, ValidationError
from django.db.models import Sum, Avg, F, Q
from django.db.models.functions import Extract
from django.utils import timezone
from django.db import transaction
from django.views.decorators.http import require_POST
from django.core.files.storage import default_storage
import json
from django.views.decorators.http import require_POST, require_GET
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, Http404, HttpResponseForbidden
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, Count, Q, Avg, F
from django.db.models.functions import Extract
from django.utils import timezone
from django.db import transaction
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.urls import reverse
from datetime import datetime, timedelta
from django.views.decorators.http import require_http_methods
from django.core.mail import send_mail
from project import settings
from .models import (
    ActionPlanWorkflow, ClientInteraction, ClientPortfolio, ExecutionMetrics, ExecutionPlan, MutualFundScheme, PlanAction, PlanComment, PlanTemplate, PlanWorkflowHistory, PortfolioAction, PortfolioActionPlan, ServiceRequestComment, ServiceRequestDocument, ServiceRequestType, User, Lead, Client, Task, ServiceRequest, BusinessTracker, 
    Team, ProductDiscussion, ClientProfile,
    Note, NoteList
)
from .forms import (
    ActionPlanApprovalForm, ClientSearchForm, LeadForm, OperationsTaskAssignmentForm, PortfolioActionPlanForm, RedeemActionForm, SIPActionForm, STPActionForm, SWPActionForm, SwitchActionForm, TaskForm, ServiceRequestForm,
    ClientProfileForm, NoteForm, NoteListForm, QuickNoteForm
)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
from django.template.loader import render_to_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_protect

# Helper functions for role checks
def is_top_management(user):
    return user.role == 'top_management'

def is_business_head(user):
    return user.role == 'business_head'

def is_business_head_ops(user):
    return user.role == 'business_head_ops'

def is_rm_head(user):
    return user.role == 'rm_head'

def is_rm(user):
    return user.role == 'rm'

def is_ops_team_lead(user):
    return user.role == 'ops_team_lead'

def is_ops_exec(user):
    return user.role == 'ops_exec'

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.cache import never_cache
from django.core.cache import cache
from functools import lru_cache

# Role constants for faster comparisons
MANAGEMENT_ROLES = {'top_management', 'business_head'}
HIGH_PRIVILEGE_ROLES = MANAGEMENT_ROLES | {'rm_head'}
OPS_ROLES = {'ops_team_lead', 'ops_exec'}

@never_cache
@ensure_csrf_cookie
def user_login(request):
    """Ultra-fast login view optimized for minimal processing"""
    # Fast bypass for already authenticated users (no DB hit)
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        # Get raw POST data directly (avoid MultiValueDict overhead)
        post_data = request.POST
        username = post_data.get('username', '')[:150]  # Truncate to max username length
        password = post_data.get('password', '')[:128]  # Reasonable password length limit
        
        # Fast empty check (no strip() to save cycles)
        if not username or not password:
            return render(request, 'base/login.html', {'error': 'Please enter both username and password'})
        
        try:
            # Authenticate with minimal DB queries
            user = authenticate(request, username=username, password=password)
            
            if user is not None and user.is_active:
                # Pre-cache user role data before login
                _cache_user_data(user)
                
                # Login with minimal session processing
                login(request, user)
                
                # Fast redirect without additional processing
                return redirect('dashboard')
            
            # Generic error message to prevent username enumeration
            return render(request, 'base/login.html', {'error': 'Invalid credentials'})
            
        except Exception as e:
            logger.error(f"Login error for {username}: {str(e)}", exc_info=True)
            return render(request, 'base/login.html', {'error': 'Login service unavailable'})
    
    # GET request - minimal template rendering
    return render(request, 'base/login.html')

def _cache_user_data(user):
    """Cache all essential user data in one atomic operation"""
    from django.core.cache import caches
    role = getattr(user, 'role', '')
    
    cache_data = {
        'role': role,
        'is_management': role in MANAGEMENT_ROLES,
        'is_high_privilege': role in HIGH_PRIVILEGE_ROLES,
        'is_ops': role in OPS_ROLES,
        'permissions_hash': hash(getattr(user, 'permissions_hash', ''))
    }
    
    # Use direct cache access for maximum speed
    cache_backend = caches['default']
    cache_backend.set_many({
        f"user_{user.id}_data": cache_data,
        f"user_{user.id}_role": role,
    }, timeout=3600)  # Cache for 1 hour

def _cache_user_role_data(user):
    """Cache user role data to speed up subsequent requests"""
    cache_key = f"user_role_{user.id}"
    role_data = {
        'role': user.role,
        'is_management': user.role in MANAGEMENT_ROLES,
        'is_rm_head': user.role == HIGH_PRIVILEGE_ROLES,
    }
    # Cache for 30 minutes
    cache.set(cache_key, role_data, 1800)

@login_required
def user_logout(request):
    """Optimized logout with bulk cache clearing"""
    user_id = request.user.id
    cache.delete_many([
        f"user_{user_id}_data",
        f"user_{user_id}_role",
        f"accessible_users_{user_id}",
        f"accessible_data_{user_id}_*"  # Wildcard delete
    ])
    
    # Clear LRU caches
    _get_cached_user_role.cache_clear()
    _is_team_member.cache_clear()
    
    logout(request)
    return redirect('login')

@lru_cache(maxsize=200)
def _get_cached_user_role(user_id, role):
    """LRU cache for user role checks"""
    return {
        'is_management': role in MANAGEMENT_ROLES,
        'is_rm_head': role == HIGH_PRIVILEGE_ROLES,
        'role': role
    }

def can_manage_user(manager, target_user):
    """Optimized permission check with caching"""
    # Self-management is always allowed
    if manager.id == target_user.id:
        return True
    
    # Use cached role data
    manager_role_data = _get_cached_user_role(manager.id, manager.role)
    
    # Fast role-based checks
    if manager_role_data['is_management']:
        return True
    elif manager_role_data['is_rm_head']:
        # Cache team members check
        return _is_team_member(manager, target_user)
    
    return False

@lru_cache(maxsize=100)
def _is_team_member(manager, target_user):
    """Cached team membership check"""
    try:
        team_members = manager.get_team_members()
        return target_user in team_members
    except AttributeError:
        return False

def get_user_accessible_data(user, model_class, user_field='assigned_to'):
    """Optimized data access with intelligent caching"""
    # Create cache key based on user, model, and field
    cache_key = f"accessible_data_{user.id}_{model_class._meta.label}_{user_field}"
    
    # Try to get from cache first
    cached_ids = cache.get(cache_key)
    if cached_ids is not None:
        return model_class.objects.filter(id__in=cached_ids)
    
    # Get user role data from cache
    role_data = _get_cached_user_role(user.id, user.role)
    
    # Efficient query based on role
    if role_data['is_management']:
        # Management sees everything
        queryset = model_class.objects.all()
    elif role_data['is_rm_head']:
        # RM Head sees team data
        accessible_users = _get_cached_accessible_users(user)
        queryset = model_class.objects.filter(**{f"{user_field}__in": accessible_users})
    else:
        # Regular users see only their data
        queryset = model_class.objects.filter(**{user_field: user})
    
    # Cache the result IDs (limit to prevent memory issues)
    result_ids = list(queryset.values_list('id', flat=True)[:500])
    cache.set(cache_key, result_ids, 600)  # 10 minutes cache
    
    return queryset

def _get_cached_accessible_users(user):
    """Cache accessible users for RM heads"""
    cache_key = f"accessible_users_{user.id}"
    cached_users = cache.get(cache_key)
    
    if cached_users is None:
        try:
            # Get accessible users and cache their IDs
            accessible_users = user.get_accessible_users()
            cached_users = list(accessible_users.values_list('id', flat=True))
            cache.set(cache_key, cached_users, 900)  # 15 minutes cache
        except AttributeError:
            cached_users = [user.id]
            cache.set(cache_key, cached_users, 900)
    
    return cached_users

# Utility function to clear user caches when roles change
def clear_user_cache(user_id):
    """Clear all cached data for a user (call when user roles change)"""
    cache.delete(f"user_role_{user_id}")
    cache.delete(f"accessible_users_{user_id}")
    # Clear LRU cache
    _get_cached_user_role.cache_clear()
    _is_team_member.cache_clear()

# Performance monitoring decorator (optional)
def monitor_performance(func):
    """Decorator to monitor login performance"""
    import time
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        # Log slow operations (>100ms)
        if end_time - start_time > 0.1:
            print(f"Slow operation: {func.__name__} took {end_time - start_time:.3f}s")
        return result
    return wrapper

# Apply performance monitoring to login (optional)
# user_login = monitor_performance(user_login)


# Quick fix - Add this at the top of your views.py file

def patch_execution_plan():
    def approve(self, approved_by):
        if self.status != 'pending_approval':
            return False
        if self.created_by == approved_by:
            return False
        if not hasattr(approved_by, 'role'):
            return False
        if approved_by.role not in ['business_head', 'top_management', 'business_head_ops']:
            return False
        
        self.status = 'approved'
        self.approved_by = approved_by
        self.approved_at = timezone.now()
        self.save()
        return True
    
    def reject(self, rejected_by, reason):
        if self.status != 'pending_approval':
            return False
        if self.created_by == rejected_by:
            return False
        if not hasattr(rejected_by, 'role'):
            return False
        if rejected_by.role not in ['business_head', 'top_management', 'business_head_ops']:
            return False
        
        self.status = 'rejected'
        self.rejected_by = rejected_by
        self.rejected_at = timezone.now()
        if hasattr(self, 'rejection_reason'):
            self.rejection_reason = reason
        self.save()
        return True
    
    ExecutionPlan.approve = approve
    ExecutionPlan.reject = reject

# Call this function
patch_execution_plan()

@login_required
def dashboard(request):
    user = request.user
    context = {}

    # Add debug logging
    import logging
    logger = logging.getLogger(__name__)
    
    print(f"DEBUG: User role is {user.role}")

    # Check if execution plans models are available
    try:
        from .models import ExecutionPlan, PlanAction, ExecutionMetrics
        EXECUTION_PLANS_AVAILABLE = True
    except ImportError:
        EXECUTION_PLANS_AVAILABLE = False

    try:
        if user.role == 'top_management':
            print("DEBUG: Processing top_management")
            # Aggregate KPIs across entire system
            total_aum = Client.objects.aggregate(total=Sum('aum'))['total'] or 0
            total_sip = Client.objects.aggregate(total=Sum('sip_amount'))['total'] or 0
            total_clients = Client.objects.count()
            total_leads = Lead.objects.count()
            
            print("DEBUG: About to query tasks")
            total_tasks = Task.objects.filter(completed=False).count()
            print("DEBUG: Tasks query completed")
            
            open_service_requests = ServiceRequest.objects.filter(status='open').count()

            # Team metrics
            business_heads_count = User.objects.filter(role='business_head').count()
            business_heads_ops_count = User.objects.filter(role='business_head_ops').count()
            rm_heads_count = User.objects.filter(role='rm_head').count()
            rms_count = User.objects.filter(role='rm').count()
            ops_team_leads_count = User.objects.filter(role='ops_team_lead').count()
            ops_execs_count = User.objects.filter(role='ops_exec').count()

            # Recent activities
            recent_leads = Lead.objects.order_by('-created_at')[:5]
            recent_service_requests = ServiceRequest.objects.order_by('-created_at')[:5]
            recent_notes = Note.objects.filter(user=user).order_by('-updated_at')[:3]

            # Execution Plans metrics (if available)
            execution_plans_stats = {}
            if EXECUTION_PLANS_AVAILABLE:
                current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                monthly_plans = ExecutionPlan.objects.filter(created_at__gte=current_month_start)
                completed_monthly = monthly_plans.filter(status='completed')
                
                execution_plans_stats = {
                    'total_plans': ExecutionPlan.objects.count(),
                    'pending_approval': ExecutionPlan.objects.filter(status='pending_approval').count(),
                    'in_execution': ExecutionPlan.objects.filter(status='in_execution').count(),
                    'completed_plans': ExecutionPlan.objects.filter(status='completed').count(),
                    'recent_plans': ExecutionPlan.objects.order_by('-created_at')[:5],
                    'monthly_completion_rate': round((completed_monthly.count() / monthly_plans.count()) * 100, 2) if monthly_plans.count() > 0 else 0,
                    'plans_this_month': monthly_plans.count(),
                }

            context.update({
                'total_aum': total_aum,
                'total_sip': total_sip,
                'total_clients': total_clients,
                'total_leads': total_leads,
                'total_tasks': total_tasks,
                'open_service_requests': open_service_requests,
                'business_heads_count': business_heads_count,
                'business_heads_ops_count': business_heads_ops_count,
                'rm_heads_count': rm_heads_count,
                'rms_count': rms_count,
                'ops_team_leads_count': ops_team_leads_count,
                'ops_execs_count': ops_execs_count,
                'recent_leads': recent_leads,
                'recent_service_requests': recent_service_requests,
                'recent_notes': recent_notes,
                'execution_plans_stats': execution_plans_stats,
            })
            template_name = 'base/dashboard_top_management.html'

        elif user.role == 'business_head':
            print("DEBUG: Processing business_head")
            # Monitor all RM Heads and their performance
            rm_heads = User.objects.filter(role='rm_head')
            all_rms = User.objects.filter(role='rm')
            
            # System-wide metrics
            total_leads = Lead.objects.count()
            converted_leads = Lead.objects.filter(status='converted').count()
            open_service_requests = ServiceRequest.objects.filter(status='open')
            
            # Performance metrics
            lead_conversion_rate = (converted_leads / total_leads * 100) if total_leads > 0 else 0
            
            # Calculate average response time using Python datetime operations
            resolved_requests = ServiceRequest.objects.filter(
                status__in=['resolved', 'closed'],
                resolved_at__isnull=False
            ).values('created_at', 'resolved_at')
            
            if resolved_requests.exists():
                total_response_seconds = 0
                request_count = 0
                
                for request in resolved_requests:
                    response_time = request['resolved_at'] - request['created_at']
                    total_response_seconds += response_time.total_seconds()
                    request_count += 1
                
                avg_response_time_days = (total_response_seconds / request_count) / (24 * 60 * 60) if request_count > 0 else 0
            else:
                avg_response_time_days = 0

            recent_notes = Note.objects.filter(user=user).order_by('-updated_at')[:3]

            # Execution Plans for business head
            execution_plans_stats = {}
            if EXECUTION_PLANS_AVAILABLE:
                current_month = timezone.now().month
                current_year = timezone.now().year
                
                execution_plans_stats = {
                    'pending_approval': ExecutionPlan.objects.filter(status='pending_approval').count(),
                    'approved_plans': ExecutionPlan.objects.filter(status='approved').count(),
                    'recent_approvals': ExecutionPlan.objects.filter(
                        approved_by=user
                    ).order_by('-approved_at')[:5],
                    'plans_this_month': ExecutionPlan.objects.filter(
                        created_at__month=current_month,
                        created_at__year=current_year
                    ).count(),
                    'approval_pending_count': ExecutionPlan.objects.filter(
                        status='pending_approval'
                    ).count(),
                    'completed_this_month': ExecutionPlan.objects.filter(
                        status='completed',
                        completed_at__month=current_month,
                        completed_at__year=current_year
                    ).count(),
                }

            context.update({
                'rm_heads': rm_heads,
                'all_rms': all_rms,
                'total_leads': total_leads,
                'converted_leads': converted_leads,
                'lead_conversion_rate': round(lead_conversion_rate, 2),
                'open_service_requests': open_service_requests,
                'avg_response_time': round(avg_response_time_days, 2),
                'recent_notes': recent_notes,
                'execution_plans_stats': execution_plans_stats,
            })
            template_name = 'base/dashboard_business_head.html'

        elif user.role == 'business_head_ops':
            print("DEBUG: Processing business_head_ops")
            # Operations oversight dashboard
            ops_team_leads = User.objects.filter(role='ops_team_lead')
            ops_execs = User.objects.filter(role='ops_exec')
            
            # Operations metrics
            total_client_profiles = ClientProfile.objects.count()
            active_profiles = ClientProfile.objects.filter(status='active').count()
            muted_profiles = ClientProfile.objects.filter(status='muted').count()
            
            # Service requests related to operations
            print("DEBUG: About to query ops_service_requests")
            ops_service_requests = ServiceRequest.objects.filter(
                Q(assigned_to__role__in=['ops_team_lead', 'ops_exec']) |
                Q(raised_by__role__in=['ops_team_lead', 'ops_exec'])
            )
            print("DEBUG: ops_service_requests query completed")
            
            # Task metrics for operations team
            print("DEBUG: About to query ops_tasks")
            ops_tasks = Task.objects.filter(assigned_to__role__in=['ops_team_lead', 'ops_exec'])
            print("DEBUG: ops_tasks query completed")
            
            pending_ops_tasks = ops_tasks.filter(completed=False).count()
            overdue_ops_tasks = ops_tasks.filter(
                completed=False, 
                due_date__lt=timezone.now()
            ).count()

            # Team performance data
            print("DEBUG: About to process team performance")
            team_performance = []
            for lead in ops_team_leads:
                print(f"DEBUG: Processing lead {lead.id}")
                team_members = lead.get_team_members()
                print(f"DEBUG: Got team members: {team_members}")
                
                # Convert to list of IDs to avoid subquery issues
                team_member_ids = list(team_members.values_list('id', flat=True))
                print(f"DEBUG: Team member IDs: {team_member_ids}")
                
                print("DEBUG: About to query team_tasks")
                team_tasks = Task.objects.filter(assigned_to_id__in=team_member_ids)
                print("DEBUG: team_tasks query completed")
                
                print("DEBUG: About to query team_service_requests")
                team_service_requests = ServiceRequest.objects.filter(
                    Q(assigned_to_id__in=team_member_ids) | Q(raised_by_id__in=team_member_ids)
                )
                print("DEBUG: team_service_requests query completed")
                
                team_performance.append({
                    'lead': lead,
                    'team_size': len(team_member_ids),
                    'pending_tasks': team_tasks.filter(completed=False).count(),
                    'total_service_requests': team_service_requests.count(),
                    'open_service_requests': team_service_requests.filter(status='open').count(),
                })

            recent_notes = Note.objects.filter(user=user).order_by('-updated_at')[:3]

            # Execution Plans for operations head
            execution_plans_stats = {}
            if EXECUTION_PLANS_AVAILABLE:
                today = timezone.now().date()
                
                execution_plans_stats = {
                    'ready_for_execution': ExecutionPlan.objects.filter(status='client_approved').count(),
                    'in_execution': ExecutionPlan.objects.filter(status='in_execution').count(),
                    'completed_today': ExecutionPlan.objects.filter(
                        status='completed',
                        completed_at__date=today
                    ).count(),
                    'pending_actions': PlanAction.objects.filter(
                        execution_plan__status='in_execution',
                        status='pending'
                    ).count(),
                    'total_value_in_execution': PlanAction.objects.filter(
                        execution_plan__status='in_execution',
                        status='pending'
                    ).aggregate(total=Sum('amount'))['total'] or 0,
                }

            context.update({
                'ops_team_leads': ops_team_leads,
                'ops_execs': ops_execs,
                'total_client_profiles': total_client_profiles,
                'active_profiles': active_profiles,
                'muted_profiles': muted_profiles,
                'ops_service_requests': ops_service_requests,
                'pending_ops_tasks': pending_ops_tasks,
                'overdue_ops_tasks': overdue_ops_tasks,
                'team_performance': team_performance,
                'recent_notes': recent_notes,
                'execution_plans_stats': execution_plans_stats,
            })
            template_name = 'base/dashboard_business_head_ops.html'

        elif user.role == 'rm_head':
            print("DEBUG: Processing rm_head")
            print("DEBUG: About to get team members")
            team_members = user.get_team_members()
            print(f"DEBUG: Team members: {team_members}")
            
            print("DEBUG: About to get accessible users")
            accessible_users = user.get_accessible_users()
            print(f"DEBUG: Accessible users: {accessible_users}")
            
            # Convert querysets to lists of IDs to avoid subquery issues
            accessible_user_ids = list(accessible_users.values_list('id', flat=True))
            print(f"DEBUG: Accessible user IDs: {accessible_user_ids}")
            
            # Team metrics
            print("DEBUG: About to query team_leads")
            team_leads = Lead.objects.filter(assigned_to_id__in=accessible_user_ids)
            print("DEBUG: team_leads query completed")
            
            print("DEBUG: About to query team_clients")
            team_clients = Client.objects.filter(user_id__in=accessible_user_ids)
            print("DEBUG: team_clients query completed")
            
            print("DEBUG: About to query team_tasks")
            team_tasks = Task.objects.filter(assigned_to_id__in=accessible_user_ids)
            print("DEBUG: team_tasks query completed")
            
            print("DEBUG: About to query team_service_requests")
            team_service_requests = ServiceRequest.objects.filter(
                Q(raised_by_id__in=accessible_user_ids) | Q(client__user_id__in=accessible_user_ids)
            )
            print("DEBUG: team_service_requests query completed")
            
            # Performance metrics
            pending_tasks = team_tasks.filter(completed=False).count()
            overdue_tasks = team_tasks.filter(
                completed=False, 
                due_date__lt=timezone.now()
            ).count()
            
            team_aum = team_clients.aggregate(total=Sum('aum'))['total'] or 0
            team_sip = team_clients.aggregate(total=Sum('sip_amount'))['total'] or 0
            
            # Prepare detailed member statistics
            team_members_data = []
            for member in team_members:
                member_leads = team_leads.filter(assigned_to=member)
                member_clients = team_clients.filter(user=member)
                member_tasks = team_tasks.filter(assigned_to=member)
                
                # Add execution plans data for each member
                member_execution_stats = {}
                if EXECUTION_PLANS_AVAILABLE:
                    member_plans = ExecutionPlan.objects.filter(created_by=member)
                    member_execution_stats = {
                        'total_plans': member_plans.count(),
                        'pending_approval': member_plans.filter(status='pending_approval').count(),
                        'approved_plans': member_plans.filter(status='approved').count(),
                        'completed_plans': member_plans.filter(status='completed').count(),
                    }
                
                team_members_data.append({
                    'member': member,
                    'lead_count': member_leads.count(),
                    'client_count': member_clients.count(),
                    'aum': member_clients.aggregate(total=Sum('aum'))['total'] or 0,
                    'sip': member_clients.aggregate(total=Sum('sip_amount'))['total'] or 0,
                    'pending_tasks': member_tasks.filter(completed=False).count(),
                    'overdue_tasks': member_tasks.filter(completed=False, due_date__lt=timezone.now()).count(),
                    'performance_score': getattr(member, 'performance_score', 0),
                    'execution_plans': member_execution_stats,
                })

            recent_notes = Note.objects.filter(user=user).order_by('-updated_at')[:3]

            # Execution Plans for RM Head
            execution_plans_stats = {}
            

            context.update({
                'team_members': team_members,
                'team_leads': team_leads,
                'team_clients': team_clients,
                'team_tasks': team_tasks,
                'team_service_requests': team_service_requests,
                'pending_tasks': pending_tasks,
                'overdue_tasks': overdue_tasks,
                'team_aum': team_aum,
                'team_sip': team_sip,
                'leads_count': team_leads.count(),
                'clients_count': team_clients.count(),
                'team_members_data': team_members_data,
                'recent_notes': recent_notes,
                'execution_plans_stats': execution_plans_stats,
            })
            template_name = 'base/dashboard_rm_head.html'

        elif user.role == 'ops_team_lead':
            print("DEBUG: Processing ops_team_lead")
            # Operations Team Lead dashboard
            team_members = user.get_team_members()  # Ops Executives under this lead
            team_member_ids = list(team_members.values_list('id', flat=True))
            team_with_lead_ids = team_member_ids + [user.id]
            
            # Team metrics
            team_tasks = Task.objects.filter(assigned_to_id__in=team_with_lead_ids)
            team_service_requests = ServiceRequest.objects.filter(
                Q(assigned_to_id__in=team_with_lead_ids) |
                Q(raised_by_id__in=team_with_lead_ids)
            )
            
            # Client profiles managed by team
            team_client_profiles = ClientProfile.objects.filter(
                Q(mapped_ops_exec_id__in=team_member_ids) | Q(created_by_id__in=team_with_lead_ids)
            )
            
            # Performance metrics
            pending_tasks = team_tasks.filter(completed=False).count()
            overdue_tasks = team_tasks.filter(
                completed=False, 
                due_date__lt=timezone.now()
            ).count()
            
            # Service request metrics
            open_requests = team_service_requests.filter(status='open').count()
            in_progress_requests = team_service_requests.filter(status='in_progress').count()
            
            pending_conversion_approvals = LeadStatusChange.objects.filter(
                needs_approval=True,
                new_status='conversion_requested',
                approval_by=user,
                approved__isnull=True
                ).count()
            recent_conversion_decisions = LeadStatusChange.objects.filter(
                approval_by=user,
                new_status='conversion_requested',
                approved__isnull=False
                ).select_related('lead', 'changed_by').order_by('-approved_at')[:5]
    
            # Team member performance
            team_members_data = []
            for member in team_members:
                member_tasks = team_tasks.filter(assigned_to=member)
                member_service_requests = team_service_requests.filter(
                    Q(assigned_to=member) | Q(raised_by=member)
                )
                member_client_profiles = team_client_profiles.filter(mapped_ops_exec=member)
                
                # Add execution plans data for ops members
                member_execution_stats = {}
                if EXECUTION_PLANS_AVAILABLE:
                    today = timezone.now().date()
                    member_actions = PlanAction.objects.filter(
                        execution_plan__status__in=['client_approved', 'in_execution']
                    )
                    member_execution_stats = {
                        'pending_actions': member_actions.filter(status='pending').count(),
                        'completed_actions': PlanAction.objects.filter(
                            executed_by=member,
                            status='completed'
                        ).count(),
                        'actions_today': PlanAction.objects.filter(
                            executed_by=member,
                            executed_at__date=today,
                            status='completed'
                        ).count(),
                    }
                
                team_members_data.append({
                    'member': member,
                    'task_count': member_tasks.count(),
                    'pending_tasks': member_tasks.filter(completed=False).count(),
                    'service_requests': member_service_requests.count(),
                    'open_requests': member_service_requests.filter(status='open').count(),
                    'client_profiles': member_client_profiles.count(),
                    'execution_stats': member_execution_stats,
                })

            recent_notes = Note.objects.filter(user=user).order_by('-updated_at')[:3]

            # Execution Plans for Ops Team Lead
            execution_plans_stats = {}
            if EXECUTION_PLANS_AVAILABLE:
                today = timezone.now().date()
                
                execution_plans_stats = {
                    'ready_for_execution': ExecutionPlan.objects.filter(status='client_approved').count(),
                    'in_execution': ExecutionPlan.objects.filter(status='in_execution').count(),
                    'pending_actions': PlanAction.objects.filter(
                        execution_plan__status='in_execution',
                        status='pending'
                    ).count(),
                    'team_actions_today': PlanAction.objects.filter(
                        executed_by_id__in=team_member_ids,
                        executed_at__date=today,
                        status='completed'
                    ).count(),
                    'team_efficiency': 0,  # Can calculate based on completion rate
                }

            context.update({
                'team_members': team_members,
                'team_tasks': team_tasks,
                'team_service_requests': team_service_requests,
                'team_client_profiles': team_client_profiles,
                'pending_tasks': pending_tasks,
                'overdue_tasks': overdue_tasks,
                'open_requests': open_requests,
                'in_progress_requests': in_progress_requests,
                'team_members_data': team_members_data,
                'recent_notes': recent_notes,
                'execution_plans_stats': execution_plans_stats,
                'pending_conversion_approvals': pending_conversion_approvals,
                'recent_conversion_decisions': recent_conversion_decisions,
            })
            template_name = 'base/dashboard_ops_team_lead.html'

        elif user.role == 'ops_exec':
            print("DEBUG: Processing ops_exec")
            # Operations Executive dashboard
            my_tasks = Task.objects.filter(assigned_to=user)
            my_service_requests = ServiceRequest.objects.filter(
                Q(assigned_to=user) | Q(raised_by=user)
            )
            my_client_profiles = ClientProfile.objects.filter(mapped_ops_exec=user)
            
            # Personal metrics
            pending_tasks = my_tasks.filter(completed=False).count()
            overdue_tasks = my_tasks.filter(completed=False, due_date__lt=timezone.now()).count()
            open_requests = my_service_requests.filter(status='open').count()
            
            # Recent activities
            recent_tasks = my_tasks.order_by('-created_at')[:5]
            recent_service_requests = my_service_requests.order_by('-created_at')[:5]
            recent_client_profiles = my_client_profiles.order_by('-updated_at')[:5]
            recent_notes = Note.objects.filter(user=user).order_by('-updated_at')[:3]

            # Execution Plans for Ops Exec
            execution_plans_stats = {}
            if EXECUTION_PLANS_AVAILABLE:
                today = timezone.now().date()
                last_30_days = timezone.now() - timedelta(days=30)
                
                my_actions = PlanAction.objects.filter(
                    execution_plan__status__in=['client_approved', 'in_execution']
                )
                
                execution_plans_stats = {
                    'pending_actions': my_actions.filter(status='pending').count(),
                    'completed_today': PlanAction.objects.filter(
                        executed_by=user,
                        executed_at__date=today,
                        status='completed'
                    ).count(),
                    'total_completed': PlanAction.objects.filter(
                        executed_by=user,
                        status='completed'
                    ).count(),
                    'recent_actions': PlanAction.objects.filter(
                        executed_by=user
                    ).order_by('-executed_at')[:5],
                    'monthly_performance': PlanAction.objects.filter(
                        executed_by=user,
                        executed_at__gte=last_30_days,
                        status='completed'
                    ).count(),
                }

            context.update({
                'my_tasks': my_tasks,
                'my_service_requests': my_service_requests,
                'my_client_profiles': my_client_profiles,
                'pending_tasks': pending_tasks,
                'overdue_tasks': overdue_tasks,
                'open_requests': open_requests,
                'recent_tasks': recent_tasks,
                'recent_service_requests': recent_service_requests,
                'recent_client_profiles': recent_client_profiles,
                'recent_notes': recent_notes,
                'execution_plans_stats': execution_plans_stats,
            })
            template_name = 'base/dashboard_ops_exec.html'

        else:  # Relationship Manager
            print("DEBUG: Processing relationship manager")
            # Personal dashboard
            leads = Lead.objects.filter(assigned_to=user)
            clients = Client.objects.filter(user=user)
            
            print("DEBUG: About to query tasks for RM")
            tasks = Task.objects.filter(assigned_to=user)
            print("DEBUG: RM tasks query completed")
            
            reminders = user.reminders.filter(is_done=False, remind_at__gte=timezone.now()) if hasattr(user, 'reminders') else []
            service_requests = ServiceRequest.objects.filter(raised_by=user)
            
            # Personal metrics
            pending_tasks = tasks.filter(completed=False).count()
            overdue_tasks = tasks.filter(completed=False, due_date__lt=timezone.now()).count()
            my_aum = clients.aggregate(total=Sum('aum'))['total'] or 0
            my_sip = clients.aggregate(total=Sum('sip_amount'))['total'] or 0
            
            # Recent activities
            recent_clients = clients.order_by('-created_at')[:3]
            recent_notes = Note.objects.filter(user=user).order_by('-updated_at')[:3]

            # Execution Plans for RM
            execution_plans_stats = {}
            if EXECUTION_PLANS_AVAILABLE:
                current_month = timezone.now().month
                current_year = timezone.now().year
                last_90_days = timezone.now() - timedelta(days=90)
                
                my_plans = ExecutionPlan.objects.filter(created_by=user)
                
                # Calculate approval success rate
                submitted_plans = my_plans.filter(created_at__gte=last_90_days)
                approved_plans = submitted_plans.filter(status__in=['approved', 'client_approved', 'in_execution', 'completed'])
                approval_success_rate = round((approved_plans.count() / submitted_plans.count()) * 100, 2) if submitted_plans.count() > 0 else 0
                
                # FIX: Calculate total plan value using PlanAction model directly
                # Instead of using Sum('actions__amount') which causes the subquery error
                plan_ids = my_plans.filter(
                    status__in=['approved', 'client_approved', 'in_execution', 'completed']
                ).values_list('id', flat=True)
                
                total_plan_value = PlanAction.objects.filter(
                    execution_plan_id__in=plan_ids
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                execution_plans_stats = {
                    'my_plans': my_plans.count(),
                    'draft_plans': my_plans.filter(status='draft').count(),
                    'pending_approval': my_plans.filter(status='pending_approval').count(),
                    'approved_plans': my_plans.filter(status='approved').count(),
                    'in_execution': my_plans.filter(status='in_execution').count(),
                    'completed_plans': my_plans.filter(status='completed').count(),
                    'recent_plans': my_plans.order_by('-created_at')[:5],
                    'plans_this_month': my_plans.filter(
                        created_at__month=current_month,
                        created_at__year=current_year
                    ).count(),
                    'approval_success_rate': approval_success_rate,
                    'total_plan_value': total_plan_value,
                }

            context.update({
                'leads': leads,
                'clients': clients,
                'tasks': tasks,
                'reminders': reminders,
                'service_requests': service_requests,
                'pending_tasks': pending_tasks,
                'overdue_tasks': overdue_tasks,
                'my_aum': my_aum,
                'my_sip': my_sip,
                'recent_clients': recent_clients,
                'recent_notes': recent_notes,
                'leads_count': leads.count(),
                'clients_count': clients.count(),
                'execution_plans_stats': execution_plans_stats,
            })
            template_name = 'base/dashboard_rm.html'

        # Add global execution plans availability flag
        context['execution_plans_available'] = EXECUTION_PLANS_AVAILABLE
        
        print(f"DEBUG: About to render template {template_name}")
        return render(request, template_name, context)
        
    except Exception as e:
        print(f"DEBUG: Error occurred: {e}")
        import traceback
        traceback.print_exc()
        raise

# Notes System Views
@login_required
def notes_dashboard(request):
    """Main notes dashboard with overview"""
    user = request.user
    
    # Get user's note lists and notes
    note_lists = NoteList.objects.filter(user=user)
    notes = Note.objects.filter(user=user)
    
    # Notes statistics
    total_notes = notes.count()
    completed_notes = notes.filter(is_completed=True).count()
    pending_notes = notes.filter(is_completed=False).count()
    overdue_notes = notes.filter(is_completed=False, due_date__lt=timezone.now().date()).count()
    
    # Recent notes
    recent_notes = notes.order_by('-updated_at')[:10]
    
    # Upcoming reminders
    upcoming_reminders = notes.filter(
        reminder_date__gte=timezone.now(),
        is_completed=False
    ).order_by('reminder_date')[:5]
    
    # Due soon (next 7 days)
    from datetime import timedelta
    due_soon = notes.filter(
        due_date__lte=timezone.now().date() + timedelta(days=7),
        due_date__gte=timezone.now().date(),
        is_completed=False
    ).order_by('due_date')[:5]
    
    context = {
        'note_lists': note_lists,
        'total_notes': total_notes,
        'completed_notes': completed_notes,
        'pending_notes': pending_notes,
        'overdue_notes': overdue_notes,
        'recent_notes': recent_notes,
        'upcoming_reminders': upcoming_reminders,
        'due_soon': due_soon,
    }
    
    return render(request, 'notes/dashboard.html', context)

@login_required
def note_list_view(request):
    """View all notes with filtering options"""
    user = request.user
    notes = Note.objects.filter(user=user)
    
    # Filtering
    status_filter = request.GET.get('status')
    if status_filter == 'completed':
        notes = notes.filter(is_completed=True)
    elif status_filter == 'pending':
        notes = notes.filter(is_completed=False)
    elif status_filter == 'overdue':
        notes = notes.filter(is_completed=False, due_date__lt=timezone.now().date())
    
    list_filter = request.GET.get('list')
    if list_filter:
        notes = notes.filter(note_list_id=list_filter)
    
    # Search
    search_query = request.GET.get('search')
    if search_query:
        notes = notes.filter(
            Q(heading__icontains=search_query) |
            Q(content__icontains=search_query)
        )
    
    # Sorting
    sort_by = request.GET.get('sort', '-updated_at')
    notes = notes.order_by(sort_by)
    
    # Get note lists for filter dropdown
    note_lists = NoteList.objects.filter(user=user)
    
    context = {
        'notes': notes,
        'note_lists': note_lists,
        'status_filter': status_filter,
        'list_filter': list_filter,
        'search_query': search_query,
        'sort_by': sort_by,
    }
    
    return render(request, 'notes/note_list.html', context)

@login_required
def note_create(request):
    """Create a new note"""
    if request.method == 'POST':
        form = NoteForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            note = form.save(commit=False)
            note.user = request.user
            try:
                note.save()
                messages.success(request, "Note created successfully.")
                return redirect('note_detail', pk=note.pk)
            except ValidationError as e:
                messages.error(request, str(e))
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = NoteForm(user=request.user)
    
    return render(request, 'notes/note_form.html', {
        'form': form, 
        'action': 'Create',
        'title': 'Create New Note'
    })

@login_required
def note_detail(request, pk):
    """View note details"""
    note = get_object_or_404(Note, pk=pk, user=request.user)
    
    context = {
        'note': note,
        'can_edit': True,  # User can always edit their own notes
        'can_delete': True,  # User can always delete their own notes
    }
    
    return render(request, 'notes/note_detail.html', context)

@login_required
def note_update(request, pk):
    """Update a note"""
    note = get_object_or_404(Note, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = NoteForm(request.POST, request.FILES, instance=note, user=request.user)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Note updated successfully.")
                return redirect('note_detail', pk=note.pk)
            except ValidationError as e:
                messages.error(request, str(e))
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = NoteForm(instance=note, user=request.user)
    
    return render(request, 'notes/note_form.html', {
        'form': form, 
        'action': 'Update',
        'title': 'Update Note',
        'note': note
    })

@login_required
def note_delete(request, pk):
    """Delete a note"""
    note = get_object_or_404(Note, pk=pk, user=request.user)
    
    if request.method == 'POST':
        note_heading = note.heading
        
        # Delete associated file if exists
        if note.attachment:
            try:
                default_storage.delete(note.attachment.name)
            except:
                pass  # File might not exist
        
        note.delete()
        messages.success(request, f"Note '{note_heading}' has been deleted.")
        return redirect('note_list')
    
    return render(request, 'notes/note_confirm_delete.html', {'note': note})

@login_required
@require_POST
def note_toggle_complete(request, pk):
    """Toggle note completion status via AJAX"""
    note = get_object_or_404(Note, pk=pk, user=request.user)
    
    note.is_completed = not note.is_completed
    if note.is_completed:
        note.completed_at = timezone.now()
    else:
        note.completed_at = None
    note.save()
    
    return JsonResponse({
        'success': True,
        'is_completed': note.is_completed,
        'completed_at': note.completed_at.isoformat() if note.completed_at else None
    })
    
import logging
logger = logging.getLogger(__name__)


@login_required
@require_POST
def note_toggle_complete(request, pk):
    """Toggle note completion status via AJAX with enhanced error handling"""
    try:
        note = get_object_or_404(Note, pk=pk, user=request.user)
        
        # Handle both AJAX and form submissions
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                force_status = data.get('is_completed')
                if force_status is not None:
                    note.is_completed = bool(force_status)
                else:
                    note.is_completed = not note.is_completed
            except (json.JSONDecodeError, KeyError):
                note.is_completed = not note.is_completed
        else:
            note.is_completed = not note.is_completed
        
        if note.is_completed:
            note.completed_at = timezone.now()
        else:
            note.completed_at = None
        
        note.save(update_fields=['is_completed', 'completed_at', 'updated_at'])
        
        response_data = {
            'success': True,
            'is_completed': note.is_completed,
            'completed_at': note.completed_at.isoformat() if note.completed_at else None,
            'message': 'Note marked as completed!' if note.is_completed else 'Note marked as incomplete!'
        }
        
        return JsonResponse(response_data)
    
    except Note.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Note not found or access denied',
            'message': 'Note not found or you do not have permission to modify it.'
        }, status=404)
    
    except Exception as e:
        logger.error(f"Error toggling completion for note {pk}, user {request.user.id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Server error',
            'message': 'An error occurred while updating the note status.'
        }, status=500)
        
@login_required
def note_list_management(request):
    """Manage note lists"""
    user = request.user
    note_lists = NoteList.objects.filter(user=user)
    
    context = {
        'note_lists': note_lists,
    }
    
    return render(request, 'notes/note_list_management.html', context)

@login_required
def note_list_create(request):
    """Create a new note list"""
    if request.method == 'POST':
        form = NoteListForm(request.POST)
        if form.is_valid():
            note_list = form.save(commit=False)
            note_list.user = request.user
            note_list.save()
            messages.success(request, f"Note list '{note_list.name}' created successfully.")
            return redirect('note_list_management')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = NoteListForm()
    
    return render(request, 'notes/note_list_form.html', {
        'form': form, 
        'action': 'Create',
        'title': 'Create Note List'
    })

@login_required
def note_list_update(request, pk):
    """Update a note list"""
    note_list = get_object_or_404(NoteList, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = NoteListForm(request.POST, instance=note_list)
        if form.is_valid():
            form.save()
            messages.success(request, f"Note list '{note_list.name}' updated successfully.")
            return redirect('note_list_management')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = NoteListForm(instance=note_list)
    
    return render(request, 'notes/note_list_form.html', {
        'form': form, 
        'action': 'Update',
        'title': 'Update Note List',
        'note_list': note_list
    })

@login_required
def note_list_delete(request, pk):
    """Delete a note list"""
    note_list = get_object_or_404(NoteList, pk=pk, user=request.user)
    
    # Check if list has notes
    notes_count = note_list.notes.count()
    if notes_count > 0 and request.method == 'GET':
        messages.warning(request, f"This list contains {notes_count} notes. Deleting the list will also delete all notes in it.")
    
    if request.method == 'POST':
        list_name = note_list.name
        note_list.delete()  # This will cascade delete all notes in the list
        messages.success(request, f"Note list '{list_name}' and all its notes have been deleted.")
        return redirect('note_list_management')
    
    return render(request, 'notes/note_list_confirm_delete.html', {
        'note_list': note_list,
        'notes_count': notes_count
    })

# API endpoints for notes
@login_required
def get_notes_by_list(request, list_id):
    """AJAX endpoint to get notes by list"""
    note_list = get_object_or_404(NoteList, pk=list_id, user=request.user)
    notes = note_list.notes.all().values('id', 'heading', 'is_completed', 'due_date')
    
    return JsonResponse(list(notes), safe=False)

@login_required
def get_upcoming_reminders(request):
    """AJAX endpoint for upcoming reminders"""
    notes = Note.objects.filter(
        user=request.user,
        reminder_date__gte=timezone.now(),
        reminder_date__lte=timezone.now() + timezone.timedelta(days=1),
        is_completed=False
    ).values('id', 'heading', 'reminder_date')
    
    return JsonResponse(list(notes), safe=False)



# Helper decorators for role-based access
def operations_team_required(user):
    """Check if user is part of operations team"""
    return user.role in ['business_head_ops', 'ops_team_lead', 'ops_exec']

def ops_team_lead_required(user):
    """Check if user is operations team lead"""
    return user.role == 'ops_team_lead'

def business_head_ops_required(user):
    """Check if user is business head operations"""
    return user.role == 'business_head_ops'

def ops_management_required(user):
    """Check if user can manage operations"""
    return user.role in ['business_head_ops', 'ops_team_lead']

# Operations Team Lead Views
@login_required
@user_passes_test(ops_team_lead_required)
def ops_team_performance(request):
    """Operations Team Lead view for team performance metrics"""
    user = request.user
    team_members = user.get_team_members()
    
    # Get date range for filtering
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        date_from = timezone.now().date() - timedelta(days=30)
    else:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    
    if not date_to:
        date_to = timezone.now().date()
    else:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Team performance metrics
    team_stats = []
    for member in team_members:
        # Tasks metrics
        member_tasks = Task.objects.filter(
            assigned_to=member,
            created_at__date__range=[date_from, date_to]
        )
        
        # Service requests metrics
        member_requests = ServiceRequest.objects.filter(
            assigned_to=member,
            created_at__date__range=[date_from, date_to]
        )
        
        # Client profiles metrics
        member_profiles = ClientProfile.objects.filter(
            mapped_ops_exec=member,
            created_at__date__range=[date_from, date_to]
        )
        
        # Calculate completion rates
        total_tasks = member_tasks.count()
        completed_tasks = member_tasks.filter(completed=True).count()
        task_completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        
        total_requests = member_requests.count()
        resolved_requests = member_requests.filter(status__in=['resolved', 'closed']).count()
        request_resolution_rate = (resolved_requests / total_requests * 100) if total_requests > 0 else 0
        
        # Average response time for service requests
        resolved_with_time = member_requests.filter(
            status__in=['resolved', 'closed'],
            resolved_at__isnull=False
        )
        
        if resolved_with_time.exists():
            avg_response_time = resolved_with_time.aggregate(
                avg_time=Avg(
                    Extract('epoch', F('resolved_at')) - Extract('epoch', F('created_at'))
                )
            )['avg_time']
            avg_response_hours = avg_response_time / 3600 if avg_response_time else 0
        else:
            avg_response_hours = 0
        
        team_stats.append({
            'member': member,
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'pending_tasks': total_tasks - completed_tasks,
            'task_completion_rate': round(task_completion_rate, 2),
            'total_requests': total_requests,
            'resolved_requests': resolved_requests,
            'request_resolution_rate': round(request_resolution_rate, 2),
            'avg_response_hours': round(avg_response_hours, 2),
            'client_profiles': member_profiles.count(),
            'performance_score': round((task_completion_rate + request_resolution_rate) / 2, 2)
        })
    
    # Overall team metrics
    team_totals = {
        'total_tasks': sum(stat['total_tasks'] for stat in team_stats),
        'completed_tasks': sum(stat['completed_tasks'] for stat in team_stats),
        'total_requests': sum(stat['total_requests'] for stat in team_stats),
        'resolved_requests': sum(stat['resolved_requests'] for stat in team_stats),
        'total_profiles': sum(stat['client_profiles'] for stat in team_stats),
    }
    
    context = {
        'team_members': team_members,
        'team_stats': team_stats,
        'team_totals': team_totals,
        'date_from': date_from,
        'date_to': date_to,
        'avg_team_performance': round(sum(stat['performance_score'] for stat in team_stats) / len(team_stats), 2) if team_stats else 0,
    }
    
    return render(request, 'operations/team_performance.html', context)


@login_required
@user_passes_test(ops_team_lead_required)
def ops_client_profiles(request):
    """Operations Team Lead view for managing client profiles"""
    user = request.user
    team_members = user.get_team_members()
    
    # Get client profiles assigned to team
    client_profiles = ClientProfile.objects.filter(
        mapped_ops_exec__in=team_members
    ).select_related('mapped_rm', 'mapped_ops_exec')
    
    # Apply filters
    search_form = ClientSearchForm(request.GET)
    if search_form.is_valid():
        search_query = search_form.cleaned_data.get('search_query')
        status = search_form.cleaned_data.get('status')
        ops_exec = search_form.cleaned_data.get('ops_exec')
        
        if search_query:
            client_profiles = client_profiles.filter(
                Q(client_full_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(mobile_number__icontains=search_query) |
                Q(pan_number__icontains=search_query)
            )
        
        if status:
            client_profiles = client_profiles.filter(status=status)
        
        if ops_exec:
            client_profiles = client_profiles.filter(mapped_ops_exec=ops_exec)
    
    # Pagination
    paginator = Paginator(client_profiles, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    stats = {
        'total_profiles': client_profiles.count(),
        'active_profiles': client_profiles.filter(status='active').count(),
        'muted_profiles': client_profiles.filter(status='muted').count(),
        'pending_kyc': client_profiles.filter(
            # Add your KYC pending logic here based on your business rules
        ).count() if hasattr(ClientProfile, 'kyc_status') else 0,
    }
    
    context = {
        'client_profiles': page_obj,
        'search_form': search_form,
        'team_members': team_members,
        'stats': stats,
        'page_obj': page_obj,
    }
    
    return render(request, 'operations/client_profiles.html', context)


@login_required
@user_passes_test(ops_team_lead_required)
def ops_task_assignment(request):
    """Operations Team Lead view for task assignment"""
    user = request.user
    team_members = user.get_team_members()
    
    if request.method == 'POST':
        form = OperationsTaskAssignmentForm(request.POST, current_user=user)
        if form.is_valid():
            task = form.save(commit=False)
            task.assigned_by = user
            task.save()
            messages.success(request, f"Task assigned to {task.assigned_to.get_full_name()}")
            return redirect('ops_task_assignment')
    else:
        form = OperationsTaskAssignmentForm(current_user=user)
    
    # Get team tasks
    team_tasks = Task.objects.filter(
        assigned_to__in=[user] + list(team_members)
    ).select_related('assigned_to', 'assigned_by').order_by('-created_at')
    
    # Task statistics
    task_stats = {
        'total_tasks': team_tasks.count(),
        'pending_tasks': team_tasks.filter(completed=False).count(),
        'completed_tasks': team_tasks.filter(completed=True).count(),
        'overdue_tasks': team_tasks.filter(
            completed=False,
            due_date__lt=timezone.now()
        ).count(),
        'high_priority': team_tasks.filter(
            completed=False,
            priority__in=['high', 'urgent']
        ).count(),
    }
    
    # Individual member task breakdown
    member_task_breakdown = []
    for member in [user] + list(team_members):
        member_tasks = team_tasks.filter(assigned_to=member)
        member_task_breakdown.append({
            'member': member,
            'total': member_tasks.count(),
            'pending': member_tasks.filter(completed=False).count(),
            'completed': member_tasks.filter(completed=True).count(),
            'overdue': member_tasks.filter(
                completed=False,
                due_date__lt=timezone.now()
            ).count(),
        })
    
    context = {
        'form': form,
        'team_tasks': team_tasks[:20],  # Show latest 20 tasks
        'task_stats': task_stats,
        'member_task_breakdown': member_task_breakdown,
        'team_members': team_members,
    }
    
    return render(request, 'operations/task_assignment.html', context)


# Operations Executive Views
@login_required
@user_passes_test(lambda u: u.role == 'ops_exec')
def ops_my_tasks(request):
    """Operations Executive view for personal tasks"""
    user = request.user
    
    # Get user's tasks
    my_tasks = Task.objects.filter(assigned_to=user).select_related('assigned_by')
    
    # Apply filters
    status_filter = request.GET.get('status', 'pending')
    priority_filter = request.GET.get('priority')
    
    if status_filter == 'pending':
        my_tasks = my_tasks.filter(completed=False)
    elif status_filter == 'completed':
        my_tasks = my_tasks.filter(completed=True)
    elif status_filter == 'overdue':
        my_tasks = my_tasks.filter(completed=False, due_date__lt=timezone.now())
    
    if priority_filter:
        my_tasks = my_tasks.filter(priority=priority_filter)
    
    my_tasks = my_tasks.order_by('-created_at')
    
    # Task statistics
    task_stats = {
        'total_tasks': Task.objects.filter(assigned_to=user).count(),
        'pending_tasks': Task.objects.filter(assigned_to=user, completed=False).count(),
        'completed_tasks': Task.objects.filter(assigned_to=user, completed=True).count(),
        'overdue_tasks': Task.objects.filter(
            assigned_to=user,
            completed=False,
            due_date__lt=timezone.now()
        ).count(),
        'due_today': Task.objects.filter(
            assigned_to=user,
            completed=False,
            due_date__date=timezone.now().date()
        ).count(),
        'due_this_week': Task.objects.filter(
            assigned_to=user,
            completed=False,
            due_date__date__range=[
                timezone.now().date(),
                timezone.now().date() + timedelta(days=7)
            ]
        ).count(),
    }
    
    # Performance metrics
    last_30_days = timezone.now() - timedelta(days=30)
    recent_tasks = Task.objects.filter(
        assigned_to=user,
        created_at__gte=last_30_days
    )
    
    performance_metrics = {
        'tasks_last_30_days': recent_tasks.count(),
        'completed_last_30_days': recent_tasks.filter(completed=True).count(),
        'avg_completion_time': 0,  # Calculate based on your business logic
        'completion_rate': 0,
    }
    
    if recent_tasks.exists():
        performance_metrics['completion_rate'] = round(
            (performance_metrics['completed_last_30_days'] / performance_metrics['tasks_last_30_days']) * 100, 2
        )
    
    context = {
        'my_tasks': my_tasks,
        'task_stats': task_stats,
        'performance_metrics': performance_metrics,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'priority_choices': Task.PRIORITY_CHOICES,
    }
    
    return render(request, 'operations/my_tasks.html', context)


@login_required
@user_passes_test(lambda u: u.role == 'ops_exec')
def ops_my_clients(request):
    """Operations Executive view for assigned client profiles"""
    user = request.user
    
    # Get assigned client profiles
    my_clients = ClientProfile.objects.filter(
        mapped_ops_exec=user
    ).select_related('mapped_rm')
    
    # Apply filters
    status_filter = request.GET.get('status')
    search_query = request.GET.get('search')
    
    if status_filter:
        my_clients = my_clients.filter(status=status_filter)
    
    if search_query:
        my_clients = my_clients.filter(
            Q(client_full_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(pan_number__icontains=search_query)
        )
    
    my_clients = my_clients.order_by('-updated_at')
    
    # Pagination
    paginator = Paginator(my_clients, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Client statistics
    client_stats = {
        'total_clients': ClientProfile.objects.filter(mapped_ops_exec=user).count(),
        'active_clients': ClientProfile.objects.filter(mapped_ops_exec=user, status='active').count(),
        'muted_clients': ClientProfile.objects.filter(mapped_ops_exec=user, status='muted').count(),
        'new_this_month': ClientProfile.objects.filter(
            mapped_ops_exec=user,
            created_at__month=timezone.now().month,
            created_at__year=timezone.now().year
        ).count(),
        'updated_today': ClientProfile.objects.filter(
            mapped_ops_exec=user,
            updated_at__date=timezone.now().date()
        ).count(),
    }
    
    # Recent activities
    recent_updates = ClientProfile.objects.filter(
        mapped_ops_exec=user
    ).order_by('-updated_at')[:5]
    
    context = {
        'my_clients': page_obj,
        'client_stats': client_stats,
        'recent_updates': recent_updates,
        'status_filter': status_filter,
        'search_query': search_query,
        'page_obj': page_obj,
    }
    
    return render(request, 'operations/my_clients.html', context)



# Business Head Operations Views
@login_required
@user_passes_test(business_head_ops_required)
def bh_ops_overview(request):
    """Business Head Operations overview dashboard"""
    user = request.user
    
    # Get all operations team members
    ops_team_leads = User.objects.filter(role='ops_team_lead')
    ops_executives = User.objects.filter(role='ops_exec')
    
    # Overall operations metrics
    total_client_profiles = ClientProfile.objects.count()
    active_profiles = ClientProfile.objects.filter(status='active').count()
    muted_profiles = ClientProfile.objects.filter(status='muted').count()
    
    # Task metrics
    ops_tasks = Task.objects.filter(assigned_to__role__in=['ops_team_lead', 'ops_exec'])
    pending_ops_tasks = ops_tasks.filter(completed=False).count()
    overdue_ops_tasks = ops_tasks.filter(
        completed=False,
        due_date__lt=timezone.now()
    ).count()
    
    # Service request metrics
    ops_service_requests = ServiceRequest.objects.filter(
        assigned_to__role__in=['ops_team_lead', 'ops_exec']
    )
    open_requests = ops_service_requests.filter(status='open').count()
    in_progress_requests = ops_service_requests.filter(status='in_progress').count()
    
    # Team performance summary
    team_performance = []
    for lead in ops_team_leads:
        team_members = lead.get_team_members()
        team_tasks = Task.objects.filter(assigned_to__in=team_members)
        team_requests = ServiceRequest.objects.filter(assigned_to__in=team_members)
        
        team_performance.append({
            'lead': lead,
            'team_size': team_members.count(),
            'pending_tasks': team_tasks.filter(completed=False).count(),
            'open_requests': team_requests.filter(status='open').count(),
            'client_profiles': ClientProfile.objects.filter(mapped_ops_exec__in=team_members).count(),
        })
    
    # Recent activities
    recent_profiles = ClientProfile.objects.order_by('-created_at')[:10]
    recent_requests = ServiceRequest.objects.filter(
        assigned_to__role__in=['ops_team_lead', 'ops_exec']
    ).order_by('-created_at')[:10]
    
    # Monthly trends (last 6 months)
    monthly_data = []
    for i in range(6):
        month_start = (timezone.now().replace(day=1) - timedelta(days=30*i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'profiles_created': ClientProfile.objects.filter(
                created_at__range=[month_start, month_end]
            ).count(),
            'requests_resolved': ServiceRequest.objects.filter(
                resolved_at__range=[month_start, month_end],
                assigned_to__role__in=['ops_team_lead', 'ops_exec']
            ).count(),
        })
    
    context = {
        'ops_team_leads': ops_team_leads,
        'ops_executives': ops_executives,
        'total_client_profiles': total_client_profiles,
        'active_profiles': active_profiles,
        'muted_profiles': muted_profiles,
        'pending_ops_tasks': pending_ops_tasks,
        'overdue_ops_tasks': overdue_ops_tasks,
        'open_requests': open_requests,
        'in_progress_requests': in_progress_requests,
        'team_performance': team_performance,
        'recent_profiles': recent_profiles,
        'recent_requests': recent_requests,
        'monthly_data': list(reversed(monthly_data)),
    }
    
    return render(request, 'operations/bh_ops_overview.html', context)


@login_required
@user_passes_test(business_head_ops_required)
def bh_ops_team_management(request):
    """Business Head Operations team management"""
    user = request.user
    
    # Get operations teams
    ops_teams = Team.objects.filter(is_ops_team=True).select_related('leader')
    ops_users = User.objects.filter(role__in=['ops_team_lead', 'ops_exec']).select_related('manager')
    
    # Team statistics
    team_stats = []
    for team in ops_teams:
        team_members = team.members.all()
        team_tasks = Task.objects.filter(assigned_to__in=team_members)
        team_requests = ServiceRequest.objects.filter(assigned_to__in=team_members)
        
        # Calculate team performance metrics
        total_tasks = team_tasks.count()
        completed_tasks = team_tasks.filter(completed=True).count()
        task_completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        
        total_requests = team_requests.count()
        resolved_requests = team_requests.filter(status__in=['resolved', 'closed']).count()
        request_resolution_rate = (resolved_requests / total_requests * 100) if total_requests > 0 else 0
        
        team_stats.append({
            'team': team,
            'member_count': team_members.count(),
            'task_completion_rate': round(task_completion_rate, 2),
            'request_resolution_rate': round(request_resolution_rate, 2),
            'pending_tasks': team_tasks.filter(completed=False).count(),
            'open_requests': team_requests.filter(status='open').count(),
            'performance_score': round((task_completion_rate + request_resolution_rate) / 2, 2),
        })
    
    # Individual user performance
    user_performance = []
    for ops_user in ops_users:
        user_tasks = Task.objects.filter(assigned_to=ops_user)
        user_requests = ServiceRequest.objects.filter(assigned_to=ops_user)
        user_profiles = ClientProfile.objects.filter(mapped_ops_exec=ops_user)
        
        # Calculate individual metrics
        total_tasks = user_tasks.count()
        completed_tasks = user_tasks.filter(completed=True).count()
        task_completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        
        total_requests = user_requests.count()
        resolved_requests = user_requests.filter(status__in=['resolved', 'closed']).count()
        request_resolution_rate = (resolved_requests / total_requests * 100) if total_requests > 0 else 0
        
        user_performance.append({
            'user': ops_user,
            'task_completion_rate': round(task_completion_rate, 2),
            'request_resolution_rate': round(request_resolution_rate, 2),
            'client_profiles': user_profiles.count(),
            'pending_tasks': user_tasks.filter(completed=False).count(),
            'open_requests': user_requests.filter(status='open').count(),
            'performance_score': round((task_completion_rate + request_resolution_rate) / 2, 2),
        })
    
    # Sort by performance score
    user_performance.sort(key=lambda x: x['performance_score'], reverse=True)
    
    context = {
        'ops_teams': ops_teams,
        'team_stats': team_stats,
        'user_performance': user_performance,
        'total_ops_users': ops_users.count(),
        'ops_team_leads_count': ops_users.filter(role='ops_team_lead').count(),
        'ops_executives_count': ops_users.filter(role='ops_exec').count(),
    }
    
    return render(request, 'operations/team_management.html', context)


@login_required
@user_passes_test(business_head_ops_required)
def bh_ops_performance_metrics(request):
    """Business Head Operations performance metrics and analytics"""
    user = request.user
    
    # Get date range for filtering
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        date_from = timezone.now().date() - timedelta(days=90)  # Last 3 months
    else:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    
    if not date_to:
        date_to = timezone.now().date()
    else:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Operations team members
    ops_users = User.objects.filter(role__in=['ops_team_lead', 'ops_exec'])
    
    # Task performance metrics
    ops_tasks = Task.objects.filter(
        assigned_to__in=ops_users,
        created_at__date__range=[date_from, date_to]
    )
    
    task_metrics = {
        'total_tasks': ops_tasks.count(),
        'completed_tasks': ops_tasks.filter(completed=True).count(),
        'pending_tasks': ops_tasks.filter(completed=False).count(),
        'overdue_tasks': ops_tasks.filter(
            completed=False,
            due_date__lt=timezone.now()
        ).count(),
        'high_priority_tasks': ops_tasks.filter(priority__in=['high', 'urgent']).count(),
    }
    
    # Service request performance metrics
    ops_requests = ServiceRequest.objects.filter(
        assigned_to__in=ops_users,
        created_at__date__range=[date_from, date_to]
    )
    
    request_metrics = {
        'total_requests': ops_requests.count(),
        'open_requests': ops_requests.filter(status='open').count(),
        'in_progress_requests': ops_requests.filter(status='in_progress').count(),
        'resolved_requests': ops_requests.filter(status__in=['resolved', 'closed']).count(),
        'urgent_requests': ops_requests.filter(priority='urgent').count(),
    }
    
    # Client profile metrics
    client_metrics = {
        'total_profiles': ClientProfile.objects.filter(
            created_at__date__range=[date_from, date_to]
        ).count(),
        'active_profiles': ClientProfile.objects.filter(
            created_at__date__range=[date_from, date_to],
            status='active'
        ).count(),
        'profiles_by_ops': ClientProfile.objects.filter(
            created_at__date__range=[date_from, date_to],
            mapped_ops_exec__in=ops_users
        ).count(),
    }
    
    # Calculate performance rates
    task_completion_rate = (task_metrics['completed_tasks'] / task_metrics['total_tasks'] * 100) if task_metrics['total_tasks'] > 0 else 0
    request_resolution_rate = (request_metrics['resolved_requests'] / request_metrics['total_requests'] * 100) if request_metrics['total_requests'] > 0 else 0
    
    # Weekly performance trends
    weekly_trends = []
    current_date = date_from
    while current_date <= date_to:
        week_end = min(current_date + timedelta(days=6), date_to)
        
        week_tasks = ops_tasks.filter(created_at__date__range=[current_date, week_end])
        week_requests = ops_requests.filter(created_at__date__range=[current_date, week_end])
        
        weekly_trends.append({
            'week_start': current_date,
            'tasks_created': week_tasks.count(),
            'tasks_completed': week_tasks.filter(completed=True).count(),
            'requests_created': week_requests.count(),
            'requests_resolved': week_requests.filter(status__in=['resolved', 'closed']).count(),
        })
        
        current_date = week_end + timedelta(days=1)
    
    # Top performers
    top_performers = []
    for ops_user in ops_users:
        user_tasks = ops_tasks.filter(assigned_to=ops_user)
        user_requests = ops_requests.filter(assigned_to=ops_user)
        
        user_task_rate = (user_tasks.filter(completed=True).count() / user_tasks.count() * 100) if user_tasks.count() > 0 else 0
        user_request_rate = (user_requests.filter(status__in=['resolved', 'closed']).count() / user_requests.count() * 100) if user_requests.count() > 0 else 0
        
        overall_score = (user_task_rate + user_request_rate) / 2
        
        top_performers.append({
            'user': ops_user,
            'task_completion_rate': round(user_task_rate, 2),
            'request_resolution_rate': round(user_request_rate, 2),
            'overall_score': round(overall_score, 2),
            'total_tasks': user_tasks.count(),
            'total_requests': user_requests.count(),
        })
    
    # Sort by overall score
    top_performers.sort(key=lambda x: x['overall_score'], reverse=True)
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'task_metrics': task_metrics,
        'request_metrics': request_metrics,
        'client_metrics': client_metrics,
        'task_completion_rate': round(task_completion_rate, 2),
        'request_resolution_rate': round(request_resolution_rate, 2),
        'weekly_trends': weekly_trends,
        'top_performers': top_performers[:10],  # Top 10 performers
        'ops_users_count': ops_users.count(),
    }
    
    return render(request, 'operations/performance_metrics.html', context)


@login_required
@user_passes_test(business_head_ops_required)
def bh_ops_compliance(request):
    """Business Head Operations compliance monitoring"""
    user = request.user
    
    # Get operations team
    ops_users = User.objects.filter(role__in=['ops_team_lead', 'ops_exec'])
    
    # Compliance metrics
    compliance_metrics = {
        'total_client_profiles': ClientProfile.objects.count(),
        'complete_profiles': ClientProfile.objects.exclude(
            Q(client_full_name='') | Q(email='') | Q(mobile_number='') | Q(pan_number='')
        ).count(),
        'incomplete_profiles': ClientProfile.objects.filter(
            Q(client_full_name='') | Q(email='') | Q(mobile_number='') | Q(pan_number='')
        ).count(),
        'missing_ops_assignment': ClientProfile.objects.filter(mapped_ops_exec__isnull=True).count(),
        'missing_rm_assignment': ClientProfile.objects.filter(mapped_rm__isnull=True).count(),
    }
    
    # Calculate compliance rate
    compliance_rate = (compliance_metrics['complete_profiles'] / compliance_metrics['total_client_profiles'] * 100) if compliance_metrics['total_client_profiles'] > 0 else 0
    
    # Overdue tasks compliance
    overdue_tasks = Task.objects.filter(
        assigned_to__in=ops_users,
        completed=False,
        due_date__lt=timezone.now()
    ).select_related('assigned_to')
    
    # Service request SLA compliance
    sla_breached_requests = ServiceRequest.objects.filter(
        assigned_to__in=ops_users,
        status='open',
        created_at__lt=timezone.now() - timedelta(hours=24)  # Assuming 24-hour SLA
    ).select_related('assigned_to', 'client')
    
    # Data quality issues
    data_quality_issues = []
    
    # Check for duplicate PAN numbers
    duplicate_pans = ClientProfile.objects.values('pan_number').annotate(
        count=Count('pan_number')
    ).filter(count__gt=1)
    
    if duplicate_pans.exists():
        data_quality_issues.append({
            'type': 'Duplicate PAN Numbers',
            'count': duplicate_pans.count(),
            'severity': 'High',
            'details': f"{duplicate_pans.count()} PAN numbers have multiple entries"
        })
    
    # Check for missing email addresses
    missing_emails = ClientProfile.objects.filter(email='').count()
    if missing_emails > 0:
        data_quality_issues.append({
            'type': 'Missing Email Addresses',
            'count': missing_emails,
            'severity': 'Medium',
            'details': f"{missing_emails} client profiles missing email addresses"
        })
    
    # Check for invalid mobile numbers
    invalid_mobiles = ClientProfile.objects.exclude(
        mobile_number__regex=r'^[6-9]\d{9}$'
    ).exclude(mobile_number='').count()
    
    if invalid_mobiles > 0:
        data_quality_issues.append({
            'type': 'Invalid Mobile Numbers',
            'count': invalid_mobiles,
            'severity': 'Medium',
            'details': f"{invalid_mobiles} client profiles have invalid mobile numbers"
        })
    
    # Team compliance scores
    team_compliance = []
    for ops_user in ops_users:
        user_profiles = ClientProfile.objects.filter(mapped_ops_exec=ops_user)
        user_tasks = Task.objects.filter(assigned_to=ops_user)
        user_requests = ServiceRequest.objects.filter(assigned_to=ops_user)
        
        # Calculate individual compliance scores
        complete_profiles = user_profiles.exclude(
            Q(client_full_name='') | Q(email='') | Q(mobile_number='') | Q(pan_number='')
        ).count()
        
        profile_completion_rate = (complete_profiles / user_profiles.count() * 100) if user_profiles.count() > 0 else 100
        
        overdue_user_tasks = user_tasks.filter(
            completed=False,
            due_date__lt=timezone.now()
        ).count()
        
        task_compliance_rate = ((user_tasks.count() - overdue_user_tasks) / user_tasks.count() * 100) if user_tasks.count() > 0 else 100
        
        sla_breached_user_requests = user_requests.filter(
            status='open',
            created_at__lt=timezone.now() - timedelta(hours=24)
        ).count()
        
        request_compliance_rate = ((user_requests.count() - sla_breached_user_requests) / user_requests.count() * 100) if user_requests.count() > 0 else 100
        
        overall_compliance = (profile_completion_rate + task_compliance_rate + request_compliance_rate) / 3
        
        team_compliance.append({
            'user': ops_user,
            'profile_completion_rate': round(profile_completion_rate, 2),
            'task_compliance_rate': round(task_compliance_rate, 2),
            'request_compliance_rate': round(request_compliance_rate, 2),
            'overall_compliance': round(overall_compliance, 2),
            'assigned_profiles': user_profiles.count(),
            'overdue_tasks': overdue_user_tasks,
            'sla_breached_requests': sla_breached_user_requests,
        })
    
    # Sort by overall compliance
    team_compliance.sort(key=lambda x: x['overall_compliance'], reverse=True)
    
    context = {
        'compliance_metrics': compliance_metrics,
        'compliance_rate': round(compliance_rate, 2),
        'overdue_tasks': overdue_tasks,
        'sla_breached_requests': sla_breached_requests,
        'data_quality_issues': data_quality_issues,
        'team_compliance': team_compliance,
        'ops_users_count': ops_users.count(),
    }
    
    return render(request, 'operations/compliance.html', context)


# API Endpoints for AJAX calls
@login_required
@require_GET
def get_dashboard_stats(request):
    """API endpoint to get dashboard statistics"""
    user = request.user
    
    try:
        if user.role == 'top_management':
            stats = {
                'total_users': User.objects.count(),
                'total_clients': Client.objects.count(),
                'total_leads': Lead.objects.count(),
                'total_tasks': Task.objects.filter(completed=False).count(),
                'total_service_requests': ServiceRequest.objects.filter(status='open').count(),
            }
        elif user.role in ['business_head', 'business_head_ops']:
            stats = {
                'total_clients': Client.objects.count(),
                'total_leads': Lead.objects.count(),
                'pending_tasks': Task.objects.filter(completed=False).count(),
                'open_service_requests': ServiceRequest.objects.filter(status='open').count(),
            }
        elif user.role in ['rm_head', 'ops_team_lead']:
            accessible_users = user.get_accessible_users()
            stats = {
                'team_members': accessible_users.count(),
                'team_tasks': Task.objects.filter(assigned_to__in=accessible_users, completed=False).count(),
                'team_leads': Lead.objects.filter(assigned_to__in=accessible_users).count(),
                'team_clients': Client.objects.filter(user__in=accessible_users).count(),
            }
        else:  # RM, Ops Exec
            stats = {
                'my_tasks': Task.objects.filter(assigned_to=user, completed=False).count(),
                'my_leads': Lead.objects.filter(assigned_to=user).count(),
                'my_clients': Client.objects.filter(user=user).count(),
            }
        
        return JsonResponse({'success': True, 'stats': stats})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_GET
def get_user_hierarchy(request):
    """API endpoint to get user hierarchy data"""
    user = request.user
    
    try:
        hierarchy_data = []
        
        if user.role in ['top_management', 'business_head', 'business_head_ops']:
            # Get all users with their hierarchy
            all_users = User.objects.select_related('manager').all()
            
            for u in all_users:
                hierarchy_data.append({
                    'id': u.id,
                    'name': u.get_full_name() or u.username,
                    'role': u.get_role_display(),
                    'manager_id': u.manager.id if u.manager else None,
                    'email': u.email,
                })
        
        elif user.role in ['rm_head', 'ops_team_lead']:
            # Get accessible users
            accessible_users = user.get_accessible_users()
            
            for u in accessible_users:
                hierarchy_data.append({
                    'id': u.id,
                    'name': u.get_full_name() or u.username,
                    'role': u.get_role_display(),
                    'manager_id': u.manager.id if u.manager else None,
                    'email': u.email,
                })
        
        return JsonResponse({'success': True, 'hierarchy': hierarchy_data})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_GET
def get_team_performance(request):
    """API endpoint to get team performance data"""
    user = request.user
    
    try:
        if user.role in ['rm_head', 'ops_team_lead']:
            team_members = user.get_team_members()
            performance_data = []
            
            for member in team_members:
                if user.role == 'rm_head':
                    # RM team performance
                    member_leads = Lead.objects.filter(assigned_to=member)
                    member_clients = Client.objects.filter(user=member)
                    
                    performance_data.append({
                        'user_id': member.id,
                        'name': member.get_full_name() or member.username,
                        'leads': member_leads.count(),
                        'clients': member_clients.count(),
                        'aum': float(member_clients.aggregate(total=Sum('aum'))['total'] or 0),
                    })
                
                elif user.role == 'ops_team_lead':
                    # Operations team performance
                    member_tasks = Task.objects.filter(assigned_to=member)
                    member_requests = ServiceRequest.objects.filter(assigned_to=member)
                    member_profiles = ClientProfile.objects.filter(mapped_ops_exec=member)
                    
                    performance_data.append({
                        'user_id': member.id,
                        'name': member.get_full_name() or member.username,
                        'total_tasks': member_tasks.count(),
                        'completed_tasks': member_tasks.filter(completed=True).count(),
                        'service_requests': member_requests.count(),
                        'client_profiles': member_profiles.count(),
                    })
            
            return JsonResponse({'success': True, 'performance': performance_data})
        
        else:
            return JsonResponse({'success': False, 'error': 'Insufficient permissions'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def quick_create_note(request):
    """API endpoint for quick note creation"""
    try:
        data = json.loads(request.body)
        
        # Get or create default note list
        note_list, created = NoteList.objects.get_or_create(
            user=request.user,
            name='Quick Notes',
            defaults={'description': 'Quick notes created from dashboard'}
        )
        
        # Create note
        note = Note.objects.create(
            user=request.user,
            note_list=note_list,
            heading=data.get('heading', 'Quick Note'),
            content=data.get('content', ''),
            creation_date=timezone.now().date()
        )
        
        return JsonResponse({
            'success': True,
            'note_id': note.id,
            'message': 'Note created successfully'
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def quick_assign_task(request):
    """API endpoint for quick task assignment"""
    try:
        data = json.loads(request.body)
        
        # Get assignee
        assignee_id = data.get('assignee_id')
        if not assignee_id:
            assignee = request.user
        else:
            assignee = User.objects.get(id=assignee_id)
            
            # Check if user can assign to this person
            if not request.user.can_access_user_data(assignee):
                return JsonResponse({'success': False, 'error': 'Permission denied'})
        
        # Create task
        task = Task.objects.create(
            assigned_to=assignee,
            assigned_by=request.user,
            title=data.get('title', 'Quick Task'),
            description=data.get('description', ''),
            priority=data.get('priority', 'medium'),
            due_date=data.get('due_date')
        )
        
        return JsonResponse({
            'success': True,
            'task_id': task.id,
            'message': f'Task assigned to {assignee.get_full_name()}'
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# Error handling views
def permission_denied(request, exception=None):
    """403 error handler"""
    return render(request, 'errors/403.html', status=403)


def not_found(request, exception=None):
    """404 error handler"""
    return render(request, 'errors/404.html', status=404)


def server_error(request):
    """500 error handler"""
    return render(request, 'errors/500.html', status=500)


# Quick access views
@login_required
def quick_note_create(request):
    """Quick note creation for modal/popup"""
    if request.method == 'POST':
        form = QuickNoteForm(request.POST, user=request.user)
        if form.is_valid():
            note = form.save()
            messages.success(request, 'Note created successfully')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'note_id': note.id,
                    'redirect_url': reverse('note_detail', args=[note.id])
                })
            
            return redirect('note_detail', pk=note.id)
    else:
        form = QuickNoteForm(user=request.user)
    
    context = {'form': form}
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'notes/quick_note_form.html', context)
    
    return render(request, 'notes/note_form.html', context)


@login_required
def quick_task_create(request):
    """Quick task creation"""
    if request.method == 'POST':
        form = TaskForm(request.POST, current_user=request.user)
        if form.is_valid():
            task = form.save()
            messages.success(request, f'Task assigned to {task.assigned_to.get_full_name()}')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'task_id': task.id,
                    'message': f'Task assigned to {task.assigned_to.get_full_name()}'
                })
            
            return redirect('task_list')
    else:
        form = TaskForm(current_user=request.user)
    
    context = {'form': form}
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'tasks/quick_task_form.html', context)
    
    return render(request, 'tasks/task_form.html', context)


@login_required
@require_GET
def quick_client_search(request):
    """Quick client search for autocomplete"""
    query = request.GET.get('q', '')
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    # Filter clients based on user access
    if request.user.role in ['top_management', 'business_head', 'business_head_ops']:
        clients = ClientProfile.objects.all()
    elif request.user.role in ['rm_head', 'ops_team_lead']:
        accessible_users = request.user.get_accessible_users()
        clients = ClientProfile.objects.filter(
            Q(mapped_rm__in=accessible_users) | Q(mapped_ops_exec__in=accessible_users)
        )
    elif request.user.role == 'rm':
        clients = ClientProfile.objects.filter(mapped_rm=request.user)
    elif request.user.role == 'ops_exec':
        clients = ClientProfile.objects.filter(mapped_ops_exec=request.user)
    else:
        clients = ClientProfile.objects.none()
    
    # Search clients
    clients = clients.filter(
        Q(client_full_name__icontains=query) |
        Q(email__icontains=query) |
        Q(mobile_number__icontains=query) |
        Q(pan_number__icontains=query)
    )[:10]
    
    results = []
    for client in clients:
        results.append({
            'id': client.id,
            'text': client.client_full_name,
            'email': client.email,
            'mobile': client.mobile_number,
            'pan': client.pan_number,
            'status': client.get_status_display(),
        })
    
    return JsonResponse({'results': results})


@require_GET
def quick_lead_search(request):
    """Quick lead search for autocomplete"""
    query = request.GET.get('q', '')
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    # Filter leads based on user access
    if request.user.role in ['top_management', 'business_head']:
        leads = Lead.objects.all()
    elif request.user.role == 'rm_head':
        accessible_users = request.user.get_accessible_users()
        leads = Lead.objects.filter(assigned_to__in=accessible_users)
    elif request.user.role == 'rm':
        leads = Lead.objects.filter(assigned_to=request.user)
    else:
        leads = Lead.objects.none()
    
    # Search leads
    leads = leads.filter(
        Q(name__icontains=query) |
        Q(email__icontains=query) |
        Q(mobile__icontains=query) |
        Q(lead_id__icontains=query)
    )[:10]
    
    results = []
    for lead in leads:
        results.append({
            'id': lead.id,
            'text': lead.name,
            'email': lead.email or '',
            'mobile': lead.mobile or '',
            'status': lead.get_status_display(),
            'probability': lead.probability,
            'assigned_to': lead.assigned_to.get_full_name() if lead.assigned_to else '',
        })
    
    return JsonResponse({'results': results})


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone

from .models import Lead, LeadInteraction, ProductDiscussion, LeadStatusChange, Client
from .forms import (
    LeadForm, 
    LeadInteractionForm, 
    ProductDiscussionForm,
    LeadConversionForm,
    LeadStatusChangeForm,
    LeadReassignmentForm,
    ClientCreationForm
)
from .models import User

# Update the lead_list view in views.py

@login_required
def lead_list(request):
    user = request.user
    
    # Role-based lead access
    if user.role == 'ops_team_lead':
        # Ops team lead can see converted leads and leads awaiting conversion approval
        # IMPORTANT: Include leads where conversion is requested but not yet approved
        leads = Lead.objects.filter(
            Q(status='converted') | 
            Q(status='conversion_requested') |
            Q(conversion_requested_at__isnull=False, converted=False)  # Additional safety check
        ).distinct()
    elif user.role in ['top_management', 'business_head']:
        # Top management and business head can see all leads in the system
        leads = Lead.objects.all()
    elif user.role == 'rm_head':
        # RM Head can see leads assigned to themselves and their RMs
        subordinate_rms = get_user_accessible_data(user, User, 'manager')
        
        leads = Lead.objects.filter(
            Q(assigned_to=user) |
            Q(assigned_to__in=subordinate_rms) |
            Q(final_assigned_rm=user) |
            Q(final_assigned_rm__in=subordinate_rms) |
            Q(created_by=user)
        ).distinct()
    else:
        # For other roles (RM, ops_exec, etc.)
        accessible_leads = get_user_accessible_data(user, Lead, 'assigned_to')
        created_leads = Lead.objects.filter(created_by=user)
        final_assigned_leads = Lead.objects.filter(final_assigned_rm=user)
        
        leads = (accessible_leads | created_leads | final_assigned_leads).distinct()
    
    # Filtering options
    status_filter = request.GET.get('status')
    if status_filter:
        leads = leads.filter(status=status_filter)
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        leads = leads.filter(
            Q(name__icontains=search_query) | 
            Q(email__icontains=search_query) |
            Q(mobile__icontains=search_query) |
            Q(source__icontains=search_query) |
            Q(lead_id__icontains=search_query) |
            Q(client_id__icontains=search_query)
        )

    # Filter by conversion status
    converted_filter = request.GET.get('converted')
    if converted_filter == 'true':
        leads = leads.filter(converted=True)
    elif converted_filter == 'false':
        leads = leads.filter(converted=False)

    # Add data specific to different roles
    pending_conversion_approvals = 0
    recent_conversion_decisions = []
    
    if user.role == 'ops_team_lead':
        # Get pending conversion approvals count - both those needing approval and those that can be self-assigned
        pending_conversion_approvals = LeadStatusChange.objects.filter(
            Q(needs_approval=True, new_status='conversion_requested', approved__isnull=True) |
            Q(lead__status='conversion_requested', lead__converted=False)
        ).distinct().count()
        
        # Also count leads that are in conversion_requested status but don't have approval records
        leads_without_approval = Lead.objects.filter(
            status='conversion_requested',
            converted=False
        ).exclude(
            status_changes__needs_approval=True,
            status_changes__new_status='conversion_requested'
        ).count()
        
        pending_conversion_approvals += leads_without_approval
        
        # Get recent conversion decisions
        recent_conversion_decisions = LeadStatusChange.objects.filter(
            Q(approval_by=user) | Q(changed_by=user),
            new_status__in=['conversion_requested', 'converted'],
            approved__isnull=False
        ).select_related('lead', 'changed_by').order_by('-approved_at')[:5]

    # Order leads based on role
    if user.role == 'ops_team_lead':
        # Show conversion_requested leads first, then converted leads by conversion date
        leads = leads.extra(
            select={'priority': "CASE WHEN status = 'conversion_requested' THEN 0 ELSE 1 END"}
        ).order_by('priority', '-created_at')
    elif user.role in ['top_management', 'business_head']:
        leads = leads.order_by('-created_at')
    elif user.role == 'rm_head':
        leads = leads.extra(
            select={'priority': """
                CASE 
                    WHEN status = 'conversion_requested' THEN 0
                    WHEN status IN ('new', 'contacted', 'qualified') THEN 1
                    ELSE 2 
                END
            """}
        ).order_by('priority', '-created_at')
    else:
        leads = leads.order_by('-created_at')

    # Debug information
    debug_info = {}
    if user.role in ['ops_team_lead', 'top_management', 'business_head', 'rm_head']:
        debug_info = {
            'total_leads': Lead.objects.count(),
            'converted_leads': Lead.objects.filter(status='converted').count(),
            'conversion_requested': Lead.objects.filter(status='conversion_requested').count(),
            'accessible_leads_count': leads.count(),
            'user_role': user.role,
            'current_filters': {
                'status': status_filter,
                'search': search_query,
                'converted': converted_filter,
            }
        }
        
        # Role-specific debug info
        if user.role == 'ops_team_lead':
            debug_info['ops_relevant_leads'] = Lead.objects.filter(
                Q(status='converted') | Q(status='conversion_requested')
            ).count()
            debug_info['pending_approvals'] = pending_conversion_approvals
        elif user.role in ['top_management', 'business_head']:
            debug_info['role_privilege'] = 'Can view all leads'
        elif user.role == 'rm_head':
            subordinate_rms = get_user_accessible_data(user, User, 'manager')
            debug_info['team_leads'] = Lead.objects.filter(
                Q(assigned_to=user) | Q(assigned_to__in=subordinate_rms)
            ).count()
            debug_info['team_size'] = subordinate_rms.count() + 1

    # Get status choices for filter dropdown
    status_choices = Lead.STATUS_CHOICES

    context = {
        'leads': leads,
        'status_choices': status_choices,
        'current_status': status_filter,
        'search_query': search_query,
        'converted_filter': converted_filter,
        'pending_conversion_approvals': pending_conversion_approvals,
        'recent_conversion_decisions': recent_conversion_decisions,
        'debug_info': debug_info,
    }
    
    return render(request, 'base/leads.html', context)


@login_required
def lead_detail(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    # Simplified permission check
    has_permission = False
    
    # Check based on user role and relationship to lead
    if request.user.role in ['top_management', 'business_head', 'business_head_ops']:
        # Top level roles can access all leads
        has_permission = True
    elif request.user.role in ['ops_team_lead', 'ops_exec']:
        # Ops team can access leads pending their approval or previously decided by them
        pending_approval = LeadStatusChange.objects.filter(
            lead=lead,
            needs_approval=True,
            new_status='conversion_requested',
            approved__isnull=True
        ).exists()
        
        previous_decision = LeadStatusChange.objects.filter(
            lead=lead,
            approval_by=request.user,
            approved__isnull=False
        ).exists()
        
        # Allow access to conversion_requested leads
        conversion_requested = (lead.status == 'conversion_requested')
        
        has_permission = pending_approval or previous_decision or conversion_requested
    elif request.user.role == 'rm_head':
        # RM Head can access leads assigned to their team members
        # Simple check: if lead is assigned to an RM, allow access
        has_permission = (lead.assigned_to.role == 'rm') or (lead.assigned_to == request.user)
    elif request.user.role == 'rm':
        # RM can access their own leads
        has_permission = (lead.assigned_to == request.user)
    elif hasattr(lead, 'created_by') and lead.created_by == request.user:
        # Creator can always access
        has_permission = True
    
    if not has_permission:
        raise PermissionDenied("You don't have permission to view this lead.")
    
    interactions = lead.interactions.all().order_by('-interaction_date')
    product_discussions = lead.product_discussions.all()
    status_changes = lead.status_changes.all().order_by('-changed_at')
    
    # Forms for interaction and status change
    interaction_form = LeadInteractionForm()
    status_change_form = LeadStatusChangeForm()
    product_discussion_form = ProductDiscussionForm()
    
    # Conversion form (only for managers and assigned RM) - REMOVED direct conversion
    conversion_form = None
    # Direct conversion is no longer allowed - all must go through ops verification
    
    # Reassignment form (only for managers)
    reassignment_form = None
    if request.user.role in ['rm_head', 'business_head', 'top_management']:
        reassignment_form = LeadReassignmentForm(user=request.user)
    
    # Check if there's a pending approval for this lead
    pending_approval = None
    if request.user.role in ['ops_team_lead', 'ops_exec']:
        pending_approval = LeadStatusChange.objects.filter(
            lead=lead,
            needs_approval=True,
            new_status='conversion_requested',
            approved__isnull=True
        ).first()
    
    # Get available RMs for approval form (if ops_team_lead or ops_exec)
    available_rms = None
    if request.user.role in ['ops_team_lead', 'ops_exec'] and (pending_approval or lead.status == 'conversion_requested'):
        available_rms = User.objects.filter(role='rm', is_active=True).order_by('first_name', 'last_name')
    
    product_choices = ProductDiscussion.PRODUCT_CHOICES
    context = {
        'lead': lead,
        'interactions': interactions,
        'product_choices': product_choices,
        'product_discussions': product_discussions,
        'status_changes': status_changes,
        'interaction_form': interaction_form,
        'status_change_form': status_change_form,
        'product_discussion_form': product_discussion_form,
        'conversion_form': conversion_form,  # This will be None now
        'reassignment_form': reassignment_form,
        'auto_client_id': lead.generate_client_id(),
        'pending_approval': pending_approval,
        'available_rms': available_rms,
        'user_role': request.user.role,
        'can_request_conversion': (
            request.user.role in ['rm', 'rm_head', 'business_head', 'top_management'] and 
            not lead.converted and 
            lead.status != 'conversion_requested' and
            request.user.can_access_user_data(lead.assigned_to)
        ),
    }
    
    return render(request, 'base/lead_detail.html', context)

@login_required
def lead_create(request):
    if not request.user.role in ['rm', 'rm_head', 'business_head', 'top_management','ops_team_lead','ops_exec']:
        messages.error(request, "You don't have permission to create leads.")
        return redirect('lead_list')
        
    if request.method == 'POST':
        form = LeadForm(request.POST)
        if form.is_valid():
            lead = form.save(commit=False)
            lead.created_by = request.user
            lead.status = 'new'  # Default status
            lead.lead_id = generate_lead_id()  # Custom function to generate ID
            lead.save()
            
            # Create initial status change record
            LeadStatusChange.objects.create(
                lead=lead,
                changed_by=request.user,
                old_status='',
                new_status='new',
                notes='Lead created'
            )
            
            messages.success(request, "Lead created successfully.")
            return redirect('lead_detail', pk=lead.pk)
    else:
        form = LeadForm()
        
        # Limit assignee choices based on user's role and team
        if request.user.role == 'rm':
            # RM can only assign leads to themselves
            form.fields['assigned_to'].queryset = User.objects.filter(id=request.user.id)
        elif request.user.role == 'rm_head':
            accessible_users = request.user.get_accessible_users()
            form.fields['assigned_to'].queryset = User.objects.filter(
                id__in=[u.id for u in accessible_users],
                role__in=['rm', 'rm_head']
            )
    
    return render(request, 'base/lead_form.html', {'form': form, 'action': 'Create'})

@login_required
def lead_update(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    # Check permissions
    if not request.user.can_access_user_data(lead.assigned_to):
        raise PermissionDenied("You don't have permission to edit this lead.")
    
    if request.method == 'POST':
        form = LeadForm(request.POST, instance=lead)
        if form.is_valid():
            old_status = lead.status
            lead = form.save()
            
            # Record status change if it was modified
            if old_status != lead.status:
                LeadStatusChange.objects.create(
                    lead=lead,
                    changed_by=request.user,
                    old_status=old_status,
                    new_status=lead.status,
                    notes='Status updated via lead edit'
                )
                
            messages.success(request, "Lead updated successfully.")
            return redirect('lead_detail', pk=lead.pk)
    else:
        form = LeadForm(instance=lead)
        
        # Limit assignee choices based on user's role
        if request.user.role == 'rm':
            # RM can only assign leads to themselves
            form.fields['assigned_to'].queryset = User.objects.filter(id=request.user.id)
        elif request.user.role == 'rm_head':
            accessible_users = request.user.get_accessible_users()
            form.fields['assigned_to'].queryset = User.objects.filter(
                id__in=[u.id for u in accessible_users],
                role__in=['rm', 'rm_head']
            )
    
    return render(request, 'base/lead_form.html', {'form': form, 'action': 'Update', 'lead': lead})

@login_required
def lead_delete(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    # Check permissions
    if not request.user.can_access_user_data(lead.assigned_to):
        raise PermissionDenied("You don't have permission to delete this lead.")
    
    if request.method == 'POST':
        lead.delete()
        messages.success(request, "Lead deleted successfully.")
        return redirect('lead_list')
    
    return render(request, 'base/lead_confirm_delete.html', {'lead': lead})

@login_required
@require_POST
def add_interaction(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    # Check permissions
    if not request.user.can_access_user_data(lead.assigned_to):
        raise PermissionDenied("You don't have permission to add interactions for this lead.")
    
    form = LeadInteractionForm(request.POST)
    
    if not form.is_valid():
        # Collect all form errors for display
        error_list = []
        for field, errors in form.errors.items():
            field_name = form.fields[field].label or field
            error_list.append(f"{field_name}: {' '.join(errors)}")
        
        messages.error(request, "Please correct the errors below: " + ", ".join(error_list))
        return redirect('lead_detail', pk=pk)
    
    try:
        with transaction.atomic():
            interaction = form.save(commit=False)
            interaction.lead = lead
            interaction.interacted_by = request.user
            interaction.save()
            
            # Handle first interaction logic
            if not lead.first_interaction_date:
                lead.first_interaction_date = interaction.interaction_date
                lead.status = 'contacted'
                lead.save()
                
                LeadStatusChange.objects.create(
                    lead=lead,
                    changed_by=request.user,
                    old_status=lead.status,
                    new_status='contacted',
                    notes='First interaction completed'
                )
            
            messages.success(request, f"{interaction.get_interaction_type_display()} interaction added successfully.")
    
    except Exception as e:
        messages.error(request, f"Failed to save interaction: {str(e)}")
    
    return redirect('lead_detail', pk=pk)

@login_required
@require_POST
def add_product_discussion(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    # Check permissions
    if not request.user.can_access_user_data(lead.assigned_to):
        raise PermissionDenied("You don't have permission to add product discussions for this lead.")
    
    # Check if this is the first interaction
    if not lead.interactions.exists():
        messages.error(request, "You must have at least one interaction before adding product discussions.")
        return redirect('lead_detail', pk=lead.pk)
    
    form = ProductDiscussionForm(request.POST)
    if form.is_valid():
        discussion = form.save(commit=False)
        discussion.lead = lead
        discussion.discussed_by = request.user
        discussion.save()
        messages.success(request, "Product discussion added successfully.")
    else:
        messages.error(request, "Error adding product discussion.")
    
    return redirect('lead_detail', pk=lead.pk)

@login_required
@require_POST
def change_lead_status(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    # Check permissions
    if not request.user.can_access_user_data(lead.assigned_to):
        raise PermissionDenied("You don't have permission to change status for this lead.")
    
    form = LeadStatusChangeForm(request.POST)
    if form.is_valid():
        new_status = form.cleaned_data['new_status']
        notes = form.cleaned_data['notes']
        
        # Record status change
        LeadStatusChange.objects.create(
            lead=lead,
            changed_by=request.user,
            old_status=lead.status,
            new_status=new_status,
            notes=notes
        )
        
        # Update lead status
        lead.status = new_status
        lead.save()
        
        messages.success(request, f"Lead status changed to {new_status}.")
    else:
        messages.error(request, "Error changing lead status.")
    
    return redirect('lead_detail', pk=lead.pk)

@login_required
@require_POST
def request_conversion(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    # Check permissions
    if not request.user.can_access_user_data(lead.assigned_to):
        raise PermissionDenied("You don't have permission to request conversion for this lead.")
    
    if lead.converted:
        messages.error(request, "This lead is already converted.")
        return redirect('lead_detail', pk=pk)
    
    # Check if conversion is already requested
    if lead.status == 'conversion_requested':
        messages.warning(request, "Conversion request is already pending ops team verification.")
        return redirect('lead_detail', pk=pk)
    
    # MANDATORY: All conversion requests must go through ops_team_lead for verification
    # Find ops_team_lead for business verification
    ops_team_lead = User.objects.filter(role='ops_team_lead', is_active=True).first()
    
    if not ops_team_lead:
        messages.error(request, 
            "No operations team lead found for business verification. "
            "Please contact your administrator."
        )
        return redirect('lead_detail', pk=pk)
    
    try:
        with transaction.atomic():
            # Store the old status before changing
            old_status = lead.status
            
            # Update lead status FIRST
            lead.status = 'conversion_requested'
            lead.conversion_requested_at = timezone.now()
            lead.conversion_requested_by = request.user
            lead.save()
            
            # Create a status change record as conversion request - THIS IS THE KEY FIX
            status_change = LeadStatusChange.objects.create(
                lead=lead,
                changed_by=request.user,
                old_status=old_status,  # Use the actual old status
                new_status='conversion_requested',
                notes=f'Conversion requested by {request.user.get_full_name()} - pending business verification from ops team',
                needs_approval=True,  # THIS IS CRITICAL
                approval_by=ops_team_lead  # THIS IS CRITICAL
            )
            
            messages.success(request, 
                f"Conversion request sent to ops team for business verification. "
                f"The request will be reviewed by {ops_team_lead.get_full_name()}."
            )
                
    except Exception as e:
        messages.error(request, f"Error requesting conversion: {str(e)}")
    
    return redirect('lead_detail', pk=pk)

@login_required
def debug_ops_workflow(request):
    """Debug view to check why ops team isn't seeing pending approvals"""
    
    if not request.user.role in ['ops_team_lead', 'ops_exec', 'top_management']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    debug_data = {}
    
    # 1. Check leads with conversion_requested status
    conversion_requested_leads = Lead.objects.filter(status='conversion_requested')
    debug_data['conversion_requested_leads'] = {
        'count': conversion_requested_leads.count(),
        'leads': [
            {
                'id': lead.id,
                'name': lead.name,
                'lead_id': lead.lead_id,
                'status': lead.status,
                'conversion_requested_at': lead.conversion_requested_at,
                'conversion_requested_by': lead.conversion_requested_by.get_full_name() if lead.conversion_requested_by else None,
                'converted': lead.converted
            }
            for lead in conversion_requested_leads[:10]  # Show first 10
        ]
    }
    
    # 2. Check LeadStatusChange records that need approval
    pending_status_changes = LeadStatusChange.objects.filter(
        needs_approval=True,
        new_status='conversion_requested',
        approved__isnull=True
    )
    debug_data['pending_status_changes'] = {
        'count': pending_status_changes.count(),
        'records': [
            {
                'id': change.id,
                'lead_name': change.lead.name,
                'lead_id': change.lead.lead_id,
                'changed_by': change.changed_by.get_full_name(),
                'changed_at': change.changed_at,
                'approval_by': change.approval_by.get_full_name() if change.approval_by else None,
                'notes': change.notes
            }
            for change in pending_status_changes[:10]
        ]
    }
    
    # 3. Check ops team users
    ops_users = User.objects.filter(role__in=['ops_team_lead', 'ops_exec'], is_active=True)
    debug_data['ops_users'] = {
        'count': ops_users.count(),
        'users': [
            {
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'role': user.role,
                'is_active': user.is_active
            }
            for user in ops_users
        ]
    }
    
    # 4. Check recent conversions
    recent_conversions = Lead.objects.filter(converted=True).order_by('-converted_at')[:5]
    debug_data['recent_conversions'] = [
        {
            'name': lead.name,
            'converted_at': lead.converted_at,
            'converted_by': lead.converted_by.get_full_name() if lead.converted_by else None,
            'client_id': lead.client_id
        }
        for lead in recent_conversions
    ]
    
    # 5. Check all status changes for conversion_requested
    all_conversion_requests = LeadStatusChange.objects.filter(
        new_status='conversion_requested'
    ).order_by('-changed_at')[:10]
    debug_data['all_conversion_requests'] = [
        {
            'lead_name': change.lead.name,
            'changed_by': change.changed_by.get_full_name(),
            'changed_at': change.changed_at,
            'needs_approval': change.needs_approval,
            'approved': change.approved,
            'approval_by': change.approval_by.get_full_name() if change.approval_by else None,
            'approved_at': change.approved_at
        }
        for change in all_conversion_requests
    ]
    
    # 6. Check the current user's pending approvals specifically
    user_pending_approvals = LeadStatusChange.objects.filter(
        needs_approval=True,
        new_status='conversion_requested',
        approval_by=request.user,
        approved__isnull=True
    )
    debug_data['user_pending_approvals'] = {
        'count': user_pending_approvals.count(),
        'approvals': [
            {
                'lead_name': approval.lead.name,
                'changed_by': approval.changed_by.get_full_name(),
                'changed_at': approval.changed_at,
                'notes': approval.notes
            }
            for approval in user_pending_approvals
        ]
    }
    
    # 7. Check if there are any unassigned conversion requests
    unassigned_conversions = Lead.objects.filter(
        status='conversion_requested',
        converted=False
    ).exclude(
        id__in=LeadStatusChange.objects.filter(
            needs_approval=True,
            new_status='conversion_requested'
        ).values_list('lead_id', flat=True)
    )
    debug_data['unassigned_conversions'] = {
        'count': unassigned_conversions.count(),
        'leads': [
            {
                'name': lead.name,
                'lead_id': lead.lead_id,
                'conversion_requested_at': lead.conversion_requested_at,
                'conversion_requested_by': lead.conversion_requested_by.get_full_name() if lead.conversion_requested_by else None
            }
            for lead in unassigned_conversions
        ]
    }
    
    return JsonResponse(debug_data, indent=2)

def ops_pending_approvals(request):
    """Operations Team Lead view for pending conversion approvals - FIXED VERSION"""
    if request.user.role not in ['ops_team_lead', 'ops_exec']:
        messages.error(request, "Access denied.")
        return redirect('dashboard')
    
    # FIXED QUERY: Use approved=False instead of approved__isnull=True
    if request.user.role == 'ops_team_lead':
        # Show all pending conversion requests for ops_team_lead
        pending_approvals = LeadStatusChange.objects.filter(
            needs_approval=True,
            new_status='conversion_requested',
            approved=False  # FIXED: was approved__isnull=True
        ).select_related('lead', 'changed_by').order_by('-changed_at')
        
        # Also auto-assign unassigned conversion requests to ops_team_lead
        unassigned_conversions = Lead.objects.filter(
            status='conversion_requested',
            converted=False
        ).exclude(
            id__in=LeadStatusChange.objects.filter(
                needs_approval=True,
                new_status='conversion_requested'
            ).values_list('lead_id', flat=True)
        )
        
        # Create approval records for unassigned conversions
        for lead in unassigned_conversions:
            LeadStatusChange.objects.get_or_create(
                lead=lead,
                new_status='conversion_requested',
                needs_approval=True,
                defaults={
                    'changed_by': lead.conversion_requested_by or request.user,
                    'old_status': 'qualified',
                    'notes': f'Auto-assigned conversion verification to {request.user.get_full_name()}',
                    'approval_by': request.user,
                    'approved': False  # EXPLICIT: Set to False
                }
            )
        
        # Refresh pending approvals after creating new records
        pending_approvals = LeadStatusChange.objects.filter(
            needs_approval=True,
            new_status='conversion_requested',
            approved=False  # FIXED: was approved__isnull=True
        ).select_related('lead', 'changed_by').order_by('-changed_at')
        
    else:  # ops_exec
        # Show only conversion requests specifically assigned to this ops_exec
        pending_approvals = LeadStatusChange.objects.filter(
            needs_approval=True,
            new_status='conversion_requested',
            approval_by=request.user,
            approved=False  # FIXED: was approved__isnull=True
        ).select_related('lead', 'changed_by').order_by('-changed_at')
    
    # Get all RMs for assignment dropdown
    available_rms = User.objects.filter(role='rm', is_active=True).order_by('first_name', 'last_name')
    
    # Get ops executives if current user is ops_team_lead
    available_ops_execs = []
    if request.user.role == 'ops_team_lead':
        available_ops_execs = User.objects.filter(role='ops_exec', is_active=True).order_by('first_name', 'last_name')
    
    # Get recent approvals/rejections for reference - FIXED QUERY
    recent_decisions = LeadStatusChange.objects.filter(
        approval_by=request.user,
        approved__in=[True, False],  # FIXED: Include both approved and rejected
        approved_at__isnull=False    # FIXED: Only get actually processed ones
    ).select_related('lead', 'changed_by').order_by('-approved_at')[:10]
    
    context = {
        'pending_approvals': pending_approvals,
        'available_rms': available_rms,
        'available_ops_execs': available_ops_execs,
        'recent_decisions': recent_decisions,
        'pending_count': pending_approvals.count(),
        'user_role': request.user.role,
    }
    
    return render(request, 'base/pending_approvals.html', context)


@login_required
@require_POST
def assign_verification_task(request, pk):
    """Allow ops_team_lead to assign verification task to ops_exec"""
    if request.user.role != 'ops_team_lead':
        messages.error(request, "You don't have permission to assign verification tasks.")
        return redirect('lead_detail', pk=pk)
    
    lead = get_object_or_404(Lead, pk=pk)
    ops_exec_id = request.POST.get('ops_exec_id')
    
    if not ops_exec_id:
        messages.error(request, "Please select an operations executive.")
        return redirect('ops_pending_approvals')
    
    try:
        ops_exec = User.objects.get(id=ops_exec_id, role='ops_exec', is_active=True)
    except User.DoesNotExist:
        messages.error(request, "Invalid operations executive selected.")
        return redirect('ops_pending_approvals')
    
    try:
        with transaction.atomic():
            # Find the existing approval record
            approval = LeadStatusChange.objects.filter(
                lead=lead,
                needs_approval=True,
                new_status='conversion_requested',
                approved=False  # FIXED: Use approved=False
            ).first()
            
            if approval:
                # Update existing approval record
                approval.approval_by = ops_exec
                approval.notes += f" | Verification task assigned to {ops_exec.get_full_name()} by {request.user.get_full_name()}"
                approval.save()
                messages.success(request, f"Verification task assigned to {ops_exec.get_full_name()}.")
            else:
                messages.error(request, "No pending approval found for this lead.")
            
    except Exception as e:
        messages.error(request, f"Error assigning verification task: {str(e)}")
    
    return redirect('ops_pending_approvals')

# Add this new view to handle self-assignment of verification
@login_required
@require_POST
def self_assign_verification(request, pk):
    """Allow ops_team_lead to self-assign verification task"""
    if request.user.role != 'ops_team_lead':
        messages.error(request, "You don't have permission to self-assign verification tasks.")
        return redirect('ops_pending_approvals')
    
    lead = get_object_or_404(Lead, pk=pk)
    
    try:
        with transaction.atomic():
            # Find or create approval record
            approval = LeadStatusChange.objects.filter(
                lead=lead,
                needs_approval=True,
                new_status='conversion_requested',
                approved=False  # FIXED: Use approved=False
            ).first()
            
            if approval:
                # Update existing record
                approval.approval_by = request.user
                approval.notes += f" | Self-assigned by {request.user.get_full_name()}"
                approval.save()
                messages.success(request, "Verification task self-assigned successfully.")
            else:
                messages.error(request, "No pending approval found for this lead.")
            
    except Exception as e:
        messages.error(request, f"Error self-assigning verification task: {str(e)}")
    
    return redirect('ops_pending_approvals')

@login_required
@require_POST
def approve_conversion(request, pk):
    """Only ops_team_lead can approve conversions after business verification"""
    if request.user.role not in ['ops_team_lead', 'ops_exec']:
        messages.error(request, "You don't have permission to verify business details.")
        return redirect('ops_pending_approvals')
    
    lead = get_object_or_404(Lead, pk=pk)
    approval_id = request.POST.get('approval_id')
    
    if not approval_id:
        messages.error(request, "No approval ID provided.")
        return redirect('ops_pending_approvals')
    
    try:
        approval = LeadStatusChange.objects.get(
            id=approval_id, 
            lead=lead, 
            needs_approval=True,
            approved=False  # FIXED: Use approved=False instead of approved__isnull=True
        )
    except LeadStatusChange.DoesNotExist:
        messages.error(request, "Approval request not found or already processed.")
        return redirect('ops_pending_approvals')
    
    # Get business verification data from form
    assigned_rm_id = request.POST.get('assigned_rm')
    business_verification_notes = request.POST.get('business_verification_notes', '')
    client_id = request.POST.get('client_id')
    
    if not assigned_rm_id:
        messages.error(request, "Please select an RM to assign the client to.")
        return redirect('ops_pending_approvals')
    
    try:
        assigned_rm = User.objects.get(id=assigned_rm_id, role__in=['rm', 'rm_head'])
    except User.DoesNotExist:
        messages.error(request, "Invalid RM selected for client assignment.")
        return redirect('ops_pending_approvals')
    
    try:
        with transaction.atomic():
            # Generate client ID if not provided
            if not client_id or not client_id.strip():
                client_id = generate_client_id()
            else:
                client_id = client_id.strip()
                # Validate uniqueness
                if Lead.objects.filter(client_id=client_id).exclude(pk=lead.pk).exists():
                    client_id = generate_client_id()
                    messages.warning(request, f"Provided client ID was not unique. Generated new ID: {client_id}")
            
            # Update the approval record
            approval.approved = True
            approval.approved_at = timezone.now()
            approval.approved_by = request.user
            approval.needs_approval = False
            approval.notes += f" | Business verified by {request.user.get_full_name()}: {business_verification_notes}"
            approval.save()
            
            # Create Client entry
            client = Client.objects.create(
                name=f"{client_id} - {lead.name}",
                contact_info=f"{lead.email or ''} | {lead.mobile or ''}".strip(' |'),
                user=assigned_rm,
                created_by=request.user,
                converted_from_lead=lead,
                business_verification_notes=business_verification_notes,
                original_rm=lead.assigned_to,
                conversion_approved_by=request.user,
                aum=0.00,
                sip_amount=0.00,
                demat_count=0,
            )
            
            # Convert the lead
            lead.converted = True
            lead.converted_at = timezone.now()
            lead.converted_by = request.user
            lead.status = 'converted'
            lead.business_verified_by = request.user
            lead.business_verification_notes = business_verification_notes
            lead.final_assigned_rm = assigned_rm
            lead.client_id = client_id
            lead.generated_client = client
            lead.save()
            
            # Create a new status change record for the conversion
            LeadStatusChange.objects.create(
                lead=lead,
                changed_by=request.user,
                old_status=approval.new_status,
                new_status='converted',
                notes=f"Business verified and converted to client {client_id} by {request.user.get_full_name()}. Assigned to {assigned_rm.get_full_name()}.",
                approved=True,
                approved_by=request.user,
                approved_at=timezone.now()
            )
            
            messages.success(request, 
                f"Lead successfully converted to client {client_id} and assigned to {assigned_rm.get_full_name()}."
            )
            
    except Exception as e:
        messages.error(request, f"Error converting lead: {str(e)}")
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Lead conversion failed for lead {pk}: {str(e)}")
    
    return redirect('ops_pending_approvals')

@login_required
@require_POST
def reject_conversion(request, pk):
    """Only ops_team_lead can reject conversions"""
    if request.user.role not in ['ops_team_lead', 'ops_exec']:
        messages.error(request, "You don't have permission to reject conversions.")
        return redirect('ops_pending_approvals')
    
    lead = get_object_or_404(Lead, pk=pk)
    approval_id = request.POST.get('approval_id')
    
    if not approval_id:
        messages.error(request, "No approval ID provided.")
        return redirect('ops_pending_approvals')
    
    try:
        approval = LeadStatusChange.objects.get(
            id=approval_id, 
            lead=lead, 
            needs_approval=True,
            approved=False  # FIXED: Use approved=False instead of approved__isnull=True
        )
    except LeadStatusChange.DoesNotExist:
        messages.error(request, "Approval request not found or already processed.")
        return redirect('ops_pending_approvals')
    
    rejection_reason = request.POST.get('rejection_reason', 'No reason provided')
    
    if not rejection_reason.strip():
        messages.error(request, "Please provide a rejection reason.")
        return redirect('ops_pending_approvals')
    
    try:
        with transaction.atomic():
            # Update the approval record
            approval.approved = False  # Keep it False but mark as processed
            approval.approved_at = timezone.now()
            approval.approved_by = request.user
            approval.needs_approval = False  # Mark as processed
            approval.notes = f"Business verification failed. Conversion rejected by {request.user.get_full_name()}. Reason: {rejection_reason}"
            approval.save()
            
            # Revert lead status to previous status
            lead.status = approval.old_status
            lead.business_verification_notes = f"Rejected: {rejection_reason}"
            lead.conversion_requested_at = None
            lead.conversion_requested_by = None
            lead.save()
            
            # Create a new status change record for the rejection
            LeadStatusChange.objects.create(
                lead=lead,
                changed_by=request.user,
                old_status='conversion_requested',
                new_status=approval.old_status,
                notes=f"Business verification failed. Conversion rejected by {request.user.get_full_name()}. Reason: {rejection_reason}",
                approved=False,
                approved_by=request.user,
                approved_at=timezone.now()
            )
            
            messages.warning(request, "Conversion request has been rejected after business verification.")
    except Exception as e:
        messages.error(request, f"Error rejecting conversion: {str(e)}")
    
    return redirect('ops_pending_approvals')

@login_required
def converted_leads_list(request):
    """View for showing all converted leads with their interactions - maintains lead history"""
    user = request.user
    
    # Get converted leads accessible to user
    if user.role in ['top_management', 'business_head', 'ops_team_lead']:
        converted_leads = Lead.objects.filter(converted=True)
    elif user.role == 'rm_head':
        accessible_users = user.get_accessible_users()
        # Show leads that were assigned to team members OR were converted and assigned to team members
        converted_leads = Lead.objects.filter(
            converted=True
        ).filter(
            Q(assigned_to__in=accessible_users) | 
            Q(final_assigned_rm__in=accessible_users)
        )
    elif user.role == 'rm':
        # Show leads that were assigned to the RM OR are now assigned to the RM as client
        converted_leads = Lead.objects.filter(
            converted=True
        ).filter(
            Q(assigned_to=user) | 
            Q(final_assigned_rm=user)
        )
    else:
        converted_leads = Lead.objects.none()
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        converted_leads = converted_leads.filter(
            Q(name__icontains=search_query) | 
            Q(email__icontains=search_query) |
            Q(mobile__icontains=search_query) |
            Q(lead_id__icontains=search_query) |
            Q(client_id__icontains=search_query)
        )
    
    # Prefetch related data for performance
    converted_leads = converted_leads.select_related(
        'assigned_to', 'final_assigned_rm', 'converted_by', 'business_verified_by', 'generated_client'
    ).prefetch_related(
        'interactions', 'product_discussions', 'status_changes'
    ).order_by('-converted_at')
    
    context = {
        'converted_leads': converted_leads,
        'search_query': search_query,
    }
    
    return render(request, 'base/converted_leads.html', context)


@login_required
@require_POST
def reassign_lead(request, pk):
    if not request.user.role in ['rm_head', 'business_head', 'top_management']:
        messages.error(request, "You don't have permission to reassign leads.")
        return redirect('lead_detail', pk=pk)
    
    lead = get_object_or_404(Lead, pk=pk)
    form = LeadReassignmentForm(request.user, request.POST)
    
    if form.is_valid():
        new_rm = form.cleaned_data['assigned_to']
        old_rm = lead.assigned_to
        
        # Record reassignment
        LeadStatusChange.objects.create(
            lead=lead,
            changed_by=request.user,
            old_status=f"assigned_to:{old_rm.id}",
            new_status=f"assigned_to:{new_rm.id}",
            notes=f"Lead reassigned from {old_rm.get_full_name()} to {new_rm.get_full_name()}"
        )
        
        # Update lead
        lead.assigned_to = new_rm
        lead.save()
        
        # TODO: Send notification to new RM
        
        messages.success(request, f"Lead successfully reassigned to {new_rm.get_full_name()}.")
    else:
        messages.error(request, "Error reassigning lead.")
    
    return redirect('lead_detail', pk=lead.pk)

# Helper functions
def generate_lead_id():
    """Generate a unique lead ID"""
    from datetime import datetime
    prefix = "LD"
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"{prefix}{timestamp}"

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from datetime import datetime
import random
import string

def generate_client_id():
    """Standalone function to generate unique client ID"""
    date_part = datetime.now().strftime("%Y%m%d")
    random_part = ''.join(random.choices(string.digits, k=4))
    
    client_id = f"CL{date_part}{random_part}"
    
    # Check uniqueness against Lead model's client_id field
    from .models import Lead, Client
    
    # Check against both models if Client has client_id field, otherwise just Lead
    try:
        while (Lead.objects.filter(client_id=client_id).exists() or 
               Client.objects.filter(client_id=client_id).exists()):
            random_part = ''.join(random.choices(string.digits, k=4))
            client_id = f"CL{date_part}{random_part}"
    except Exception:
        # Fallback: only check against Lead model
        while Lead.objects.filter(client_id=client_id).exists():
            random_part = ''.join(random.choices(string.digits, k=4))
            client_id = f"CL{date_part}{random_part}"
    
    return client_id

@login_required
@require_http_methods(["GET"])
def generate_client_id_view(request):
    """Generate a new client ID via AJAX"""
    client_id = generate_client_id()
    return JsonResponse({'client_id': client_id})

@login_required
def get_rm_details(request, rm_id):
    """AJAX endpoint to get RM details for assignment"""
    if request.user.role != 'ops_team_lead':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        rm = User.objects.get(id=rm_id, role='rm', is_active=True)
        
        # Get RM's current client count and other relevant metrics
        client_count = Client.objects.filter(user=rm).count()
        
        rm_details = {
            'id': rm.id,
            'name': rm.get_full_name() or rm.username,
            'email': rm.email,
            'current_clients': client_count,
            'department': getattr(rm, 'department', 'Not specified'),
            'phone': getattr(rm, 'phone', 'Not available'),
        }
        
        return JsonResponse({'success': True, 'rm_details': rm_details})
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'RM not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
@login_required
@require_POST
def delete_interaction(request, interaction_id):
    interaction = get_object_or_404(LeadInteraction, pk=interaction_id)
    lead_pk = interaction.lead.pk
    
    # Check permissions
    if not (request.user == interaction.interacted_by or 
            request.user.can_access_user_data(interaction.lead.assigned_to)):
        raise PermissionDenied("You don't have permission to delete this interaction.")
    
    interaction.delete()
    messages.success(request, "Interaction deleted successfully.")
    return redirect('leads:lead_detail', pk=lead_pk)

@login_required
@require_POST
def request_reassignment(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    # Check permissions
    if not request.user.can_access_user_data(lead.assigned_to):
        raise PermissionDenied("You don't have permission to request reassignment for this lead.")
    
    new_rm_id = request.POST.get('new_rm')
    try:
        new_rm = User.objects.get(pk=new_rm_id, role__in=['rm', 'rm_head'])
    except User.DoesNotExist:
        messages.error(request, "Invalid RM selected.")
        return redirect('leads:lead_detail', pk=pk)
    
    if lead.assigned_to == new_rm:
        messages.warning(request, "Lead is already assigned to this RM.")
        return redirect('leads:lead_detail', pk=pk)
    
    # Request reassignment
    lead.needs_reassignment_approval = True
    lead.reassignment_requested_to = request.user.get_line_manager()
    lead.save()
    
    # Create status change record
    LeadStatusChange.objects.create(
        lead=lead,
        changed_by=request.user,
        old_status=f"assigned_to:{lead.assigned_to.id}",
        new_status=f"assigned_to:{new_rm.id}",
        notes=f"Reassignment requested to {new_rm.get_full_name()}",
        needs_approval=True,
        approval_by=request.user.get_line_manager()
    )
    
    messages.success(request, "Reassignment request sent to your manager.")
    return redirect('leads:lead_detail', pk=pk)

@login_required
@require_POST
def approve_reassignment(request, pk):
    lead = get_object_or_404(Lead, pk=pk)
    
    # Check if current user is the approver
    if request.user != lead.reassignment_requested_to:
        raise PermissionDenied("You don't have permission to approve this reassignment.")
    
    # Get the latest reassignment request
    status_change = lead.status_changes.filter(needs_approval=True).latest('changed_at')
    
    # Extract new RM ID from status change
    new_rm_id = status_change.new_status.split(':')[-1]
    try:
        new_rm = User.objects.get(pk=new_rm_id)
    except User.DoesNotExist:
        messages.error(request, "The requested RM no longer exists.")
        return redirect('leads:lead_detail', pk=pk)
    
    # Approve the reassignment
    lead.assigned_to = new_rm
    lead.needs_reassignment_approval = False
    lead.reassignment_requested_to = None
    lead.save()
    
    # Update status change record
    status_change.approved = True
    status_change.approved_by = request.user
    status_change.approved_at = timezone.now()
    status_change.save()
    
    messages.success(request, f"Lead successfully reassigned to {new_rm.get_full_name()}.")
    return redirect('leads:lead_detail', pk=pk)

@login_required
def get_reference_clients(request):
    """AJAX view to get reference clients for autocomplete"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Not authenticated'}, status=403)
    
    query = request.GET.get('q', '')
    clients = Lead.objects.filter(
        Q(converted=True),
        Q(name__icontains=query) | Q(client_id__icontains=query)
    ).values('id', 'name', 'client_id')[:10]
    
    return JsonResponse(list(clients), safe=False)

@login_required
def get_accessible_users(request):
    """AJAX view to get accessible users for current user"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Not authenticated'}, status=403)
    
    query = request.GET.get('q', '')
    accessible_users = request.user.get_accessible_users()
    
    users = User.objects.filter(
        Q(id__in=[u.id for u in accessible_users]),
        Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(email__icontains=query)
    ).values('id', 'first_name', 'last_name', 'email', 'role')[:10]
    
    return JsonResponse(list(users), safe=False)
# Client Views with Hierarchy
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.db.models import Q, Sum, Count, IntegerField, Case, When
from django.db.models.functions import Coalesce
from django.forms import ModelForm
from django import forms
from .models import ClientProfile, User, MFUCANAccount
from .forms import ClientProfileForm  # Import the form

# Add this to your views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django import forms

# Form for converting client profile to full client
class ConvertToClientForm(forms.ModelForm):
    class Meta:
        model = Client
        fields = ['aum', 'sip_amount', 'demat_count', 'contact_info']
        widgets = {
            'aum': forms.NumberInput(attrs={
                'class': 'form-control',
                'placeholder': 'Enter AUM amount',
                'step': '0.01'
            }),
            'sip_amount': forms.NumberInput(attrs={
                'class': 'form-control',
                'placeholder': 'Enter SIP amount',
                'step': '0.01'
            }),
            'demat_count': forms.NumberInput(attrs={
                'class': 'form-control',
                'placeholder': 'Number of demat accounts',
                'min': '0'
            }),
            'contact_info': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Additional contact information'
            })
        }
        labels = {
            'aum': 'Assets Under Management (AUM)',
            'sip_amount': 'SIP Amount',
            'demat_count': 'Number of Demat Accounts',
            'contact_info': 'Additional Contact Information'
        }

@login_required
def convert_to_client(request, profile_id):
    """Convert a client profile to a full client"""
    profile = get_object_or_404(ClientProfile, id=profile_id)
    
    # Check if user has permission to convert
    if not request.user.role in ['rm', 'rm_head', 'business_head', 'top_management']:
        messages.error(request, "You don't have permission to convert client profiles.")
        return redirect('client_profile_list')
    
    # Check if profile is already converted
    if hasattr(profile, 'legacy_client') and profile.legacy_client:
        messages.warning(request, "This client profile is already converted to a full client.")
        return redirect('client_profile_detail', pk=profile.id)
    
    # Role-based access control
    if request.user.role == 'rm' and profile.mapped_rm != request.user:
        messages.error(request, "You can only convert your own client profiles.")
        return redirect('client_profile_list')
    
    if request.method == 'POST':
        form = ConvertToClientForm(request.POST)
        if form.is_valid():
            # Create new client instance
            client = form.save(commit=False)
            client.name = profile.client_full_name
            client.user = profile.mapped_rm
            client.client_profile = profile
            client.created_by = request.user
            client.save()
            
            messages.success(
                request, 
                f"Client profile '{profile.client_full_name}' has been successfully converted to a full client."
            )
            return redirect('client_profile_list')
    else:
        form = ConvertToClientForm()
    
    context = {
        'form': form,
        'profile': profile,
        'page_title': f'Convert {profile.client_full_name} to Full Client'
    }
    
    return render(request, 'base/convert_to_client.html', context)


from django.db.models import Count, Q, Max
# Update your client_profile_list view to include the convert permission and stats
@login_required
def client_profile_list(request):
    """Optimized client profile list with database-level calculations"""
    user = request.user
    
    # Get base queryset with optimized select_related and prefetch_related
    base_queryset = ClientProfile.objects.select_related(
        'mapped_rm', 'mapped_ops_exec', 'legacy_client'
    ).prefetch_related(
        'interactions'  # Only prefetch if needed for immediate display
    )
    
    # Role-based filtering with hierarchy support
    if user.role in ['top_management', 'business_head']:
        client_profiles = base_queryset.all()
    elif user.role == 'rm_head':
        # Get accessible users once and use in filter
        accessible_user_ids = list(user.get_accessible_users().values_list('id', flat=True))
        client_profiles = base_queryset.filter(
            Q(mapped_rm_id__in=accessible_user_ids) | Q(created_by=user)
        )
    elif user.role == 'rm':
        client_profiles = base_queryset.filter(mapped_rm=user)
    elif user.role in ['ops_team_lead', 'ops_exec']:
        client_profiles = base_queryset.all()
    else:
        client_profiles = base_queryset.none()

    # Apply filters early to reduce dataset
    search_query = request.GET.get('search')
    if search_query:
        client_profiles = client_profiles.filter(
            Q(client_full_name__icontains=search_query) | 
            Q(email__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(pan_number__icontains=search_query)
        )

    status_filter = request.GET.get('status')
    if status_filter:
        client_profiles = client_profiles.filter(status=status_filter)

    rm_filter = request.GET.get('rm')
    if rm_filter and user.role in ['rm_head', 'business_head', 'top_management']:
        client_profiles = client_profiles.filter(mapped_rm_id=rm_filter)

    interaction_filter = request.GET.get('interaction_filter')
    
    # Apply interaction filters at database level using annotations
    if interaction_filter:
        from django.utils import timezone
        today = timezone.now().date()
        month_start = today.replace(day=1)
        
        if interaction_filter == 'no_interactions':
            client_profiles = client_profiles.annotate(
                interaction_count=Count('interactions')
            ).filter(interaction_count=0)
        elif interaction_filter == 'recent_interactions':
            client_profiles = client_profiles.annotate(
                recent_interactions_count=Count(
                    'interactions',
                    filter=Q(interactions__interaction_date__gte=month_start)
                )
            ).filter(recent_interactions_count__gt=0)
        elif interaction_filter == 'overdue_followups':
            client_profiles = client_profiles.annotate(
                overdue_followups=Count(
                    'interactions',
                    filter=Q(
                        interactions__follow_up_date__lt=today,
                        interactions__follow_up_completed=False
                    )
                )
            ).filter(overdue_followups__gt=0)
        elif interaction_filter == 'due_today':
            client_profiles = client_profiles.annotate(
                due_today_followups=Count(
                    'interactions',
                    filter=Q(
                        interactions__follow_up_date=today,
                        interactions__follow_up_completed=False
                    )
                )
            ).filter(due_today_followups__gt=0)

    # Order by most commonly used field first
    client_profiles = client_profiles.order_by('-created_at')

    # Use pagination early to limit database queries
    paginator = Paginator(client_profiles, 25)
    page_number = request.GET.get('page')
    client_profiles_paginated = paginator.get_page(page_number)

    # Calculate statistics only for filtered results, not all data
    filtered_profiles = client_profiles
    stats = {
        'total_clients': filtered_profiles.count(),
        'active_clients': filtered_profiles.filter(status='active').count(),
        'muted_clients': filtered_profiles.filter(status='muted').count(),
        'converted_clients': filtered_profiles.filter(legacy_client__isnull=False).count(),
    }

    # Get RM list only if needed for filter dropdown
    rm_list = None
    if user.role in ['rm_head', 'business_head', 'top_management']:
        if user.role == 'rm_head':
            accessible_user_ids = list(user.get_accessible_users().values_list('id', flat=True))
            rm_list = User.objects.filter(
                id__in=accessible_user_ids,
                role='rm'
            ).only('id', 'first_name', 'last_name', 'username')
        else:
            rm_list = User.objects.filter(role='rm').only('id', 'first_name', 'last_name', 'username')

    context = {
        'client_profiles': client_profiles_paginated,
        'search_query': search_query,
        'status_filter': status_filter,
        'rm_filter': rm_filter,
        'interaction_filter': interaction_filter,
        'stats': stats,
        'rm_list': rm_list,
        'can_create': user.role in ['rm', 'rm_head', 'business_head', 'top_management'],
        'can_modify': user.can_modify_client_profile(),
        'can_convert_to_client': user.role in ['rm', 'rm_head', 'business_head', 'top_management'],
    }
    
    return render(request, 'base/client_profiles.html', context)


def add_interaction_data_safely(profiles):
    """
    Add interaction data to profiles using only Python calculations.
    This avoids any database annotation issues.
    """
    from django.utils import timezone
    
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    for profile in profiles:
        # Initialize all fields with safe defaults
        profile.interaction_count = 0
        profile.recent_interactions_count = 0
        profile.high_priority_count = 0
        profile.last_interaction_date = None
        profile.last_interaction_type = None
        profile.days_since_last_interaction = None
        profile.overdue_followups = 0
        profile.due_today_followups = 0
        profile.upcoming_followups = 0
        
        try:
            # Try to find the interaction relationship
            interactions = None
            
            # Try common relationship names
            for relation_name in ['interactions', 'clientinteraction_set', 'client_interactions']:
                if hasattr(profile, relation_name):
                    interactions = getattr(profile, relation_name).all()
                    break
            
            # If we still don't have interactions, try to find them dynamically
            if interactions is None:
                # Import here to avoid circular imports
                from django.apps import apps
                
                # Look for models that might be ClientInteraction
                for model in apps.get_models():
                    model_name = model.__name__.lower()
                    if 'interaction' in model_name and 'client' in model_name:
                        # Try to find the foreign key field
                        for field in model._meta.fields:
                            if (hasattr(field, 'related_model') and 
                                field.related_model == profile.__class__):
                                interactions = model.objects.filter(**{field.name: profile})
                                break
                        if interactions is not None:
                            break
            
            if interactions is None:
                continue  # Skip this profile if we can't find interactions
            
            # Convert to list to avoid repeated database hits
            interaction_list = list(interactions)
            
            # Basic count
            profile.interaction_count = len(interaction_list)
            
            if profile.interaction_count == 0:
                continue
            
            # Process each interaction
            recent_count = 0
            high_priority_count = 0
            last_interaction = None
            last_date = None
            overdue_count = 0
            due_today_count = 0
            upcoming_count = 0
            
            for interaction in interaction_list:
                # Get interaction date
                interaction_date = None
                for date_field in ['interaction_date', 'date', 'created_at']:
                    if hasattr(interaction, date_field):
                        date_value = getattr(interaction, date_field)
                        if date_value:
                            if hasattr(date_value, 'date'):  # datetime object
                                interaction_date = date_value.date()
                            else:  # date object
                                interaction_date = date_value
                            break
                
                if interaction_date:
                    # Check if recent (this month)
                    if interaction_date >= month_start:
                        recent_count += 1
                    
                    # Check if this is the latest interaction
                    if last_date is None or interaction_date > last_date:
                        last_date = interaction_date
                        last_interaction = interaction
                
                # Check priority
                for priority_field in ['priority', 'importance', 'urgency']:
                    if hasattr(interaction, priority_field):
                        priority_value = getattr(interaction, priority_field)
                        if priority_value and str(priority_value).lower() == 'high':
                            high_priority_count += 1
                        break
                
                # Check follow-ups
                follow_up_date = None
                for followup_field in ['follow_up_date', 'followup_date', 'next_follow_up']:
                    if hasattr(interaction, followup_field):
                        follow_up_date = getattr(interaction, followup_field)
                        break
                
                if follow_up_date:
                    # Check if completed
                    is_completed = False
                    for completed_field in ['follow_up_completed', 'completed', 'is_completed']:
                        if hasattr(interaction, completed_field):
                            is_completed = bool(getattr(interaction, completed_field))
                            break
                    
                    if not is_completed:
                        if follow_up_date < today:
                            overdue_count += 1
                        elif follow_up_date == today:
                            due_today_count += 1
                        elif follow_up_date > today:
                            upcoming_count += 1
            
            # Set calculated values
            profile.recent_interactions_count = recent_count
            profile.high_priority_count = high_priority_count
            profile.overdue_followups = overdue_count
            profile.due_today_followups = due_today_count
            profile.upcoming_followups = upcoming_count
            
            # Set last interaction data
            if last_interaction and last_date:
                profile.last_interaction_date = last_date
                profile.days_since_last_interaction = (today - last_date).days
                
                # Get interaction type
                for type_field in ['interaction_type', 'type', 'category']:
                    if hasattr(last_interaction, type_field):
                        profile.last_interaction_type = getattr(last_interaction, type_field)
                        break
                
                if not profile.last_interaction_type:
                    profile.last_interaction_type = 'Contact'
        
        except Exception as e:
            print(f"Error processing interactions for profile {profile.id}: {str(e)}")
            # Keep the default values we set at the beginning
            continue
    
    return profiles

@login_required
def client_profile_create(request):
    """Create a new client profile"""
    if not request.user.role in ['rm', 'rm_head', 'business_head', 'top_management']:
        messages.error(request, "You don't have permission to create client profiles.")
        return redirect('client_profile_list')
        
    if request.method == 'POST':
        form = ClientProfileForm(request.POST, current_user=request.user)
        if form.is_valid():
            client_profile = form.save(commit=False)
            client_profile.created_by = request.user
            
            # Auto-assign RM if user is an RM and no RM is mapped
            if request.user.role == 'rm' and not client_profile.mapped_rm:
                client_profile.mapped_rm = request.user
                
            client_profile.save()
            messages.success(request, "Client profile created successfully.")
            return redirect('client_profile_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ClientProfileForm(current_user=request.user)
    
    return render(request, 'base/client_profile_form.html', {
        'form': form, 
        'action': 'Create',
        'title': 'Create Client Profile'
    })

@login_required
def client_profile_detail(request, pk):
    """View client profile details"""
    client_profile = get_object_or_404(ClientProfile, pk=pk)
    
    # Check permissions
    if not request.user.can_view_client_profile():
        raise PermissionDenied("You don't have permission to view client profiles.")
    
    # Check if user can access this specific client's data
    if not request.user.can_access_user_data(client_profile.mapped_rm):
        raise PermissionDenied("You don't have permission to view this client's profile.")
    
    # Get related accounts using the correct related names from your models
    mfu_accounts = client_profile.mfu_accounts.all()
    
    # Get modification history if available
    try:
        modifications = client_profile.modifications.select_related('requested_by', 'approved_by').order_by('-requested_at')[:10]
    except AttributeError:
        modifications = []
    
    # Get client interactions - recent 10 interactions
    try:
        interactions = client_profile.interactions.select_related('created_by').order_by('-interaction_date')[:10]
    except AttributeError:
        interactions = []
    
    # Calculate interaction counts for the stats section
    try:
        total_interactions = client_profile.interactions.count()
        follow_up_interactions = client_profile.interactions.filter(follow_up_required=True).count()
        # Additional stats that might be useful
        recent_interactions = client_profile.interactions.filter(
            interaction_date__gte=timezone.now() - timedelta(days=30)
        ).count()
        urgent_interactions = client_profile.interactions.filter(
            priority='urgent'
        ).count()
        high_priority_interactions = client_profile.interactions.filter(
            priority='high'
        ).count()
    except AttributeError:
        total_interactions = 0
        follow_up_interactions = 0
        recent_interactions = 0
        urgent_interactions = 0
        high_priority_interactions = 0
    
    # Check if user can add interactions (only assigned RM)
    can_add_interaction = (
        request.user.role == 'rm' and 
        client_profile.mapped_rm == request.user
    )
    
    # Check if user can modify this client's profile
    can_modify = request.user.can_modify_client_profile()
    
    # Check if user can mute notifications for this client
    can_mute = request.user.role in ['business_head', 'top_management', 'ops_team_lead']
    
    context = {
        'client_profile': client_profile,
        'mfu_accounts': mfu_accounts,
        'modifications': modifications,
        'interactions': interactions,
        'can_modify': can_modify,
        'can_mute': can_mute,
        'can_add_interaction': can_add_interaction,
        # Interaction statistics - these will be available in the template
        'total_interactions': total_interactions,
        'follow_up_interactions': follow_up_interactions,
        'recent_interactions': recent_interactions,
        'urgent_interactions': urgent_interactions,
        'high_priority_interactions': high_priority_interactions,
    }
    
    return render(request, 'base/client_profile_detail.html', context)

@login_required
def client_profile_update(request, pk):
    """Update client profile"""
    client_profile = get_object_or_404(ClientProfile, pk=pk)
    
    # Check permissions
    if not request.user.can_modify_client_profile():
        messages.error(request, "You don't have permission to modify client profiles.")
        return redirect('client_profile_detail', pk=pk)
    
    # Check if user can access this specific client's data
    if not request.user.can_access_user_data(client_profile.mapped_rm):
        raise PermissionDenied("You don't have permission to edit this client's profile.")
    
    if request.method == 'POST':
        # FIXED: Use current_user instead of user
        form = ClientProfileForm(request.POST, instance=client_profile, current_user=request.user)
        if form.is_valid():
            # For critical fields, create modification request instead of direct update
            critical_fields = ['pan_number', 'client_full_name', 'date_of_birth']
            has_critical_changes = any(
                form.cleaned_data.get(field) != getattr(client_profile, field) 
                for field in critical_fields if field in form.cleaned_data
            )
            
            if has_critical_changes and request.user.role not in ['top_management']:
                # Create modification request
                try:
                    from .models import ClientProfileModification
                    import json
                    
                    modification_data = {}
                    for field in critical_fields:
                        if field in form.cleaned_data and form.cleaned_data.get(field) != getattr(client_profile, field):
                            modification_data[field] = form.cleaned_data.get(field)
                    
                    ClientProfileModification.objects.create(
                        client=client_profile,
                        requested_by=request.user,
                        modification_data=modification_data,
                        reason=request.POST.get('modification_reason', 'Profile update'),
                        requires_top_management=True
                    )
                    
                    messages.info(request, "Critical field changes require approval. Modification request submitted.")
                    return redirect('client_profile_detail', pk=pk)
                except ImportError:
                    # If ClientProfileModification model doesn't exist, proceed with direct update
                    form.save()
                    messages.success(request, "Client profile updated successfully.")
                    return redirect('client_profile_detail', pk=pk)
            else:
                form.save()
                messages.success(request, "Client profile updated successfully.")
                return redirect('client_profile_detail', pk=pk)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        # FIXED: Use current_user instead of user
        form = ClientProfileForm(instance=client_profile, current_user=request.user)
    
    return render(request, 'base/client_profile_form.html', {
        'form': form, 
        'action': 'Update',
        'title': 'Update Client Profile',
        'client_profile': client_profile
    })

@login_required
def client_profile_mute(request, pk):
    """Mute/Unmute client profile"""
    client_profile = get_object_or_404(ClientProfile, pk=pk)
    
    # Check permissions
    if not request.user.role in ['business_head', 'top_management', 'ops_team_lead']:
        messages.error(request, "You don't have permission to mute/unmute clients.")
        return redirect('client_profile_detail', pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'mute' and client_profile.status == 'active':
            reason = request.POST.get('mute_reason')
            if not reason:
                messages.error(request, "Mute reason is required.")
                return redirect('client_profile_detail', pk=pk)
            
            try:
                client_profile.mute_client(reason, request.user)
                messages.success(request, "Client has been muted successfully.")
            except AttributeError:
                # If mute_client method doesn't exist, update status directly
                client_profile.status = 'muted'
                client_profile.save()
                messages.success(request, "Client has been muted successfully.")
            
        elif action == 'unmute' and client_profile.status == 'muted':
            try:
                client_profile.unmute_client(request.user)
                messages.success(request, "Client has been unmuted successfully.")
            except AttributeError:
                # If unmute_client method doesn't exist, update status directly
                client_profile.status = 'active'
                client_profile.save()
                messages.success(request, "Client has been unmuted successfully.")
        
        return redirect('client_profile_detail', pk=pk)
    
    return render(request, 'base/client_profile_mute.html', {
        'client_profile': client_profile
    })

@login_required
def client_profile_delete(request, pk):
    """Delete client profile (restricted to top management)"""
    client_profile = get_object_or_404(ClientProfile, pk=pk)
    
    # Only top management can delete client profiles
    if request.user.role != 'top_management':
        messages.error(request, "You don't have permission to delete client profiles.")
        return redirect('client_profile_detail', pk=pk)
    
    if request.method == 'POST':
        client_name = client_profile.client_full_name
        client_profile.delete()
        messages.success(request, f"Client profile for {client_name} has been deleted.")
        return redirect('client_profile_list')
    
    return render(request, 'base/client_profile_confirm_delete.html', {
        'client_profile': client_profile
    })

# NEW: Client Interaction Views
@login_required
def client_interaction_create(request, profile_id):
    """Create a new client interaction - only assigned RM can create"""
    client_profile = get_object_or_404(ClientProfile, id=profile_id)
    
    # Only assigned RM can create interactions
    if request.user.role != 'rm' or client_profile.mapped_rm != request.user:
        messages.error(request, "Only the assigned RM can add client interactions.")
        return redirect('client_profile_detail', pk=profile_id)
    
    if request.method == 'POST':
        try:
            from .models import ClientInteraction
            from .forms import ClientInteractionForm
            
            form = ClientInteractionForm(request.POST)
            if form.is_valid():
                interaction = form.save(commit=False)
                interaction.client_profile = client_profile
                interaction.created_by = request.user
                interaction.save()
                
                messages.success(request, "Client interaction recorded successfully.")
                return redirect('client_profile_detail', pk=profile_id)
            else:
                messages.error(request, "Please correct the errors below.")
        except ImportError:
            messages.error(request, "Client interaction feature is not available.")
            return redirect('client_profile_detail', pk=profile_id)
    else:
        try:
            from .forms import ClientInteractionForm
            form = ClientInteractionForm()
        except ImportError:
            messages.error(request, "Client interaction feature is not available.")
            return redirect('client_profile_detail', pk=profile_id)
    
    context = {
        'form': form,
        'client_profile': client_profile,
        'title': f'Add Interaction - {client_profile.client_full_name}'
    }
    
    return render(request, 'base/client_interaction_form.html', context)

@login_required
def client_interaction_list(request, profile_id):
    """List all interactions for a client profile"""
    client_profile = get_object_or_404(ClientProfile, id=profile_id)
    
    # Check if user can access this client's data
    if not request.user.can_access_user_data(client_profile.mapped_rm):
        raise PermissionDenied("You don't have permission to view this client's interactions.")
    
    try:
        # Get all interactions for this client
        interactions = client_profile.interactions.select_related('created_by').order_by('-interaction_date')
        
        # Filter by interaction type if specified
        interaction_type = request.GET.get('type')
        if interaction_type:
            interactions = interactions.filter(interaction_type=interaction_type)
        
        # Search in interaction notes
        search_query = request.GET.get('search')
        if search_query:
            interactions = interactions.filter(
                Q(notes__icontains=search_query) |
                Q(interaction_type__icontains=search_query)
            )
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(interactions, 20)  # Show 20 interactions per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Check if user can add interactions (only assigned RM)
        can_add_interaction = (
            request.user.role == 'rm' and 
            client_profile.mapped_rm == request.user
        )
        
        context = {
            'client_profile': client_profile,
            'interactions': page_obj,
            'search_query': search_query,
            'interaction_type': interaction_type,
            'can_add_interaction': can_add_interaction,
        }
        
        return render(request, 'base/client_interaction_list.html', context)
        
    except AttributeError:
        messages.error(request, "Client interaction feature is not available.")
        return redirect('client_profile_detail', pk=profile_id)

@login_required
def client_interaction_detail(request, client_pk, interaction_pk):
    """View for displaying detailed information about a specific client interaction"""
    
    # Get the client profile and interaction
    client_profile = get_object_or_404(ClientProfile, pk=client_pk)
    interaction = get_object_or_404(ClientInteraction, pk=interaction_pk, client_profile=client_profile)
    
    # Check permissions
    if not request.user.can_view_client_profile():
        raise PermissionDenied("You don't have permission to view client profiles.")
    
    # Check if user can access this specific client's data
    if not request.user.can_access_user_data(client_profile.mapped_rm):
        raise PermissionDenied("You don't have permission to view this client's profile.")
    
    # Check if the user can edit this interaction
    # Users can edit interactions within 24 hours of creation and only if they created it
    can_edit = False
    if interaction.created_by == request.user:
        time_diff = timezone.now() - interaction.created_at
        if time_diff.total_seconds() < 24 * 60 * 60:  # 24 hours in seconds
            can_edit = True
    
    # Calculate interaction statistics for the sidebar
    try:
        total_interactions = client_profile.interactions.count()
        follow_up_interactions = client_profile.interactions.filter(follow_up_required=True).count()
        # Additional stats that might be useful
        recent_interactions = client_profile.interactions.filter(
            interaction_date__gte=timezone.now() - timedelta(days=30)
        ).count()
        urgent_interactions = client_profile.interactions.filter(
            priority='urgent'
        ).count()
        high_priority_interactions = client_profile.interactions.filter(
            priority='high'
        ).count()
        
        # Get recent interactions list (excluding current one)
        recent_interactions_list = client_profile.interactions.exclude(
            id=interaction.id
        ).select_related('created_by').order_by('-interaction_date')[:5]
        
    except AttributeError:
        # Handle case where the interactions relationship doesn't exist
        total_interactions = 0
        follow_up_interactions = 0
        recent_interactions = 0
        urgent_interactions = 0
        high_priority_interactions = 0
        recent_interactions_list = []
    
    context = {
        'client_profile': client_profile,
        'interaction': interaction,
        'can_edit': can_edit,
        # Pass the calculated statistics to the template
        'total_interactions': total_interactions,
        'follow_up_interactions': follow_up_interactions,
        'recent_interactions': recent_interactions,
        'urgent_interactions': urgent_interactions,
        'high_priority_interactions': high_priority_interactions,
        'recent_interactions_list': recent_interactions_list,
    }
    
    return render(request, 'base/client_interaction_detail.html', context)

@login_required
def client_interaction_update(request, profile_id, interaction_id):
    """Update a client interaction - only creator can update within 24 hours"""
    client_profile = get_object_or_404(ClientProfile, id=profile_id)
    
    try:
        interaction = get_object_or_404(client_profile.interactions, id=interaction_id)
        
        # Check if user can edit this interaction
        from django.utils import timezone
        from datetime import timedelta
        
        if request.user != interaction.created_by:
            messages.error(request, "You can only edit your own interactions.")
            return redirect('client_interaction_detail', profile_id=profile_id, interaction_id=interaction_id)
        
        if timezone.now() - interaction.created_at > timedelta(hours=24):
            messages.error(request, "Interactions can only be edited within 24 hours of creation.")
            return redirect('client_interaction_detail', profile_id=profile_id, interaction_id=interaction_id)
        
        if request.method == 'POST':
            from .forms import ClientInteractionForm
            form = ClientInteractionForm(request.POST, instance=interaction)
            if form.is_valid():
                form.save()
                messages.success(request, "Interaction updated successfully.")
                return redirect('client_interaction_detail', profile_id=profile_id, interaction_id=interaction_id)
            else:
                messages.error(request, "Please correct the errors below.")
        else:
            from .forms import ClientInteractionForm
            form = ClientInteractionForm(instance=interaction)
        
        context = {
            'form': form,
            'client_profile': client_profile,
            'interaction': interaction,
            'title': f'Edit Interaction - {client_profile.client_full_name}'
        }
        
        return render(request, 'base/client_interaction_form.html', context)
        
    except (AttributeError, ImportError):
        messages.error(request, "Client interaction feature is not available.")
        return redirect('client_profile_detail', pk=profile_id)

@login_required
def client_interaction_delete(request, profile_id, interaction_id):
    """Delete a client interaction - only creator can delete within 24 hours"""
    client_profile = get_object_or_404(ClientProfile, id=profile_id)
    
    try:
        interaction = get_object_or_404(client_profile.interactions, id=interaction_id)
        
        # Check if user can delete this interaction
        from django.utils import timezone
        from datetime import timedelta
        
        if request.user != interaction.created_by:
            messages.error(request, "You can only delete your own interactions.")
            return redirect('client_interaction_detail', profile_id=profile_id, interaction_id=interaction_id)
        
        if timezone.now() - interaction.created_at > timedelta(hours=24):
            messages.error(request, "Interactions can only be deleted within 24 hours of creation.")
            return redirect('client_interaction_detail', profile_id=profile_id, interaction_id=interaction_id)
        
        if request.method == 'POST':
            interaction.delete()
            messages.success(request, "Interaction deleted successfully.")
            return redirect('client_profile_detail', pk=profile_id)
        
        context = {
            'client_profile': client_profile,
            'interaction': interaction,
        }
        
        return render(request, 'base/client_interaction_confirm_delete.html', context)
        
    except AttributeError:
        messages.error(request, "Client interaction feature is not available.")
        return redirect('client_profile_detail', pk=profile_id)

@login_required
def modification_requests(request):
    """View pending modification requests"""
    if not request.user.role in ['top_management', 'business_head']:
        messages.error(request, "You don't have permission to view modification requests.")
        return redirect('client_profile_list')
    
    try:
        from .models import ClientProfileModification
        
        # Get pending modifications
        pending_modifications = ClientProfileModification.objects.filter(
            status='pending'
        ).select_related('client', 'requested_by').order_by('-requested_at')
        
        # Get recent approved/rejected modifications
        recent_modifications = ClientProfileModification.objects.filter(
            status__in=['approved', 'rejected']
        ).select_related('client', 'requested_by', 'approved_by').order_by('-approved_at')[:20]
        
        context = {
            'pending_modifications': pending_modifications,
            'recent_modifications': recent_modifications,
        }
        
        return render(request, 'base/client_modification_requests.html', context)
    except ImportError:
        messages.error(request, "Modification requests feature is not available.")
        return redirect('client_profile_list')

@login_required
def approve_modification(request, pk):
    """Approve or reject modification request"""
    if not request.user.role in ['top_management', 'business_head']:
        messages.error(request, "You don't have permission to approve modifications.")
        return redirect('client_profile_list')
    
    try:
        from .models import ClientProfileModification
        modification = get_object_or_404(ClientProfileModification, pk=pk)
        
        if request.method == 'POST':
            action = request.POST.get('action')
            
            if action == 'approve':
                if modification.approve(request.user):
                    messages.success(request, "Modification approved and applied successfully.")
                else:
                    messages.error(request, "Failed to approve modification.")
            elif action == 'reject':
                if modification.reject(request.user):
                    messages.success(request, "Modification rejected.")
                else:
                    messages.error(request, "Failed to reject modification.")
            
            return redirect('modification_requests')
        
        return render(request, 'base/client_approve_modification.html', {
            'modification': modification
        })
    except ImportError:
        messages.error(request, "Modification requests feature is not available.")
        return redirect('client_profile_list')

# Legacy Client Views (for backward compatibility)
@login_required
def client_list(request):
    """Legacy client list view - redirects to new client profile list"""
    messages.info(request, "Redirected to new Client Profile system.")
    return redirect('client_profile_list')

@login_required
def client_create(request):
    """Legacy client create view - redirects to new client profile create"""
    messages.info(request, "Redirected to new Client Profile system.")
    return redirect('client_profile_create')

@login_required
def client_update(request, pk):
    """Legacy client update view"""
    # Try to find corresponding client profile
    try:
        from .models import Client
        client = get_object_or_404(Client, pk=pk)
        if hasattr(client, 'client_profile') and client.client_profile:
            return redirect('client_profile_update', pk=client.client_profile.pk)
        else:
            messages.error(request, "No corresponding client profile found.")
            return redirect('client_profile_list')
    except ImportError:
        messages.error(request, "Legacy client system is not available.")
        return redirect('client_profile_list')

@login_required
def client_delete(request, pk):
    """Legacy client delete view"""
    try:
        from .models import Client
        client = get_object_or_404(Client, pk=pk)
        if hasattr(client, 'client_profile') and client.client_profile:
            return redirect('client_profile_delete', pk=client.client_profile.pk)
        else:
            # Handle legacy client without profile
            if request.user.role != 'top_management':
                messages.error(request, "You don't have permission to delete clients.")
                return redirect('client_profile_list')
            
            if request.method == 'POST':
                client_name = client.name
                client.delete()
                messages.success(request, f"Legacy client {client_name} has been deleted.")
                return redirect('client_profile_list')
            
            return render(request, 'base/client_confirm_delete.html', {'client': client})
    except ImportError:
        messages.error(request, "Legacy client system is not available.")
        return redirect('client_profile_list')

# Task Views with Hierarchy
from django.views.decorators.csrf import csrf_protect
from django.db.models import Q

@login_required
def task_list(request):
    user = request.user
    
    # Get tasks accessible to user (assigned to them or users they can access)
    accessible_tasks = get_user_accessible_data(user, Task, 'assigned_to')
    
    # Also include tasks created by the current user (regardless of assignee)
    created_tasks = Task.objects.filter(assigned_by=user)
    
    # Combine both querysets and remove duplicates
    tasks = (accessible_tasks | created_tasks).distinct()
    
    # Add filtering options
    status_filter = request.GET.get('status')
    if status_filter == 'completed':
        tasks = tasks.filter(completed=True)
    elif status_filter == 'pending':
        tasks = tasks.filter(completed=False)
    elif status_filter == 'overdue':
        tasks = tasks.filter(completed=False, due_date__lt=timezone.now())
    
    # Add search functionality
    search_query = request.GET.get('search')
    if search_query:
        tasks = tasks.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query)
        )

    context = {
        'tasks': tasks.order_by('-created_at'),
        'current_status': status_filter,
        'search_query': search_query,
        'now': timezone.now(),
    }
    
    return render(request, 'base/tasks.html', context)

@login_required
def task_create(request):
    # Only RM Heads and above can assign tasks to others
    if not request.user.role in ['rm_head', 'business_head', 'top_management', 'ops_team_lead']:
        messages.error(request, "You don't have permission to create tasks for others.")
        return redirect('task_list')
        
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.assigned_by = request.user
            task.save()
            messages.success(request, "Task created successfully.")
            return redirect('task_list')
    else:
        form = TaskForm()
        
        # Limit assignee choices based on user's role and hierarchy
        if request.user.role == 'rm_head':
            accessible_users = request.user.get_accessible_users()
            form.fields['assigned_to'].queryset = User.objects.filter(
                id__in=[u.id for u in accessible_users]
            )
        elif request.user.role == 'ops_team_lead':
            # Can only assign to ops_exec users
            form.fields['assigned_to'].queryset = User.objects.filter(role='ops_exec')
        elif request.user.role in ['business_head', 'top_management']:
            # Can assign to anyone below them in hierarchy
            form.fields['assigned_to'].queryset = User.objects.exclude(role='top_management')
    
    return render(request, 'base/task_form.html', {'form': form, 'action': 'Create'})

@login_required
def task_update(request, pk):
    task = get_object_or_404(Task, pk=pk)
    
    # Check permissions - can edit if assigned to them or they can manage the assignee
    if not (task.assigned_to == request.user or request.user.can_access_user_data(task.assigned_to)):
        raise PermissionDenied("You don't have permission to edit this task.")
    
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            messages.success(request, "Task updated successfully.")
            return redirect('task_list')
    else:
        form = TaskForm(instance=task)
        
        # Limit assignee choices based on user's role
        if request.user.role == 'rm_head':
            accessible_users = request.user.get_accessible_users()
            form.fields['assigned_to'].queryset = User.objects.filter(
                id__in=[u.id for u in accessible_users]
            )
        elif request.user.role == 'rm':
            # RMs can only see tasks assigned to them, so disable changing assignee
            form.fields['assigned_to'].widget.attrs['readonly'] = True
    
    return render(request, 'base/task_form.html', {'form': form, 'action': 'Update', 'task': task})

@login_required
def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    
    # Check permissions
    if not (task.assigned_to == request.user or request.user.can_access_user_data(task.assigned_to)):
        raise PermissionDenied("You don't have permission to delete this task.")
    
    if request.method == 'POST':
        task.delete()
        messages.success(request, "Task deleted successfully.")
        return redirect('task_list')
    
    return render(request, 'base/task_confirm_delete.html', {'task': task})

# Enhanced version of your existing task_toggle_complete function
@login_required
@require_POST
@csrf_protect
def task_toggle_complete(request, pk):
    """AJAX endpoint to toggle task completion"""
    try:
        task = get_object_or_404(Task, pk=pk)
        
        # Check permissions
        if not (task.assigned_to == request.user or request.user.can_access_user_data(task.assigned_to)):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to modify this task.'
            }, status=403)
        
        # Toggle completion status
        task.completed = not task.completed
        if task.completed:
            task.completed_at = timezone.now()
        else:
            task.completed_at = None
        
        task.save()
        
        # Return success response
        return JsonResponse({
            'success': True,
            'completed': task.completed,
            'completed_at': task.completed_at.isoformat() if task.completed_at else None,
            'message': 'Task marked as completed!' if task.completed else 'Task marked as incomplete!',
            'task_id': task.id,
            'task_title': task.title
        })
        
    except Task.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Task not found.'
        }, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error toggling task completion {pk}: {str(e)}")
        
        return JsonResponse({
            'success': False,
            'error': 'An unexpected error occurred. Please try again.'
        }, status=500)

# New dedicated function for marking as done (non-toggle)
@login_required
@require_POST
@csrf_protect
def mark_task_done(request, pk):
    """AJAX endpoint to mark task as completed (one-way action)"""
    try:
        task = get_object_or_404(Task, pk=pk)
        
        # Check permissions
        if not (task.assigned_to == request.user or request.user.can_access_user_data(task.assigned_to)):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to modify this task.'
            }, status=403)
        
        # Check if already completed
        if task.completed:
            return JsonResponse({
                'success': False,
                'error': 'Task is already completed.'
            })
        
        # Mark as completed
        task.completed = True
        task.completed_at = timezone.now()
        task.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Task "{task.title}" has been marked as completed!',
            'task_id': task.id,
            'completed_at': task.completed_at.isoformat(),
            'assigned_to': task.assigned_to.get_full_name()
        })
        
    except Task.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Task not found.'
        }, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error marking task as done {pk}: {str(e)}")
        
        return JsonResponse({
            'success': False,
            'error': 'An unexpected error occurred. Please try again.'
        }, status=500)

# Task reopen function (opposite of mark as done)
@login_required
@require_POST
@csrf_protect
def reopen_task(request, pk):
    """Reopen a completed task"""
    try:
        task = get_object_or_404(Task, pk=pk)
        
        # Check permissions
        if not (task.assigned_to == request.user or request.user.can_access_user_data(task.assigned_to)):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to modify this task.'
            }, status=403)
        
        # Check if task is completed
        if not task.completed:
            return JsonResponse({
                'success': False,
                'error': 'Task is not completed.'
            })
        
        # Reopen task
        task.completed = False
        task.completed_at = None
        task.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Task "{task.title}" has been reopened.',
            'task_id': task.id
        })
        
    except Task.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Task not found.'
        }, status=404)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error reopening task {pk}: {str(e)}")
        
        return JsonResponse({
            'success': False,
            'error': 'An unexpected error occurred.'
        }, status=500)

# Bulk operations for multiple tasks
@login_required
@require_POST
@csrf_protect
def bulk_mark_tasks_done(request):
    """Mark multiple tasks as completed"""
    try:
        data = json.loads(request.body)
        task_ids = data.get('task_ids', [])
        
        if not task_ids:
            return JsonResponse({
                'success': False,
                'error': 'No task IDs provided.'
            })
        
        # Get tasks and check permissions
        tasks = Task.objects.filter(id__in=task_ids)
        accessible_tasks = []
        
        for task in tasks:
            if (task.assigned_to == request.user or 
                request.user.can_access_user_data(task.assigned_to)):
                accessible_tasks.append(task)
        
        if not accessible_tasks:
            return JsonResponse({
                'success': False,
                'error': 'No accessible tasks found or permission denied.'
            })
        
        # Update tasks
        updated_count = 0
        for task in accessible_tasks:
            if not task.completed:
                task.completed = True
                task.completed_at = timezone.now()
                task.save()
                updated_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} task(s) marked as completed.',
            'updated_count': updated_count,
            'total_requested': len(task_ids)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data.'
        })
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in bulk mark tasks done: {str(e)}")
        
        return JsonResponse({
            'success': False,
            'error': 'An unexpected error occurred.'
        }, status=500)
        
        # Task statistics view
@login_required
def task_stats(request):
    """Get task statistics for the current user"""
    user = request.user
    tasks = get_user_accessible_data(user, Task, 'assigned_to')
    
    # Calculate statistics
    total_tasks = tasks.count()
    completed_tasks = tasks.filter(completed=True).count()
    pending_tasks = tasks.filter(completed=False).count()
    overdue_tasks = tasks.filter(
        completed=False, 
        due_date__lt=timezone.now()
    ).count()
    
    # Tasks due in next 7 days
    next_week = timezone.now() + timezone.timedelta(days=7)
    due_soon = tasks.filter(
        completed=False,
        due_date__range=[timezone.now(), next_week]
    ).count()
    
    # Priority breakdown
    priority_stats = {}
    for priority, _ in Task.PRIORITY_CHOICES:
        priority_stats[priority] = tasks.filter(
            priority=priority, 
            completed=False
        ).count()
    
    stats = {
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
        'pending_tasks': pending_tasks,
        'overdue_tasks': overdue_tasks,
        'due_soon': due_soon,
        'completion_rate': round((completed_tasks / total_tasks * 100), 1) if total_tasks > 0 else 0,
        'priority_breakdown': priority_stats
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse(stats)
    
    return render(request, 'base/task_stats.html', {'stats': stats})

# Utility function to check if task is overdue
def is_task_overdue(task):
    """Check if a task is overdue"""
    if task.completed or not task.due_date:
        return False
    return task.due_date < timezone.now()

# Helper function to get task status
def get_task_status(task):
    """Get human-readable task status"""
    if task.completed:
        return 'completed'
    elif is_task_overdue(task):
        return 'overdue'
    else:
        return 'pending'
    
# Solution 3: Quick fix - modify your service_request_list view
# Remove or comment out the problematic line

@login_required
def service_request_list(request):
    """Enhanced service request list view with proper hierarchy-based access"""
    user = request.user
    
    # Base queryset with optimized queries
    base_queryset = ServiceRequest.objects.select_related(
        'client', 'raised_by', 'assigned_to', 'current_owner', 'request_type'
    ).prefetch_related('comments', 'documents')
    
    # Role-based filtering with hierarchy support
    if user.role in ['top_management']:
        service_requests = base_queryset.all()
    elif user.role == 'operations_head':
        # Operations head can see all requests in their business unit
        accessible_users = user.get_accessible_users() if hasattr(user, 'get_accessible_users') else []
        service_requests = base_queryset.filter(
            Q(raised_by__in=accessible_users) | 
            Q(assigned_to__in=accessible_users) |
            Q(current_owner__in=accessible_users)
        ).distinct()
    elif user.role == 'ops_team_lead':
        # Team lead can see requests from their team and mapped RMs
        team_members = user.get_team_members() if hasattr(user, 'get_team_members') else []
        mapped_rms = user.get_mapped_rms() if hasattr(user, 'get_mapped_rms') else []
        service_requests = base_queryset.filter(
            Q(assigned_to__in=team_members) |
            Q(raised_by__in=mapped_rms) |
            Q(assigned_to=user) |
            Q(raised_by=user)
        ).distinct()
    elif user.role == 'business_head':
        # Business head can see requests from their entire hierarchy
        accessible_users = user.get_accessible_users() if hasattr(user, 'get_accessible_users') else []
        service_requests = base_queryset.filter(
            Q(raised_by__in=accessible_users) | 
            Q(client__user__in=accessible_users) |
            Q(assigned_to__in=accessible_users)
        ).distinct()
    elif user.role == 'rm_head':
        # RM head can see requests from their team
        accessible_users = user.get_accessible_users() if hasattr(user, 'get_accessible_users') else []
        service_requests = base_queryset.filter(
            Q(raised_by__in=accessible_users) | 
            Q(client__user__in=accessible_users)
        ).distinct()
    elif user.role in ['rm', 'ops_exec']:
        # RM and Ops Exec can see their own requests
        service_requests = base_queryset.filter(
            Q(raised_by=user) | 
            Q(assigned_to=user) |
            Q(current_owner=user) |
            Q(client__user=user)
        ).distinct()
    else:
        service_requests = base_queryset.none()
    
    # Enhanced filtering options
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    request_type_filter = request.GET.get('request_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    assigned_to_filter = request.GET.get('assigned_to', '')
    
    if status_filter:
        service_requests = service_requests.filter(status=status_filter)
    
    if priority_filter:
        service_requests = service_requests.filter(priority=priority_filter)
        
    if request_type_filter:
        service_requests = service_requests.filter(request_type_id=request_type_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            service_requests = service_requests.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            service_requests = service_requests.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            pass
    
    if assigned_to_filter:
        service_requests = service_requests.filter(assigned_to_id=assigned_to_filter)
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        service_requests = service_requests.filter(
            Q(request_id__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(client__name__icontains=search_query) |
            Q(client__email__icontains=search_query) |
            Q(request_type__name__icontains=search_query)
        )
    
    # Sorting
    sort_by = request.GET.get('sort', '-created_at')
    valid_sorts = [
        'created_at', '-created_at', 'status', '-status',
        'priority', '-priority', 'updated_at', '-updated_at'
    ]
    if sort_by in valid_sorts:
        service_requests = service_requests.order_by(sort_by)
    else:
        service_requests = service_requests.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(service_requests, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Dashboard statistics
    stats = get_service_request_stats(user, service_requests.model.objects)
    
    # Get filter options with hierarchy-based assignment restrictions
    request_types = ServiceRequestType.objects.filter(is_active=True)
    
    # Uncomment the line below for debugging user roles
    # check_users_and_roles()
    
    assignable_users = get_assignable_users_by_hierarchy(user)  # Modified function
    
    context = {
        'page_obj': page_obj,
        'service_requests': page_obj.object_list,
        'stats': stats,
        'status_choices': ServiceRequest.STATUS_CHOICES,
        'priority_choices': ServiceRequest.PRIORITY_CHOICES,
        'request_types': request_types,
        'assignable_users': assignable_users,
        # Filter values
        'current_status': status_filter,
        'current_priority': priority_filter,
        'current_request_type': request_type_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'assigned_to_filter': assigned_to_filter,
        'sort_by': sort_by,
        'user_role': user.role,
    }
    
    return render(request, 'base/service_requests.html', context)

def get_assignable_users_by_hierarchy(user):
    """Get users that can be assigned service requests based on hierarchy"""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    # Debug logging
    print(f"DEBUG: Current user: {user.username}, Role: {user.role}")
    
    if user.role == 'rm_head':
        # RM Head can only assign to ops_team_lead
        assignable_users = User.objects.filter(role='ops_team_lead', is_active=True)
        print(f"DEBUG: RM Head assignable users: {list(assignable_users.values_list('username', 'role'))}")
        return assignable_users
    elif user.role == 'ops_team_lead':
        # Ops Team Lead can only assign to ops_exec
        assignable_users = User.objects.filter(role='ops_exec', is_active=True)
        print(f"DEBUG: Ops Team Lead assignable users: {list(assignable_users.values_list('username', 'role'))}")
        return assignable_users
    elif user.role in ['business_head_ops', 'business_head', 'top_management']:
        # Higher roles can assign to multiple levels
        assignable_users = User.objects.filter(
            role__in=['ops_team_lead', 'ops_exec'], 
            is_active=True
        )
        print(f"DEBUG: Higher role assignable users: {list(assignable_users.values_list('username', 'role'))}")
        return assignable_users
    else:
        # Other roles cannot assign (or have limited assignment)
        print(f"DEBUG: Role {user.role} has no assignment privileges")
        return User.objects.none()
    
def check_users_and_roles():
    """Helper function to check what users exist in the database"""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    print("\n=== USER ROLES DEBUG ===")
    print("All active users and their roles:")
    for user in User.objects.filter(is_active=True).order_by('role', 'username'):
        print(f"  {user.username} ({user.get_full_name() or 'No name'}): {user.role}")
    
    print("\nActive users by role:")
    for role_code, role_name in User.ROLE_CHOICES:
        count = User.objects.filter(role=role_code, is_active=True).count()
        if count > 0:
            users = User.objects.filter(role=role_code, is_active=True).values_list('username', flat=True)
            print(f"  {role_code} ({role_name}): {count} users - {', '.join(users)}")
    print("=" * 30)
    
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_http_methods, require_POST
from django.core.exceptions import ValidationError
from django.utils import timezone
import json


@login_required
@require_POST
def service_request_submit(request, pk):
    """Submit a service request from draft to operations"""
    service_request = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    
    # Check permissions - only the person who raised it can submit
    if service_request.raised_by != user:
        messages.error(request, 'You can only submit your own service requests.')
        return redirect('service_request_detail', pk=pk)
    
    # Check if request can be submitted
    if service_request.status != 'draft':
        messages.error(request, f'Request cannot be submitted. Current status: {service_request.get_status_display()}')
        return redirect('service_request_detail', pk=pk)
    
    try:
        # Use the model method to submit
        service_request.submit_request(user)
        messages.success(request, f'Service request {service_request.request_id} submitted successfully to operations team.')
        
        # Check if it's an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Request submitted successfully',
                'new_status': service_request.get_status_display()
            })
            
    except ValidationError as e:
        messages.error(request, f'Error submitting request: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': str(e)})
    
    return redirect('service_request_detail', pk=pk)


@login_required
@require_POST
def service_request_action(request, pk, action):
    """Handle various service request workflow actions"""
    service_request = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    
    # Parse request data
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST.dict()
    except:
        data = {}
    
    try:
        # Handle different workflow actions
        if action == 'request_documents':
            if not can_request_documents(user, service_request):
                raise ValidationError('You do not have permission to request documents.')
            
            document_list = data.get('document_list', [])
            if isinstance(document_list, str):
                document_list = [doc.strip() for doc in document_list.split(',') if doc.strip()]
            
            service_request.request_documents(document_list, user)
            message = 'Document request sent to RM successfully.'
            
        elif action == 'submit_documents':
            if not can_submit_documents(user, service_request):
                raise ValidationError('You do not have permission to submit documents.')
            
            service_request.submit_documents(user)
            message = 'Documents submitted to operations team successfully.'
            
        elif action == 'start_processing':
            if not can_start_processing(user, service_request):
                raise ValidationError('You do not have permission to start processing.')
            
            service_request.start_processing(user)
            message = 'Request processing started successfully.'
            
        elif action == 'resolve':
            if not can_resolve_request(user, service_request):
                raise ValidationError('You do not have permission to resolve this request.')
            
            resolution_summary = data.get('resolution_summary', data.get('remarks', ''))
            if not resolution_summary:
                raise ValidationError('Resolution summary is required.')
            
            service_request.resolve_request(resolution_summary, user)
            message = 'Request resolved successfully. Sent to RM for verification.'
            
        elif action == 'verify_resolution':
            if not can_verify_resolution(user, service_request):
                raise ValidationError('You do not have permission to verify this resolution.')
            
            approved = data.get('approved', 'true').lower() == 'true'
            service_request.client_verification_complete(approved, user)
            
            if approved:
                message = 'Resolution verified successfully.'
            else:
                message = 'Resolution rejected. Request sent back for rework.'
                
        elif action == 'close':
            if not can_close_request(user, service_request):
                raise ValidationError('You do not have permission to close this request.')
            
            service_request.close_request(user)
            message = 'Service request closed successfully.'
            
        elif action == 'escalate':
            if not can_escalate_request(user, service_request):
                raise ValidationError('You do not have permission to escalate this request.')
            
            reason = data.get('reason', 'Request escalated by user')
            service_request.escalate_to_manager(user, reason)
            message = 'Request escalated to manager successfully.'
            
        elif action == 'reassign':
            if not can_reassign_request(user, service_request):
                raise ValidationError('You do not have permission to reassign this request.')
            
            new_assignee_id = data.get('assigned_to')
            if not new_assignee_id:
                raise ValidationError('New assignee is required.')
            
            from django.contrib.auth import get_user_model
            User = get_user_model()
            new_assignee = get_object_or_404(User, id=new_assignee_id)
            
            old_assignee = service_request.assigned_to
            service_request.assigned_to = new_assignee
            service_request.current_owner = new_assignee
            service_request.save()
            
            # Add comment
            ServiceRequestComment.objects.create(
                service_request=service_request,
                comment=f"Request reassigned from {old_assignee} to {new_assignee}",
                commented_by=user,
                is_internal=True
            )
            
            message = f'Request reassigned to {new_assignee.get_full_name() or new_assignee.username} successfully.'
            
        else:
            raise ValidationError(f'Unknown action: {action}')
        
        # Return response
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': message,
                'new_status': service_request.get_status_display(),
                'current_owner': service_request.current_owner.get_full_name() if service_request.current_owner else None
            })
        else:
            messages.success(request, message)
            return redirect('service_request_detail', pk=pk)
            
    except ValidationError as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': str(e)})
        else:
            messages.error(request, str(e))
            return redirect('service_request_detail', pk=pk)
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': f'An unexpected error occurred: {str(e)}'})
        else:
            messages.error(request, f'An unexpected error occurred: {str(e)}')
            return redirect('service_request_detail', pk=pk)


@login_required
@require_POST
def service_request_upload_document(request, pk):
    """Upload documents for a service request"""
    service_request = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    
    # Check permissions
    if not can_upload_documents(user, service_request):
        messages.error(request, 'You do not have permission to upload documents for this request.')
        return redirect('service_request_detail', pk=pk)
    
    # Handle file uploads
    uploaded_files = request.FILES.getlist('documents')
    if not uploaded_files:
        messages.error(request, 'No files were uploaded.')
        return redirect('service_request_detail', pk=pk)
    
    try:
        uploaded_count = 0
        for uploaded_file in uploaded_files:
            # Validate file
            if uploaded_file.size > 10 * 1024 * 1024:  # 10MB limit
                messages.warning(request, f'File {uploaded_file.name} is too large (max 10MB).')
                continue
            
            # Check file type
            allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx', '.xls', '.xlsx']
            file_extension = uploaded_file.name.lower().split('.')[-1]
            if f'.{file_extension}' not in allowed_extensions:
                messages.warning(request, f'File {uploaded_file.name} has an unsupported format.')
                continue
            
            # Create document record
            ServiceRequestDocument.objects.create(
                service_request=service_request,
                document=uploaded_file,
                document_name=uploaded_file.name,
                uploaded_by=user
            )
            uploaded_count += 1
        
        if uploaded_count > 0:
            messages.success(request, f'{uploaded_count} document(s) uploaded successfully.')
            
            # Add comment
            ServiceRequestComment.objects.create(
                service_request=service_request,
                comment=f"{uploaded_count} document(s) uploaded by {user.get_full_name() or user.username}",
                commented_by=user,
                is_internal=False
            )
            
            # If this is in response to document request, update status
            if service_request.status == 'documents_requested':
                # You might want to automatically change status or let user do it manually
                pass
        else:
            messages.error(request, 'No valid files were uploaded.')
            
    except Exception as e:
        messages.error(request, f'Error uploading documents: {str(e)}')
    
    return redirect('service_request_detail', pk=pk)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse


@login_required
def service_request_detail(request, pk):
    """View service request details with comments and documents"""
    service_request = get_object_or_404(
        ServiceRequest.objects.select_related(
            'client', 'raised_by', 'assigned_to', 'current_owner', 'request_type'
        ).prefetch_related('comments', 'documents'),
        pk=pk
    )
    
    user = request.user
    
    # Check permission to view
    if not can_view_service_request(user, service_request):
        messages.error(request, 'You do not have permission to view this service request.')
        return redirect('service_request_list')
    
    # Get comments (filter internal comments based on role)
    if user.role in ['ops_exec', 'ops_team_lead', 'operations_head']:
        comments = service_request.comments.all().order_by('-created_at')
    else:
        comments = service_request.comments.filter(is_internal=False).order_by('-created_at')
    
    # Get documents
    documents = service_request.documents.all().order_by('-uploaded_at')
    
    # Get available actions
    available_actions = []
    if user == service_request.raised_by and service_request.status == 'draft':
        available_actions.append('submit')
    if user == service_request.assigned_to:
        if service_request.status == 'submitted':
            available_actions.extend(['request_documents', 'start_processing'])
        elif service_request.status == 'in_progress':
            available_actions.append('resolve')
    if user == service_request.raised_by and service_request.status == 'resolved':
        available_actions.append('verify')
    
    # User permissions
    permissions = {
        'can_edit': can_edit_service_request(user, service_request),
        'can_delete': can_delete_service_request(user, service_request),
        'can_add_comment': can_add_comment(user, service_request),
        'can_upload_document': can_upload_documents(user, service_request),
    }
    
    context = {
        'service_request': service_request,
        'comments': comments,
        'documents': documents,
        'available_actions': available_actions,
        'permissions': permissions,
        'status_choices': ServiceRequest.STATUS_CHOICES,
    }
    
    return render(request, 'base/service_request_detail.html', context)


# Permission helper functions (essential ones only)
def can_view_service_request(user, service_request):
    """Check if user can view the service request"""
    return (user == service_request.raised_by or 
            user == service_request.assigned_to or 
            user == service_request.current_owner or
            user.role in ['operations_head', 'business_head', 'top_management','ops_team_lead'])


def can_edit_service_request(user, service_request):
    """Check if user can edit the service request"""
    return (user == service_request.raised_by and service_request.status in ['draft', 'submitted'] or
            user.role in ['operations_head', 'business_head', 'top_management','ops_team_lead'])


def can_delete_service_request(user, service_request):
    """Check if user can delete the service request"""
    return (user == service_request.raised_by and service_request.status == 'draft' or
            user.role in ['operations_head', 'business_head', 'top_management','ops_team_lead'])


def can_add_comment(user, service_request):
    """Check if user can add comment"""
    return (user == service_request.raised_by or
            user == service_request.assigned_to or
            user == service_request.current_owner or
            user.role in ['operations_head', 'business_head', 'top_management','ops_team_lead'])


def can_upload_documents(user, service_request):
    """Check if user can upload documents"""
    return (user == service_request.raised_by or
            user == service_request.assigned_to or
            user.role in ['operations_head', 'business_head', 'top_management','ops_team_lead'])


def can_request_documents(user, service_request):
    """Check if user can request documents"""
    return (user == service_request.assigned_to or
            user.role in ['ops_team_lead', 'operations_head'])


def can_submit_documents(user, service_request):
    """Check if user can submit documents"""
    return user == service_request.raised_by


def can_start_processing(user, service_request):
    """Check if user can start processing"""
    return (user == service_request.assigned_to or
            user.role in ['ops_team_lead', 'operations_head'])


def can_resolve_request(user, service_request):
    """Check if user can resolve request"""
    return (user == service_request.assigned_to or
            user.role in ['ops_team_lead', 'operations_head'])


def can_verify_resolution(user, service_request):
    """Check if user can verify resolution"""
    return user == service_request.raised_by


def can_close_request(user, service_request):
    """Check if user can close request"""
    return user == service_request.raised_by


def can_escalate_request(user, service_request):
    """Check if user can escalate request"""
    return (user == service_request.assigned_to or
            user == service_request.raised_by or
            user.role in ['ops_team_lead', 'operations_head'])


def can_reassign_request(user, service_request):
    """Check if user can reassign request"""
    return user.role in ['ops_team_lead', 'operations_head', 'business_head', 'top_management']


@login_required
@require_POST
def service_request_delete_document(request, doc_id):
    """Delete a document from service request"""
    document = get_object_or_404(ServiceRequestDocument, id=doc_id)
    service_request = document.service_request
    user = request.user
    
    # Check permissions
    if not can_delete_document(user, document):
        messages.error(request, 'You do not have permission to delete this document.')
        return redirect('service_request_detail', pk=service_request.pk)
    
    try:
        document_name = document.document_name
        document.delete()
        
        # Add comment
        ServiceRequestComment.objects.create(
            service_request=service_request,
            comment=f"Document '{document_name}' deleted by {user.get_full_name() or user.username}",
            commented_by=user,
            is_internal=True
        )
        
        messages.success(request, f'Document "{document_name}" deleted successfully.')
        
    except Exception as e:
        messages.error(request, f'Error deleting document: {str(e)}')
    
    return redirect('service_request_detail', pk=service_request.pk)


@login_required
@require_POST
def service_request_add_comment(request, pk):
    """Add a comment to service request"""
    service_request = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    
    # Check permissions
    if not can_add_comment(user, service_request):
        messages.error(request, 'You do not have permission to add comments to this request.')
        return redirect('service_request_detail', pk=pk)
    
    comment_text = request.POST.get('comment', '').strip()
    is_internal = request.POST.get('is_internal') == 'on'
    
    if not comment_text:
        messages.error(request, 'Comment cannot be empty.')
        return redirect('service_request_detail', pk=pk)
    
    try:
        ServiceRequestComment.objects.create(
            service_request=service_request,
            comment=comment_text,
            commented_by=user,
            is_internal=is_internal
        )
        
        messages.success(request, 'Comment added successfully.')
        
        # Check if it's an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Comment added successfully'
            })
            
    except Exception as e:
        messages.error(request, f'Error adding comment: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': str(e)})
    
    return redirect('service_request_detail', pk=pk)


# Permission helper functions
def can_request_documents(user, service_request):
    """Check if user can request documents"""
    return (user == service_request.assigned_to or 
            user == service_request.current_owner or
            user.role in ['ops_team_lead', 'operations_head', 'top_management'])


def can_submit_documents(user, service_request):
    """Check if user can submit documents"""
    return (user == service_request.raised_by or
            user.role in ['rm_head', 'business_head', 'top_management'])


def can_start_processing(user, service_request):
    """Check if user can start processing"""
    return (user == service_request.assigned_to or
            user == service_request.current_owner or
            user.role in ['ops_team_lead', 'operations_head', 'top_management'])


def can_resolve_request(user, service_request):
    """Check if user can resolve request"""
    return (user == service_request.assigned_to or
            user == service_request.current_owner or
            user.role in ['ops_team_lead', 'operations_head', 'top_management'])


def can_verify_resolution(user, service_request):
    """Check if user can verify resolution"""
    return (user == service_request.raised_by or
            user.role in ['rm_head', 'business_head', 'top_management'])


def can_close_request(user, service_request):
    """Check if user can close request"""
    return (user == service_request.raised_by or
            user.role in ['rm_head', 'business_head', 'top_management'])


def can_escalate_request(user, service_request):
    """Check if user can escalate request"""
    return (user == service_request.assigned_to or
            user == service_request.raised_by or
            user == service_request.current_owner or
            user.role in ['ops_team_lead', 'rm_head', 'operations_head', 'business_head'])


def can_reassign_request(user, service_request):
    """Check if user can reassign request"""
    return user.role in ['ops_team_lead', 'operations_head', 'business_head', 'top_management']


def can_upload_documents(user, service_request):
    """Check if user can upload documents"""
    return (user == service_request.raised_by or
            user == service_request.assigned_to or
            user == service_request.current_owner or
            user.role in ['rm_head', 'ops_team_lead', 'operations_head', 'business_head', 'top_management'])


def can_delete_document(user, document):
    """Check if user can delete document"""
    return (user == document.uploaded_by or
            user == document.service_request.raised_by or
            user == document.service_request.assigned_to or
            user.role in ['ops_team_lead', 'operations_head', 'business_head', 'top_management'])


def can_add_comment(user, service_request):
    """Check if user can add comment"""
    return (user == service_request.raised_by or
            user == service_request.assigned_to or
            user == service_request.current_owner or
            user.role in ['rm_head', 'ops_team_lead', 'operations_head', 'business_head', 'top_management'])


@login_required
@user_passes_test(lambda u: u.role == 'ops_exec')
def ops_service_requests(request):
    """Enhanced Operations Executive view for service requests"""
    user = request.user
    
    # Base queryset with optimized queries
    base_queryset = ServiceRequest.objects.filter(
        Q(assigned_to=user) | Q(current_owner=user)
    ).select_related(
        'client', 'raised_by', 'request_type'
    ).prefetch_related('comments', 'documents')
    
    # Apply filters
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    request_type_filter = request.GET.get('request_type', '')
    overdue_filter = request.GET.get('overdue', '')
    
    my_requests = base_queryset
    
    if status_filter:
        my_requests = my_requests.filter(status=status_filter)
    
    if priority_filter:
        my_requests = my_requests.filter(priority=priority_filter)
        
    if request_type_filter:
        my_requests = my_requests.filter(request_type_id=request_type_filter)
    
    if overdue_filter == 'true':
        my_requests = my_requests.filter(
            expected_completion_date__lt=timezone.now(),
            status__in=['submitted', 'documents_received', 'in_progress']
        )
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        my_requests = my_requests.filter(
            Q(request_id__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(client__name__icontains=search_query)
        )
    
    # Sorting with priority boost for urgent requests
    sort_by = request.GET.get('sort', 'priority_sort')
    if sort_by == 'priority_sort':
        my_requests = my_requests.annotate(
            priority_order=Case(
                When(priority='urgent', then=1),
                When(priority='high', then=2),
                When(priority='medium', then=3),
                When(priority='low', then=4),
                default=5,
                output_field=IntegerField()
            )
        ).order_by('priority_order', '-created_at')
    else:
        valid_sorts = ['created_at', '-created_at', 'status', '-status', 'updated_at', '-updated_at']
        if sort_by in valid_sorts:
            my_requests = my_requests.order_by(sort_by)
        else:
            my_requests = my_requests.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(my_requests, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Enhanced statistics
    request_stats = get_ops_request_stats(user)
    
    # Performance metrics
    performance_metrics = get_ops_performance_metrics(user)
    
    # Workload distribution
    workload_stats = get_ops_workload_stats(user)
    
    # SLA tracking
    sla_stats = get_ops_sla_stats(user)
    
    # Quick actions data
    quick_actions = get_ops_quick_actions(user)
    
    context = {
        'page_obj': page_obj,
        'my_requests': page_obj.object_list,
        'request_stats': request_stats,
        'performance_metrics': performance_metrics,
        'workload_stats': workload_stats,
        'sla_stats': sla_stats,
        'quick_actions': quick_actions,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'request_type_filter': request_type_filter,
        'overdue_filter': overdue_filter,
        'search_query': search_query,
        'sort_by': sort_by,
        'status_choices': ServiceRequest.STATUS_CHOICES,
        'priority_choices': ServiceRequest.PRIORITY_CHOICES,
        'request_types': ServiceRequestType.objects.filter(is_active=True),
    }
    
    return render(request, 'operations/service_requests.html', context)


@login_required
def rm_service_requests(request):
    """RM view for managing their service requests"""
    user = request.user
    
    if user.role not in ['rm', 'rm_head']:
        return HttpResponseForbidden("Access denied")
    
    # Get RM's service requests
    if user.role == 'rm_head':
        team_members = user.get_team_members() if hasattr(user, 'get_team_members') else [user]
        my_requests = ServiceRequest.objects.filter(
            raised_by__in=team_members
        )
    else:
        my_requests = ServiceRequest.objects.filter(raised_by=user)
    
    my_requests = my_requests.select_related(
        'client', 'assigned_to', 'current_owner', 'request_type'
    ).prefetch_related('comments', 'documents')
    
    # Apply filters
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    
    if status_filter:
        my_requests = my_requests.filter(status=status_filter)
    
    if priority_filter:
        my_requests = my_requests.filter(priority=priority_filter)
    
    # Search
    search_query = request.GET.get('search', '')
    if search_query:
        my_requests = my_requests.filter(
            Q(request_id__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(client__name__icontains=search_query)
        )
    
    my_requests = my_requests.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(my_requests, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # RM-specific statistics
    rm_stats = get_rm_request_stats(user)
    
    context = {
        'page_obj': page_obj,
        'my_requests': page_obj.object_list,
        'rm_stats': rm_stats,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'search_query': search_query,
        'status_choices': ServiceRequest.STATUS_CHOICES,
        'priority_choices': ServiceRequest.PRIORITY_CHOICES,
        'request_types': ServiceRequestType.objects.filter(is_active=True),
    }
    
    return render(request, 'rm/service_requests.html', context)


# Utility functions for statistics and data

def get_service_request_stats(user, queryset):
    """Get general service request statistics"""
    if user.role in ['top_management']:
        base_qs = queryset.all()
    elif user.role in ['operations_head', 'business_head']:
        accessible_users = user.get_accessible_users() if hasattr(user, 'get_accessible_users') else []
        base_qs = queryset.filter(
            Q(raised_by__in=accessible_users) | 
            Q(assigned_to__in=accessible_users)
        )
    else:
        base_qs = queryset.filter(
            Q(raised_by=user) | 
            Q(assigned_to=user) |
            Q(current_owner=user)
        )
    
    stats = {
        'total_requests': base_qs.count(),
        'open_requests': base_qs.filter(status__in=['submitted', 'documents_requested']).count(),
        'in_progress': base_qs.filter(status__in=['documents_received', 'in_progress']).count(),
        'resolved_requests': base_qs.filter(status='resolved').count(),
        'closed_requests': base_qs.filter(status='closed').count(),
        'urgent_requests': base_qs.filter(
            priority='urgent',
            status__in=['submitted', 'documents_requested', 'documents_received', 'in_progress']
        ).count(),
        'overdue_requests': base_qs.filter(
            expected_completion_date__lt=timezone.now(),
            status__in=['submitted', 'documents_requested', 'documents_received', 'in_progress']
        ).count(),
    }
    
    return stats


def get_ops_request_stats(user):
    """Get operations-specific statistics"""
    base_qs = ServiceRequest.objects.filter(assigned_to=user)
    
    stats = {
        'total_assigned': base_qs.count(),
        'pending_documents': base_qs.filter(status='documents_requested').count(),
        'ready_to_process': base_qs.filter(status='documents_received').count(),
        'in_progress': base_qs.filter(status='in_progress').count(),
        'awaiting_verification': base_qs.filter(status='resolved').count(),
        'completed_today': base_qs.filter(
            status__in=['resolved', 'closed'],
            updated_at__date=timezone.now().date()
        ).count(),
        'urgent_pending': base_qs.filter(
            priority='urgent',
            status__in=['submitted', 'documents_received', 'in_progress']
        ).count(),
    }
    
    return stats


def get_ops_performance_metrics(user):
    """Get performance metrics for operations executive"""
    last_30_days = timezone.now() - timedelta(days=30)
    base_qs = ServiceRequest.objects.filter(assigned_to=user)
    
    recent_requests = base_qs.filter(created_at__gte=last_30_days)
    resolved_requests = recent_requests.filter(
        status__in=['resolved', 'closed'],
        resolved_at__isnull=False
    )
    
    metrics = {
        'requests_last_30_days': recent_requests.count(),
        'resolved_last_30_days': resolved_requests.count(),
        'resolution_rate': 0,
        'avg_resolution_time_hours': 0,
        'client_satisfaction': 0,  # Can be calculated from feedback if available
    }
    
    if recent_requests.exists():
        metrics['resolution_rate'] = round(
            (metrics['resolved_last_30_days'] / metrics['requests_last_30_days']) * 100, 2
        )
    
    if resolved_requests.exists():
        avg_resolution_time = resolved_requests.aggregate(
            avg_time=Avg(
                Extract('epoch', F('resolved_at')) - Extract('epoch', F('created_at'))
            )
        )['avg_time']
        metrics['avg_resolution_time_hours'] = round(
            avg_resolution_time / 3600, 2
        ) if avg_resolution_time else 0
    
    return metrics


def get_ops_workload_stats(user):
    """Get current workload statistics"""
    base_qs = ServiceRequest.objects.filter(assigned_to=user)
    
    workload = {
        'total_active': base_qs.filter(
            status__in=['submitted', 'documents_received', 'in_progress']
        ).count(),
        'high_priority': base_qs.filter(
            priority__in=['urgent', 'high'],
            status__in=['submitted', 'documents_received', 'in_progress']
        ).count(),
        'due_today': base_qs.filter(
            expected_completion_date__date=timezone.now().date(),
            status__in=['submitted', 'documents_received', 'in_progress']
        ).count(),
        'overdue': base_qs.filter(
            expected_completion_date__lt=timezone.now(),
            status__in=['submitted', 'documents_received', 'in_progress']
        ).count(),
    }
    
    return workload


def get_ops_sla_stats(user):
    """Get SLA-related statistics"""
    base_qs = ServiceRequest.objects.filter(assigned_to=user)
    
    sla_stats = {
        'sla_breached': base_qs.filter(sla_breached=True).count(),
        'at_risk': base_qs.filter(
            expected_completion_date__lte=timezone.now() + timedelta(hours=4),
            expected_completion_date__gt=timezone.now(),
            status__in=['submitted', 'documents_received', 'in_progress']
        ).count(),
        'compliance_rate': 0,
    }
    
    total_completed = base_qs.filter(status__in=['resolved', 'closed']).count()
    if total_completed > 0:
        breached_completed = base_qs.filter(
            status__in=['resolved', 'closed'],
            sla_breached=True
        ).count()
        sla_stats['compliance_rate'] = round(
            ((total_completed - breached_completed) / total_completed) * 100, 2
        )
    
    return sla_stats


def get_ops_quick_actions(user):
    """Get data for quick action buttons"""
    base_qs = ServiceRequest.objects.filter(assigned_to=user)
    
    actions = {
        'document_requests_to_send': base_qs.filter(
            status='submitted'
        ).count(),
        'ready_to_process': base_qs.filter(
            status='documents_received'
        ).count(),
        'pending_resolution': base_qs.filter(
            status='in_progress'
        ).count(),
    }
    
    return actions


def get_rm_request_stats(user):
    """Get RM-specific statistics"""
    if user.role == 'rm_head':
        team_members = user.get_team_members() if hasattr(user, 'get_team_members') else [user]
        base_qs = ServiceRequest.objects.filter(raised_by__in=team_members)
    else:
        base_qs = ServiceRequest.objects.filter(raised_by=user)
    
    stats = {
        'total_raised': base_qs.count(),
        'pending_documents': base_qs.filter(status='documents_requested').count(),
        'with_operations': base_qs.filter(
            status__in=['documents_received', 'in_progress']
        ).count(),
        'awaiting_client_approval': base_qs.filter(status='resolved').count(),
        'completed_this_month': base_qs.filter(
            status='closed',
            closed_at__month=timezone.now().month,
            closed_at__year=timezone.now().year
        ).count(),
        'client_satisfaction_pending': base_qs.filter(
            status='client_verification'
        ).count(),
    }
    
    return stats


def get_assignable_users(user):
    """Get list of users that can be assigned requests based on hierarchy"""
    # This should implement your actual hierarchy logic
    # For now, returning a basic implementation
    if user.role in ['top_management', 'operations_head']:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        return User.objects.filter(role='ops_exec', is_active=True)
    elif user.role == 'ops_team_lead':
        # Return team members
        return user.get_team_members() if hasattr(user, 'get_team_members') else []
    else:
        return []


# API Views for AJAX requests

@login_required
@require_http_methods(["POST"])
def update_service_request_status(request, request_id):
    """API endpoint to update service request status"""
    service_request = get_object_or_404(ServiceRequest, request_id=request_id)
    
    # Check permissions
    if not can_update_service_request(request.user, service_request):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    data = json.loads(request.body)
    new_status = data.get('status')
    remarks = data.get('remarks', '')
    
    try:
        # Handle status transitions based on workflow
        if new_status == 'documents_requested':
            document_list = data.get('document_list', [])
            service_request.request_documents(document_list, request.user)
        elif new_status == 'documents_received':
            service_request.submit_documents(request.user)
        elif new_status == 'in_progress':
            service_request.start_processing(request.user)
        elif new_status == 'resolved':
            resolution_summary = data.get('resolution_summary', remarks)
            service_request.resolve_request(resolution_summary, request.user)
        elif new_status == 'closed':
            service_request.close_request(request.user)
        else:
            # Direct status update for other cases
            service_request.status = new_status
            service_request.save()
            
            # Add comment
            ServiceRequestComment.objects.create(
                service_request=service_request,
                comment=f"Status updated to {new_status}. {remarks}",
                commented_by=request.user,
                is_internal=True
            )
        
        return JsonResponse({
            'success': True,
            'new_status': service_request.get_status_display(),
            'message': 'Status updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def can_update_service_request(user, service_request):
    """Check if user can update the service request"""
    # Implement your permission logic here
    if user.role in ['top_management', 'operations_head']:
        return True
    elif user == service_request.assigned_to or user == service_request.current_owner:
        return True
    elif user == service_request.raised_by and service_request.status in ['documents_requested', 'resolved']:
        return True
    else:
        return False

@login_required
def service_request_create(request):
    """Create service request with hierarchy-based assignment restrictions"""
    if request.method == 'POST':
        form = ServiceRequestForm(request.POST)
        
        # CRITICAL: Apply hierarchy restrictions to the form
        assignable_users = get_assignable_users_by_hierarchy(request.user)
        form.fields['assigned_to'].queryset = assignable_users
        
        if form.is_valid():
            service_request = form.save(commit=False)
            service_request.raised_by = request.user
            
            # Set current owner to assigned user if provided, otherwise to the requester
            assigned_to = form.cleaned_data.get('assigned_to')
            if assigned_to:
                service_request.current_owner = assigned_to
            else:
                service_request.current_owner = request.user
            
            service_request.save()
            
            messages.success(request, 'Service request created successfully.')
            return redirect('service_request_detail', pk=service_request.pk)
        else:
            # Debug form errors
            print(f"DEBUG: Form errors: {form.errors}")
    else:
        form = ServiceRequestForm()
        
        # CRITICAL: Apply hierarchy restrictions to the form queryset
        assignable_users = get_assignable_users_by_hierarchy(request.user)
        form.fields['assigned_to'].queryset = assignable_users
        
        # Debug output
        print(f"DEBUG: CREATE - User {request.user.username} ({request.user.role}) can assign to:")
        for user in assignable_users:
            print(f"  - {user.username} ({user.role}) - {user.get_full_name()}")
    
    context = {
        'form': form,
        'service_request': None,  # For new requests
    }
    
    return render(request, 'base/service_request_form.html', context)

@login_required 
def service_request_update(request, pk):
    """Update service request with hierarchy-based assignment restrictions"""
    service_request = get_object_or_404(ServiceRequest, pk=pk)
    
    # Basic permission check - you can expand this
    if not (request.user == service_request.raised_by or 
            request.user == service_request.assigned_to or
            request.user == service_request.current_owner or
            request.user.role in ['business_head', 'business_head_ops', 'top_management', 'ops_team_lead']):
        messages.error(request, "You don't have permission to edit this service request.")
        return redirect('service_request_detail', pk=pk)
    
    if request.method == 'POST':
        form = ServiceRequestForm(request.POST, instance=service_request)
        
        # CRITICAL: Apply hierarchy restrictions to the form
        assignable_users = get_assignable_users_by_hierarchy(request.user)
        form.fields['assigned_to'].queryset = assignable_users
        
        if form.is_valid():
            updated_request = form.save(commit=False)
            
            # Update current owner if assigned_to changed
            if 'assigned_to' in form.changed_data:
                assigned_to = form.cleaned_data.get('assigned_to')
                if assigned_to:
                    updated_request.current_owner = assigned_to
            
            updated_request.save()
            messages.success(request, 'Service request updated successfully.')
            return redirect('service_request_detail', pk=service_request.pk)
        else:
            # Debug form errors
            print(f"DEBUG: Update form errors: {form.errors}")
    else:
        form = ServiceRequestForm(instance=service_request)
        
        # CRITICAL: Apply hierarchy restrictions to the form queryset
        assignable_users = get_assignable_users_by_hierarchy(request.user)
        form.fields['assigned_to'].queryset = assignable_users
        
        # Debug output
        print(f"DEBUG: UPDATE - User {request.user.username} ({request.user.role}) can assign to:")
        for user in assignable_users:
            print(f"  - {user.username} ({user.role}) - {user.get_full_name()}")
    
    context = {
        'form': form,
        'service_request': service_request,
    }
    
    return render(request, 'base/service_request_form.html', context)

@login_required
def service_request_delete(request, pk):
    service_request = get_object_or_404(ServiceRequest, pk=pk)
    
    # Check permissions
    accessible_users = request.user.get_accessible_users()
    if not (service_request.raised_by in accessible_users or 
            service_request.client.user in accessible_users):
        raise PermissionDenied("You don't have permission to delete this service request.")
    
    if request.method == 'POST':
        service_request.delete()
        messages.success(request, "Service request deleted successfully.")
        return redirect('service_request_list')
    
    return render(request, 'base/service_request_confirm_delete.html', {'service_request': service_request})


# Team Management Views (for RM Heads and above)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from .models import Team, User
from .forms import TeamForm, UserEditForm

@login_required
@user_passes_test(lambda u: u.role in ['rm_head', 'business_head', 'top_management'])
def team_management(request):
    user = request.user
    
    if user.role == 'top_management':
        teams = Team.objects.all()
        all_users = User.objects.all()
    elif user.role == 'business_head':
        teams = Team.objects.all()
        all_users = User.objects.all()
    else:  # rm_head
        teams = user.led_teams.all()
        all_users = user.get_accessible_users()
    
    context = {
        'teams': teams,
        'all_users': all_users,
        'user_role': user.role,
    }
    return render(request, 'base/team_management.html', context)

@login_required
@user_passes_test(lambda u: u.role in ['business_head', 'top_management'])
def create_team(request):
    if request.method == 'POST':
        form = TeamForm(request.POST)
        if form.is_valid():
            team = form.save()
            messages.success(request, f"Team '{team.name}' created successfully.")
            return redirect('team_detail', team_id=team.id)
    else:
        form = TeamForm(initial={'leader': request.user})
    
    context = {
        'form': form,
        'action': 'Create',
    }
    return render(request, 'base/team_form.html', context)

@login_required
@user_passes_test(lambda u: u.role in ['rm_head', 'business_head', 'top_management'])
def team_detail(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    
    # Check if user has permission to view this team
    if (request.user.role not in ['business_head', 'top_management'] and 
        request.user not in team.members.all() and 
        request.user != team.leader):
        messages.error(request, "You don't have permission to view this team.")
        return redirect('team_management')
    
    context = {
        'team': team,
        'members': team.members.all(),
    }
    return render(request, 'base/team_detail.html', context)

@login_required
def edit_team(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    
    # Check permissions - business_head/top_management or team leader
    if (request.user.role not in ['business_head', 'top_management'] and 
        request.user != team.leader):
        messages.error(request, "You don't have permission to edit this team.")
        return redirect('team_management')
    
    if request.method == 'POST':
        form = TeamForm(request.POST, instance=team)
        if form.is_valid():
            form.save()
            messages.success(request, f"Team '{team.name}' updated successfully.")
            return redirect('team_detail', team_id=team.id)
    else:
        form = TeamForm(instance=team)
    
    context = {
        'form': form,
        'team': team,
        'action': 'Edit',
    }
    return render(request, 'base/team_form.html', context)

@login_required
def user_profile(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    # Check if current user has permission to view this profile
    if (request.user.role not in ['business_head', 'top_management'] and 
        request.user != user and 
        not request.user.led_teams.filter(members=user).exists()):
        messages.error(request, "You don't have permission to view this profile.")
        return redirect('team_management')
    
    context = {
        'profile_user': user,
        'teams': user.teams.all(),
    }
    return render(request, 'base/user_profile.html', context)

@login_required
@user_passes_test(lambda u: u.role in ['business_head', 'top_management'])
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f"User {user.get_full_name()} updated successfully.")
            return redirect('user_profile', user_id=user.id)
    else:
        form = UserEditForm(instance=user)
    
    context = {
        'form': form,
        'profile_user': user,
    }
    return render(request, 'base/user_form.html', context)


# Analytics and Reports (role-based access)
@login_required
def analytics_dashboard(request):
    user = request.user
    context = {}
    
    if user.role in ['top_management', 'business_head']:
        # System-wide analytics
        total_aum = Client.objects.aggregate(total=Sum('aum'))['total'] or 0
        total_sip = Client.objects.aggregate(total=Sum('sip_amount'))['total'] or 0
        total_clients = Client.objects.count()
        
        # Performance metrics
        total_leads = Lead.objects.count()
        converted_leads = Lead.objects.filter(status='converted').count()
        lead_conversion_rate = (converted_leads / max(total_leads, 1)) * 100
        
        # Team performance
        team_performance = User.objects.filter(role='rm').annotate(
            client_count=Count('clients'),
            total_aum=Sum('clients__aum'),
            total_sip=Sum('clients__sip_amount')
        ).order_by('-total_aum')
        
        context.update({
            'total_aum': total_aum,
            'total_sip': total_sip,
            'total_clients': total_clients,
            'lead_conversion_rate': round(lead_conversion_rate, 2),
            'team_performance': team_performance[:10],  # Top 10 performers
        })
        
    elif user.role == 'rm_head':
        # Team analytics
        accessible_users = user.get_accessible_users()
        team_clients = Client.objects.filter(user__in=accessible_users)
        
        team_aum = team_clients.aggregate(total=Sum('aum'))['total'] or 0
        team_sip = team_clients.aggregate(total=Sum('sip_amount'))['total'] or 0
        
        # Individual team member performance
        team_performance = accessible_users.filter(role='rm').annotate(
            client_count=Count('clients'),
            total_aum=Sum('clients__aum'),
            total_sip=Sum('clients__sip_amount')
        ).order_by('-total_aum')
        
        context.update({
            'team_aum': team_aum,
            'team_sip': team_sip,
            'team_clients_count': team_clients.count(),
            'team_performance': team_performance,
        })
        
    else:  # RM
        # Personal analytics
        my_clients = Client.objects.filter(user=user)
        my_aum = my_clients.aggregate(total=Sum('aum'))['total'] or 0
        my_sip = my_clients.aggregate(total=Sum('sip_amount'))['total'] or 0
        
        # Monthly performance
        from django.utils import timezone
        from datetime import timedelta
        
        last_30_days = timezone.now() - timedelta(days=30)
        recent_clients = my_clients.filter(created_at__gte=last_30_days).count()
        
        context.update({
            'my_aum': my_aum,
            'my_sip': my_sip,
            'my_clients_count': my_clients.count(),
            'recent_clients': recent_clients,
        })
    
    template_name = f'base/analytics_{user.role}.html'
    return render(request, template_name, context)


@login_required
def create_plan(request):
    """Optimized create plan view with efficient client loading"""
    if request.user.role not in ['rm', 'rm_head']:
        messages.error(request, "Only RMs and RM Heads can create execution plans.")
        return redirect('dashboard')
    
    # Get ONLY essential client data with optimized queries
    if request.user.role == 'rm':
        # Get clients with select_related first, then only fetch needed fields
        legacy_clients = Client.objects.filter(user=request.user).select_related(
            'client_profile', 'user', 'created_by'
        ).only(
            'id', 'name', 'aum', 'sip_amount', 'demat_count', 'contact_info', 'created_at',
            'client_profile__id', 'client_profile__pan_number', 'client_profile__email', 
            'client_profile__mobile_number', 'client_profile__address_kyc',
            'user__id', 'user__first_name', 'user__last_name', 'user__username',
            'created_by__id', 'created_by__first_name', 'created_by__last_name', 'created_by__username'
        )
    else:
        # RM Head - get team members first, then their clients
        team_rm_ids = list(User.objects.filter(
            manager=request.user, role='rm'
        ).values_list('id', flat=True))
        
        legacy_clients = Client.objects.filter(
            user_id__in=team_rm_ids
        ).select_related(
            'client_profile', 'user', 'created_by'
        ).only(
            'id', 'name', 'aum', 'sip_amount', 'demat_count', 'contact_info', 'created_at',
            'client_profile__id', 'client_profile__pan_number', 'client_profile__email', 
            'client_profile__mobile_number', 'client_profile__address_kyc',
            'user__id', 'user__first_name', 'user__last_name', 'user__username',
            'created_by__id', 'created_by__first_name', 'created_by__last_name', 'created_by__username'
        )

    # Order by AUM descending to show high-value clients first
    legacy_clients = legacy_clients.order_by('-aum', 'name')

    # Build optimized client data - load portfolio summary for clients with profiles
    clients_data = []
    
    # Get portfolio data in bulk for clients with profiles
    client_profiles = [client.client_profile for client in legacy_clients if client.client_profile]
    
    portfolio_summaries = {}
    if client_profiles:
        # Get portfolio summaries in one query
        portfolio_data = ClientPortfolio.objects.filter(
            client_profile__in=client_profiles,
            is_active=True
        ).values('client_profile_id').annotate(
            total_aum=Sum('total_value'),
            scheme_count=Count('id')
        )
        
        # Convert to dictionary for quick lookup
        for data in portfolio_data:
            portfolio_summaries[data['client_profile_id']] = {
                'total_aum': float(data['total_aum'] or 0),
                'scheme_count': data['scheme_count'] or 0
            }
    
    for client in legacy_clients:
        # Get basic client info
        client_info = {
            'id': client.id,
            'name': client.name,
            'type': 'legacy',
            'client_id': client.id,
            'contact_info': client.contact_info or '',
            'aum': float(client.aum) if client.aum else 0.0,
            'sip_amount': float(client.sip_amount) if client.sip_amount else 0.0,
            'demat_count': client.demat_count or 0,
            'created_at': client.created_at,
            'mapped_rm': client.user.get_full_name() if client.user else 'Not Mapped',
            'created_by': client.created_by.get_full_name() if client.created_by else 'Unknown',
            'has_profile': bool(client.client_profile),
        }
        
        # Add profile data if exists
        if client.client_profile:
            portfolio_summary = portfolio_summaries.get(client.client_profile.id, {
                'total_aum': 0.0, 'scheme_count': 0
            })
            
            client_info.update({
                'pan': client.client_profile.pan_number,
                'profile_id': client.client_profile.id,
                'email': client.client_profile.email,
                'mobile': client.client_profile.mobile_number,
                'full_address': client.client_profile.address_kyc,
                'total_aum': portfolio_summary['total_aum'],
                'scheme_count': portfolio_summary['scheme_count'],
                'has_portfolio': portfolio_summary['scheme_count'] > 0,
            })
        else:
            client_info.update({
                'pan': 'N/A',
                'profile_id': None,
                'email': '',
                'mobile': '',
                'full_address': '',
                'total_aum': 0.0,
                'scheme_count': 0,
                'has_portfolio': False,
            })
        
        clients_data.append(client_info)

    # Calculate statistics including portfolio data
    total_clients = len(clients_data)
    clients_with_profiles = len([c for c in clients_data if c['has_profile']])
    clients_with_portfolio = len([c for c in clients_data if c['has_portfolio']])
    total_aum = sum(c['aum'] for c in clients_data)
    total_portfolio_aum = sum(c['total_aum'] for c in clients_data)
    average_aum = (total_aum + total_portfolio_aum) / total_clients if total_clients > 0 else 0

    context = {
        'clients': clients_data,
        'total_clients': total_clients,
        'clients_with_profiles': clients_with_profiles,
        'clients_with_portfolio': clients_with_portfolio,
        'total_aum': total_aum,
        'total_portfolio_aum': total_portfolio_aum,
        'average_aum': average_aum,
        'user_role': request.user.role,
    }
    
    return render(request, 'execution_plans/create_plan.html', context)


@login_required
def client_portfolio_ajax(request, client_id):
    """Optimized AJAX endpoint for client portfolio - PORTFOLIO INDEPENDENT VERSION"""
    try:
        # Handle the client_id format (remove any prefixes)
        if isinstance(client_id, str):
            if client_id.startswith('legacy_'):
                client_id = client_id.replace('legacy_', '')
            elif client_id.startswith('profile_'):
                messages.error(request, "Profile clients are not supported. Please select a legacy client.")
                return JsonResponse({'error': 'Profile clients not supported'}, status=400)
        
        client_id = int(client_id)
        client = get_object_or_404(Client, id=client_id)
        
        # Check access permission
        if request.user.role == 'rm' and client.user != request.user:
            return JsonResponse({'error': 'Access denied'}, status=403)
        elif request.user.role == 'rm_head':
            if client.user and client.user.manager != request.user:
                return JsonResponse({'error': 'Access denied'}, status=403)
        
        # Get client profile if exists
        client_profile = getattr(client, 'client_profile', None)
        
        if client_profile:
            # Get portfolio with optimized query - NO SCHEME MAPPING
            portfolio_holdings = ClientPortfolio.objects.filter(
                client_profile=client_profile,
                is_active=True
            ).only(
                'id', 'scheme_name', 'isin_number', 'folio_number', 'units',
                'total_value', 'cost_value', 'equity_value', 'debt_value',
                'hybrid_value', 'liquid_ultra_short_value', 'other_value',
                'arbitrage_value', 'allocation_percentage', 'gain_loss',
                'gain_loss_percentage', 'primary_asset_class', 'nav_price'
            ).order_by('-total_value')  # Order by value descending
            
            portfolio_data = []
            total_aum = 0
            
            for holding in portfolio_holdings:
                portfolio_data.append({
                    'id': holding.id,
                    'scheme_name': holding.scheme_name,  # Use scheme name as-is from portfolio
                    'isin_number': holding.isin_number or 'N/A',
                    'folio_number': holding.folio_number or 'N/A',
                    'units': float(holding.units),
                    'current_value': float(holding.total_value),
                    'cost_value': float(holding.cost_value) if holding.cost_value else 0,
                    'equity_value': float(holding.equity_value),
                    'debt_value': float(holding.debt_value),
                    'hybrid_value': float(holding.hybrid_value),
                    'liquid_value': float(holding.liquid_ultra_short_value),
                    'other_value': float(holding.other_value),
                    'arbitrage_value': float(holding.arbitrage_value),
                    'allocation_percentage': float(holding.allocation_percentage),
                    'gain_loss': float(holding.gain_loss) if holding.gain_loss else 0,
                    'gain_loss_percentage': round(holding.gain_loss_percentage, 2) if holding.gain_loss_percentage else 0,
                    'primary_asset_class': holding.primary_asset_class,
                    'nav_price': float(holding.nav_price) if holding.nav_price else None,
                    'can_redeem': holding.total_value > 0,
                    'can_switch': holding.total_value > 0,
                    'can_stp': holding.total_value >= 1000,
                    'can_swp': holding.total_value >= 5000,
                    # Portfolio-specific identifiers (not mapped to MutualFundScheme)
                    'portfolio_holding_id': holding.id,
                    'is_portfolio_scheme': True,  # Flag to indicate this is from portfolio
                })
                total_aum += holding.total_value
            
            # Optimized asset allocation using database aggregation
            asset_allocation = portfolio_holdings.aggregate(
                total_equity=Sum('equity_value'),
                total_debt=Sum('debt_value'),
                total_hybrid=Sum('hybrid_value'),
                total_liquid=Sum('liquid_ultra_short_value'),
                total_other=Sum('other_value'),
                total_arbitrage=Sum('arbitrage_value')
            )
            
            return JsonResponse({
                'success': True,
                'portfolio': portfolio_data,
                'client_name': client_profile.client_full_name,
                'client_pan': client_profile.pan_number,
                'total_aum': total_aum,
                'asset_allocation': asset_allocation,
                'portfolio_count': len(portfolio_data),
                'client_type': 'profile',
                'has_portfolio_data': True
            })
        else:
            # No portfolio data available for legacy client
            return JsonResponse({
                'success': True,
                'portfolio': [],
                'client_name': client.name,
                'client_type': 'legacy',
                'total_aum': 0,
                'portfolio_count': 0,
                'has_portfolio_data': False,
                'message': 'No detailed portfolio data available'
            })
        
    except Exception as e:
        logger.error(f"Error in client_portfolio_ajax: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def create_plan_step2(request, client_id):
    """Create execution plan - Step 2: Design plan with Portfolio Actions (Legacy Clients Only) - FIXED VERSION"""
    
    # Debug logging
    logger.info(f"create_plan_step2 called with client_id: {client_id} (type: {type(client_id)})")
    
    try:
        # Ensure client_id is an integer (removing any prefixes if they exist)
        if isinstance(client_id, str):
            # Remove any legacy prefixes that might still be in URLs
            if client_id.startswith('legacy_'):
                client_id = client_id.replace('legacy_', '')
            elif client_id.startswith('profile_'):
                messages.error(request, "Profile clients are not supported. Please select a legacy client.")
                return redirect('create_plan')
            
            try:
                client_id = int(client_id)
            except ValueError:
                messages.error(request, f"Invalid client ID format: {client_id}")
                return redirect('create_plan')
        
        # Get the legacy client
        client = get_object_or_404(Client, id=client_id)
        logger.info(f"Found legacy client: {client.name} (ID: {client.id})")
        
        # FIXED ACCESS CHECK - Same logic as create_plan view
        user_can_access = False
        
        if request.user.role == 'rm':
            # Check BOTH direct assignment AND profile mapping (like create_plan)
            directly_assigned = (client.user == request.user)
            profile_mapped = (hasattr(client, 'client_profile') and 
                            client.client_profile and 
                            client.client_profile.mapped_rm == request.user)
            
            user_can_access = directly_assigned or profile_mapped
            
            logger.info(f"RM {request.user.username} access check: Direct={directly_assigned}, Profile={profile_mapped}, CanAccess={user_can_access}")
            
        elif request.user.role == 'rm_head':
            # Check if client belongs to team member (both direct and profile)
            team_rms = User.objects.filter(manager=request.user, role='rm')
            
            team_direct = (client.user in team_rms) or (client.user == request.user)
            team_profile = (hasattr(client, 'client_profile') and 
                          client.client_profile and 
                          client.client_profile.mapped_rm in team_rms) or \
                         (hasattr(client, 'client_profile') and 
                          client.client_profile and 
                          client.client_profile.mapped_rm == request.user)
            
            user_can_access = team_direct or team_profile
            
            logger.info(f"RM Head {request.user.username} access check: TeamDirect={team_direct}, TeamProfile={team_profile}, CanAccess={user_can_access}")
            
        elif request.user.role in ['business_head', 'business_head_ops', 'top_management']:
            user_can_access = True
        
        # Deny access if user cannot access this client
        if not user_can_access:
            logger.warning(f"Access denied for {request.user.username} to client {client_id} ({client.name})")
            messages.error(request, f"Access denied - You don't have permission to create execution plans for {client.name}")
            return redirect('create_plan')
        
        logger.info(f"✅ Access granted for {request.user.username} to client {client.name}")
        
        # Get linked client profile if exists
        client_profile = None
        try:
            if hasattr(client, 'client_profile') and client.client_profile:
                client_profile = client.client_profile
                logger.info(f"Found linked profile for client {client.id}")
        except Exception as e:
            logger.warning(f"No linked profile for client {client_id}: {e}")
        
        # Get portfolio data from linked profile (if exists) - NO SCHEME MAPPING
        portfolio = ClientPortfolio.objects.none()  # Empty queryset
        portfolio_data = []
        total_portfolio_value = 0.0
        asset_allocation = None
        
        if client_profile:
            portfolio = ClientPortfolio.objects.filter(
                client_profile=client_profile,
                is_active=True
            ).order_by('scheme_name')
            
            logger.info(f"Found {portfolio.count()} portfolio holdings for client {client.id}")
            
            if portfolio.exists():
                # Use safe utility function for asset allocation
                from django.db.models import Sum
                
                # Get aggregated values and convert to float safely
                aggregation = portfolio.aggregate(
                    total_equity=Sum('equity_value'),
                    total_debt=Sum('debt_value'),
                    total_hybrid=Sum('hybrid_value'),
                    total_liquid=Sum('liquid_ultra_short_value'),
                    total_other=Sum('other_value'),
                    total_arbitrage=Sum('arbitrage_value'),
                    total_aum=Sum('total_value')
                )
                
                # Convert all Decimal results to float safely
                asset_allocation = {key: float(value or 0.0) for key, value in aggregation.items()}
                total_portfolio_value = asset_allocation.get('total_aum', 0.0)
                
                # Prepare portfolio data for action planning - NO SCHEME MAPPING
                for holding in portfolio:
                    # Convert Decimal fields to float consistently
                    total_value = float(holding.total_value) if holding.total_value else 0.0
                    units = float(holding.units) if holding.units else 0.0
                    equity_value = float(holding.equity_value) if holding.equity_value else 0.0
                    debt_value = float(holding.debt_value) if holding.debt_value else 0.0
                    hybrid_value = float(holding.hybrid_value) if holding.hybrid_value else 0.0
                    liquid_value = float(holding.liquid_ultra_short_value) if holding.liquid_ultra_short_value else 0.0
                    other_value = float(holding.other_value) if holding.other_value else 0.0
                    arbitrage_value = float(holding.arbitrage_value) if holding.arbitrage_value else 0.0
                    allocation_percentage = float(holding.allocation_percentage) if holding.allocation_percentage else 0.0
                    cost_value = float(holding.cost_value) if holding.cost_value else 0.0
                    nav_price = float(holding.nav_price) if holding.nav_price else 0.0
                    gain_loss = float(holding.gain_loss) if holding.gain_loss is not None else None
                    gain_loss_percentage = float(holding.gain_loss_percentage) if holding.gain_loss_percentage is not None else None
                    
                    portfolio_data.append({
                        'id': holding.id,
                        'scheme_name': holding.scheme_name,  # Use scheme name as-is from portfolio
                        'isin_number': holding.isin_number or '',  # Use ISIN as-is from portfolio
                        'current_value': total_value,
                        'units': units,
                        'equity_value': equity_value,
                        'debt_value': debt_value,
                        'hybrid_value': hybrid_value,
                        'liquid_value': liquid_value,
                        'other_value': other_value,
                        'arbitrage_value': arbitrage_value,
                        'allocation_percentage': allocation_percentage,
                        'primary_asset_class': holding.primary_asset_class,
                        'cost_value': cost_value,
                        'gain_loss': gain_loss,
                        'gain_loss_percentage': gain_loss_percentage,
                        # Action capabilities based on current value
                        'can_redeem': total_value > 0,
                        'can_switch': total_value > 0,
                        'can_stp': total_value >= 1000,  # Minimum for STP
                        'can_swp': total_value >= 5000,  # Minimum for SWP
                        # Additional info
                        'nav_price': nav_price,
                        'folio_number': holding.folio_number or '',
                        # Portfolio-specific fields (not mapped to MutualFundScheme)
                        'portfolio_scheme_name': holding.scheme_name,  # Original scheme name from portfolio
                        'portfolio_isin': holding.isin_number or '',    # Original ISIN from portfolio
                    })
        
        # If no portfolio data from profile, try to get any legacy portfolio data
        if not portfolio_data:
            logger.info(f"No portfolio data from profile, checking for any legacy data for client {client.id}")
            # You can add logic here if you have any legacy portfolio data directly linked to Client model
            # For now, we'll just note that no portfolio exists
            client_aum = float(client.aum) if client.aum else 0.0
            asset_allocation = {
                'total_aum': client_aum,
                'total_equity': 0.0,
                'total_debt': 0.0,
                'total_hybrid': 0.0,
                'total_liquid': 0.0,
                'total_other': 0.0,
                'total_arbitrage': 0.0,
            }
            total_portfolio_value = 0.0
        
        # Get all available schemes for adding new investments - ONLY FOR NEW INVESTMENTS
        all_schemes = MutualFundScheme.objects.filter(is_active=True).order_by('amc_name', 'scheme_name')
        
        # Get plan templates accessible to user
        templates = PlanTemplate.objects.filter(
            Q(is_public=True) | Q(created_by=request.user)
        ).filter(is_active=True)
        
        # Portfolio Action Types with descriptions - UPDATED FOR PORTFOLIO INDEPENDENCE
        portfolio_action_types = [
            {
                'type': 'redeem',
                'name': 'Redeem',
                'icon': 'fas fa-minus-circle',
                'color': 'danger',
                'description': 'Withdraw units or amount from existing portfolio schemes',
                'options': [
                    {'value': 'all_units', 'label': 'All Units'},
                    {'value': 'specific_amount', 'label': 'Specific Amount'},
                    {'value': 'specific_units', 'label': 'Specific Units'},
                ]
            },
            {
                'type': 'switch',
                'name': 'Switch',
                'icon': 'fas fa-exchange-alt',
                'color': 'primary',
                'description': 'Switch investment from existing portfolio scheme to new scheme',
                'options': [
                    {'value': 'all_units', 'label': 'All Units'},
                    {'value': 'specific_amount', 'label': 'Specific Amount'},
                    {'value': 'specific_units', 'label': 'Specific Units'},
                ]
            },
            {
                'type': 'stp',
                'name': 'STP',
                'icon': 'fas fa-arrow-right',
                'color': 'info',
                'description': 'Systematic Transfer Plan - Regular transfer from portfolio scheme to new scheme',
                'frequencies': [
                    {'value': 'monthly', 'label': 'Monthly'},
                    {'value': 'weekly', 'label': 'Weekly'},
                ]
            },
            {
                'type': 'sip',
                'name': 'SIP',
                'icon': 'fas fa-plus-circle',
                'color': 'success',
                'description': 'Systematic Investment Plan - New regular investments (independent of portfolio)',
                'frequencies': [
                    {'value': 'daily', 'label': 'Daily'},
                    {'value': 'weekly', 'label': 'Weekly'},
                    {'value': 'fortnightly', 'label': 'Fortnightly'},
                    {'value': 'monthly', 'label': 'Monthly'},
                ]
            },
            {
                'type': 'swp',
                'name': 'SWP',
                'icon': 'fas fa-arrow-down',
                'color': 'warning',
                'description': 'Systematic Withdrawal Plan - Regular withdrawals from portfolio schemes',
                'frequencies': [
                    {'value': 'monthly', 'label': 'Monthly'},
                    {'value': 'weekly', 'label': 'Weekly'},
                ]
            },
            {
                'type': 'purchase',
                'name': 'Purchase',
                'icon': 'fas fa-shopping-cart',
                'color': 'success',
                'description': 'One-time purchase of new mutual fund schemes (independent of portfolio)',
            }
        ]
        
        # Get existing execution plans for this legacy client
        existing_plans = ExecutionPlan.objects.filter(
            client=client
        ).order_by('-created_at')[:5]
        
        # Get existing portfolio action plans if client has portfolio
        existing_action_plans = []
        if client_profile and portfolio.exists():
            existing_action_plans = PortfolioActionPlan.objects.filter(
                client_portfolio__client_profile=client_profile
            ).order_by('-created_at')[:5]
        
        # Summary statistics for UI
        client_aum = float(client.aum) if client.aum else 0.0
        client_sip = float(client.sip_amount) if client.sip_amount else 0.0
        client_demat = int(client.demat_count) if client.demat_count else 0
        
        portfolio_summary = {
            'total_schemes': len(portfolio_data),
            'total_value': total_portfolio_value,
            'average_scheme_value': total_portfolio_value / len(portfolio_data) if portfolio_data else 0.0,
            'redeemable_schemes': sum(1 for p in portfolio_data if p.get('can_redeem', False)),
            'switchable_schemes': sum(1 for p in portfolio_data if p.get('can_switch', False)),
            'stp_eligible_schemes': sum(1 for p in portfolio_data if p.get('can_stp', False)),
            'swp_eligible_schemes': sum(1 for p in portfolio_data if p.get('can_swp', False)),
            # Client-specific data
            'client_aum': client_aum,
            'client_sip': client_sip,
            'client_demat': client_demat,
            'combined_aum': total_portfolio_value + client_aum,  # Both are now floats
        }
        
        # ALL schemes available for new investments (SIP/Purchase) - NO FILTERING
        available_investment_schemes = all_schemes
        
        # Prepare client data for template
        client_data = {
            'id': client.id,  # Use the integer client ID
            'name': client.name,
            'pan': client_profile.pan_number if client_profile else 'N/A',
            'email': client_profile.email if client_profile else '',
            'mobile': client_profile.mobile_number if client_profile else '',
            'type': 'legacy',
            'client_id': client.id,
            'profile_id': client_profile.id if client_profile else None,
            'contact_info': client.contact_info or '',
            'has_profile': client_profile is not None,
            'has_portfolio': len(portfolio_data) > 0,
            'mapped_rm': client.user.get_full_name() if client.user else 'Not Mapped',
            'created_by': client.created_by.get_full_name() if client.created_by else 'Unknown',
        }
        
        context = {
            'client': client_data,
            'client_object': client,  # Pass the actual client object for forms
            'client_profile': client_profile,
            'portfolio': portfolio,
            'portfolio_data': portfolio_data,
            'portfolio_summary': portfolio_summary,
            'all_schemes': all_schemes,
            'available_investment_schemes': available_investment_schemes,  # All schemes for new investments
            'templates': templates,
            'action_types': PlanAction.ACTION_TYPES if 'PlanAction' in globals() else [],
            'portfolio_action_types': portfolio_action_types,
            'asset_allocation': asset_allocation,
            'existing_plans': existing_plans,
            'existing_action_plans': existing_action_plans,
            'total_portfolio_value': total_portfolio_value,
            'has_portfolio_data': len(portfolio_data) > 0,
        }
        
        # Handle POST request for creating execution plan
        if request.method == 'POST':
            plan_name = request.POST.get('plan_name', '').strip()
            description = request.POST.get('description', '').strip()
            
            if not plan_name:
                messages.error(request, "Plan name is required.")
                return render(request, 'execution_plans/create_plan_step2.html', context)
            
            try:
                # Create execution plan linked to legacy client
                execution_plan = ExecutionPlan.objects.create(
                    client=client,  # Use the legacy Client object
                    plan_name=plan_name,
                    description=description,
                    created_by=request.user,
                    status='draft'
                )
                
                logger.info(f"Created execution plan {execution_plan.plan_id} for client {client.id}")
                messages.success(request, f"Execution plan '{plan_name}' created successfully!")
                
                # Redirect to plan detail or edit page
                return redirect('execution_plan_detail', plan_id=execution_plan.plan_id)
                
            except Exception as e:
                logger.error(f"Error creating execution plan for client {client_id}: {e}")
                messages.error(request, f"Error creating execution plan: {str(e)}")
        
        return render(request, 'execution_plans/create_plan_step2.html', context)
        
    except Exception as e:
        logger.error(f"Unexpected error in create_plan_step2 for client {client_id}: {e}")
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('create_plan')

# Add this new view to handle portfolio action creation from the execution plan interface
@login_required 
@require_http_methods(["POST"])
def create_portfolio_action_from_plan(request, client_id):
    """Create portfolio action from execution plan interface - FIXED VERSION"""
    try:
        data = json.loads(request.body)
        action_type = data.get('action_type')
        portfolio_id = data.get('portfolio_id')
        
        # Get the portfolio
        portfolio = get_object_or_404(ClientPortfolio, id=portfolio_id)
        
        # Check permissions
        if not can_create_action_plan(request.user, portfolio):
            return JsonResponse({'error': 'Access denied'}, status=403)
        
        # Create action plan
        plan_name = f"{action_type.upper()} - {portfolio.scheme_name}"
        
        # Check if PortfolioActionPlan model exists
        try:
            action_plan = PortfolioActionPlan.objects.create(
                client_portfolio=portfolio,
                plan_name=plan_name,
                description=f"{action_type.title()} action for {portfolio.client_name}",
                created_by=request.user
            )
        except Exception as e:
            logger.error(f"PortfolioActionPlan model not available: {str(e)}")
            return JsonResponse({'error': 'Portfolio action plan feature not available'}, status=500)
        
        # Create the action based on type
        action_data = {
            'action_plan': action_plan,
            'action_type': action_type,
            'source_scheme': portfolio.scheme_name,
            'priority': 1
        }
        
        # Handle different action types
        if action_type == 'redeem':
            action_data.update({
                'redeem_by': data.get('redeem_by', 'all_units'),
                'redeem_amount': data.get('redeem_amount') if data.get('redeem_by') == 'specific_amount' else None,
                'redeem_units': data.get('redeem_units') if data.get('redeem_by') == 'specific_units' else None,
            })
        elif action_type == 'switch':
            action_data.update({
                'target_scheme': data.get('target_scheme'),
                'switch_by': data.get('switch_by', 'all_units'),
                'switch_amount': data.get('switch_amount') if data.get('switch_by') == 'specific_amount' else None,
                'switch_units': data.get('switch_units') if data.get('switch_by') == 'specific_units' else None,
            })
        elif action_type == 'stp':
            action_data.update({
                'target_scheme': data.get('target_scheme'),
                'stp_amount': data.get('stp_amount'),
                'stp_frequency': data.get('stp_frequency', 'monthly'),
            })
        elif action_type == 'sip':
            action_data.update({
                'target_scheme': data.get('target_scheme'),
                'sip_amount': data.get('sip_amount'),
                'sip_frequency': data.get('sip_frequency', 'monthly'),
                'sip_date': data.get('sip_date'),
            })
        elif action_type == 'swp':
            action_data.update({
                'swp_amount': data.get('swp_amount'),
                'swp_frequency': data.get('swp_frequency', 'monthly'),
                'swp_date': data.get('swp_date'),
            })
        
        # Create the action - check if PortfolioAction model exists
        try:
            action = PortfolioAction.objects.create(**action_data)
            
            # Create workflow entry if ActionPlanWorkflow exists
            try:
                ActionPlanWorkflow.objects.create(
                    action_plan=action_plan,
                    from_status='',
                    to_status='draft',
                    changed_by=request.user,
                    notes=f'{action_type.title()} action plan created from execution plan interface'
                )
            except Exception as e:
                logger.warning(f"ActionPlanWorkflow model not available: {str(e)}")
            
            # Create simple action summary instead of calling get_action_summary
            action_summary = f"{action_type.title()} action for {portfolio.scheme_name}"
            if data.get('amount'):
                action_summary += f" - ₹{data.get('amount'):,.2f}"
            
            return JsonResponse({
                'success': True,
                'message': f'{action_type.title()} action created successfully',
                'action_plan_id': action_plan.id,
                'action_summary': action_summary
            })
            
        except Exception as e:
            logger.error(f"PortfolioAction model not available: {str(e)}")
            return JsonResponse({'error': 'Portfolio action feature not available'}, status=500)
        
    except Exception as e:
        logger.error(f"Error in create_portfolio_action_from_plan: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

# Helper function (add to your existing helper functions)
def can_create_action_plan(user, portfolio):
    """Check if user can create action plans for this portfolio"""
    if user.role in ['top_management', 'business_head']:
        return True
    elif user.role == 'rm_head':
        # RM Head can create plans for portfolios of their team members
        if hasattr(portfolio, 'mapped_rm') and portfolio.mapped_rm and portfolio.mapped_rm.manager == user:
            return True
    elif user.role == 'rm':
        # RM can create plans for their own portfolios
        if hasattr(portfolio, 'mapped_rm') and portfolio.mapped_rm == user:
            return True
    return False

@login_required
@require_http_methods(["POST"])
def save_execution_plan(request):
    """Save execution plan with actions - FIXED VERSION for Portfolio Independence"""
    try:
        data = json.loads(request.body)
        client_id = data.get('client_id')
        plan_name = data.get('plan_name')
        description = data.get('description', '')
        actions_data = data.get('actions', [])
        submit_for_approval = data.get('submit_for_approval', False)
        
        # Enhanced validation
        if not client_id:
            return JsonResponse({'error': 'Client ID is required'}, status=400)
        if not plan_name or not plan_name.strip():
            return JsonResponse({'error': 'Plan name is required'}, status=400)
        if not actions_data:
            return JsonResponse({'error': 'At least one action is required'}, status=400)
        
        # Get the client object (handle integer ID)
        try:
            client_id_int = int(client_id)
            client = get_object_or_404(Client, id=client_id_int)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid client ID format'}, status=400)
        
        # Check access permission
        if request.user.role == 'rm' and client.user != request.user:
            return JsonResponse({'error': 'Access denied - not your client'}, status=403)
        elif request.user.role == 'rm_head':
            team_members = User.objects.filter(manager=request.user, role='rm')
            if client.user not in team_members and client.user != request.user:
                return JsonResponse({'error': 'Access denied - client not in your team'}, status=403)
        elif request.user.role not in ['business_head', 'business_head_ops', 'top_management']:
            return JsonResponse({'error': 'Access denied - insufficient permissions'}, status=403)
        
        # Variables to track execution plan and actions
        execution_plan = None
        created_actions = []
        failed_actions = []
        
        try:
            with transaction.atomic():
                # Create execution plan
                execution_plan = ExecutionPlan.objects.create(
                    client=client,
                    plan_name=plan_name.strip(),
                    description=description.strip(),
                    created_by=request.user,
                    status='draft'
                )
                
                logger.info(f"Created execution plan {execution_plan.plan_id} for client {client.id}")
                
                # Create plan actions
                for i, action_data in enumerate(actions_data):
                    try:
                        # Enhanced logging
                        logger.info(f"Processing action {i + 1}: {action_data}")
                        
                        action = process_action_data_portfolio_independent(
                            action_data, execution_plan, client, i + 1
                        )
                        if action:
                            created_actions.append(action)
                            logger.info(f"Successfully created action {action.id}")
                        else:
                            failed_actions.append(f"Action {i + 1}: Unknown error")
                        
                    except Exception as e:
                        error_msg = f"Action {i + 1}: {str(e)}"
                        logger.error(f"Error creating action {i + 1}: {str(e)}")
                        logger.error(f"Action data: {action_data}")
                        failed_actions.append(error_msg)
                        continue
                
                if not created_actions:
                    return JsonResponse({
                        'error': 'No valid actions could be created',
                        'failed_actions': failed_actions
                    }, status=400)
                
                # Create workflow history
                PlanWorkflowHistory.objects.create(
                    execution_plan=execution_plan,
                    from_status='',
                    to_status='draft',
                    changed_by=request.user,
                    comments=f'Execution plan created with {len(created_actions)} actions'
                )
                
                # Submit for approval if requested
                if submit_for_approval and len(created_actions) > 0:
                    try:
                        if execution_plan.submit_for_approval():
                            PlanWorkflowHistory.objects.create(
                                execution_plan=execution_plan,
                                from_status='draft',
                                to_status='pending_approval',
                                changed_by=request.user,
                                comments='Plan submitted for approval'
                            )
                            logger.info(f"Plan {execution_plan.plan_id} submitted for approval")
                            
                            # Notify approval manager
                            try:
                                notify_approval_required(execution_plan)
                            except Exception as e:
                                logger.warning(f"Failed to send approval notification: {str(e)}")
                    except Exception as e:
                        logger.warning(f"Could not submit for approval: {str(e)}")
            
            # FIXED: Generate Excel file OUTSIDE of transaction
            excel_generated = False
            excel_error = None
            
            try:
                # Generate Excel in memory (no file system operations)
                excel_data = generate_execution_plan_excel_in_memory(execution_plan)
                if excel_data:
                    # Store Excel data reference (implementation depends on your storage strategy)
                    # For now, we'll just log success
                    excel_generated = True
                    logger.info(f"Generated Excel data for plan {execution_plan.plan_id}")
                else:
                    excel_error = "Excel generation returned no data"
            except Exception as e:
                excel_error = str(e)
                logger.warning(f"Could not generate Excel file: {excel_error}")
            
            # Prepare response
            response_data = {
                'success': True,
                'plan_id': execution_plan.plan_id,
                'plan_url': reverse('execution_plan_detail', args=[execution_plan.id]),
                'message': f'Execution plan created successfully with {len(created_actions)} actions',
                'actions_created': len(created_actions),
                'actions_failed': len(failed_actions),
                'excel_generated': excel_generated
            }
            
            if failed_actions:
                response_data['failed_actions'] = failed_actions
                response_data['message'] += f'. {len(failed_actions)} actions failed.'
            
            if excel_error:
                response_data['excel_error'] = excel_error
                response_data['message'] += f' (Excel generation issue: {excel_error})'
            
            return JsonResponse(response_data)
            
        except Exception as e:
            # If we get here, the transaction was rolled back
            logger.error(f"Transaction failed: {str(e)}")
            return JsonResponse({
                'error': f'Database transaction failed: {str(e)}',
                'actions_processed': len(created_actions),
                'actions_failed': len(failed_actions)
            }, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error saving execution plan: {str(e)}")
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


def process_action_data(action_data, execution_plan, client, priority):
    """Process individual action data and create appropriate action record - PORTFOLIO INDEPENDENT VERSION"""
    action_type = action_data.get('action_type')
    
    if not action_type:
        logger.error(f"No action type specified for action: {action_data}")
        return None
    
    # Handle scheme selection differently for portfolio vs new investment actions
    scheme_obj = None
    scheme_name = action_data.get('scheme_name', '')
    portfolio_holding_id = action_data.get('portfolio_holding_id')
    
    # For portfolio-based actions (redeem, switch from, STP from, SWP from)
    if action_type in ['redemption', 'switch', 'stp_start', 'swp_start']:
        # These actions work on existing portfolio holdings
        # We don't need to map to MutualFundScheme - just store the scheme name
        if portfolio_holding_id:
            try:
                # Get the portfolio holding to verify it exists and get accurate scheme name
                portfolio_holding = ClientPortfolio.objects.get(id=portfolio_holding_id)
                scheme_name = portfolio_holding.scheme_name
                logger.info(f"Using portfolio scheme: {scheme_name} from holding {portfolio_holding_id}")
            except ClientPortfolio.DoesNotExist:
                logger.warning(f"Portfolio holding {portfolio_holding_id} not found, using provided scheme name")
        
        # For portfolio actions, we don't need a MutualFundScheme object
        # The scheme_name from portfolio is sufficient
        if not scheme_name:
            raise ValueError(f"Scheme name is required for portfolio action {action_type}")
    
    # For new investment actions (purchase, SIP)
    elif action_type in ['purchase', 'sip_start', 'sip_modify']:
        scheme_id = action_data.get('target_scheme_id') or action_data.get('scheme_id')
        target_scheme_name = action_data.get('target_scheme_name') or action_data.get('scheme_name')
        
        if scheme_id:
            try:
                scheme_obj = MutualFundScheme.objects.get(id=scheme_id)
                scheme_name = scheme_obj.scheme_name
            except MutualFundScheme.DoesNotExist:
                logger.error(f"MutualFundScheme with ID {scheme_id} not found")
                raise ValueError(f"Selected scheme with ID {scheme_id} not found")
        elif target_scheme_name:
            # Try to find by name for new investments
            scheme_obj = MutualFundScheme.objects.filter(
                scheme_name__iexact=target_scheme_name
            ).first()
            if not scheme_obj:
                # Try partial match
                scheme_obj = MutualFundScheme.objects.filter(
                    scheme_name__icontains=target_scheme_name
                ).first()
            if scheme_obj:
                scheme_name = scheme_obj.scheme_name
            else:
                # For new investments, we need a valid MutualFundScheme
                logger.error(f"MutualFundScheme with name '{target_scheme_name}' not found")
                raise ValueError(f"Scheme '{target_scheme_name}' not found in available schemes")
        else:
            raise ValueError(f"Scheme selection is required for {action_type}")
    
    # Handle target scheme for switch/STP operations (new investment target)
    target_scheme_obj = None
    if action_type in ['switch', 'stp_start']:
        target_scheme_id = action_data.get('target_scheme_id')
        target_scheme_name = action_data.get('target_scheme_name')
        
        if target_scheme_id:
            try:
                target_scheme_obj = MutualFundScheme.objects.get(id=target_scheme_id)
            except MutualFundScheme.DoesNotExist:
                if target_scheme_name:
                    target_scheme_obj = MutualFundScheme.objects.filter(
                        scheme_name__icontains=target_scheme_name
                    ).first()
                    if not target_scheme_obj:
                        raise ValueError(f"Target scheme '{target_scheme_name}' not found")
                else:
                    raise ValueError(f"Target scheme with ID {target_scheme_id} not found")
        elif target_scheme_name:
            target_scheme_obj = MutualFundScheme.objects.filter(
                scheme_name__icontains=target_scheme_name
            ).first()
            if not target_scheme_obj:
                raise ValueError(f"Target scheme '{target_scheme_name}' not found")
    
    # Create the main plan action
    try:
        # Store portfolio-specific data in notes for portfolio actions
        notes = action_data.get('notes', '')
        if action_type in ['redemption', 'switch', 'stp_start', 'swp_start'] and portfolio_holding_id:
            portfolio_info = {
                'portfolio_holding_id': portfolio_holding_id,
                'portfolio_scheme_name': scheme_name,
                'action_type': action_type,
                'is_portfolio_action': True
            }
            notes = f"Portfolio Action: {json.dumps(portfolio_info)}\n{notes}".strip()
        
        action = PlanAction.objects.create(
            execution_plan=execution_plan,
            action_type=action_type,
            scheme=scheme_obj,  # Will be None for portfolio actions, valid for new investments
            target_scheme=target_scheme_obj,  # For switch/STP targets
            amount=safe_decimal(action_data.get('purchase_amount') or action_data.get('amount')),
            units=safe_decimal(action_data.get('units')),
            sip_date=safe_int(action_data.get('sip_date')),
            notes=notes,
            priority=priority,
            status='pending',
            # Store scheme name directly for portfolio actions
            portfolio_scheme_name=scheme_name if action_type in ['redemption', 'switch', 'stp_start', 'swp_start'] else None
        )
        
        logger.info(f"Successfully created action {action.id} for scheme '{scheme_name}' (portfolio action: {action_type in ['redemption', 'switch', 'stp_start', 'swp_start']})")
        return action
        
    except Exception as e:
        logger.error(f"Error creating PlanAction: {str(e)}")
        logger.error(f"Action data: {action_data}")
        raise


def create_sip_action(action, action_data, target_scheme):
    """Create SIP action record"""
    sip_amount = safe_decimal(action_data.get('sip_amount'))
    sip_date = safe_int(action_data.get('sip_date'))
    sip_frequency = action_data.get('sip_frequency', 'monthly')
    
    if not sip_amount or sip_amount <= 0:
        raise ValueError("Valid SIP amount is required")
    
    if not sip_date or sip_date < 1 or sip_date > 31:
        raise ValueError("Valid SIP date (1-31) is required")
    
    # Update the main action fields
    action.amount = sip_amount
    action.sip_date = sip_date
    action.frequency = sip_frequency
    action.transaction_type = 'purchase'
    action.save()
    
    # Create specific SIP record if you have a separate model
    # SIPAction.objects.create(
    #     plan_action=action,
    #     scheme=target_scheme,
    #     amount=sip_amount,
    #     frequency=sip_frequency,
    #     date=sip_date
    # )


def create_redeem_action(action, action_data, source_portfolio, source_scheme):
    """Create redeem action record"""
    redeem_by = action_data.get('redeem_by', 'all_units')
    redeem_amount = safe_decimal(action_data.get('redeem_amount'))
    redeem_units = safe_decimal(action_data.get('redeem_units'))
    
    if redeem_by == 'specific_amount' and (not redeem_amount or redeem_amount <= 0):
        raise ValueError("Valid redeem amount is required")
    
    if redeem_by == 'specific_units' and (not redeem_units or redeem_units <= 0):
        raise ValueError("Valid redeem units is required")
    
    # Update the main action fields
    action.transaction_type = 'redemption'
    action.amount = redeem_amount
    action.units = redeem_units
    action.save()
    
    # Store additional redeem-specific data in notes if needed
    redeem_details = {
        'redeem_by': redeem_by,
        'source_scheme': source_scheme,
        'portfolio_id': source_portfolio.id if source_portfolio else None
    }
    
    if action.notes:
        action.notes += f"\nRedeem details: {json.dumps(redeem_details)}"
    else:
        action.notes = f"Redeem details: {json.dumps(redeem_details)}"
    action.save()


def create_switch_action(action, action_data, source_portfolio, source_scheme, target_scheme):
    """Create switch action record"""
    switch_by = action_data.get('switch_by', 'all_units')
    switch_amount = safe_decimal(action_data.get('switch_amount'))
    switch_units = safe_decimal(action_data.get('switch_units'))
    
    if switch_by == 'specific_amount' and (not switch_amount or switch_amount <= 0):
        raise ValueError("Valid switch amount is required")
    
    if switch_by == 'specific_units' and (not switch_units or switch_units <= 0):
        raise ValueError("Valid switch units is required")
    
    # Update the main action fields
    action.transaction_type = 'switch'
    action.amount = switch_amount
    action.units = switch_units
    action.target_scheme = target_scheme
    action.save()
    
    # Store switch-specific data
    switch_details = {
        'switch_by': switch_by,
        'source_scheme': source_scheme,
        'target_scheme': target_scheme.scheme_name,
        'portfolio_id': source_portfolio.id if source_portfolio else None
    }
    
    if action.notes:
        action.notes += f"\nSwitch details: {json.dumps(switch_details)}"
    else:
        action.notes = f"Switch details: {json.dumps(switch_details)}"
    action.save()


def create_stp_action(action, action_data, source_portfolio, source_scheme, target_scheme):
    """Create STP action record"""
    stp_amount = safe_decimal(action_data.get('stp_amount'))
    stp_frequency = action_data.get('stp_frequency', 'monthly')
    
    if not stp_amount or stp_amount <= 0:
        raise ValueError("Valid STP amount is required")
    
    # Update the main action fields
    action.transaction_type = 'stp'
    action.amount = stp_amount
    action.frequency = stp_frequency
    action.target_scheme = target_scheme
    action.save()
    
    # Store STP-specific data
    stp_details = {
        'source_scheme': source_scheme,
        'target_scheme': target_scheme.scheme_name,
        'portfolio_id': source_portfolio.id if source_portfolio else None
    }
    
    if action.notes:
        action.notes += f"\nSTP details: {json.dumps(stp_details)}"
    else:
        action.notes = f"STP details: {json.dumps(stp_details)}"
    action.save()


def create_swp_action(action, action_data, source_portfolio, source_scheme):
    """Create SWP action record"""
    swp_amount = safe_decimal(action_data.get('swp_amount'))
    swp_frequency = action_data.get('swp_frequency', 'monthly')
    swp_date = safe_int(action_data.get('swp_date'))
    
    if not swp_amount or swp_amount <= 0:
        raise ValueError("Valid SWP amount is required")
    
    if not swp_date or swp_date < 1 or swp_date > 31:
        raise ValueError("Valid SWP date (1-31) is required")
    
    # Update the main action fields
    action.transaction_type = 'swp'
    action.amount = swp_amount
    action.frequency = swp_frequency
    action.sip_date = swp_date  # Reuse sip_date field for SWP date
    action.save()
    
    # Store SWP-specific data
    swp_details = {
        'source_scheme': source_scheme,
        'portfolio_id': source_portfolio.id if source_portfolio else None
    }
    
    if action.notes:
        action.notes += f"\nSWP details: {json.dumps(swp_details)}"
    else:
        action.notes = f"SWP details: {json.dumps(swp_details)}"
    action.save()


def safe_decimal(value):
    """Safely convert value to Decimal"""
    if value is None or value == '' or value == 'null':
        return None
    try:
        return Decimal(str(value))
    except (ValueError, TypeError, InvalidOperation):
        logger.warning(f"Could not convert to decimal: {value}")
        return None


def safe_int(value):
    """Safely convert value to int"""
    if value is None or value == '' or value == 'null':
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        logger.warning(f"Could not convert to int: {value}")
        return None


def notify_approval_required(execution_plan):
    """Notify managers that approval is required"""
    try:
        # Get approval managers based on your hierarchy
        approval_managers = User.objects.filter(
            role__in=['rm_head', 'business_head', 'business_head_ops']
        )
        
        for manager in approval_managers:
            # Send notification (implement your notification system)
            # send_notification(manager, f"Execution plan '{execution_plan.plan_name}' requires approval")
            pass
            
    except Exception as e:
        logger.error(f"Error sending approval notifications: {str(e)}")



def generate_execution_plan_excel(execution_plan):
    """Generate Excel file for execution plan with actual data"""
    try:
        # Initialize plan_actions first
        plan_actions = []
        
        # Get plan actions based on your model structure
        try:
            # Your model has 'actions' related name from PlanAction
            if hasattr(execution_plan, 'actions'):
                plan_actions = list(execution_plan.actions.all().order_by('priority', 'id'))
                logger.info(f"Found {len(plan_actions)} actions for execution plan {execution_plan.id}")
            else:
                logger.warning(f"No actions relationship found on ExecutionPlan {execution_plan.id}")
        except Exception as e:
            logger.error(f"Error getting plan actions: {str(e)}")
            plan_actions = []
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Execution Plan"
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20
        
        # Create header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Plan header information
        row = 1
        ws.merge_cells(f'A{row}:K{row}')
        plan_id = getattr(execution_plan, 'plan_id', f"Plan #{execution_plan.id}")
        ws[f'A{row}'] = f"EXECUTION PLAN - {plan_id}"
        ws[f'A{row}'].font = Font(name='Arial', size=16, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        
        # Plan details - Handle different client attribute names
        client_name = "Unknown Client"
        if hasattr(execution_plan, 'client') and execution_plan.client:
            if hasattr(execution_plan.client, 'name'):
                client_name = execution_plan.client.name
            elif hasattr(execution_plan.client, 'client_full_name'):
                client_name = execution_plan.client.client_full_name
            elif hasattr(execution_plan.client, 'full_name'):
                client_name = execution_plan.client.full_name
            else:
                client_name = str(execution_plan.client)
        elif hasattr(execution_plan, 'client_name'):
            client_name = execution_plan.client_name
        
        ws[f'A{row}'] = "Client:"
        ws[f'B{row}'] = client_name
        ws[f'F{row}'] = "Date:"
        ws[f'G{row}'] = execution_plan.created_at.strftime('%Y-%m-%d') if hasattr(execution_plan, 'created_at') else "N/A"
        row += 1
        
        # Get total amount from actions
        total_amount = 0
        if plan_actions:
            total_amount = sum(float(action.amount) for action in plan_actions if action.amount)
        
        ws[f'A{row}'] = "Total Amount:"
        ws[f'B{row}'] = f"₹{total_amount:,.2f}"
        ws[f'F{row}'] = "Status:"
        ws[f'G{row}'] = execution_plan.get_status_display() if hasattr(execution_plan, 'get_status_display') else (execution_plan.status if hasattr(execution_plan, 'status') else "N/A")
        row += 1
        
        # Add RM and other details if available
        if hasattr(execution_plan, 'created_by') and execution_plan.created_by:
            ws[f'A{row}'] = "Created By:"
            ws[f'B{row}'] = execution_plan.created_by.get_full_name() or execution_plan.created_by.username
            row += 1
        
        if hasattr(execution_plan, 'plan_name') and execution_plan.plan_name:
            ws[f'A{row}'] = "Plan Name:"
            ws[f'B{row}'] = execution_plan.plan_name
            row += 1
        
        if hasattr(execution_plan, 'description') and execution_plan.description:
            ws[f'A{row}'] = "Description:"
            ws[f'B{row}'] = execution_plan.description
            row += 1
        
        row += 1  # Empty row
        
        # Table headers - More comprehensive
        headers = [
            'Sr. No.',
            'Scheme Name',
            'ISIN',
            'Transaction Type',
            'Amount (₹)',
            'SIP Amount (₹)',
            'SIP Date',
            'Frequency',
            'Status',
            'Priority',
            'Remarks'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.style = header_style
        
        row += 1
        
        # Data rows for actions
        total_amount = 0
        
        for idx, action in enumerate(plan_actions, 1):
            # Get scheme information
            scheme = None
            scheme_name = "Unknown Scheme"
            isin_value = "N/A"
            
            if hasattr(action, 'scheme') and action.scheme:
                scheme = action.scheme
                # Handle different possible attribute names for scheme name
                if hasattr(scheme, 'scheme_name'):
                    scheme_name = scheme.scheme_name
                elif hasattr(scheme, 'name'):
                    scheme_name = scheme.name
                elif hasattr(scheme, 'scheme_title'):
                    scheme_name = scheme.scheme_title
                elif hasattr(scheme, 'fund_name'):
                    scheme_name = scheme.fund_name
                
                # Handle different possible attribute names for ISIN
                if hasattr(scheme, 'isin_growth'):
                    isin_value = scheme.isin_growth or "N/A"
                elif hasattr(scheme, 'isin_number'):
                    isin_value = scheme.isin_number or "N/A"
                elif hasattr(scheme, 'isin'):
                    isin_value = scheme.isin or "N/A"
                elif hasattr(scheme, 'isin_code'):
                    isin_value = scheme.isin_code or "N/A"
            
            # Get action details
            action_type = "N/A"
            if hasattr(action, 'action_type'):
                if hasattr(action, 'get_action_type_display'):
                    action_type = action.get_action_type_display()
                else:
                    action_type = action.action_type
            
            # Get amount
            amount = 0
            if hasattr(action, 'amount') and action.amount:
                amount = float(action.amount)
            
            # Get SIP details
            sip_amount = ""
            if hasattr(action, 'sip_amount') and action.sip_amount:
                sip_amount = f"₹{action.sip_amount:,.2f}"
            
            sip_date = ""
            if hasattr(action, 'sip_date') and action.sip_date:
                sip_date = str(action.sip_date)
            
            # Get frequency
            frequency = ""
            if hasattr(action, 'frequency'):
                frequency = action.frequency
            
            # Get status
            status = ""
            if hasattr(action, 'status'):
                if hasattr(action, 'get_status_display'):
                    status = action.get_status_display()
                else:
                    status = action.status
            
            # Get priority
            priority = ""
            if hasattr(action, 'priority'):
                priority = str(action.priority)
            
            # Get notes
            notes = ""
            if hasattr(action, 'notes') and action.notes:
                notes = action.notes
            
            # Build row data
            data = [
                idx,
                scheme_name,
                isin_value,
                action_type,
                f"₹{amount:,.2f}" if amount > 0 else "N/A",
                sip_amount if sip_amount else "N/A",
                sip_date if sip_date else "N/A",
                frequency if frequency else "N/A",
                status if status else "N/A",
                priority if priority else "N/A",
                notes if notes else ""
            ]
            
            # Add to total
            total_amount += amount
            
            # Write row data
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center' if col in [1, 3, 5, 7, 9, 10] else 'left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format amount columns
                if col in [5, 6] and value != "N/A" and value:
                    cell.font = Font(name='Arial', size=10, bold=True)
            
            row += 1
        
        # Add totals row
        if plan_actions:
            row += 1
            ws[f'A{row}'] = "TOTAL"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'E{row}'] = f"₹{total_amount:,.2f}"
            ws[f'E{row}'].font = Font(bold=True)
            
            # Add border to total row
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thick'),
                    bottom=Side(style='thick')
                )
        
        # Add summary information
        row += 3
        ws[f'A{row}'] = "Summary:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = f"Total Actions: {len(plan_actions)}"
        row += 1
        
        ws[f'A{row}'] = f"Total Amount: ₹{total_amount:,.2f}"
        row += 1
        
        ws[f'A{row}'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 1
        
        # Generate filename with consistent path handling
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        plan_id_clean = str(plan_id).replace('#', '').replace(' ', '_')
        filename = f"execution_plan_{plan_id_clean}_{timestamp}.xlsx"
        
        # Use consistent temporary directory - /tmp is standard for Lambda/serverless
        temp_dir = "/tmp"
        file_path = os.path.join(temp_dir, filename)
        
        # Ensure directory exists
        os.makedirs(temp_dir, exist_ok=True)
        
        # Save the workbook
        wb.save(file_path)
        
        logger.info(f"Excel file generated successfully: {filename}")
        logger.info(f"File saved at: {file_path}")
        logger.info(f"Total actions exported: {len(plan_actions)}")
        logger.info(f"Total amount: ₹{total_amount:,.2f}")
        
        # Return both filename and full path for flexibility
        return {
            'filename': filename,
            'file_path': file_path,
            'temp_dir': temp_dir
        }
        
    except Exception as e:
        logger.error(f"Error generating Excel for execution plan {execution_plan.id}: {str(e)}")
        logger.error(f"Error traceback: {traceback.format_exc()}")
        return None


def debug_scheme_attributes(scheme):
    """Debug function to check available attributes on MutualFundScheme object"""
    logger.info(f"Scheme object type: {type(scheme)}")
    logger.info(f"Scheme attributes: {dir(scheme)}")
    
    # Check for common ISIN attribute names
    isin_attributes = ['isin_number', 'isin', 'isin_code', 'scheme_isin', 'security_id']
    name_attributes = ['scheme_name', 'name', 'scheme_title', 'title']
    
    logger.info("Available ISIN-related attributes:")
    for attr in isin_attributes:
        if hasattr(scheme, attr):
            logger.info(f"  {attr}: {getattr(scheme, attr)}")
    
    logger.info("Available name-related attributes:")
    for attr in name_attributes:
        if hasattr(scheme, attr):
            logger.info(f"  {attr}: {getattr(scheme, attr)}")


def debug_scheme_attributes(scheme):
    """Debug function to check available attributes on MutualFundScheme object"""
    logger.info(f"Scheme object type: {type(scheme)}")
    logger.info(f"Scheme attributes: {dir(scheme)}")
    
    # Check for common ISIN attribute names
    isin_attributes = ['isin_number', 'isin', 'isin_code', 'scheme_isin', 'security_id']
    name_attributes = ['scheme_name', 'name', 'scheme_title', 'title']
    
    logger.info("Available ISIN-related attributes:")
    for attr in isin_attributes:
        if hasattr(scheme, attr):
            logger.info(f"  {attr}: {getattr(scheme, attr)}")
    
    logger.info("Available name-related attributes:")
    for attr in name_attributes:
        if hasattr(scheme, attr):
            logger.info(f"  {attr}: {getattr(scheme, attr)}")


def debug_scheme_attributes(scheme):
    """Debug function to check available attributes on MutualFundScheme object"""
    logger.info(f"Scheme object type: {type(scheme)}")
    logger.info(f"Scheme attributes: {dir(scheme)}")
    
    # Check for common ISIN attribute names
    isin_attributes = ['isin_number', 'isin', 'isin_code', 'scheme_isin', 'security_id']
    name_attributes = ['scheme_name', 'name', 'scheme_title', 'title']
    
    logger.info("Available ISIN-related attributes:")
    for attr in isin_attributes:
        if hasattr(scheme, attr):
            logger.info(f"  {attr}: {getattr(scheme, attr)}")
    
    logger.info("Available name-related attributes:")
    for attr in name_attributes:
        if hasattr(scheme, attr):
            logger.info(f"  {attr}: {getattr(scheme, attr)}")


def debug_execution_plan_structure(execution_plan):
    """Debug function to check available attributes on ExecutionPlan object"""
    logger.info(f"ExecutionPlan object type: {type(execution_plan)}")
    logger.info(f"ExecutionPlan ID: {execution_plan.id}")
    
    # Check for common client relationship names
    client_attributes = ['client_profile', 'client', 'client_name', 'customer']
    logger.info("Available client-related attributes:")
    for attr in client_attributes:
        if hasattr(execution_plan, attr):
            value = getattr(execution_plan, attr)
            logger.info(f"  {attr}: {value} (type: {type(value)})")
    
    # Check for plan items relationship
    items_attributes = ['execution_plan_items', 'executionplanitem_set', 'items', 'plan_items']
    logger.info("Available plan items-related attributes:")
    for attr in items_attributes:
        if hasattr(execution_plan, attr):
            try:
                value = getattr(execution_plan, attr)
                if hasattr(value, 'all'):
                    count = value.all().count()
                    logger.info(f"  {attr}: {count} items")
                else:
                    logger.info(f"  {attr}: {value}")
            except Exception as e:
                logger.info(f"  {attr}: Error accessing - {e}")
    
    # Check for common fields
    common_fields = ['total_amount', 'status', 'created_at', 'plan_id']
    logger.info("Available common fields:")
    for field in common_fields:
        if hasattr(execution_plan, field):
            value = getattr(execution_plan, field)
            logger.info(f"  {field}: {value}")

@login_required
def plan_detail(request, plan_id):
    """View execution plan details - Enhanced version"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check access permission
    if not can_access_plan(request.user, execution_plan):
        messages.error(request, "Access denied")
        return redirect('dashboard')
    
    # Get plan actions
    actions = execution_plan.actions.all().select_related('scheme', 'target_scheme', 'executed_by')
    
    # Get current portfolio if client has one
    current_portfolio = None
    if hasattr(execution_plan.client, 'client_profile') and execution_plan.client.client_profile:
        current_portfolio = ClientPortfolio.objects.filter(
            client_profile=execution_plan.client.client_profile,
            is_active=True
        ).order_by('scheme_name')
    
    # Get workflow history
    workflow_history = execution_plan.workflow_history.all().select_related('changed_by')
    
    # Get comments
    comments = execution_plan.comments.all().select_related('commented_by')
    
    # Calculate metrics if plan is completed
    metrics = None
    if execution_plan.status == 'completed':
        metrics, created = ExecutionMetrics.objects.get_or_create(execution_plan=execution_plan)
        if created or not metrics.updated_at:
            metrics.calculate_metrics()
    
    # Check permissions
    can_approve = can_approve_plan(request.user, execution_plan)
    can_execute = can_execute_plan(request.user, execution_plan)
    can_edit = can_edit_plan(request.user, execution_plan)
    
    context = {
        'execution_plan': execution_plan,
        'actions': actions,
        'current_portfolio': current_portfolio,
        'workflow_history': workflow_history,
        'comments': comments,
        'metrics': metrics,
        'can_approve': can_approve,
        'can_execute': can_execute,
        'can_edit': can_edit,
    }
    
    return render(request, 'execution_plans/plan_detail.html', context)


@login_required
@require_http_methods(["POST"])
def submit_for_approval(request, plan_id):
    """Submit execution plan for approval"""
    try:
        execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
        
        # Check permission
        if execution_plan.created_by != request.user:
            return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
        
        if execution_plan.status != 'draft':
            return JsonResponse({
                'success': False, 
                'error': f'Plan cannot be submitted for approval. Current status: {execution_plan.get_status_display()}'
            }, status=400)
        
        # Validate that plan has actions
        if not execution_plan.actions.exists():
            return JsonResponse({
                'success': False, 
                'error': 'Cannot submit plan without any actions'
            }, status=400)
        
        # Use transaction to ensure data consistency
        with transaction.atomic():
            # Update plan status
            execution_plan.status = 'pending_approval'
            execution_plan.submitted_at = timezone.now()
            execution_plan.save()
            
            # Create workflow history
            PlanWorkflowHistory.objects.create(
                execution_plan=execution_plan,
                from_status='draft',
                to_status='pending_approval',
                changed_by=request.user,
                comments='Plan submitted for approval'
            )
            
            # Notify approval managers
            try:
                notify_approval_required(execution_plan)
            except Exception as e:
                logger.warning(f"Failed to send approval notifications: {str(e)}")
        
        return JsonResponse({
            'success': True, 
            'message': 'Plan submitted for approval successfully',
            'new_status': execution_plan.status
        })
        
    except Exception as e:
        logger.error(f"Error submitting plan for approval: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': 'Failed to submit plan for approval'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def approve_plan(request, plan_id):
    """Approve execution plan - Role based approval"""
    try:
        execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
        
        # Check permission - only BH and TM can approve
        if not can_approve_plan(request.user, execution_plan):
            return JsonResponse({
                'success': False, 
                'error': 'Access denied. Only Business Head or Top Management can approve execution plans.'
            }, status=403)
        
        if execution_plan.status != 'pending_approval':
            return JsonResponse({
                'success': False,
                'error': f'Plan cannot be approved. Current status: {execution_plan.get_status_display()}'
            }, status=400)
        
        comments = request.POST.get('comments', '')
        
        # Use transaction for data consistency
        with transaction.atomic():
            # Approve the plan
            if execution_plan.approve(request.user):
                # Create workflow history
                PlanWorkflowHistory.objects.create(
                    execution_plan=execution_plan,
                    from_status='pending_approval',
                    to_status='approved',
                    changed_by=request.user,
                    comments=comments or 'Plan approved'
                )
                
                # Add comment if provided
                if comments:
                    PlanComment.objects.create(
                        execution_plan=execution_plan,
                        comment=comments,
                        commented_by=request.user,
                        is_internal=True
                    )
                
                # Notify plan creator
                try:
                    send_approval_notification(execution_plan, request.user)
                except Exception as e:
                    logger.warning(f"Failed to send approval notification: {str(e)}")
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Plan approved successfully',
                    'new_status': execution_plan.status
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'error': 'Failed to approve plan'
                }, status=400)
    
    except Exception as e:
        logger.error(f"Error approving plan: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': 'An error occurred while approving the plan'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def reject_plan(request, plan_id):
    """Reject execution plan - Role based rejection"""
    try:
        execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
        
        # Check permission - only BH and TM can reject
        if not can_approve_plan(request.user, execution_plan):
            return JsonResponse({
                'success': False, 
                'error': 'Access denied. Only Business Head or Top Management can reject execution plans.'
            }, status=403)
        
        if execution_plan.status != 'pending_approval':
            return JsonResponse({
                'success': False,
                'error': f'Plan cannot be rejected. Current status: {execution_plan.get_status_display()}'
            }, status=400)
        
        reason = request.POST.get('reason', '').strip()
        if not reason:
            return JsonResponse({
                'success': False, 
                'error': 'Rejection reason is required'
            }, status=400)
        
        # Use transaction for data consistency
        with transaction.atomic():
            # Reject the plan
            if execution_plan.reject(request.user, reason):
                # Create workflow history
                PlanWorkflowHistory.objects.create(
                    execution_plan=execution_plan,
                    from_status='pending_approval',
                    to_status='rejected',
                    changed_by=request.user,
                    comments=f'Plan rejected: {reason}'
                )
                
                # Add rejection comment
                PlanComment.objects.create(
                    execution_plan=execution_plan,
                    comment=f"Plan rejected by {request.user.get_full_name() or request.user.username}. Reason: {reason}",
                    commented_by=request.user,
                    is_internal=True
                )
                
                # Notify plan creator
                try:
                    send_rejection_notification(execution_plan, request.user, reason)
                except Exception as e:
                    logger.warning(f"Failed to send rejection notification: {str(e)}")
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Plan rejected successfully',
                    'new_status': execution_plan.status
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'error': 'Failed to reject plan'
                }, status=400)
    
    except Exception as e:
        logger.error(f"Error rejecting plan: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': 'An error occurred while rejecting the plan'
        }, status=500)

def send_approval_notification(execution_plan, approved_by):
    """Send notification when plan is approved"""
    try:
        # Notify plan creator
        creator = execution_plan.created_by
        if creator and creator.email:
            subject = f"Execution Plan Approved - {execution_plan.plan_name}"
            message = f"""
            Your execution plan "{execution_plan.plan_name}" for client {execution_plan.client.name} 
            has been approved by {approved_by.get_full_name() or approved_by.username}.
            
            Plan ID: {execution_plan.plan_id}
            Approved on: {execution_plan.approved_at.strftime('%Y-%m-%d %H:%M')}
            
            You can now proceed to get client approval.
            """
            
            # Send email notification here
            logger.info(f"Approval notification sent to {creator.email}")
    
    except Exception as e:
        logger.error(f"Error sending approval notification: {str(e)}")


def send_rejection_notification(execution_plan, rejected_by, reason):
    """Send notification when plan is rejected"""
    try:
        # Notify plan creator
        creator = execution_plan.created_by
        if creator and creator.email:
            subject = f"Execution Plan Rejected - {execution_plan.plan_name}"
            message = f"""
            Your execution plan "{execution_plan.plan_name}" for client {execution_plan.client.name} 
            has been rejected by {rejected_by.get_full_name() or rejected_by.username}.
            
            Plan ID: {execution_plan.plan_id}
            Rejected on: {execution_plan.rejected_at.strftime('%Y-%m-%d %H:%M')}
            
            Rejection Reason: {reason}
            
            Please review and resubmit the plan after making necessary changes.
            """
            
            # Send email notification here
            logger.info(f"Rejection notification sent to {creator.email}")
    
    except Exception as e:
        logger.error(f"Error sending rejection notification: {str(e)}")
        
        
@login_required
@require_http_methods(["POST"])
def send_to_client(request, plan_id):
    """Send execution plan to client via email"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check permission
    if execution_plan.created_by != request.user and not can_approve_plan(request.user, execution_plan):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if execution_plan.status != 'approved':
        return JsonResponse({'error': 'Only approved plans can be sent to client'}, status=400)
    
    try:
        # Send email to client
        client_email = execution_plan.client.contact_info  # Adjust based on your client model
        if '@' not in client_email:
            return JsonResponse({'error': 'Client email not found'}, status=400)
        
        subject = f"Investment Execution Plan - {execution_plan.plan_name}"
        
        # Prepare email content
        email_content = render_to_string('execution_plans/emails/client_plan.html', {
            'execution_plan': execution_plan,
            'client': execution_plan.client,
            'rm': execution_plan.created_by,
        })
        
        # Attach Excel file if available
        attachments = []
        if execution_plan.excel_file:
            file_path = os.path.join(settings.MEDIA_ROOT, execution_plan.excel_file.name)
            if os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    attachments.append((execution_plan.get_filename(), f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        
        send_mail(
            subject=subject,
            message='',
            html_message=email_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[client_email],
            attachments=attachments,
        )
        
        # Update plan status
        execution_plan.client_communication_sent = True
        execution_plan.save()
        
        # Add comment
        PlanComment.objects.create(
            execution_plan=execution_plan,
            comment='Execution plan sent to client via email',
            commented_by=request.user,
            is_internal=True
        )
        
        return JsonResponse({'success': True, 'message': 'Plan sent to client successfully'})
        
    except Exception as e:
        return JsonResponse({'error': f'Failed to send email: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def mark_client_approved(request, plan_id):
    """Mark plan as client approved with optional email notification"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check permission - only RM who created the plan can mark as client approved
    if execution_plan.created_by != request.user:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    # Get email parameters from request
    send_email = request.POST.get('send_email', 'false').lower() == 'true'
    include_excel = request.POST.get('include_excel', 'false').lower() == 'true'
    
    if execution_plan.mark_client_approved():
        # Create workflow history
        PlanWorkflowHistory.objects.create(
            execution_plan=execution_plan,
            from_status='approved',
            to_status='client_approved',
            changed_by=request.user,
            comments='Client approval confirmed'
        )
        
        response_data = {
            'success': True, 
            'message': 'Plan marked as client approved'
        }
        
        if send_email:
            try:
                # Send email to operations team
                send_execution_plan_email(
                    execution_plan=execution_plan,
                    email_type='client_approved',
                    include_excel=include_excel,
                    send_to_rm=True,
                    send_to_client=False
                )
                response_data['email_sent'] = True
                response_data['email_message'] = 'Operations team notified'
            except Exception as e:
                logger.error(f"Failed to send client approval email: {str(e)}")
                response_data['email_sent'] = False
                response_data['email_message'] = str(e)
        
        # Notify operations team via other channels (if needed)
        notify_ops_team(execution_plan)
        
        return JsonResponse(response_data)
    else:
        return JsonResponse({'error': 'Failed to mark as client approved'}, status=400)


# Add this modified view to your views.py file
# Make sure you have these imports at the top of your views.py:
# from django.db.models import Case, When, IntegerField, Q, Count, Sum
# from django.core.paginator import Paginator

@login_required
def ongoing_plans(request):
    """View ongoing execution plans with actual database statistics"""
    user = request.user
    
    # Get base queryset based on user role
    if user.role == 'rm':
        plans = ExecutionPlan.objects.filter(
            created_by=user,
            status__in=['pending_approval', 'approved', 'client_approved', 'in_execution']
        )
    elif user.role == 'rm_head':
        subordinate_rms = User.objects.filter(manager=user, role='rm')
        plans = ExecutionPlan.objects.filter(
            Q(created_by__in=subordinate_rms) | Q(created_by=user),
            status__in=['pending_approval', 'approved', 'client_approved', 'in_execution']
        )
    elif user.role in ['ops_exec', 'ops_team_lead']:
        plans = ExecutionPlan.objects.filter(
            status__in=['client_approved', 'in_execution']
        )
    elif user.role in ['business_head', 'business_head_ops']:
        plans = ExecutionPlan.objects.filter(
            status__in=['pending_approval', 'approved', 'client_approved', 'in_execution']
        )
    elif user.role == 'top_management':
        plans = ExecutionPlan.objects.filter(
            status__in=['pending_approval', 'approved', 'client_approved', 'in_execution']
        )
    else:
        plans = ExecutionPlan.objects.none()
    
    # Calculate actual statistics from database
    stats = {
        'pending_approval': plans.filter(status='pending_approval').count(),
        'approved': plans.filter(status='approved').count(),
        'client_approved': plans.filter(status='client_approved').count(),
        'in_execution': plans.filter(status='in_execution').count(),
    }
    
    # Apply filters
    status_filter = request.GET.get('status')
    if status_filter:
        plans = plans.filter(status=status_filter)
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        plans = plans.filter(
            Q(plan_name__icontains=search) |
            Q(client__name__icontains=search) |
            Q(plan_id__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Order by priority (pending approval first, then by creation date)
    plans = plans.annotate(
        status_priority=Case(
            When(status='pending_approval', then=1),
            When(status='approved', then=2),
            When(status='client_approved', then=3),
            When(status='in_execution', then=4),
            default=5,
            output_field=IntegerField()
        )
    ).order_by('status_priority', '-created_at')
    
    # Pagination
    paginator = Paginator(plans.select_related('client', 'created_by', 'approved_by'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Additional context for role-specific permissions
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'status_filter': status_filter,
        'search': search,
        'user_role': user.role,
        'can_create': user.role in ['rm', 'rm_head'],
        'can_approve': user.role in ['rm_head', 'business_head', 'business_head_ops', 'top_management'],
        'can_execute': user.role in ['ops_exec', 'ops_team_lead', 'business_head_ops'],
        'status_choices': ExecutionPlan.STATUS_CHOICES,
    }
    
    return render(request, 'execution_plans/ongoing_plans.html', context)


@login_required
def completed_plans(request):
    """View completed execution plans - FIXED VERSION"""
    # Get plans based on user role
    if request.user.role == 'rm':
        plans = ExecutionPlan.objects.filter(
            created_by=request.user,
            status__in=['completed', 'cancelled', 'rejected']
        )
    elif request.user.role == 'rm_head':
        subordinate_rms = User.objects.filter(manager=request.user, role='rm')
        plans = ExecutionPlan.objects.filter(
            Q(created_by__in=subordinate_rms) | Q(created_by=request.user),
            status__in=['completed', 'cancelled', 'rejected']
        )
    else:
        plans = ExecutionPlan.objects.filter(
            status__in=['completed', 'cancelled', 'rejected']
        )
    
    # Debug logging to check what plans exist
    all_completed = ExecutionPlan.objects.filter(status='completed')
    logger.info(f"Total completed plans in database: {all_completed.count()}")
    logger.info(f"User {request.user.username} can see {plans.count()} completed plans")
    
    # Filter by status if requested
    status_filter = request.GET.get('status')
    if status_filter:
        plans = plans.filter(status=status_filter)
    
    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        plans = plans.filter(created_at__gte=date_from)
    if date_to:
        plans = plans.filter(created_at__lte=date_to)
    
    # Search
    search = request.GET.get('search')
    if search:
        plans = plans.filter(
            Q(plan_name__icontains=search) |
            Q(client__name__icontains=search) |
            Q(plan_id__icontains=search)
        )
    
    # Order by completion date, then creation date (most recent first)
    plans = plans.order_by('-completed_at', '-created_at')
    
    # Calculate statistics for the filtered plans
    stats = {
        'total_completed': plans.filter(status='completed').count(),
        'total_cancelled': plans.filter(status='cancelled').count(),
        'total_rejected': plans.filter(status='rejected').count(),
        'total_all': plans.count(),
    }
    
    # Pagination
    paginator = Paginator(plans.select_related('client', 'created_by', 'approved_by'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
        'user_role': request.user.role,
        'stats': stats,
        'status_choices': [
            ('completed', 'Completed'),
            ('cancelled', 'Cancelled'),
            ('rejected', 'Rejected'),
        ],
        'title': 'Completed Plans',
    }
    
    return render(request, 'execution_plans/completed_plans.html', context)



@login_required
@require_http_methods(["POST"])
def start_execution(request, plan_id):
    """Start plan execution"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    if not can_execute_plan(request.user, execution_plan):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if execution_plan.start_execution(request.user):
        # Create workflow history
        PlanWorkflowHistory.objects.create(
            execution_plan=execution_plan,
            from_status='client_approved',
            to_status='in_execution',
            changed_by=request.user,
            comments='Plan execution started'
        )
        
        return JsonResponse({'success': True, 'message': 'Plan execution started'})
    else:
        return JsonResponse({'error': 'Failed to start execution'}, status=400)


@login_required
@require_http_methods(["POST"])
def execute_action(request, action_id):
    """Execute individual plan action"""
    action = get_object_or_404(PlanAction, id=action_id)
    
    if not can_execute_plan(request.user, action.execution_plan):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        # Get transaction details from request
        transaction_details = {
            'transaction_id': request.POST.get('transaction_id', ''),
            'amount': request.POST.get('executed_amount'),
            'units': request.POST.get('executed_units'),
            'nav_price': request.POST.get('nav_price'),
        }
        
        # Convert string values to Decimal where needed
        if transaction_details['amount']:
            transaction_details['amount'] = Decimal(str(transaction_details['amount']))
        if transaction_details['units']:
            transaction_details['units'] = Decimal(str(transaction_details['units']))
        if transaction_details['nav_price']:
            transaction_details['nav_price'] = Decimal(str(transaction_details['nav_price']))
        
        notes = request.POST.get('notes', '')
        if notes:
            action.notes = notes
            action.save()
        
        if action.execute(request.user, transaction_details):
            return JsonResponse({'success': True, 'message': 'Action executed successfully'})
        else:
            return JsonResponse({'error': 'Failed to execute action'}, status=400)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def mark_action_failed(request, action_id):
    """Mark action as failed"""
    action = get_object_or_404(PlanAction, id=action_id)
    
    if not can_execute_plan(request.user, action.execution_plan):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    reason = request.POST.get('reason', '')
    if not reason:
        return JsonResponse({'error': 'Failure reason is required'}, status=400)
    
    action.mark_failed(reason, request.user)
    
    return JsonResponse({'success': True, 'message': 'Action marked as failed'})


@login_required
def download_excel(request, plan_id):
    """Download execution plan Excel file - Vercel compatible version"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check access permission
    if not can_access_plan(request.user, execution_plan):
        raise Http404("Access denied")
    
    try:
        # Generate Excel in memory instead of saving to disk
        excel_data = generate_execution_plan_excel_in_memory(execution_plan)
        
        if not excel_data:
            return JsonResponse({'error': 'Unable to generate Excel file'}, status=500)
        
        # Create HTTP response with Excel data
        filename = f"{execution_plan.plan_id}_execution_plan.xlsx"
        
        response = HttpResponse(
            excel_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(excel_data)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        
        return response
            
    except Exception as e:
        logger.error(f"Error downloading Excel file for plan {plan_id}: {str(e)}")
        return JsonResponse({'error': f'Error downloading file: {str(e)}'}, status=500)

def generate_execution_plan_excel_portfolio_independent(execution_plan):
    """Generate Excel file for execution plan with portfolio-independent scheme handling"""
    try:
        # Get plan actions
        plan_actions = []
        try:
            if hasattr(execution_plan, 'actions'):
                plan_actions = list(execution_plan.actions.all().order_by('priority', 'id'))
                logger.info(f"Found {len(plan_actions)} actions for execution plan {execution_plan.id}")
        except Exception as e:
            logger.error(f"Error getting plan actions: {str(e)}")
            plan_actions = []
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Execution Plan"
        
        # Set column widths - improved sizing
        column_widths = {
            'A': 8,   # Sr. No.
            'B': 45,  # Scheme Name (increased)
            'C': 25,  # Source/ISIN (increased)
            'D': 18,  # Transaction Type (increased)
            'E': 18,  # Amount (increased)
            'F': 18,  # SIP Amount (increased)
            'G': 12,  # SIP Date
            'H': 15,  # Frequency
            'I': 15,  # Status
            'J': 10,  # Priority
            'K': 20   # Action Source
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Create styles
        header_style = NamedStyle(name="header")
        header_style.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title style
        title_style = NamedStyle(name="title")
        title_style.font = Font(name='Arial', size=16, bold=True, color='1F4E79')
        title_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # Info style
        info_label_style = NamedStyle(name="info_label")
        info_label_style.font = Font(name='Arial', size=11, bold=True)
        info_label_style.alignment = Alignment(horizontal='left', vertical='center')
        
        info_value_style = NamedStyle(name="info_value")
        info_value_style.font = Font(name='Arial', size=11)
        info_value_style.alignment = Alignment(horizontal='left', vertical='center')
        
        # Plan header information
        row = 1
        ws.merge_cells(f'A{row}:K{row}')
        plan_id = getattr(execution_plan, 'plan_id', f"Plan #{execution_plan.id}")
        ws[f'A{row}'] = f"EXECUTION PLAN - {plan_id}"
        ws[f'A{row}'].style = title_style
        ws.row_dimensions[row].height = 25
        row += 2  # Extra space
        
        # Plan details in a more structured format
        client_name = "Unknown Client"
        if hasattr(execution_plan, 'client') and execution_plan.client:
            if hasattr(execution_plan.client, 'name'):
                client_name = execution_plan.client.name
            elif hasattr(execution_plan.client, 'client_full_name'):
                client_name = execution_plan.client.client_full_name
        
        # Client and Date row
        ws[f'A{row}'] = "Client:"
        ws[f'A{row}'].style = info_label_style
        ws[f'B{row}'] = client_name
        ws[f'B{row}'].style = info_value_style
        
        ws[f'G{row}'] = "Date Created:"
        ws[f'G{row}'].style = info_label_style
        ws[f'H{row}'] = execution_plan.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(execution_plan, 'created_at') else "N/A"
        ws[f'H{row}'].style = info_value_style
        row += 1
        
        # Calculate totals
        total_amount = 0
        total_sip_amount = 0
        if plan_actions:
            for action in plan_actions:
                if hasattr(action, 'amount') and action.amount:
                    total_amount += float(action.amount)
                if hasattr(action, 'sip_amount') and action.sip_amount:
                    total_sip_amount += float(action.sip_amount)
        
        # Total Amount and Status row
        ws[f'A{row}'] = "Total Amount:"
        ws[f'A{row}'].style = info_label_style
        ws[f'B{row}'] = f"₹{total_amount:,.2f}"
        ws[f'B{row}'].style = info_value_style
        ws[f'B{row}'].font = Font(name='Arial', size=11, bold=True, color='0D5016')
        
        ws[f'G{row}'] = "Status:"
        ws[f'G{row}'].style = info_label_style
        status_value = execution_plan.get_status_display() if hasattr(execution_plan, 'get_status_display') else execution_plan.status
        ws[f'H{row}'] = status_value
        ws[f'H{row}'].style = info_value_style
        row += 1
        
        # Total SIP Amount and Created By row
        if total_sip_amount > 0:
            ws[f'A{row}'] = "Total SIP Amount:"
            ws[f'A{row}'].style = info_label_style
            ws[f'B{row}'] = f"₹{total_sip_amount:,.2f}"
            ws[f'B{row}'].style = info_value_style
            ws[f'B{row}'].font = Font(name='Arial', size=11, bold=True, color='7B2CBF')
        
        if hasattr(execution_plan, 'created_by') and execution_plan.created_by:
            ws[f'G{row}'] = "Created By:"
            ws[f'G{row}'].style = info_label_style
            ws[f'H{row}'] = execution_plan.created_by.get_full_name() or execution_plan.created_by.username
            ws[f'H{row}'].style = info_value_style
        row += 1
        
        # Actions count
        ws[f'A{row}'] = "Total Actions:"
        ws[f'A{row}'].style = info_label_style
        ws[f'B{row}'] = str(len(plan_actions))
        ws[f'B{row}'].style = info_value_style
        row += 2  # Extra space before table
        
        # Table headers
        headers = [
            'Sr. No.',
            'Scheme Name',
            'Source/ISIN',
            'Transaction Type',
            'Amount (₹)',
            'SIP Amount (₹)',
            'SIP Date',
            'Frequency',
            'Status',
            'Priority',
            'Action Source'
        ]
        
        # Add header row with improved styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.style = header_style
        
        ws.row_dimensions[row].height = 20
        row += 1
        
        # Data rows for actions
        for idx, action in enumerate(plan_actions, 1):
            # Get scheme information - PORTFOLIO INDEPENDENT
            scheme_name = "Unknown Scheme"
            source_info = "N/A"
            action_source = "Unknown"
            
            # Check if this is a portfolio action or new investment
            if hasattr(action, 'portfolio_scheme_name') and action.portfolio_scheme_name:
                # Portfolio-based action
                scheme_name = action.portfolio_scheme_name
                source_info = "Portfolio Holdings"
                action_source = "Existing Portfolio"
                
                # Try to get portfolio holding details from notes
                try:
                    if action.notes and 'Portfolio Action:' in action.notes:
                        import json
                        json_start = action.notes.find('{')
                        json_end = action.notes.find('}') + 1
                        if json_start >= 0 and json_end > json_start:
                            portfolio_info = json.loads(action.notes[json_start:json_end])
                            if 'portfolio_holding_id' in portfolio_info:
                                source_info = f"Portfolio ID: {portfolio_info['portfolio_holding_id']}"
                except:
                    pass
                    
            elif hasattr(action, 'scheme') and action.scheme:
                # New investment action
                scheme = action.scheme
                scheme_name = scheme.scheme_name
                action_source = "New Investment"
                
                # Get ISIN for new investments
                if hasattr(scheme, 'isin_growth') and scheme.isin_growth:
                    source_info = scheme.isin_growth
                elif hasattr(scheme, 'isin_number') and scheme.isin_number:
                    source_info = scheme.isin_number
                else:
                    source_info = "New Scheme"
            
            # Get action type
            action_type = "N/A"
            if hasattr(action, 'action_type'):
                if hasattr(action, 'get_action_type_display'):
                    action_type = action.get_action_type_display()
                else:
                    action_type = str(action.action_type).replace('_', ' ').title()
            
            # Get amounts with better formatting
            amount = 0
            if hasattr(action, 'amount') and action.amount:
                amount = float(action.amount)
            
            sip_amount = 0
            sip_amount_str = "N/A"
            if hasattr(action, 'sip_amount') and action.sip_amount:
                sip_amount = float(action.sip_amount)
                sip_amount_str = f"₹{sip_amount:,.2f}"
            
            # Get other details with better handling
            sip_date = "N/A"
            if hasattr(action, 'sip_date') and action.sip_date:
                sip_date = action.sip_date.strftime('%d-%m-%Y') if hasattr(action.sip_date, 'strftime') else str(action.sip_date)
            
            frequency = "N/A"
            if hasattr(action, 'frequency') and action.frequency:
                frequency = str(action.frequency).replace('_', ' ').title()
            
            status = "N/A"
            if hasattr(action, 'status') and action.status:
                if hasattr(action, 'get_status_display'):
                    status = action.get_status_display()
                else:
                    status = str(action.status).replace('_', ' ').title()
            
            priority = "N/A"
            if hasattr(action, 'priority') and action.priority is not None:
                priority = str(action.priority)
            
            # Build row data
            data = [
                idx,
                scheme_name,
                source_info,
                action_type,
                f"₹{amount:,.2f}" if amount > 0 else "₹0.00",
                sip_amount_str,
                sip_date,
                frequency,
                status,
                priority,
                action_source
            ]
            
            # Write row data with improved formatting
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                # Set alignment based on column type
                if col == 1:  # Serial number
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col in [5, 6]:  # Amount columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if value != "N/A" and value != "₹0.00":
                        cell.font = Font(name='Arial', size=10, bold=True)
                elif col in [7, 9, 10]:  # Date, Status, Priority
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Color coding for action source with better colors
                if col == 11:  # Action Source column
                    if value == "Existing Portfolio":
                        cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
                        cell.font = Font(name='Arial', size=10, bold=True, color='1565C0')
                    elif value == "New Investment":
                        cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                        cell.font = Font(name='Arial', size=10, bold=True, color='2E7D32')
            
            # Set row height for better readability
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Add totals row with better styling
        if plan_actions:
            row += 1
            # Create totals row
            total_cells = ['TOTAL', '', '', '', f"₹{total_amount:,.2f}", 
                          f"₹{total_sip_amount:,.2f}" if total_sip_amount > 0 else "₹0.00", 
                          '', '', '', '', '']
            
            for col, value in enumerate(total_cells, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = Font(name='Arial', size=11, bold=True)
                
                if col in [1, 5, 6]:  # Total label and amount columns
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                
                if col in [5, 6]:  # Amount columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thick'),
                    bottom=Side(style='thick')
                )
            
            ws.row_dimensions[row].height = 20
        
        # Add legend with better formatting
        row += 3
        ws[f'A{row}'] = "Legend:"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='1F4E79')
        row += 1
        
        # Portfolio actions legend
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "🔵 Existing Portfolio Actions"
        ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        ws.merge_cells(f'D{row}:K{row}')
        ws[f'D{row}'] = "Actions on current portfolio holdings (Redeem, Switch From, STP From, SWP)"
        ws[f'D{row}'].font = Font(name='Arial', size=10)
        ws[f'D{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
        
        # New investment legend
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "🟢 New Investment Actions"
        ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        ws.merge_cells(f'D{row}:K{row}')
        ws[f'D{row}'] = "Fresh investments in mutual fund schemes (Purchase, SIP, Switch To, STP To)"
        ws[f'D{row}'].font = Font(name='Arial', size=10)
        ws[f'D{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
        
        # Add summary with better formatting
        row += 2
        ws[f'A{row}'] = "Summary:"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='1F4E79')
        row += 1
        
        portfolio_actions = sum(1 for action in plan_actions if hasattr(action, 'portfolio_scheme_name') and action.portfolio_scheme_name)
        new_investment_actions = len(plan_actions) - portfolio_actions
        
        summary_data = [
            f"Total Actions: {len(plan_actions)}",
            f"Portfolio Actions: {portfolio_actions}",
            f"New Investment Actions: {new_investment_actions}",
            f"Total Investment Amount: ₹{total_amount:,.2f}",
            f"Total SIP Amount: ₹{total_sip_amount:,.2f}" if total_sip_amount > 0 else None,
            f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        
        for item in summary_data:
            if item:
                ws[f'A{row}'] = item
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                row += 1
        
        # Save to BytesIO buffer
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Portfolio-independent Excel generated for plan {execution_plan.id}")
        logger.info(f"Portfolio actions: {portfolio_actions}, New investments: {new_investment_actions}")
        logger.info(f"Total amount: ₹{total_amount:,.2f}, Total SIP: ₹{total_sip_amount:,.2f}")
        
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating portfolio-independent Excel: {str(e)}")
        return None


# Updated save_execution_plan function
@login_required
@require_http_methods(["POST"])
def save_execution_plan_portfolio_independent(request):
    """Save execution plan with portfolio-independent actions"""
    try:
        data = json.loads(request.body)
        client_id = data.get('client_id')
        plan_name = data.get('plan_name')
        description = data.get('description', '')
        actions_data = data.get('actions', [])
        submit_for_approval = data.get('submit_for_approval', False)
        
        # Enhanced validation
        if not client_id:
            return JsonResponse({'error': 'Client ID is required'}, status=400)
        if not plan_name or not plan_name.strip():
            return JsonResponse({'error': 'Plan name is required'}, status=400)
        if not actions_data:
            return JsonResponse({'error': 'At least one action is required'}, status=400)
        
        # Get the client object
        try:
            client_id_int = int(client_id)
            client = get_object_or_404(Client, id=client_id_int)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid client ID format'}, status=400)
        
        # Check access permission
        if request.user.role == 'rm' and client.user != request.user:
            return JsonResponse({'error': 'Access denied - not your client'}, status=403)
        elif request.user.role == 'rm_head':
            team_members = User.objects.filter(manager=request.user, role='rm')
            if client.user not in team_members and client.user != request.user:
                return JsonResponse({'error': 'Access denied - client not in your team'}, status=403)
        elif request.user.role not in ['business_head', 'business_head_ops', 'top_management']:
            return JsonResponse({'error': 'Access denied - insufficient permissions'}, status=403)
        
        with transaction.atomic():
            # Create execution plan
            execution_plan = ExecutionPlan.objects.create(
                client=client,
                plan_name=plan_name.strip(),
                description=description.strip(),
                created_by=request.user,
                status='draft'
            )
            
            logger.info(f"Created execution plan {execution_plan.plan_id} for client {client.id}")
            
            # Process actions with portfolio independence
            created_actions = []
            failed_actions = []
            
            for i, action_data in enumerate(actions_data):
                try:
                    logger.info(f"Processing action {i + 1}: {action_data}")
                    
                    action = process_action_data_portfolio_independent(
                        action_data, execution_plan, client, i + 1
                    )
                    if action:
                        created_actions.append(action)
                        logger.info(f"Successfully created action {action.id}")
                    else:
                        failed_actions.append(f"Action {i + 1}: Unknown error")
                    
                except Exception as e:
                    error_msg = f"Action {i + 1}: {str(e)}"
                    logger.error(f"Error creating action {i + 1}: {str(e)}")
                    failed_actions.append(error_msg)
                    continue
            
            if not created_actions:
                return JsonResponse({
                    'error': 'No valid actions could be created',
                    'failed_actions': failed_actions
                }, status=400)
            
            # Generate Excel file with portfolio independence
            try:
                excel_data = generate_execution_plan_excel_portfolio_independent(execution_plan)
                if excel_data:
                    # Save Excel data (implementation depends on your file storage)
                    logger.info(f"Generated portfolio-independent Excel for plan {execution_plan.plan_id}")
            except Exception as e:
                logger.warning(f"Could not generate Excel file: {str(e)}")
            
            # Create workflow history
            PlanWorkflowHistory.objects.create(
                execution_plan=execution_plan,
                from_status='',
                to_status='draft',
                changed_by=request.user,
                comments=f'Execution plan created with {len(created_actions)} actions (portfolio-independent)'
            )
            
            # Submit for approval if requested
            if submit_for_approval and len(created_actions) > 0:
                try:
                    if execution_plan.submit_for_approval():
                        PlanWorkflowHistory.objects.create(
                            execution_plan=execution_plan,
                            from_status='draft',
                            to_status='pending_approval',
                            changed_by=request.user,
                            comments='Plan submitted for approval'
                        )
                        logger.info(f"Plan {execution_plan.plan_id} submitted for approval")
                except Exception as e:
                    logger.warning(f"Could not submit for approval: {str(e)}")
            
            response_data = {
                'success': True,
                'plan_id': execution_plan.plan_id,
                'plan_url': reverse('execution_plan_detail', args=[execution_plan.id]),
                'message': f'Portfolio-independent execution plan created successfully with {len(created_actions)} actions',
                'actions_created': len(created_actions),
                'actions_failed': len(failed_actions),
                'portfolio_independent': True
            }
            
            if failed_actions:
                response_data['failed_actions'] = failed_actions
                response_data['message'] += f'. {len(failed_actions)} actions failed.'
            
            return JsonResponse(response_data)
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error saving portfolio-independent execution plan: {str(e)}")
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


def process_action_data_portfolio_independent(action_data, execution_plan, client, priority):
    """Process individual action data with complete portfolio independence - FIXED VERSION"""
    action_type = action_data.get('action_type')
    
    if not action_type:
        logger.error(f"No action type specified for action: {action_data}")
        raise ValueError("Action type is required")
    
    # Map action types to standardized values
    action_type_mapping = {
        'redeem': 'redemption',
        'redemption': 'redemption',
        'switch': 'switch',
        'stp': 'stp_start',
        'stp_start': 'stp_start',
        'swp': 'swp_start',
        'swp_start': 'swp_start',
        'purchase': 'purchase',
        'sip': 'sip_start',
        'sip_start': 'sip_start',
    }
    
    action_type = action_type_mapping.get(action_type, action_type)
    
    # Initialize variables
    scheme_obj = None
    target_scheme_obj = None
    portfolio_scheme_name = None
    portfolio_holding_id = None
    portfolio_isin = None
    portfolio_folio_number = None
    action_mode = 'new_investment'  # Default
    notes = action_data.get('notes', '')
    
    # Handle portfolio-based actions
    if action_type in ['redemption', 'switch', 'stp_start', 'swp_start']:
        # Check multiple possible field names for portfolio reference
        portfolio_holding_id = (
            action_data.get('portfolio_holding_id') or 
            action_data.get('portfolio_id') or 
            action_data.get('source_portfolio_id')
        )
        
        scheme_name = (
            action_data.get('portfolio_scheme_name') or
            action_data.get('source_scheme_name') or
            action_data.get('scheme_name') or
            action_data.get('source_scheme') or
            action_data.get('scheme')
        )
        
        if portfolio_holding_id:
            try:
                # Verify portfolio holding exists and get scheme details
                portfolio_holding = ClientPortfolio.objects.get(id=portfolio_holding_id)
                portfolio_scheme_name = portfolio_holding.scheme_name
                portfolio_isin = portfolio_holding.isin_number or ''
                portfolio_folio_number = portfolio_holding.folio_number or ''
                action_mode = 'portfolio'
                
                # Store portfolio information in notes
                portfolio_info = {
                    'portfolio_holding_id': portfolio_holding_id,
                    'original_scheme_name': portfolio_holding.scheme_name,
                    'original_isin': portfolio_holding.isin_number or '',
                    'folio_number': portfolio_holding.folio_number or '',
                    'original_value': float(portfolio_holding.total_value) if portfolio_holding.total_value else 0,
                    'original_units': float(portfolio_holding.units) if portfolio_holding.units else 0,
                    'action_type': action_type,
                    'is_portfolio_action': True
                }
                
                # Append to existing notes instead of replacing
                if notes:
                    notes = f"Portfolio Action: {json.dumps(portfolio_info)}\n{notes}"
                else:
                    notes = f"Portfolio Action: {json.dumps(portfolio_info)}"
                
                logger.info(f"Portfolio action for holding {portfolio_holding_id}: {portfolio_scheme_name}")
                
            except ClientPortfolio.DoesNotExist:
                logger.error(f"Portfolio holding {portfolio_holding_id} not found")
                raise ValueError(f"Portfolio holding {portfolio_holding_id} not found")
            except Exception as e:
                logger.error(f"Error accessing portfolio holding {portfolio_holding_id}: {str(e)}")
                raise ValueError(f"Error accessing portfolio holding: {str(e)}")
                
        elif scheme_name:
            # Direct scheme name provided (fallback for portfolio actions)
            portfolio_scheme_name = scheme_name
            action_mode = 'portfolio'
            logger.info(f"Using provided portfolio scheme name: {scheme_name}")
        else:
            # This might be a new investment action
            action_mode = 'new_investment'
    
    # Handle new investment actions or when no portfolio reference is found
    if action_mode == 'new_investment' or action_type in ['purchase', 'sip_start']:
        scheme_id = action_data.get('scheme_id') or action_data.get('target_scheme_id')
        
        if scheme_id:
            try:
                scheme_obj = MutualFundScheme.objects.get(id=scheme_id)
                action_mode = 'new_investment'
                logger.info(f"New investment scheme: {scheme_obj.scheme_name}")
            except MutualFundScheme.DoesNotExist:
                logger.error(f"MutualFundScheme with ID {scheme_id} not found")
                raise ValueError(f"Selected scheme with ID {scheme_id} not found")
        else:
            # If no scheme_id provided and no portfolio reference, this is an error for new investments
            if not portfolio_scheme_name and action_type in ['purchase', 'sip_start']:
                raise ValueError(f"Scheme ID required for new investment action {action_type}")
    
    # Handle target scheme for switch/STP operations
    if action_type in ['switch', 'stp_start']:
        target_scheme_id = action_data.get('target_scheme_id')
        
        if target_scheme_id:
            try:
                target_scheme_obj = MutualFundScheme.objects.get(id=target_scheme_id)
                logger.info(f"Target scheme for {action_type}: {target_scheme_obj.scheme_name}")
            except MutualFundScheme.DoesNotExist:
                raise ValueError(f"Target scheme with ID {target_scheme_id} not found")
        else:
            raise ValueError(f"Target scheme required for {action_type}")
    
    # Validate that we have either scheme or portfolio reference
    if not scheme_obj and not portfolio_scheme_name:
        raise ValueError(f"Invalid action configuration: no scheme reference found for {action_type}")
    
    # Log processed values for debugging
    logger.info(f"Processed values - Mode: {action_mode}, Scheme: {scheme_obj}, Portfolio: {portfolio_scheme_name}")
    
    # Create the plan action
    try:
        action = PlanAction.objects.create(
            execution_plan=execution_plan,
            action_type=action_type,
            scheme=scheme_obj,  # Only for new investments
            target_scheme=target_scheme_obj,  # Only for switch/STP targets
            portfolio_scheme_name=portfolio_scheme_name,  # Only for portfolio actions
            portfolio_holding_id=portfolio_holding_id,
            portfolio_isin=portfolio_isin,
            portfolio_folio_number=portfolio_folio_number,
            action_mode=action_mode,
            amount=safe_decimal(action_data.get('amount') or action_data.get('redeem_amount')),
            units=safe_decimal(action_data.get('units') or action_data.get('redeem_units')),
            sip_date=safe_int(action_data.get('sip_date')),
            frequency=action_data.get('frequency'),
            notes=notes,
            priority=priority,
            status='pending'
        )
        
        # Create portfolio mapping if this is a portfolio action
        if action_mode == 'portfolio' and portfolio_holding_id:
            try:
                portfolio_holding = ClientPortfolio.objects.get(id=portfolio_holding_id)
                
                # Check if PortfolioActionMapping model exists
                try:
                    PortfolioActionMapping.objects.create(
                        plan_action=action,
                        portfolio_holding_id=portfolio_holding_id,
                        original_scheme_name=portfolio_holding.scheme_name,
                        original_isin=portfolio_holding.isin_number or '',
                        original_folio_number=portfolio_holding.folio_number or '',
                        original_units=portfolio_holding.units,
                        original_value=portfolio_holding.total_value,
                    )
                    logger.info(f"Created portfolio mapping for action {action.id}")
                except Exception as mapping_error:
                    logger.warning(f"Could not create portfolio mapping (model might not exist): {str(mapping_error)}")
                    
            except ClientPortfolio.DoesNotExist:
                logger.warning(f"Portfolio holding {portfolio_holding_id} not found for mapping")
            except Exception as e:
                logger.warning(f"Could not create portfolio mapping: {str(e)}")
        
        # Create action summary for logging
        action_source = "Portfolio" if action_mode == 'portfolio' else "New Investment"
        scheme_display = portfolio_scheme_name or (scheme_obj.scheme_name if scheme_obj else "Unknown")
        
        # Create a simple action summary
        action_summary = f"{action_type} - {scheme_display} ({action_source})"
        if action.amount:
            action_summary += f" - ₹{action.amount:,.2f}"
        elif action.units:
            action_summary += f" - {action.units} units"
        
        logger.info(f"Created {action_source} action {action.id}: {action_summary}")
        return action
        
    except Exception as e:
        logger.error(f"Error creating PlanAction: {str(e)}")
        logger.error(f"Action data: {action_data}")
        logger.error(f"Processed values - Mode: {action_mode}, Scheme: {scheme_obj}, Portfolio: {portfolio_scheme_name}")
        raise ValueError(f"Failed to create action: {str(e)}")


def safe_decimal(value):
    """Safely convert value to Decimal"""
    if value is None or value == '' or value == 'null' or value == 'None':
        return None
    try:
        from decimal import Decimal, InvalidOperation
        return Decimal(str(value))
    except (ValueError, TypeError, InvalidOperation):
        logger.warning(f"Could not convert to decimal: {value}")
        return None


def safe_int(value):
    """Safely convert value to int"""
    if value is None or value == '' or value == 'null' or value == 'None':
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        logger.warning(f"Could not convert to int: {value}")
        return None


# Alternative process_action_data function that handles the specific error
def process_action_data(action_data, execution_plan, client, priority):
    """Process individual action data - FIXED VERSION that calls the portfolio-independent function"""
    return process_action_data_portfolio_independent(action_data, execution_plan, client, priority)


# FIXED: Add debugging function to log action data structure
def debug_action_data(action_data, index):
    """Debug function to log action data structure"""
    logger.info(f"=== DEBUG ACTION {index} ===")
    logger.info(f"Action Type: {action_data.get('action_type')}")
    logger.info(f"Portfolio Fields:")
    logger.info(f"  - portfolio_id: {action_data.get('portfolio_id')}")
    logger.info(f"  - portfolio_holding_id: {action_data.get('portfolio_holding_id')}")
    logger.info(f"  - source_portfolio_id: {action_data.get('source_portfolio_id')}")
    logger.info(f"Scheme Fields:")
    logger.info(f"  - scheme: {action_data.get('scheme')}")
    logger.info(f"  - scheme_name: {action_data.get('scheme_name')}")
    logger.info(f"  - scheme_id: {action_data.get('scheme_id')}")
    logger.info(f"  - source_scheme: {action_data.get('source_scheme')}")
    logger.info(f"  - source_scheme_name: {action_data.get('source_scheme_name')}")
    logger.info(f"Amount Fields:")
    logger.info(f"  - amount: {action_data.get('amount')}")
    logger.info(f"  - redeem_amount: {action_data.get('redeem_amount')}")
    logger.info(f"  - redeem_units: {action_data.get('redeem_units')}")
    logger.info(f"  - redeem_by: {action_data.get('redeem_by')}")
    logger.info(f"All Fields: {list(action_data.keys())}")
    logger.info(f"=== END DEBUG ACTION {index} ===")


# FIXED: Updated main save function to include debugging
@login_required
@require_http_methods(["POST"])
def save_execution_plan_with_debug(request):
    """Save execution plan with enhanced debugging for frontend data structure"""
    try:
        data = json.loads(request.body)
        client_id = data.get('client_id')
        plan_name = data.get('plan_name')
        description = data.get('description', '')
        actions_data = data.get('actions', [])
        submit_for_approval = data.get('submit_for_approval', False)
        
        # Enhanced validation
        if not client_id:
            return JsonResponse({'error': 'Client ID is required'}, status=400)
        if not plan_name or not plan_name.strip():
            return JsonResponse({'error': 'Plan name is required'}, status=400)
        if not actions_data:
            return JsonResponse({'error': 'At least one action is required'}, status=400)
        
        # DEBUG: Log all action data structures
        logger.info(f"=== FRONTEND DATA ANALYSIS ===")
        logger.info(f"Total actions received: {len(actions_data)}")
        for i, action_data in enumerate(actions_data):
            debug_action_data(action_data, i + 1)
        logger.info(f"=== END FRONTEND DATA ANALYSIS ===")
        
        # Get the client object (handle integer ID)
        try:
            client_id_int = int(client_id)
            client = get_object_or_404(Client, id=client_id_int)
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid client ID format'}, status=400)
        
        # Check access permission
        if request.user.role == 'rm' and client.user != request.user:
            return JsonResponse({'error': 'Access denied - not your client'}, status=403)
        elif request.user.role == 'rm_head':
            team_members = User.objects.filter(manager=request.user, role='rm')
            if client.user not in team_members and client.user != request.user:
                return JsonResponse({'error': 'Access denied - client not in your team'}, status=403)
        elif request.user.role not in ['business_head', 'business_head_ops', 'top_management']:
            return JsonResponse({'error': 'Access denied - insufficient permissions'}, status=403)
        
        # Variables to track execution plan and actions
        execution_plan = None
        created_actions = []
        failed_actions = []
        
        try:
            with transaction.atomic():
                # Create execution plan
                execution_plan = ExecutionPlan.objects.create(
                    client=client,
                    plan_name=plan_name.strip(),
                    description=description.strip(),
                    created_by=request.user,
                    status='draft'
                )
                
                logger.info(f"Created execution plan {execution_plan.plan_id} for client {client.id}")
                
                # Create plan actions with enhanced debugging
                for i, action_data in enumerate(actions_data):
                    try:
                        logger.info(f"Processing action {i + 1}: {action_data}")
                        
                        action = process_action_data_portfolio_independent(
                            action_data, execution_plan, client, i + 1
                        )
                        if action:
                            created_actions.append(action)
                            logger.info(f"✅ Successfully created action {action.id}: {action.get_action_summary()}")
                        else:
                            failed_actions.append(f"Action {i + 1}: Unknown error")
                        
                    except Exception as e:
                        error_msg = f"Action {i + 1}: {str(e)}"
                        logger.error(f"❌ Error creating action {i + 1}: {str(e)}")
                        logger.error(f"Action data that failed: {action_data}")
                        failed_actions.append(error_msg)
                        continue
                
                if not created_actions:
                    return JsonResponse({
                        'error': 'No valid actions could be created',
                        'failed_actions': failed_actions,
                        'debug_info': {
                            'total_actions_received': len(actions_data),
                            'sample_action': actions_data[0] if actions_data else None
                        }
                    }, status=400)
                
                # Create workflow history
                PlanWorkflowHistory.objects.create(
                    execution_plan=execution_plan,
                    from_status='',
                    to_status='draft',
                    changed_by=request.user,
                    comments=f'Execution plan created with {len(created_actions)} actions'
                )
                
                # Submit for approval if requested
                if submit_for_approval and len(created_actions) > 0:
                    try:
                        if execution_plan.submit_for_approval():
                            PlanWorkflowHistory.objects.create(
                                execution_plan=execution_plan,
                                from_status='draft',
                                to_status='pending_approval',
                                changed_by=request.user,
                                comments='Plan submitted for approval'
                            )
                            logger.info(f"Plan {execution_plan.plan_id} submitted for approval")
                    except Exception as e:
                        logger.warning(f"Could not submit for approval: {str(e)}")
            
            # Generate Excel file OUTSIDE of transaction
            excel_generated = False
            excel_error = None
            
            try:
                excel_data = generate_execution_plan_excel_in_memory(execution_plan)
                if excel_data:
                    excel_generated = True
                    logger.info(f"Generated Excel data for plan {execution_plan.plan_id}")
                else:
                    excel_error = "Excel generation returned no data"
            except Exception as e:
                excel_error = str(e)
                logger.warning(f"Could not generate Excel file: {excel_error}")
            
            # Prepare response
            response_data = {
                'success': True,
                'plan_id': execution_plan.plan_id,
                'plan_url': reverse('execution_plan_detail', args=[execution_plan.id]),
                'message': f'Execution plan created successfully with {len(created_actions)} actions',
                'actions_created': len(created_actions),
                'actions_failed': len(failed_actions),
                'excel_generated': excel_generated,
                'debug_info': {
                    'portfolio_actions': len([a for a in created_actions if a.action_mode == 'portfolio']),
                    'new_investment_actions': len([a for a in created_actions if a.action_mode == 'new_investment']),
                }
            }
            
            if failed_actions:
                response_data['failed_actions'] = failed_actions
                response_data['message'] += f'. {len(failed_actions)} actions failed.'
            
            if excel_error:
                response_data['excel_error'] = excel_error
            
            return JsonResponse(response_data)
            
        except Exception as e:
            logger.error(f"Transaction failed: {str(e)}")
            return JsonResponse({
                'error': f'Database transaction failed: {str(e)}',
                'actions_processed': len(created_actions),
                'actions_failed': len(failed_actions),
                'debug_info': {
                    'last_successful_action': created_actions[-1].get_action_summary() if created_actions else None,
                    'total_actions_received': len(actions_data)
                }
            }, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error saving execution plan: {str(e)}")
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)
    
def generate_execution_plan_excel_in_memory(execution_plan):
    """
    Generate Excel file in memory for portfolio-independent execution plan
    Returns bytes data that can be directly sent in HTTP response
    """
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
        from datetime import datetime
        
        # Get plan actions
        plan_actions = []
        try:
            if hasattr(execution_plan, 'actions'):
                plan_actions = list(execution_plan.actions.all().order_by('priority', 'id'))
                logger.info(f"Found {len(plan_actions)} actions for execution plan {execution_plan.id}")
        except Exception as e:
            logger.error(f"Error getting plan actions: {str(e)}")
            plan_actions = []
        
        # Create workbook in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Execution Plan"
        
        # Set column widths - improved sizing
        column_widths = {
            'A': 8,   # Sr. No.
            'B': 45,  # Scheme Name (increased)
            'C': 25,  # Source/ISIN (increased)
            'D': 18,  # Transaction Type (increased)
            'E': 18,  # Amount (increased)
            'F': 18,  # SIP Amount (increased)
            'G': 12,  # SIP Date
            'H': 15,  # Frequency
            'I': 15,  # Status
            'J': 10,  # Priority
            'K': 20   # Action Source
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Create styles
        header_style = NamedStyle(name="header")
        header_style.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Title style
        title_style = NamedStyle(name="title")
        title_style.font = Font(name='Arial', size=16, bold=True, color='1F4E79')
        title_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # Info style
        info_label_style = NamedStyle(name="info_label")
        info_label_style.font = Font(name='Arial', size=11, bold=True)
        info_label_style.alignment = Alignment(horizontal='left', vertical='center')
        
        info_value_style = NamedStyle(name="info_value")
        info_value_style.font = Font(name='Arial', size=11)
        info_value_style.alignment = Alignment(horizontal='left', vertical='center')
        
        # Plan header information
        row = 1
        ws.merge_cells(f'A{row}:K{row}')
        plan_id = getattr(execution_plan, 'plan_id', f"Plan #{execution_plan.id}")
        ws[f'A{row}'] = f"EXECUTION PLAN - {plan_id}"
        ws[f'A{row}'].style = title_style
        ws.row_dimensions[row].height = 25
        row += 2  # Extra space
        
        # Plan details in a more structured format
        client_name = "Unknown Client"
        if hasattr(execution_plan, 'client') and execution_plan.client:
            if hasattr(execution_plan.client, 'name'):
                client_name = execution_plan.client.name
            elif hasattr(execution_plan.client, 'client_full_name'):
                client_name = execution_plan.client.client_full_name
        
        # Client and Date row
        ws[f'A{row}'] = "Client:"
        ws[f'A{row}'].style = info_label_style
        ws[f'B{row}'] = client_name
        ws[f'B{row}'].style = info_value_style
        
        ws[f'G{row}'] = "Date Created:"
        ws[f'G{row}'].style = info_label_style
        ws[f'H{row}'] = execution_plan.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(execution_plan, 'created_at') else "N/A"
        ws[f'H{row}'].style = info_value_style
        row += 1
        
        # Calculate totals
        total_amount = 0
        total_sip_amount = 0
        if plan_actions:
            for action in plan_actions:
                if hasattr(action, 'amount') and action.amount:
                    total_amount += float(action.amount)
                if hasattr(action, 'sip_amount') and action.sip_amount:
                    total_sip_amount += float(action.sip_amount)
        
        # Total Amount and Status row
        ws[f'A{row}'] = "Total Amount:"
        ws[f'A{row}'].style = info_label_style
        ws[f'B{row}'] = f"₹{total_amount:,.2f}"
        ws[f'B{row}'].style = info_value_style
        ws[f'B{row}'].font = Font(name='Arial', size=11, bold=True, color='0D5016')
        
        ws[f'G{row}'] = "Status:"
        ws[f'G{row}'].style = info_label_style
        status_value = execution_plan.get_status_display() if hasattr(execution_plan, 'get_status_display') else execution_plan.status
        ws[f'H{row}'] = status_value
        ws[f'H{row}'].style = info_value_style
        row += 1
        
        # Total SIP Amount and Created By row
        if total_sip_amount > 0:
            ws[f'A{row}'] = "Total SIP Amount:"
            ws[f'A{row}'].style = info_label_style
            ws[f'B{row}'] = f"₹{total_sip_amount:,.2f}"
            ws[f'B{row}'].style = info_value_style
            ws[f'B{row}'].font = Font(name='Arial', size=11, bold=True, color='7B2CBF')
        
        if hasattr(execution_plan, 'created_by') and execution_plan.created_by:
            ws[f'G{row}'] = "Created By:"
            ws[f'G{row}'].style = info_label_style
            ws[f'H{row}'] = execution_plan.created_by.get_full_name() or execution_plan.created_by.username
            ws[f'H{row}'].style = info_value_style
        row += 1
        
        # Actions count
        ws[f'A{row}'] = "Total Actions:"
        ws[f'A{row}'].style = info_label_style
        ws[f'B{row}'] = str(len(plan_actions))
        ws[f'B{row}'].style = info_value_style
        row += 2  # Extra space before table
        
        # Table headers for portfolio independence
        headers = [
            'Sr. No.', 'Scheme Name', 'Source/ISIN', 'Transaction Type', 'Amount (₹)',
            'SIP Amount (₹)', 'SIP Date', 'Frequency', 'Status', 'Priority', 'Action Source'
        ]
        
        # Add header row with improved styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.style = header_style
        
        ws.row_dimensions[row].height = 20
        row += 1
        
        # Data rows for actions
        for idx, action in enumerate(plan_actions, 1):
            # Get scheme information - PORTFOLIO INDEPENDENT
            scheme_name = "Unknown Scheme"
            source_info = "N/A"
            action_source = "Unknown"
            
            # Check action mode and get appropriate scheme information
            if hasattr(action, 'action_mode') and action.action_mode == 'portfolio':
                # Portfolio-based action
                scheme_name = action.portfolio_scheme_name or "Unknown Portfolio Scheme"
                source_info = action.portfolio_isin or "Portfolio Holdings"
                action_source = "Existing Portfolio"
                
                if action.portfolio_holding_id:
                    source_info = f"Portfolio ID: {action.portfolio_holding_id}"
                    
            elif hasattr(action, 'action_mode') and action.action_mode == 'new_investment':
                # New investment action
                if action.scheme:
                    scheme_name = action.scheme.scheme_name
                    action_source = "New Investment"
                    
                    # Get ISIN for new investments
                    if hasattr(action.scheme, 'isin_growth') and action.scheme.isin_growth:
                        source_info = action.scheme.isin_growth
                    elif hasattr(action.scheme, 'isin_number') and action.scheme.isin_number:
                        source_info = action.scheme.isin_number
                    else:
                        source_info = "New Scheme"
            else:
                # Fallback for actions without action_mode (legacy)
                if hasattr(action, 'portfolio_scheme_name') and action.portfolio_scheme_name:
                    scheme_name = action.portfolio_scheme_name
                    source_info = "Portfolio Holdings"
                    action_source = "Existing Portfolio"
                elif hasattr(action, 'scheme') and action.scheme:
                    scheme_name = action.scheme.scheme_name
                    source_info = getattr(action.scheme, 'isin_growth', 'New Scheme')
                    action_source = "New Investment"
            
            # Get action type with better formatting
            action_type = "N/A"
            if hasattr(action, 'action_type') and action.action_type:
                if hasattr(action, 'get_action_type_display'):
                    action_type = action.get_action_type_display()
                else:
                    action_type = str(action.action_type).replace('_', ' ').title()
            
            # Get amounts with better formatting
            amount = 0
            if hasattr(action, 'amount') and action.amount:
                amount = float(action.amount)
            
            sip_amount = 0
            sip_amount_str = "N/A"
            if hasattr(action, 'sip_amount') and action.sip_amount:
                sip_amount = float(action.sip_amount)
                sip_amount_str = f"₹{sip_amount:,.2f}"
            
            # Get other details with better handling
            sip_date = "N/A"
            if hasattr(action, 'sip_date') and action.sip_date:
                sip_date = action.sip_date.strftime('%d-%m-%Y') if hasattr(action.sip_date, 'strftime') else str(action.sip_date)
            
            frequency = "N/A"
            if hasattr(action, 'frequency') and action.frequency:
                frequency = str(action.frequency).replace('_', ' ').title()
            
            status = "N/A"
            if hasattr(action, 'status') and action.status:
                if hasattr(action, 'get_status_display'):
                    status = action.get_status_display()
                else:
                    status = str(action.status).replace('_', ' ').title()
            
            priority = "N/A"
            if hasattr(action, 'priority') and action.priority is not None:
                priority = str(action.priority)
            
            # Build row data
            data = [
                idx,
                scheme_name,
                source_info,
                action_type,
                f"₹{amount:,.2f}" if amount > 0 else "₹0.00",
                sip_amount_str,
                sip_date,
                frequency,
                status,
                priority,
                action_source
            ]
            
            # Write row data with improved formatting
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                # Set alignment based on column type
                if col == 1:  # Serial number
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col in [5, 6]:  # Amount columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if value != "N/A" and value != "₹0.00":
                        cell.font = Font(name='Arial', size=10, bold=True)
                elif col in [7, 9, 10]:  # Date, Status, Priority
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Color coding for action source with better colors
                if col == 11:  # Action Source column
                    if value == "Existing Portfolio":
                        cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
                        cell.font = Font(name='Arial', size=10, bold=True, color='1565C0')
                    elif value == "New Investment":
                        cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                        cell.font = Font(name='Arial', size=10, bold=True, color='2E7D32')
            
            # Set row height for better readability
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Add totals row with better styling
        if plan_actions:
            row += 1
            # Create totals row
            total_cells = ['TOTAL', '', '', '', f"₹{total_amount:,.2f}", 
                          f"₹{total_sip_amount:,.2f}" if total_sip_amount > 0 else "₹0.00", 
                          '', '', '', '', '']
            
            for col, value in enumerate(total_cells, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = Font(name='Arial', size=11, bold=True)
                
                if col in [1, 5, 6]:  # Total label and amount columns
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                
                if col in [5, 6]:  # Amount columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thick'), bottom=Side(style='thick')
                )
            
            ws.row_dimensions[row].height = 20
        
        # Add legend with better formatting
        row += 3
        ws[f'A{row}'] = "Legend:"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='1F4E79')
        row += 1
        
        # Portfolio actions legend
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "🔵 Existing Portfolio Actions"
        ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        ws.merge_cells(f'D{row}:K{row}')
        ws[f'D{row}'] = "Actions on current portfolio holdings (Redeem, Switch From, STP From, SWP)"
        ws[f'D{row}'].font = Font(name='Arial', size=10)
        ws[f'D{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
        
        # New investment legend
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "🟢 New Investment Actions"
        ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        ws.merge_cells(f'D{row}:K{row}')
        ws[f'D{row}'] = "Fresh investments in mutual fund schemes (Purchase, SIP, Switch To, STP To)"
        ws[f'D{row}'].font = Font(name='Arial', size=10)
        ws[f'D{row}'].alignment = Alignment(horizontal='left', vertical='center')
        row += 1
        
        # Add summary with better formatting
        row += 2
        ws[f'A{row}'] = "Summary:"
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color='1F4E79')
        row += 1
        
        # Calculate portfolio actions more accurately
        portfolio_actions = sum(1 for action in plan_actions 
                              if (getattr(action, 'action_mode', None) == 'portfolio' or 
                                  getattr(action, 'portfolio_scheme_name', None)))
        new_investment_actions = len(plan_actions) - portfolio_actions
        
        summary_data = [
            f"Total Actions: {len(plan_actions)}",
            f"Portfolio Actions: {portfolio_actions}",
            f"New Investment Actions: {new_investment_actions}",
            f"Total Investment Amount: ₹{total_amount:,.2f}",
            f"Total SIP Amount: ₹{total_sip_amount:,.2f}" if total_sip_amount > 0 else None,
            f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        
        for item in summary_data:
            if item:
                ws[f'A{row}'] = item
                ws[f'A{row}'].font = Font(name='Arial', size=10)
                row += 1
        
        # Save to BytesIO buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Portfolio-independent Excel generated in memory for plan {execution_plan.id}")
        logger.info(f"Portfolio actions: {portfolio_actions}, New investments: {new_investment_actions}")
        logger.info(f"Total amount: ₹{total_amount:,.2f}, Total SIP: ₹{total_sip_amount:,.2f}")
        
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating portfolio-independent Excel in memory: {str(e)}")
        import traceback
        logger.error(f"Excel generation traceback: {traceback.format_exc()}")
        return None


def attach_excel_file_from_memory(email, execution_plan):
    """
    Attach Excel file to email from memory - Vercel compatible
    """
    try:
        # Generate Excel in memory
        excel_data = generate_execution_plan_excel_in_memory(execution_plan)
        
        if excel_data:
            filename = f"{execution_plan.plan_id}_execution_plan.xlsx"
            email.attach(
                filename, 
                excel_data, 
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            return True
        
        return False
        
    except Exception as e:
        logger.error(f"Error attaching Excel from memory: {str(e)}")
        return False


# Update the email sending function to use in-memory Excel generation
@login_required
@require_http_methods(["POST"])
def send_to_client_enhanced_vercel(request, plan_id):
    """Enhanced send to client with Vercel-compatible Excel handling"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check access permission
    if not can_access_plan(request.user, execution_plan):
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        # Get form data
        email_to = request.POST.get('email_to', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        include_excel = request.POST.get('include_excel', 'false').lower() == 'true'
        copy_to_rm = request.POST.get('copy_to_rm', 'false').lower() == 'true'
        
        # Auto-detect client email if not provided
        if not email_to:
            email_to = get_client_email(execution_plan)
            if not email_to:
                return JsonResponse({
                    'success': False, 
                    'error': 'No client email provided or found in database.'
                })
        
        # Validate email
        from django.core.validators import validate_email
        try:
            validate_email(email_to)
        except ValidationError:
            return JsonResponse({
                'success': False, 
                'error': f'Invalid email address: {email_to}'
            })
        
        # Set defaults
        if not subject:
            subject = f"📋 Investment Plan - {execution_plan.plan_name}"
        
        if not message:
            client_name = execution_plan.client.name if hasattr(execution_plan.client, 'name') else 'Valued Client'
            message = f"Dear {client_name},\n\nPlease find attached your investment execution plan.\n\nBest regards,\n{execution_plan.created_by.get_full_name() or execution_plan.created_by.username}"
        
        # Prepare recipients
        recipients = [email_to]
        cc_recipients = []
        
        if copy_to_rm and execution_plan.created_by and execution_plan.created_by.email:
            if execution_plan.created_by.email not in recipients:
                cc_recipients.append(execution_plan.created_by.email)
        
        # Create email
        email = EmailMultiAlternatives(
            subject=subject,
            body=message,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@yourcompany.com'),
            to=recipients,
            cc=cc_recipients if cc_recipients else None,
            reply_to=[execution_plan.created_by.email] if execution_plan.created_by and execution_plan.created_by.email else None
        )
        
        # Attach Excel from memory if requested
        excel_attached = False
        if include_excel:
            excel_attached = attach_excel_file_from_memory(email, execution_plan)
        
        # Send email
        email.send(fail_silently=False)
        
        # Create audit record
        attachment_note = " with Excel attachment" if excel_attached else ""
        PlanComment.objects.create(
            execution_plan=execution_plan,
            comment=f"Plan sent to client: {email_to}{attachment_note}",
            commented_by=request.user,
            is_internal=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Email sent successfully to {email_to}',
            'excel_attached': excel_attached
        })
        
    except Exception as e:
        logger.error(f"Error in send_to_client_enhanced_vercel: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["POST"])
def generate_excel(request, plan_id):
    """Generate fresh Excel file for execution plan - Vercel compatible"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check access permission
    if not can_access_plan(request.user, execution_plan):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        # Generate Excel in memory
        excel_data = generate_execution_plan_excel_in_memory(execution_plan)
        
        if excel_data:
            filename = f"{execution_plan.plan_id}_execution_plan.xlsx"
            
            return JsonResponse({
                'success': True,
                'message': 'Excel file generated successfully',
                'filename': filename,
                'download_url': reverse('download_excel', args=[plan_id])
            })
        else:
            return JsonResponse({'error': 'Failed to generate Excel file'}, status=500)
            
    except Exception as e:
        logger.error(f"Error generating Excel for plan {plan_id}: {str(e)}")
        return JsonResponse({'error': f'Error generating Excel: {str(e)}'}, status=500)
    
    
@login_required
@require_http_methods(["POST"])
def email_plan(request, plan_id):
    """Email execution plan to specified recipients"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check access permission
    if not can_access_plan(request.user, execution_plan):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        email_to = data.get('email_to', '').strip()
        email_cc = data.get('email_cc', '').strip()
        subject = data.get('subject', '').strip()
        message = data.get('message', '').strip()
        attach_excel = data.get('attach_excel', False)
        attach_pdf = data.get('attach_pdf', False)
        
        if not email_to:
            return JsonResponse({'error': 'Recipient email is required'}, status=400)
        
        if not subject:
            return JsonResponse({'error': 'Email subject is required'}, status=400)
        
        # Create email
        email = EmailMultiAlternatives(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email_to],
            cc=[email_cc] if email_cc else []
        )
        
        # Attach Excel file
        if attach_excel:
            if execution_plan.excel_file and os.path.exists(execution_plan.excel_file.path):
                excel_filename = f"{execution_plan.plan_id}_execution_plan.xlsx"
                with open(execution_plan.excel_file.path, 'rb') as f:
                    email.attach(excel_filename, f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            else:
                # Generate Excel file if not exists
                excel_file_path = generate_execution_plan_excel(execution_plan)
                if excel_file_path:
                    execution_plan.excel_file = excel_file_path
                    execution_plan.save()
                    excel_filename = f"{execution_plan.plan_id}_execution_plan.xlsx"
                    with open(os.path.join(settings.MEDIA_ROOT, excel_file_path), 'rb') as f:
                        email.attach(excel_filename, f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Send email
        email.send()
        
        # Add comment to execution plan
        PlanComment.objects.create(
            execution_plan=execution_plan,
            comment=f"Execution plan emailed to {email_to} by {request.user.get_full_name() or request.user.username}",
            commented_by=request.user,
            is_internal=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Email sent successfully to {email_to}'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error emailing plan {plan_id}: {str(e)}")
        return JsonResponse({'error': f'Error sending email: {str(e)}'}, status=500)
    


@login_required
@require_http_methods(["POST"])
def add_comment(request, plan_id):
    """Add comment to execution plan"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    if not can_access_plan(request.user, execution_plan):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    comment_text = request.POST.get('comment', '').strip()
    is_internal = request.POST.get('is_internal', 'true').lower() == 'true'
    
    if not comment_text:
        return JsonResponse({'error': 'Comment cannot be empty'}, status=400)
    
    comment = PlanComment.objects.create(
        execution_plan=execution_plan,
        comment=comment_text,
        commented_by=request.user,
        is_internal=is_internal
    )
    
    return JsonResponse({
        'success': True,
        'comment': {
            'id': comment.id,
            'comment': comment.comment,
            'commented_by': comment.commented_by.get_full_name() or comment.commented_by.username,
            'created_at': comment.created_at.strftime('%d/%m/%Y %H:%M'),
            'is_internal': comment.is_internal,
        }
    })


@login_required
@require_http_methods(["GET"])
def search_schemes_ajax(request):
    """
    AJAX endpoint to search mutual fund schemes from database - FOR NEW INVESTMENTS ONLY
    """
    try:
        query = request.GET.get('q', '').strip()
        action_type = request.GET.get('action_type', '')
        
        # Debug logging
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Search query: '{query}', action_type: '{action_type}'")
        
        # Reduced minimum query length for better UX
        if len(query) < 1:
            return JsonResponse({
                'success': False,
                'message': 'Query cannot be empty'
            })
        
        # More flexible action type validation
        valid_actions = ['purchase', 'sip_start', 'switch_target', 'stp_target']
        if action_type and action_type not in valid_actions:
            logger.warning(f"Invalid action_type: {action_type}")
            # Don't return error, just log warning and continue
        
        # Build search query with case-insensitive search
        search_filter = Q()
        
        # Split query into words for better matching
        query_words = query.split()
        
        for word in query_words:
            word_filter = (
                Q(scheme_name__icontains=word) |
                Q(amc_name__icontains=word) |
                Q(category__icontains=word) |
                Q(scheme_code__icontains=word)
            )
            
            # Only search ISIN if the word looks like an ISIN (12 characters, alphanumeric)
            if len(word) >= 6:
                word_filter |= Q(isin_growth__icontains=word) | Q(isin_div_reinvestment__icontains=word)
            
            search_filter &= word_filter
        
        # If no words found, fallback to original search
        if not query_words:
            search_filter = (
                Q(scheme_name__icontains=query) |
                Q(amc_name__icontains=query) |
                Q(category__icontains=query) |
                Q(scheme_code__icontains=query) |
                Q(isin_growth__icontains=query) |
                Q(isin_div_reinvestment__icontains=query)
            )
        
        # Get total count before filtering by is_active
        total_schemes = MutualFundScheme.objects.filter(search_filter).count()
        logger.info(f"Total schemes matching search (before active filter): {total_schemes}")
        
        # Search in MutualFundScheme model
        schemes_queryset = MutualFundScheme.objects.filter(search_filter)
        
        # Apply active filter if the field exists
        try:
            active_schemes = schemes_queryset.filter(is_active=True)
            if active_schemes.exists():
                schemes_queryset = active_schemes
            else:
                # If no active schemes found, show all matching schemes
                logger.warning("No active schemes found, showing all matching schemes")
        except Exception as filter_error:
            logger.warning(f"is_active filter failed: {filter_error}, showing all schemes")
        
        # Order and limit results
        schemes = schemes_queryset.order_by('scheme_name')[:100]  # Increased limit
        
        logger.info(f"Final schemes count: {schemes.count()}")
        
        scheme_list = []
        for scheme in schemes:
            try:
                scheme_data = {
                    'id': scheme.id,
                    'scheme_name': scheme.scheme_name or 'N/A',
                    'amc_name': scheme.amc_name or 'N/A',
                    'scheme_type': getattr(scheme, 'scheme_type', None) or 'other',
                    'category': scheme.category or 'N/A',
                    'risk_category': getattr(scheme, 'risk_category', None) or 'moderate',
                    'isin_growth': getattr(scheme, 'isin_growth', None) or '',
                    'isin_div_reinvestment': getattr(scheme, 'isin_div_reinvestment', None) or '',
                    'scheme_code': getattr(scheme, 'scheme_code', None) or '',
                    'is_active': getattr(scheme, 'is_active', True),
                    'is_new_investment_scheme': True,
                }
                
                # Handle optional fields that might not exist
                try:
                    scheme_data['last_nav_price'] = float(scheme.last_nav_price) if scheme.last_nav_price else None
                except (AttributeError, TypeError, ValueError):
                    scheme_data['last_nav_price'] = None
                
                try:
                    scheme_data['last_nav_date'] = scheme.last_nav_date.strftime('%Y-%m-%d') if scheme.last_nav_date else None
                except (AttributeError, TypeError):
                    scheme_data['last_nav_date'] = None
                
                try:
                    scheme_data['minimum_investment'] = float(scheme.minimum_investment) if scheme.minimum_investment else 500
                except (AttributeError, TypeError, ValueError):
                    scheme_data['minimum_investment'] = 500
                
                try:
                    scheme_data['minimum_sip'] = float(scheme.minimum_sip) if scheme.minimum_sip else 500
                except (AttributeError, TypeError, ValueError):
                    scheme_data['minimum_sip'] = 500
                
                # Set default availability flags
                scheme_data.update({
                    'sip_available': True,
                    'stp_available': True,
                    'swp_available': True,
                })
                
                scheme_list.append(scheme_data)
                
            except Exception as scheme_error:
                logger.error(f"Error processing scheme {scheme.id}: {scheme_error}")
                continue
        
        response_data = {
            'success': True,
            'schemes': scheme_list,
            'count': len(scheme_list),
            'total_found': total_schemes,
            'query': query,
            'action_type': action_type,
            'debug_info': {
                'query_length': len(query),
                'words_searched': query_words,
                'schemes_before_limit': schemes_queryset.count(),
            }
        }
        
        logger.info(f"Returning {len(scheme_list)} schemes for query '{query}'")
        return JsonResponse(response_data)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in search_schemes_ajax: {str(e)}", exc_info=True)
        
        # Return more detailed error information in development
        return JsonResponse({
            'success': False,
            'message': f'An error occurred while searching schemes: {str(e)}',
            'error_type': type(e).__name__,
            'query': request.GET.get('q', ''),
            'action_type': request.GET.get('action_type', ''),
        })


@login_required
def execution_reports(request):
    """Execution plans reports dashboard"""
    if request.user.role not in ['rm_head', 'business_head', 'business_head_ops', 'top_management']:
        messages.error(request, "Access denied")
        return redirect('dashboard')
    
    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        date_from = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = timezone.now().strftime('%Y-%m-%d')
    
    # Get plans in date range
    plans = ExecutionPlan.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    )
    
    # Filter by RM if specified
    rm_filter = request.GET.get('rm')
    if rm_filter:
        plans = plans.filter(created_by_id=rm_filter)
    
    # Calculate statistics
    stats = {
        'total_plans': plans.count(),
        'draft_plans': plans.filter(status='draft').count(),
        'pending_approval': plans.filter(status='pending_approval').count(),
        'approved_plans': plans.filter(status='approved').count(),
        'completed_plans': plans.filter(status='completed').count(),
        'rejected_plans': plans.filter(status='rejected').count(),
    }
    
    # Get RMs for filter
    rms = User.objects.filter(role='rm').order_by('first_name', 'last_name', 'username')
    
    # Performance metrics
    completed_plans = plans.filter(status='completed')
    metrics_data = []
    
    for plan in completed_plans:
        try:
            metrics = ExecutionMetrics.objects.get(execution_plan=plan)
            metrics_data.append({
                'plan': plan,
                'metrics': metrics,
            })
        except ExecutionMetrics.DoesNotExist:
            continue
    
    # Average metrics
    avg_metrics = {}
    if metrics_data:
        avg_metrics = {
            'avg_time_to_approval': sum(m['metrics'].time_to_approval or 0 for m in metrics_data) / len(metrics_data),
            'avg_time_to_execution': sum(m['metrics'].time_to_execution or 0 for m in metrics_data) / len(metrics_data),
            'avg_execution_time': sum(m['metrics'].total_execution_time or 0 for m in metrics_data) / len(metrics_data),
            'avg_success_rate': sum(m['metrics'].success_rate for m in metrics_data) / len(metrics_data),
        }
    
    context = {
        'stats': stats,
        'date_from': date_from,
        'date_to': date_to,
        'rm_filter': rm_filter,
        'rms': rms,
        'metrics_data': metrics_data,
        'avg_metrics': avg_metrics,
    }
    
    return render(request, 'execution_plans/reports.html', context)


@login_required
def plan_templates(request):
    """Manage plan templates"""
    if request.user.role not in ['rm', 'rm_head', 'business_head']:
        messages.error(request, "Access denied")
        return redirect('dashboard')
    
    # Get templates accessible to user
    templates = PlanTemplate.objects.filter(
        Q(is_public=True) | Q(created_by=request.user)
    ).filter(is_active=True).order_by('-created_at')
    
    context = {
        'templates': templates,
    }
    
    return render(request, 'execution_plans/templates.html', context)


@login_required
@require_http_methods(["POST"])
def save_template(request):
    """Save plan as template - FIXED VERSION"""
    try:
        data = json.loads(request.body)
        
        template_name = data.get('template_name')
        description = data.get('description', '')
        is_public = data.get('is_public', False)
        template_data = data.get('template_data', {})
        
        if not template_name:
            return JsonResponse({'error': 'Template name is required'}, status=400)
        
        # Check if user can create public templates
        if is_public and request.user.role not in ['rm_head', 'business_head', 'top_management']:
            is_public = False
        
        # Check for duplicate template names
        existing_template = PlanTemplate.objects.filter(
            name=template_name,
            created_by=request.user
        ).first()
        
        if existing_template:
            return JsonResponse({'error': 'Template with this name already exists'}, status=400)
        
        template = PlanTemplate.objects.create(
            name=template_name,
            description=description,
            created_by=request.user,
            is_public=is_public,
            template_data=template_data
        )
        
        return JsonResponse({
            'success': True,
            'template_id': template.id,
            'message': 'Template saved successfully'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error saving template: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def load_template_ajax(request, template_id):
    """AJAX endpoint to load template data - FIXED VERSION"""
    try:
        template = get_object_or_404(PlanTemplate, id=template_id)
        
        # Check access permission
        if not (template.is_public or template.created_by == request.user):
            return JsonResponse({'error': 'Access denied'}, status=403)
        
        return JsonResponse({
            'success': True,
            'template_data': template.template_data,
            'template_name': template.name,
            'description': template.description,
        })
        
    except PlanTemplate.DoesNotExist:
        return JsonResponse({'error': 'Template not found'}, status=404)
    except Exception as e:
        logger.error(f"Error loading template {template_id}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def execution_plan_detail(request, plan_id):
    """View execution plan details - FIXED VERSION"""
    try:
        execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    except ValueError:
        # Handle case where plan_id is not a valid integer
        raise Http404("Invalid plan ID")
    
    # Check access permission
    if not can_access_plan(request.user, execution_plan):
        messages.error(request, "Access denied")
        return redirect('dashboard')
    
    # Get plan actions
    actions = execution_plan.actions.all().select_related('scheme', 'target_scheme', 'executed_by')
    
    # Get current portfolio if client has one
    current_portfolio = None
    if hasattr(execution_plan.client, 'client_profile') and execution_plan.client.client_profile:
        current_portfolio = ClientPortfolio.objects.filter(
            client_profile=execution_plan.client.client_profile,
            is_active=True
        ).order_by('scheme_name')
    
    # Get workflow history
    workflow_history = execution_plan.workflow_history.all().select_related('changed_by')
    
    # Get comments
    comments = execution_plan.comments.all().select_related('commented_by')
    
    # Calculate metrics if plan is completed
    metrics = None
    if execution_plan.status == 'completed':
        try:
            metrics, created = ExecutionMetrics.objects.get_or_create(execution_plan=execution_plan)
            if created or not metrics.updated_at:
                metrics.calculate_metrics()
        except:
            metrics = None
    
    # Check permissions
    can_approve = can_approve_plan(request.user, execution_plan)
    can_execute = can_execute_plan(request.user, execution_plan)
    can_edit = can_edit_plan(request.user, execution_plan)
    
    context = {
        'execution_plan': execution_plan,
        'actions': actions,
        'current_portfolio': current_portfolio,
        'workflow_history': workflow_history,
        'comments': comments,
        'metrics': metrics,
        'can_approve': can_approve,
        'can_execute': can_execute,
        'can_edit': can_edit,
    }
    
    return render(request, 'execution_plans/plan_detail.html', context)

# Utility functions

def can_access_plan(user, execution_plan):
    """Check if user can access the execution plan"""
    if user.role == 'top_management':
        return True
    elif user.role in ['business_head', 'business_head_ops']:
        return True
    elif user.role == 'rm_head':
        # Can access plans from subordinate RMs and own plans
        return (execution_plan.created_by == user or 
                execution_plan.created_by.manager == user)
    elif user.role == 'rm':
        # Can only access own plans
        return execution_plan.created_by == user
    elif user.role in ['ops_exec', 'ops_team_lead']:
        # Can access plans that are ready for or in execution
        return execution_plan.status in ['client_approved', 'in_execution', 'completed']
    
    return False


def can_approve_plan(user, execution_plan):
    """Check if user can approve the execution plan"""
    if execution_plan.status != 'pending_approval':
        return False
    
    return execution_plan.can_be_approved_by(user)


def can_execute_plan(user, execution_plan):
    """Check if user can execute the plan"""
    if user.role not in ['ops_exec', 'ops_team_lead']:
        return False
    
    return execution_plan.status in ['client_approved', 'in_execution']


def can_edit_plan(user, execution_plan):
    """Check if user can edit the plan"""
    if execution_plan.status not in ['draft', 'rejected']:
        return False
    
    return execution_plan.created_by == user


def notify_approval_required(execution_plan):
    """Notify line manager about approval requirement"""
    try:
        # Find the appropriate approver
        line_manager = execution_plan.created_by.get_line_manager()
        
        if line_manager and hasattr(line_manager, 'email'):
            subject = f"Execution Plan Approval Required - {execution_plan.plan_id}"
            
            email_content = render_to_string('execution_plans/emails/approval_required.html', {
                'execution_plan': execution_plan,
                'approver': line_manager,
                'rm': execution_plan.created_by,
            })
            
            send_mail(
                subject=subject,
                message='',
                html_message=email_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[line_manager.email],
            )
            
    except Exception as e:
        print(f"Error sending approval notification: {str(e)}")


def notify_ops_team(execution_plan):
    """Notify operations team about plan ready for execution"""
    try:
        # Get operations team members
        ops_users = User.objects.filter(
            role__in=['ops_exec', 'ops_team_lead', 'business_head_ops']
        ).exclude(email='')
        
        if ops_users.exists():
            subject = f"New Execution Plan Ready - {execution_plan.plan_id}"
            
            email_content = render_to_string('execution_plans/emails/ops_notification.html', {
                'execution_plan': execution_plan,
                'client': execution_plan.client,
                'rm': execution_plan.created_by,
            })
            
            recipient_list = [user.email for user in ops_users if user.email]
            
            send_mail(
                subject=subject,
                message='',
                html_message=email_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=recipient_list,
            )
            
    except Exception as e:
        print(f"Error sending ops notification: {str(e)}")


@login_required
def bulk_action_plans(request):
    """Bulk actions on execution plans"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        plan_ids = data.get('plan_ids', [])
        action = data.get('action')
        
        if not plan_ids or not action:
            return JsonResponse({'error': 'Missing required data'}, status=400)
        
        plans = ExecutionPlan.objects.filter(id__in=plan_ids)
        
        # Check permissions for all plans
        for plan in plans:
            if not can_access_plan(request.user, plan):
                return JsonResponse({'error': 'Access denied for one or more plans'}, status=403)
        
        success_count = 0
        error_count = 0
        
        if action == 'approve' and request.user.role in ['rm_head', 'business_head', 'top_management']:
            for plan in plans:
                if plan.status == 'pending_approval' and plan.can_be_approved_by(request.user):
                    if plan.approve(request.user):
                        PlanWorkflowHistory.objects.create(
                            execution_plan=plan,
                            from_status='pending_approval',
                            to_status='approved',
                            changed_by=request.user,
                            comments='Bulk approval'
                        )
                        success_count += 1
                    else:
                        error_count += 1
                else:
                    error_count += 1
        
        elif action == 'start_execution' and request.user.role in ['ops_exec', 'ops_team_lead']:
            for plan in plans:
                if plan.status == 'client_approved':
                    if plan.start_execution(request.user):
                        PlanWorkflowHistory.objects.create(
                            execution_plan=plan,
                            from_status='client_approved',
                            to_status='in_execution',
                            changed_by=request.user,
                            comments='Bulk execution start'
                        )
                        success_count += 1
                    else:
                        error_count += 1
                else:
                    error_count += 1
        
        else:
            return JsonResponse({'error': 'Invalid action or insufficient permissions'}, status=400)
        
        return JsonResponse({
            'success': True,
            'message': f'Action completed. Success: {success_count}, Errors: {error_count}'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def plan_analytics(request, plan_id):
    """Detailed analytics for a specific plan"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    if not can_access_plan(request.user, execution_plan):
        messages.error(request, "Access denied")
        return redirect('dashboard')
    
    # Get or create metrics
    metrics, created = ExecutionMetrics.objects.get_or_create(execution_plan=execution_plan)
    if created or not metrics.updated_at:
        metrics.calculate_metrics()
    
    # Get action-wise performance
    actions = execution_plan.actions.all().select_related('scheme', 'executed_by')
    
    # Calculate timeline data
    timeline_data = []
    
    if execution_plan.created_at:
        timeline_data.append({
            'event': 'Plan Created',
            'timestamp': execution_plan.created_at,
            'user': execution_plan.created_by.get_full_name() or execution_plan.created_by.username,
        })
    
    if execution_plan.submitted_at:
        timeline_data.append({
            'event': 'Submitted for Approval',
            'timestamp': execution_plan.submitted_at,
            'user': execution_plan.created_by.get_full_name() or execution_plan.created_by.username,
        })
    
    if execution_plan.approved_at:
        timeline_data.append({
            'event': 'Approved',
            'timestamp': execution_plan.approved_at,
            'user': execution_plan.approved_by.get_full_name() if execution_plan.approved_by else 'Unknown',
        })
    
    if execution_plan.client_approved_at:
        timeline_data.append({
            'event': 'Client Approved',
            'timestamp': execution_plan.client_approved_at,
            'user': execution_plan.created_by.get_full_name() or execution_plan.created_by.username,
        })
    
    if execution_plan.execution_started_at:
        timeline_data.append({
            'event': 'Execution Started',
            'timestamp': execution_plan.execution_started_at,
            'user': 'Operations Team',
        })
    
    if execution_plan.completed_at:
        timeline_data.append({
            'event': 'Execution Completed',
            'timestamp': execution_plan.completed_at,
            'user': 'Operations Team',
        })
    
    # Sort timeline by timestamp
    timeline_data.sort(key=lambda x: x['timestamp'])
    
    context = {
        'execution_plan': execution_plan,
        'metrics': metrics,
        'actions': actions,
        'timeline_data': timeline_data,
    }
    
    return render(request, 'execution_plans/plan_analytics.html', context)


# views.py - Add these missing portfolio views to your existing views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from .models import PortfolioActionMapping
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Count, Q
from django.core.paginator import Paginator
from django.utils import timezone
from .models import PortfolioUpload, ClientPortfolio, ClientProfile
from .forms import PortfolioUploadForm
import csv
import json

# Portfolio Dashboard
@login_required
def portfolio_dashboard(request):
    """Main portfolio dashboard"""
    # Summary statistics
    total_portfolios = ClientPortfolio.objects.filter(is_active=True).count()
    mapped_portfolios = ClientPortfolio.objects.filter(is_active=True, is_mapped=True).count()
    total_aum = ClientPortfolio.objects.filter(is_active=True).aggregate(
        total=Sum('total_value')
    )['total'] or 0
    
    # Recent uploads
    recent_uploads = PortfolioUpload.objects.order_by('-uploaded_at')[:5]
    
    # Portfolio distribution by asset class
    asset_distribution = ClientPortfolio.objects.filter(is_active=True).aggregate(
        total_equity=Sum('equity_value'),
        total_debt=Sum('debt_value'),
        total_hybrid=Sum('hybrid_value'),
        total_liquid=Sum('liquid_ultra_short_value'),
        total_other=Sum('other_value'),
        total_arbitrage=Sum('arbitrage_value')
    )
    
    # Top schemes by AUM
    top_schemes = ClientPortfolio.objects.filter(is_active=True).values(
        'scheme_name'
    ).annotate(
        total_aum=Sum('total_value'),
        investor_count=Count('client_pan', distinct=True)
    ).order_by('-total_aum')[:10]
    
    # Unmapped portfolios needing attention
    unmapped_portfolios = ClientPortfolio.objects.filter(
        is_active=True, 
        is_mapped=False
    ).count()
    
    context = {
        'total_portfolios': total_portfolios,
        'mapped_portfolios': mapped_portfolios,
        'mapping_percentage': (mapped_portfolios / total_portfolios * 100) if total_portfolios > 0 else 0,
        'total_aum': total_aum,
        'recent_uploads': recent_uploads,
        'asset_distribution': asset_distribution,
        'top_schemes': top_schemes,
        'unmapped_portfolios': unmapped_portfolios,
    }
    
    return render(request, 'portfolio/dashboard.html', context)

# Portfolio Upload
@login_required
def upload_portfolio(request):
    """Upload portfolio Excel file"""
    if request.method == 'POST':
        form = PortfolioUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                upload = form.save(commit=False)
                upload.uploaded_by = request.user
                upload.save()
                
                messages.success(
                    request, 
                    f"Portfolio file uploaded successfully. Upload ID: {upload.upload_id}"
                )
                
                # Process immediately if requested
                if form.cleaned_data.get('process_immediately'):
                    try:
                        # CHANGED: Use utility function instead of management command
                        from .utils import process_portfolio_upload
                        
                        results = process_portfolio_upload(upload_id=upload.upload_id, auto_map=True)
                        
                        # Refresh upload object to get updated status
                        upload.refresh_from_db()
                        
                        if upload.status == 'completed':
                            messages.success(
                                request, 
                                f"Portfolio processed successfully! {upload.successful_rows} rows processed, "
                                f"{results.get('mapping_successful', 0)} clients mapped."
                            )
                        elif upload.status == 'partial':
                            messages.warning(
                                request, 
                                f"Portfolio partially processed. {upload.successful_rows} of {upload.total_rows} "
                                f"rows processed successfully. {results.get('mapping_successful', 0)} clients mapped."
                            )
                        else:
                            messages.error(
                                request, 
                                "Portfolio processing failed. Please check the upload details for errors."
                            )
                            
                    except Exception as e:
                        messages.error(request, f"Error processing portfolio: {str(e)}")
                
                return redirect('upload_detail', upload_id=upload.upload_id)
                
            except Exception as e:
                messages.error(request, f"Error saving upload: {str(e)}")
                
    else:
        form = PortfolioUploadForm()
    
    # Get recent uploads for reference
    recent_uploads = PortfolioUpload.objects.filter(
        uploaded_by=request.user
    ).order_by('-uploaded_at')[:5]
    
    context = {
        'form': form,
        'recent_uploads': recent_uploads,
    }
    
    return render(request, 'portfolio/upload.html', context)

# Upload Detail
@login_required
def upload_detail(request, upload_id):
    """View details of a specific upload"""
    upload = get_object_or_404(PortfolioUpload, upload_id=upload_id)
    
    # Get associated portfolios
    portfolios = ClientPortfolio.objects.filter(upload_batch=upload)
    
    # Pagination
    paginator = Paginator(portfolios, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Summary stats for this upload
    upload_stats = {
        'total_portfolios': portfolios.count(),
        'mapped_portfolios': portfolios.filter(is_mapped=True).count(),
        'total_aum': portfolios.aggregate(total=Sum('total_value'))['total'] or 0,
        'unique_clients': portfolios.values('client_pan').distinct().count(),
        'unique_schemes': portfolios.values('scheme_name').distinct().count(),
    }
    
    context = {
        'upload': upload,
        'portfolios': page_obj,
        'upload_stats': upload_stats,
    }
    
    return render(request, 'portfolio/upload_detail.html', context)

# Portfolio List
@login_required
def portfolio_list(request):
    """List all portfolios with filtering"""
    portfolios = ClientPortfolio.objects.filter(is_active=True).select_related(
        'client_profile', 'upload_batch'
    )
    
    # Filtering
    search_query = request.GET.get('search')
    if search_query:
        portfolios = portfolios.filter(
            Q(client_name__icontains=search_query) |
            Q(client_pan__icontains=search_query) |
            Q(scheme_name__icontains=search_query)
        )
    
    mapped_filter = request.GET.get('mapped')
    if mapped_filter:
        portfolios = portfolios.filter(is_mapped=(mapped_filter == 'true'))
    
    asset_class_filter = request.GET.get('asset_class')
    if asset_class_filter:
        if asset_class_filter == 'equity':
            portfolios = portfolios.filter(equity_value__gt=0)
        elif asset_class_filter == 'debt':
            portfolios = portfolios.filter(debt_value__gt=0)
        elif asset_class_filter == 'hybrid':
            portfolios = portfolios.filter(hybrid_value__gt=0)
    
    # Sorting
    sort_by = request.GET.get('sort', '-total_value')
    if sort_by in ['total_value', '-total_value', 'client_name', '-client_name', 'scheme_name', '-scheme_name']:
        portfolios = portfolios.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(portfolios, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'portfolios': page_obj,
        'search_query': search_query,
        'mapped_filter': mapped_filter,
        'asset_class_filter': asset_class_filter,
        'sort_by': sort_by,
    }
    
    return render(request, 'portfolio/list.html', context)

# Portfolio Detail
@login_required
def portfolio_detail(request, portfolio_id):
    """View details of a specific portfolio"""
    portfolio = get_object_or_404(ClientPortfolio, id=portfolio_id)
    
    # Get related portfolios for the same client
    related_portfolios = ClientPortfolio.objects.filter(
        client_pan=portfolio.client_pan,
        is_active=True
    ).exclude(id=portfolio.id)
    
    # Calculate client's total portfolio if mapped
    if portfolio.client_profile:
        client_total_aum = ClientPortfolio.objects.filter(
            client_profile=portfolio.client_profile,
            is_active=True
        ).aggregate(total=Sum('total_value'))['total'] or 0
    else:
        client_total_aum = None
    
    context = {
        'portfolio': portfolio,
        'related_portfolios': related_portfolios,
        'client_total_aum': client_total_aum,
    }
    
    return render(request, 'portfolio/detail.html', context)

# Map Portfolio to Client (AJAX)
@login_required
@require_http_methods(["POST"])
def map_portfolio_to_client(request, portfolio_id):
    """AJAX endpoint to map a portfolio to a client profile"""
    portfolio = get_object_or_404(ClientPortfolio, id=portfolio_id)
    
    if portfolio.is_mapped:
        return JsonResponse({'success': False, 'message': 'Portfolio is already mapped'})
    
    try:
        mapped, message = portfolio.map_to_client_profile()
        
        if mapped:
            # Also try to map personnel
            personnel_mapped = portfolio.map_personnel()
            
            return JsonResponse({
                'success': True, 
                'message': f'{message}. Personnel mapped: {personnel_mapped}',
                'client_name': portfolio.client_profile.client_full_name if portfolio.client_profile else None
            })
        else:
            return JsonResponse({'success': False, 'message': message})
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Unmapped Portfolios
@login_required
def unmapped_portfolios(request):
    """View for managing unmapped portfolios"""
    unmapped = ClientPortfolio.objects.filter(
        is_active=True, 
        is_mapped=False
    ).order_by('client_pan', 'scheme_name')
    
    # Group by PAN for easier management
    pan_groups = {}
    for portfolio in unmapped:
        if portfolio.client_pan not in pan_groups:
            pan_groups[portfolio.client_pan] = {
                'client_name': portfolio.client_name,
                'portfolios': [],
                'total_aum': 0
            }
        pan_groups[portfolio.client_pan]['portfolios'].append(portfolio)
        pan_groups[portfolio.client_pan]['total_aum'] += portfolio.total_value
    
    # Check which PANs exist in ClientProfile
    existing_pans = set(ClientProfile.objects.values_list('pan_number', flat=True))
    
    for pan, data in pan_groups.items():
        data['exists_in_system'] = pan in existing_pans
    
    context = {
        'pan_groups': pan_groups,
        'total_unmapped': unmapped.count(),
        'total_unmapped_aum': unmapped.aggregate(total=Sum('total_value'))['total'] or 0,
    }
    
    return render(request, 'portfolio/unmapped.html', context)

# Portfolio Analytics
@login_required
def portfolio_analytics(request):
    """Portfolio analytics and reporting"""
    # Asset allocation summary
    asset_allocation = ClientPortfolio.objects.filter(is_active=True).aggregate(
        equity=Sum('equity_value'),
        debt=Sum('debt_value'),
        hybrid=Sum('hybrid_value'),
        liquid=Sum('liquid_ultra_short_value'),
        other=Sum('other_value'),
        arbitrage=Sum('arbitrage_value')
    )
    
    total_aum = sum(value for value in asset_allocation.values() if value)
    
    # Calculate percentages
    asset_percentages = {}
    for asset, value in asset_allocation.items():
        asset_percentages[asset] = (value / total_aum * 100) if total_aum > 0 and value else 0
    
    # Top clients by AUM
    top_clients = ClientPortfolio.objects.filter(
        is_active=True, 
        is_mapped=True
    ).values(
        'client_profile__client_full_name', 
        'client_pan'
    ).annotate(
        total_aum=Sum('total_value'),
        scheme_count=Count('scheme_name', distinct=True)
    ).order_by('-total_aum')[:20]
    
    # RM-wise distribution
    rm_distribution = ClientPortfolio.objects.filter(
        is_active=True,
        mapped_rm__isnull=False
    ).values(
        'mapped_rm__username',
        'mapped_rm__first_name',
        'mapped_rm__last_name'
    ).annotate(
        total_aum=Sum('total_value'),
        client_count=Count('client_pan', distinct=True),
        scheme_count=Count('scheme_name', distinct=True)
    ).order_by('-total_aum')
    
    # Scheme-wise analysis
    scheme_analysis = ClientPortfolio.objects.filter(is_active=True).values(
        'scheme_name'
    ).annotate(
        total_aum=Sum('total_value'),
        investor_count=Count('client_pan', distinct=True),
        avg_investment=Sum('total_value') / Count('client_pan', distinct=True)
    ).order_by('-total_aum')[:30]
    
    context = {
        'asset_allocation': asset_allocation,
        'asset_percentages': asset_percentages,
        'total_aum': total_aum,
        'top_clients': top_clients,
        'rm_distribution': rm_distribution,
        'scheme_analysis': scheme_analysis,
    }
    
    return render(request, 'portfolio/analytics.html', context)

# Bulk Map Portfolios
@login_required
@require_http_methods(["POST"])
def bulk_map_portfolios(request):
    """Bulk map portfolios to client profiles"""
    try:
        portfolio_ids = request.POST.getlist('portfolio_ids')
        
        if not portfolio_ids:
            messages.error(request, "No portfolios selected")
            return redirect('portfolio_list')
        
        portfolios = ClientPortfolio.objects.filter(
            id__in=portfolio_ids,
            is_mapped=False
        )
        
        mapped_count = 0
        for portfolio in portfolios:
            mapped, message = portfolio.map_to_client_profile()
            if mapped:
                portfolio.map_personnel()
                mapped_count += 1
        
        messages.success(
            request,
            f"Successfully mapped {mapped_count} out of {portfolios.count()} portfolios"
        )
        
    except Exception as e:
        messages.error(request, f"Error during bulk mapping: {str(e)}")
    
    return redirect('portfolio_list')

# Export Portfolios CSV
@login_required
def export_portfolios_csv(request):
    """Export portfolios to CSV"""
    portfolios = ClientPortfolio.objects.filter(is_active=True)
    
    # Apply filters if provided
    search = request.GET.get('search')
    if search:
        portfolios = portfolios.filter(
            Q(client_name__icontains=search) |
            Q(client_pan__icontains=search) |
            Q(scheme_name__icontains=search)
        )
    
    mapped_filter = request.GET.get('mapped')
    if mapped_filter:
        portfolios = portfolios.filter(is_mapped=(mapped_filter == 'true'))
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="portfolios_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Client Name', 'Client PAN', 'Scheme Name', 'Total Value',
        'Equity Value', 'Debt Value', 'Hybrid Value', 'Liquid Value',
        'Other Value', 'Arbitrage Value', 'Units', 'Allocation %',
        'Is Mapped', 'Client Profile', 'RM', 'Operations',
        'Upload Batch', 'Created Date'
    ])
    
    for portfolio in portfolios:
        writer.writerow([
            portfolio.client_name,
            portfolio.client_pan,
            portfolio.scheme_name,
            portfolio.total_value,
            portfolio.equity_value,
            portfolio.debt_value,
            portfolio.hybrid_value,
            portfolio.liquid_ultra_short_value,
            portfolio.other_value,
            portfolio.arbitrage_value,
            portfolio.units,
            portfolio.allocation_percentage,
            portfolio.is_mapped,
            portfolio.client_profile.client_full_name if portfolio.client_profile else '',
            portfolio.relationship_manager,
            portfolio.operations_personnel,
            portfolio.upload_batch.upload_id if portfolio.upload_batch else '',
            portfolio.created_at.strftime('%Y-%m-%d')
        ])
    
    return response

# Add these to your forms.py file if you don't have PortfolioUploadForm
class PortfolioUploadForm(forms.ModelForm):
    process_immediately = forms.BooleanField(
        required=False,
        initial=True,
        help_text="Process the file immediately after upload"
    )
    
    class Meta:
        model = PortfolioUpload
        fields = ['file']
        widgets = {
            'file': forms.FileInput(attrs={
                'accept': '.xlsx,.xls',
                'class': 'form-control'
            })
        }
    
    def clean_file(self):
        file = self.cleaned_data['file']
        
        # Check file extension
        if not file.name.lower().endswith(('.xlsx', '.xls')):
            raise forms.ValidationError("Please upload an Excel file (.xlsx or .xls)")
        
        # Check file size (max 10MB)
        if file.size > 10 * 1024 * 1024:
            raise forms.ValidationError("File size cannot exceed 10MB")
        
        return file
    
    
@login_required
def create_action_plan(request, portfolio_id):
    """Create new action plan for a portfolio"""
    portfolio = get_object_or_404(ClientPortfolio, id=portfolio_id)
    
    
    if request.method == 'POST':
        form = PortfolioActionPlanForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create action plan
                    action_plan = form.save(commit=False)
                    action_plan.client_portfolio = portfolio
                    action_plan.created_by = request.user
                    action_plan.save()
                    
                    # Create workflow entry
                    ActionPlanWorkflow.objects.create(
                        action_plan=action_plan,
                        from_status='',
                        to_status='draft',
                        changed_by=request.user,
                        notes='Action plan created'
                    )
                    
                    messages.success(
                        request, 
                        f"Action plan '{action_plan.plan_name}' created successfully"
                    )
                    return redirect('action_plan_detail', plan_id=action_plan.id)
                    
            except Exception as e:
                messages.error(request, f"Error creating action plan: {str(e)}")
    else:
        form = PortfolioActionPlanForm()
    
    context = {
        'form': form,
        'portfolio': portfolio,
    }
    
    return render(request, 'portfolio/create_action_plan.html', context)

@login_required
def action_plan_detail(request, plan_id):
    """View action plan details"""
    action_plan = get_object_or_404(PortfolioActionPlan, id=plan_id)

    
    # Get actions for this plan
    actions = action_plan.actions.all().order_by('priority', 'created_at')
    
    # Get comments
    comments = action_plan.comments.all().order_by('-created_at')
    
    # Get workflow history
    workflow_history = action_plan.workflow_history.all().order_by('-changed_at')
    
    # Check permissions
    can_edit = action_plan.can_be_edited() and action_plan.created_by == request.user
    can_approve = (action_plan.can_be_approved() and 
                   request.user.role in ['rm_head', 'business_head', 'top_management'] and
                   action_plan.created_by != request.user)
    
    context = {
        'action_plan': action_plan,
        'actions': actions,
        'comments': comments,
        'workflow_history': workflow_history,
        'can_edit': can_edit,
        'can_approve': can_approve,
    }
    
    return render(request, 'portfolio/action_plan_detail.html', context)

@login_required
def add_action_to_plan(request, plan_id):
    """Add action to existing action plan"""
    action_plan = get_object_or_404(PortfolioActionPlan, id=plan_id)
    
    # Check if user can edit this plan
    if not (action_plan.can_be_edited() and action_plan.created_by == request.user):
        messages.error(request, "Cannot add actions to this plan")
        return redirect('action_plan_detail', plan_id=plan_id)
    
    action_type = request.GET.get('type', 'redeem')
    
    if request.method == 'POST':
        try:
            action_data = {
                'action_plan': action_plan,
                'action_type': action_type,
                'source_scheme': action_plan.client_portfolio.scheme_name,
                'priority': action_plan.actions.count() + 1
            }
            
            if action_type == 'redeem':
                form = RedeemActionForm(request.POST)
                if form.is_valid():
                    action_data.update({
                        'redeem_by': form.cleaned_data['redeem_by'],
                        'redeem_amount': form.cleaned_data.get('redeem_amount'),
                        'redeem_units': form.cleaned_data.get('redeem_units'),
                    })
                else:
                    raise ValueError("Invalid redeem form data")
            
            elif action_type == 'switch':
                form = SwitchActionForm(request.POST)
                if form.is_valid():
                    action_data.update({
                        'target_scheme': form.cleaned_data['target_scheme'],
                        'switch_by': form.cleaned_data['switch_by'],
                        'switch_amount': form.cleaned_data.get('switch_amount'),
                        'switch_units': form.cleaned_data.get('switch_units'),
                    })
                else:
                    raise ValueError("Invalid switch form data")
            
            elif action_type == 'stp':
                form = STPActionForm(request.POST)
                if form.is_valid():
                    action_data.update({
                        'target_scheme': form.cleaned_data['target_scheme'],
                        'stp_amount': form.cleaned_data['stp_amount'],
                        'stp_frequency': form.cleaned_data['stp_frequency'],
                    })
                else:
                    raise ValueError("Invalid STP form data")
            
            elif action_type == 'sip':
                form = SIPActionForm(request.POST)
                if form.is_valid():
                    action_data.update({
                        'target_scheme': form.cleaned_data['target_scheme'],
                        'sip_amount': form.cleaned_data['sip_amount'],
                        'sip_frequency': form.cleaned_data['sip_frequency'],
                        'sip_date': form.cleaned_data['sip_date'],
                    })
                else:
                    raise ValueError("Invalid SIP form data")
            
            elif action_type == 'swp':
                form = SWPActionForm(request.POST)
                if form.is_valid():
                    action_data.update({
                        'swp_amount': form.cleaned_data['swp_amount'],
                        'swp_frequency': form.cleaned_data['swp_frequency'],
                        'swp_date': form.cleaned_data['swp_date'],
                    })
                else:
                    raise ValueError("Invalid SWP form data")
            
            # Create the action
            action = PortfolioAction.objects.create(**action_data)
            
            messages.success(request, f"{action.get_action_type_display()} action added successfully")
            return redirect('action_plan_detail', plan_id=plan_id)
            
        except Exception as e:
            messages.error(request, f"Error adding action: {str(e)}")
    
    # Prepare form based on action type
    if action_type == 'redeem':
        form = RedeemActionForm()
    elif action_type == 'switch':
        form = SwitchActionForm()
    elif action_type == 'stp':
        form = STPActionForm()
    elif action_type == 'sip':
        form = SIPActionForm()
    elif action_type == 'swp':
        form = SWPActionForm()
    else:
        messages.error(request, "Invalid action type")
        return redirect('action_plan_detail', plan_id=plan_id)
    
    context = {
        'action_plan': action_plan,
        'form': form,
        'action_type': action_type,
        'action_type_display': dict(PortfolioAction.ACTION_TYPE_CHOICES)[action_type],
    }
    
    return render(request, 'portfolio/add_action.html', context)

@login_required
def quick_action(request, portfolio_id, action_type):
    """Quick action creation from portfolio detail page"""
    portfolio = get_object_or_404(ClientPortfolio, id=portfolio_id)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Create action plan
            action_plan = PortfolioActionPlan.objects.create(
                client_portfolio=portfolio,
                plan_name=f"Quick {action_type.upper()} - {portfolio.scheme_name}",
                description=f"Quick {action_type} action for {portfolio.client_name}",
                created_by=request.user
            )
            
            # Create action based on type
            action_data = {
                'action_plan': action_plan,
                'action_type': action_type,
                'source_scheme': portfolio.scheme_name,
                'priority': 1
            }
            
            if action_type == 'redeem':
                action_data.update({
                    'redeem_by': data.get('redeem_by'),
                    'redeem_amount': data.get('redeem_amount'),
                    'redeem_units': data.get('redeem_units'),
                })
            elif action_type == 'switch':
                action_data.update({
                    'target_scheme': data.get('target_scheme'),
                    'switch_by': data.get('switch_by'),
                    'switch_amount': data.get('switch_amount'),
                    'switch_units': data.get('switch_units'),
                })
            # Add other action types as needed
            
            action = PortfolioAction.objects.create(**action_data)
            
            return JsonResponse({
                'success': True,
                'message': f'{action_type.title()} action created successfully',
                'action_plan_id': action_plan.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)

@login_required
@require_http_methods(["POST"])
def submit_action_plan(request, plan_id):
    """Submit action plan for approval"""
    action_plan = get_object_or_404(PortfolioActionPlan, id=plan_id)
    
    if action_plan.created_by != request.user:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if not action_plan.actions.exists():
        return JsonResponse({'error': 'Cannot submit plan without actions'}, status=400)
    
    try:
        if action_plan.submit_for_approval():
            # Create workflow entry
            ActionPlanWorkflow.objects.create(
                action_plan=action_plan,
                from_status='draft',
                to_status='pending_approval',
                changed_by=request.user,
                notes='Plan submitted for approval'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Action plan submitted for approval successfully'
            })
        else:
            return JsonResponse({'error': 'Cannot submit this plan'}, status=400)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def approve_action_plan(request, plan_id):
    """Approve or reject action plan"""
    action_plan = get_object_or_404(PortfolioActionPlan, id=plan_id)
    
    # Check permission
    if not (request.user.role in ['rm_head', 'business_head', 'top_management'] and
            action_plan.created_by != request.user):
        messages.error(request, "Access denied")
        return redirect('action_plan_detail', plan_id=plan_id)
    
    if request.method == 'POST':
        form = ActionPlanApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            notes = form.cleaned_data['notes']
            
            try:
                if action == 'approve':
                    if action_plan.approve(request.user, notes):
                        # Create workflow entry
                        ActionPlanWorkflow.objects.create(
                            action_plan=action_plan,
                            from_status='pending_approval',
                            to_status='approved',
                            changed_by=request.user,
                            notes=notes or 'Plan approved'
                        )
                        messages.success(request, "Action plan approved successfully")
                    else:
                        messages.error(request, "Cannot approve this plan")
                        
                elif action == 'reject':
                    if action_plan.reject(request.user, notes):
                        # Create workflow entry
                        ActionPlanWorkflow.objects.create(
                            action_plan=action_plan,
                            from_status='pending_approval',
                            to_status='rejected',
                            changed_by=request.user,
                            notes=notes or 'Plan rejected'
                        )
                        messages.success(request, "Action plan rejected")
                    else:
                        messages.error(request, "Cannot reject this plan")
                        
                return redirect('action_plan_detail', plan_id=plan_id)
                
            except Exception as e:
                messages.error(request, f"Error processing approval: {str(e)}")
    else:
        form = ActionPlanApprovalForm()
    
    context = {
        'action_plan': action_plan,
        'form': form,
    }
    
    return render(request, 'portfolio/approve_action_plan.html', context)

@login_required
def action_plan_list(request):
    """List all action plans with filtering"""
    # Get accessible action plans based on user role
    if request.user.role in ['top_management', 'business_head']:
        action_plans = PortfolioActionPlan.objects.all()
    elif request.user.role == 'rm_head':
        team_members = request.user.subordinates.all()
        action_plans = PortfolioActionPlan.objects.filter(
            created_by__in=list(team_members) + [request.user]
        )
    else:
        action_plans = PortfolioActionPlan.objects.filter(created_by=request.user)
    
    # Filtering
    status_filter = request.GET.get('status')
    if status_filter:
        action_plans = action_plans.filter(status=status_filter)
    
    search_query = request.GET.get('search')
    if search_query:
        action_plans = action_plans.filter(
            Q(plan_name__icontains=search_query) |
            Q(client_portfolio__client_name__icontains=search_query) |
            Q(client_portfolio__scheme_name__icontains=search_query)
        )
    
    # Sorting
    sort_by = request.GET.get('sort', '-created_at')
    if sort_by in ['-created_at', 'created_at', 'plan_name', '-plan_name', 'status']:
        action_plans = action_plans.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(action_plans.select_related('client_portfolio', 'created_by'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'action_plans': page_obj,
        'status_filter': status_filter,
        'search_query': search_query,
        'sort_by': sort_by,
        'status_choices': PortfolioActionPlan.STATUS_CHOICES,
    }
    
    return render(request, 'portfolio/action_plan_list.html', context)



@login_required
@require_http_methods(["GET"])
def get_action_details(request, action_id):
    """Get action details for execution modal"""
    try:
        action = get_object_or_404(PlanAction, id=action_id)
        execution_plan = action.execution_plan
        
        # Check permissions
        if not can_execute_plan(request.user, execution_plan):
            return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
        
        # Check if action can be executed
        if action.status != 'pending':
            return JsonResponse({
                'success': False, 
                'error': f'Action cannot be executed. Current status: {action.status}'
            }, status=400)
        
        # Prepare action details
        action_details = {
            'id': action.id,
            'action_type': action.action_type,
            'scheme_name': action.scheme.scheme_name if action.scheme else '',
            'scheme_id': action.scheme.id if action.scheme else None,
            'target_scheme_name': action.target_scheme.scheme_name if action.target_scheme else '',
            'target_scheme_id': action.target_scheme.id if action.target_scheme else None,
            'amount': str(action.amount) if action.amount else '',
            'units': str(action.units) if action.units else '',
            'sip_date': action.sip_date,
            'frequency': action.frequency,
            'transaction_type': action.transaction_type,
            'notes': action.notes or '',
            'priority': action.priority,
        }
        
        return JsonResponse({
            'success': True,
            'action': action_details,
        })
        
    except Exception as e:
        logger.error(f"Error getting action details: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to load action details'}, status=500)
    

@login_required
@require_http_methods(["POST"])
def track_workflow(request, plan_id):
    """Track workflow status changes for execution plan"""
    try:
        execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
        
        # Check permissions
        if not can_access_plan(request.user, execution_plan):
            return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
        
        # Get form data
        from_status = request.POST.get('from_status', '')
        to_status = request.POST.get('to_status', '')
        comments = request.POST.get('comments', '')
        
        # Validate status transition
        valid_transitions = {
            '': ['draft'],
            'draft': ['pending_approval', 'rejected'],
            'pending_approval': ['approved', 'rejected'],
            'approved': ['client_approved', 'rejected'],
            'client_approved': ['in_execution'],
            'in_execution': ['completed', 'paused'],
            'paused': ['in_execution', 'cancelled'],
            'rejected': ['draft'],
            'completed': [],  # Completed is final
            'cancelled': []   # Cancelled is final
        }
        
        if to_status not in valid_transitions.get(from_status, []):
            return JsonResponse({
                'success': False, 
                'error': f'Invalid status transition from {from_status} to {to_status}'
            }, status=400)
        
        with transaction.atomic():
            # Create workflow history record
            workflow_history = PlanWorkflowHistory.objects.create(
                execution_plan=execution_plan,
                from_status=from_status,
                to_status=to_status,
                changed_by=request.user,
                comments=comments,
                changed_at=timezone.now()
            )
            
            # Update execution plan status if it's different
            if execution_plan.status != to_status:
                old_status = execution_plan.status
                execution_plan.status = to_status
                
                # Set timestamp fields based on status
                if to_status == 'pending_approval':
                    execution_plan.submitted_at = timezone.now()
                elif to_status == 'approved':
                    execution_plan.approved_at = timezone.now()
                    execution_plan.approved_by = request.user
                elif to_status == 'client_approved':
                    execution_plan.client_approved_at = timezone.now()
                elif to_status == 'in_execution':
                    execution_plan.execution_started_at = timezone.now()
                elif to_status == 'completed':
                    execution_plan.completed_at = timezone.now()
                elif to_status == 'rejected':
                    execution_plan.rejected_at = timezone.now()
                    execution_plan.rejected_by = request.user
                
                execution_plan.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Workflow updated from {from_status or "initial"} to {to_status}',
            'workflow_id': workflow_history.id,
            'current_status': execution_plan.status,
            'timestamp': workflow_history.changed_at.strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        logger.error(f"Error tracking workflow: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to track workflow change'}, status=500)
    
    


####RESET PASSWORD######


from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.auth.forms import PasswordChangeForm
from django.utils.encoding import force_bytes, force_str
from django.urls import reverse
from django.contrib.sites.shortcuts import get_current_site
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
import logging
from django.middleware.csrf import get_token
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie

@ensure_csrf_cookie
@require_http_methods(["POST"])
def password_reset_request(request):
    """
    Handle password reset requests via AJAX
    """
    try:
        data = json.loads(request.body)
        username_or_email = data.get('username_or_email', '').strip().lower()
        
        if not username_or_email:
            return JsonResponse({
                'success': False,
                'message': 'Username or email is required.'
            }, status=400)
        
        User = get_user_model()
        user = None
        
        # Try to find user by email
        if '@' in username_or_email:
            try:
                user = User.objects.get(email__iexact=username_or_email, is_active=True)
            except User.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'No active account found with this email address.'
                }, status=404)
        
        # If not email, try by username
        if not user:
            try:
                user = User.objects.get(username__iexact=username_or_email, is_active=True)
            except User.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'No active account found with this username.'
                }, status=404)
        
        # Check if user has an email address
        if not user.email:
            return JsonResponse({
                'success': False,
                'message': 'This account has no email address associated with it.'
            }, status=400)
        
        # Generate password reset token
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        
        # Get current site
        current_site = get_current_site(request)
        
        # Build reset URL
        reset_url = request.build_absolute_uri(
            reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
        )
        
        # Prepare email context
        context = {
            'user': user,
            'reset_url': reset_url,
            'site_name': current_site.name,
            'domain': current_site.domain,
            'protocol': 'https' if request.is_secure() else 'http',
            'company_name': getattr(settings, 'COMPANY_NAME', 'Financial CRM'),
        }
        
        # Render email templates
        subject = f"Password Reset - {context['company_name']}"
        email_template_name = 'registration/password_reset_email.html'
        email_template_text = 'registration/password_reset_email.txt'
        
        try:
            # Try HTML template first
            html_message = render_to_string(email_template_name, context)
            text_message = render_to_string(email_template_text, context)
        except:
            # Fallback to simple text message
            html_message = f"""
            <h2>Password Reset Request</h2>
            <p>Hello {user.get_full_name() or user.username},</p>
            <p>You have requested a password reset for your {context['company_name']} account.</p>
            <p>Click the link below to reset your password:</p>
            <p><a href="{reset_url}" style="background-color: #1C64FF; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">Reset Password</a></p>
            <p>If you didn't request this password reset, please ignore this email.</p>
            <p>This link will expire in 24 hours for security reasons.</p>
            <br>
            <p>Best regards,<br>{context['company_name']} Team</p>
            """
            
            text_message = f"""
            Password Reset Request
            
            Hello {user.get_full_name() or user.username},
            
            You have requested a password reset for your {context['company_name']} account.
            
            Click the link below to reset your password:
            {reset_url}
            
            If you didn't request this password reset, please ignore this email.
            This link will expire in 24 hours for security reasons.
            
            Best regards,
            {context['company_name']} Team
            """
        
        # Send email
        try:
            send_mail(
                subject=subject,
                message=text_message,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@example.com'),
                recipient_list=[user.email],
                html_message=html_message,
                fail_silently=False,
            )
            
            logger.info(f"Password reset email sent successfully to {user.email}")
            
            return JsonResponse({
                'success': True,
                'message': 'Password reset instructions have been sent to your email address.'
            })
            
        except Exception as e:
            logger.error(f"Failed to send password reset email to {user.email}: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': 'Failed to send reset email. Please try again later or contact support.'
            }, status=500)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request format.'
        }, status=400)
    except Exception as e:
        logger.error(f"Unexpected error in password reset: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An unexpected error occurred. Please try again later.'
        }, status=500)

def password_reset_confirm(request, uidb64, token):
    """
    Handle password reset confirmation page
    """
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    # Check if token is valid
    if user is not None and default_token_generator.check_token(user, token):
        validlink = True
        
        if request.method == 'POST':
            password1 = request.POST.get('new_password1')
            password2 = request.POST.get('new_password2')
            
            # Validate passwords
            if not password1 or not password2:
                messages.error(request, 'Both password fields are required.')
            elif password1 != password2:
                messages.error(request, 'Passwords do not match.')
            elif len(password1) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
            else:
                # Set new password
                user.set_password(password1)
                user.save()
                
                # Log the user out from all sessions for security
                from django.contrib.sessions.models import Session
                for session in Session.objects.all():
                    try:
                        session_data = session.get_decoded()
                        if session_data.get('_auth_user_id') == str(user.id):
                            session.delete()
                    except:
                        pass
                
                messages.success(request, 'Your password has been reset successfully. You can now log in with your new password.')
                return redirect('login')
    else:
        validlink = False
        messages.error(request, 'This password reset link is invalid or has expired.')
    
    context = {
        'validlink': validlink,
        'user': user,
        'uidb64': uidb64,
        'token': token,
    }
    
    return render(request, 'registration/password_reset_confirm.html', context)
# Add this view to your views.py

@login_required
@require_http_methods(["GET"])
def get_action_statuses(request, plan_id):
    """Get action statuses for execution plan (AJAX endpoint)"""
    try:
        execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
        
        # Check access permission
        if not can_access_plan(request.user, execution_plan):
            return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
        
        # Get all actions for this plan
        actions = execution_plan.actions.all().select_related('scheme', 'target_scheme', 'executed_by')
        
        action_data = []
        for action in actions:
            # Get scheme information
            scheme_name = "Unknown Scheme"
            if action.scheme:
                scheme_name = action.scheme.scheme_name
            
            target_scheme_name = ""
            if action.target_scheme:
                target_scheme_name = action.target_scheme.scheme_name
            
            # Get execution details
            executed_by_name = ""
            if action.executed_by:
                executed_by_name = action.executed_by.get_full_name() or action.executed_by.username
            
            # Get action display name
            action_type_display = action.get_action_type_display() if hasattr(action, 'get_action_type_display') else action.action_type
            
            # Get status display
            status_display = action.get_status_display() if hasattr(action, 'get_status_display') else action.status
            
            action_info = {
                'id': action.id,
                'action_type': action.action_type,
                'action_type_display': action_type_display,
                'scheme_name': scheme_name,
                'target_scheme_name': target_scheme_name,
                'amount': str(action.amount) if action.amount else '',
                'units': str(action.units) if action.units else '',
                'status': action.status,
                'status_display': status_display,
                'priority': action.priority,
                'executed_by': executed_by_name,
                'executed_at': action.executed_at.isoformat() if hasattr(action, 'executed_at') and action.executed_at else None,
                'transaction_id': getattr(action, 'transaction_id', ''),
                'notes': action.notes or '',
                'can_execute': action.status == 'pending' and can_execute_plan(request.user, execution_plan),
                'can_mark_failed': action.status == 'pending' and can_execute_plan(request.user, execution_plan),
            }
            
            # Add action-specific fields
            if action.action_type == 'sip_start':
                action_info.update({
                    'sip_date': action.sip_date,
                    'frequency': getattr(action, 'frequency', ''),
                })
            elif action.action_type in ['switch', 'stp_start']:
                action_info['target_scheme_required'] = True
            
            action_data.append(action_info)
        
        # Calculate summary statistics
        total_actions = len(action_data)
        pending_actions = len([a for a in action_data if a['status'] == 'pending'])
        completed_actions = len([a for a in action_data if a['status'] == 'completed'])
        failed_actions = len([a for a in action_data if a['status'] == 'failed'])
        
        summary = {
            'total_actions': total_actions,
            'pending_actions': pending_actions,
            'completed_actions': completed_actions,
            'failed_actions': failed_actions,
            'completion_percentage': round((completed_actions / total_actions * 100), 2) if total_actions > 0 else 0,
        }
        
        return JsonResponse({
            'success': True,
            'actions': action_data,
            'summary': summary,
            'plan_status': execution_plan.status,
            'plan_id': execution_plan.plan_id if hasattr(execution_plan, 'plan_id') else plan_id,
        })
        
    except ExecutionPlan.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Execution plan not found'}, status=404)
    except Exception as e:
        logger.error(f"Error getting action statuses for plan {plan_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to load action statuses'}, status=500)


# Also add this helper view for updating action status
@login_required
@require_http_methods(["POST"])
def update_action_status(request, action_id):
    """Update individual action status (AJAX endpoint)"""
    try:
        action = get_object_or_404(PlanAction, id=action_id)
        execution_plan = action.execution_plan
        
        # Check permissions
        if not can_execute_plan(request.user, execution_plan):
            return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
        
        data = json.loads(request.body)
        new_status = data.get('status')
        notes = data.get('notes', '')
        
        # Validate status
        valid_statuses = ['pending', 'in_progress', 'completed', 'failed', 'skipped']
        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)
        
        # Update action
        old_status = action.status
        action.status = new_status
        
        if notes:
            action.notes = notes
        
        # Set timestamp fields based on status
        if new_status == 'completed':
            action.executed_at = timezone.now()
            action.executed_by = request.user
        elif new_status == 'failed':
            action.failed_at = timezone.now() if hasattr(action, 'failed_at') else None
            action.failed_by = request.user if hasattr(action, 'failed_by') else None
        
        action.save()
        
        # Check if all actions are completed to auto-complete the plan
        remaining_actions = execution_plan.actions.exclude(status__in=['completed', 'failed', 'skipped']).count()
        
        plan_status_changed = False
        if remaining_actions == 0 and execution_plan.status == 'in_execution':
            execution_plan.status = 'completed'
            execution_plan.completed_at = timezone.now()
            execution_plan.save()
            plan_status_changed = True
            
            # Create workflow history
            PlanWorkflowHistory.objects.create(
                execution_plan=execution_plan,
                from_status='in_execution',
                to_status='completed',
                changed_by=request.user,
                comments='All actions completed - plan auto-completed'
            )
        
        return JsonResponse({
            'success': True,
            'message': f'Action status updated from {old_status} to {new_status}',
            'action_id': action.id,
            'old_status': old_status,
            'new_status': new_status,
            'plan_status_changed': plan_status_changed,
            'plan_completed': plan_status_changed,
            'remaining_actions': remaining_actions,
        })
        
    except PlanAction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Action not found'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error updating action status: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to update action status'}, status=500)


# Add this view for getting execution progress
@login_required
@require_http_methods(["GET"])
def get_execution_progress(request, plan_id):
    """Get execution progress for a plan (AJAX endpoint)"""
    try:
        execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
        
        # Check access permission
        if not can_access_plan(request.user, execution_plan):
            return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
        
        # Get action counts by status
        actions = execution_plan.actions.all()
        status_counts = {
            'total': actions.count(),
            'pending': actions.filter(status='pending').count(),
            'in_progress': actions.filter(status='in_progress').count(),
            'completed': actions.filter(status='completed').count(),
            'failed': actions.filter(status='failed').count(),
            'skipped': actions.filter(status='skipped').count(),
        }
        
        # Calculate percentages
        total = status_counts['total']
        progress = {
            'completion_percentage': round((status_counts['completed'] / total * 100), 2) if total > 0 else 0,
            'failure_rate': round((status_counts['failed'] / total * 100), 2) if total > 0 else 0,
            'remaining_percentage': round(((status_counts['pending'] + status_counts['in_progress']) / total * 100), 2) if total > 0 else 0,
        }
        
        # Get recent activity
        recent_actions = actions.filter(
            status__in=['completed', 'failed']
        ).order_by('-executed_at' if hasattr(PlanAction, 'executed_at') else '-updated_at')[:5]
        
        recent_activity = []
        for action in recent_actions:
            activity = {
                'action_type': action.get_action_type_display() if hasattr(action, 'get_action_type_display') else action.action_type,
                'scheme_name': action.scheme.scheme_name if action.scheme else 'Unknown',
                'status': action.status,
                'timestamp': action.executed_at.isoformat() if hasattr(action, 'executed_at') and action.executed_at else None,
                'executed_by': action.executed_by.get_full_name() if hasattr(action, 'executed_by') and action.executed_by else None,
            }
            recent_activity.append(activity)
        
        return JsonResponse({
            'success': True,
            'plan_status': execution_plan.status,
            'status_counts': status_counts,
            'progress': progress,
            'recent_activity': recent_activity,
            'can_execute': can_execute_plan(request.user, execution_plan),
            'last_updated': timezone.now().isoformat(),
        })
        
    except Exception as e:
        logger.error(f"Error getting execution progress: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to load execution progress'}, status=500)

# Add these enhanced email functions to your views.py

import logging
from django.core.mail import EmailMessage, send_mail
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone
from datetime import datetime
import os

logger = logging.getLogger(__name__)

def get_client_email(execution_plan):
    """
    Get client email from various possible sources
    """
    client_email = None
    
    # Try to get email from different client model relationships
    try:
        # First try: Direct client email
        if hasattr(execution_plan.client, 'email') and execution_plan.client.email:
            client_email = execution_plan.client.email
        
        # Second try: Client contact_info (if it's an email)
        elif hasattr(execution_plan.client, 'contact_info') and execution_plan.client.contact_info:
            contact_info = execution_plan.client.contact_info
            if '@' in contact_info:
                client_email = contact_info
        
        # Third try: Client profile email
        elif hasattr(execution_plan.client, 'client_profile') and execution_plan.client.client_profile:
            if hasattr(execution_plan.client.client_profile, 'email') and execution_plan.client.client_profile.email:
                client_email = execution_plan.client.client_profile.email
        
        # Fourth try: User email (if client is linked to a user)
        elif hasattr(execution_plan.client, 'user') and execution_plan.client.user:
            if hasattr(execution_plan.client.user, 'email') and execution_plan.client.user.email:
                client_email = execution_plan.client.user.email
                
        logger.info(f"Found client email: {client_email} for plan {execution_plan.plan_id}")
        
    except Exception as e:
        logger.error(f"Error getting client email for plan {execution_plan.plan_id}: {str(e)}")
    
    return client_email


def send_execution_plan_email(execution_plan, email_type, include_excel=False, 
                            send_to_rm=True, send_to_client=False):
    """
    Send email notification about execution plan status change
    """
    recipients = []
    
    # Add operations team email
    ops_email = getattr(settings, 'OPERATIONS_EMAIL', None)
    if ops_email:
        recipients.append(ops_email)
    
    # Add RM email if requested
    if send_to_rm and execution_plan.created_by.email:
        recipients.append(execution_plan.created_by.email)
    
    # Add client email if requested and available
    client_email = None
    if send_to_client:
        client_email = execution_plan.client.email or (
            execution_plan.client.contact_info if '@' in execution_plan.client.contact_info else None
        )
        if client_email:
            recipients.append(client_email)
    
    if not recipients:
        raise ValueError("No valid recipients found for email")
    
    # Prepare email content
    subject = f"Execution Plan {execution_plan.plan_name} - {email_type.replace('_', ' ').title()}"
    
    context = {
        'plan': execution_plan,
        'client': execution_plan.client,
        'email_type': email_type,
    }
    
    text_message = render_to_string(f'execution_plans/emails/{email_type}.txt', context)
    html_message = render_to_string(f'execution_plans/emails/{email_type}.html', context)
    
    # Prepare attachments
    attachments = []
    if include_excel and execution_plan.excel_file:
        excel_content = execution_plan.excel_file.read()
        attachments.append((
            f'execution_plan_{execution_plan.plan_id}.xlsx',
            excel_content,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ))
    
    # Send email
    email = EmailMultiAlternatives(
        subject=subject,
        body=text_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=recipients,
        attachments=attachments
    )
    email.attach_alternative(html_message, "text/html")
    email.send()


def attach_excel_file(email, execution_plan):
    """
    Attach Excel file to email message
    """
    try:
        # Check if Excel file exists
        if execution_plan.excel_file and os.path.exists(execution_plan.excel_file.path):
            file_path = execution_plan.excel_file.path
            filename = f"{execution_plan.plan_id}_execution_plan.xlsx"
        else:
            # Generate Excel file if it doesn't exist
            excel_file_path = generate_execution_plan_excel(execution_plan)
            if excel_file_path:
                execution_plan.excel_file = excel_file_path
                execution_plan.save()
                file_path = os.path.join(settings.MEDIA_ROOT, excel_file_path)
                filename = f"{execution_plan.plan_id}_execution_plan.xlsx"
            else:
                return False
        
        # Read and attach file
        with open(file_path, 'rb') as f:
            email.attach(
                filename, 
                f.read(), 
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        return True
        
    except Exception as e:
        logger.error(f"Error attaching Excel file: {str(e)}")
        return False


def generate_default_email_content(context, email_type):
    """
    Generate default email content when templates are not available
    """
    execution_plan = context['execution_plan']
    client_name = context['client_name']
    rm_name = context['rm'].get_full_name() if context['rm'] else "Your Relationship Manager"
    
    # Email content based on type
    content_map = {
        'approved': {
            'greeting': f"Dear {client_name},",
            'main_message': f"We are pleased to inform you that your investment execution plan '{execution_plan.plan_name}' has been approved and is ready for your review.",
            'action_required': "Please review the attached execution plan details. Once you approve, we will proceed with the execution.",
            'closing': "We look forward to helping you achieve your investment goals."
        },
        'completed': {
            'greeting': f"Dear {client_name},",
            'main_message': f"Your investment execution plan '{execution_plan.plan_name}' has been successfully executed.",
            'action_required': "Please find the attached execution summary with all transaction details.",
            'closing': "Thank you for trusting us with your investments."
        },
        'updated': {
            'greeting': f"Dear {client_name},",
            'main_message': f"Your investment execution plan '{execution_plan.plan_name}' has been updated with new information.",
            'action_required': "Please review the attached updated plan details.",
            'closing': "If you have any questions, please don't hesitate to contact us."
        },
        'client_approved': {
            'greeting': f"Dear {client_name},",
            'main_message': f"Thank you for approving your investment execution plan '{execution_plan.plan_name}'.",
            'action_required': "Our operations team will now begin executing your plan. You will receive updates as each transaction is completed.",
            'closing': "We appreciate your trust in our services."
        }
    }
    
    content = content_map.get(email_type, content_map['updated'])
    
    # Generate HTML message
    html_message = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background-color: #1C64FF; color: white; padding: 20px; text-align: center; }}
            .content {{ padding: 20px; background-color: #f9f9f9; }}
            .plan-details {{ background-color: white; padding: 15px; margin: 15px 0; border-left: 4px solid #1C64FF; }}
            .footer {{ background-color: #333; color: white; padding: 15px; text-align: center; font-size: 12px; }}
            .btn {{ background-color: #1C64FF; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; display: inline-block; margin: 10px 0; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>{context['company_name']}</h1>
                <h2>Investment Execution Plan</h2>
            </div>
            
            <div class="content">
                <p>{content['greeting']}</p>
                
                <p>{content['main_message']}</p>
                
                <div class="plan-details">
                    <h3>Plan Details:</h3>
                    <ul>
                        <li><strong>Plan Name:</strong> {execution_plan.plan_name}</li>
                        <li><strong>Plan ID:</strong> {execution_plan.plan_id}</li>
                        <li><strong>Status:</strong> {execution_plan.get_status_display() if hasattr(execution_plan, 'get_status_display') else execution_plan.status}</li>
                        <li><strong>Created Date:</strong> {execution_plan.created_at.strftime('%B %d, %Y') if execution_plan.created_at else 'N/A'}</li>
                        <li><strong>Your RM:</strong> {rm_name}</li>
                    </ul>
                </div>
                
                <p>{content['action_required']}</p>
                
                {f'<p><strong>Additional Message:</strong><br>{context["custom_message"]}</p>' if context.get('custom_message') else ''}
                
                <p>If you have any questions or concerns, please don't hesitate to contact your relationship manager at {context['rm'].email if context['rm'] and context['rm'].email else 'your usual contact'}.</p>
                
                <p>{content['closing']}</p>
            </div>
            
            <div class="footer">
                <p>{context['company_name']} | Investment Management Services</p>
                <p>This is an automated message. Please do not reply to this email.</p>
                <p>Generated on {context['current_date']}</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Generate text message
    text_message = f"""
{content['greeting']}

{content['main_message']}

Plan Details:
- Plan Name: {execution_plan.plan_name}
- Plan ID: {execution_plan.plan_id}
- Status: {execution_plan.get_status_display() if hasattr(execution_plan, 'get_status_display') else execution_plan.status}
- Created Date: {execution_plan.created_at.strftime('%B %d, %Y') if execution_plan.created_at else 'N/A'}
- Your RM: {rm_name}

{content['action_required']}

{f'Additional Message: {context["custom_message"]}' if context.get('custom_message') else ''}

If you have any questions or concerns, please contact your relationship manager.

{content['closing']}

Best regards,
{context['company_name']} Team

---
This is an automated message generated on {context['current_date']}.
    """
    
    return html_message, text_message


# Enhanced workflow functions with automatic email sending

@login_required
@require_http_methods(["POST"])
def approve_plan_with_email(request, plan_id):
    """
    Approve execution plan and automatically send email to client
    """
    try:
        execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
        
        # Check permission
        if not can_approve_plan(request.user, execution_plan):
            return JsonResponse({
                'success': False, 
                'error': 'Access denied. Only Business Head or Top Management can approve execution plans.'
            }, status=403)
        
        if execution_plan.status != 'pending_approval':
            return JsonResponse({
                'success': False,
                'error': f'Plan cannot be approved. Current status: {execution_plan.get_status_display()}'
            }, status=400)
        
        comments = request.POST.get('comments', '')
        send_email = request.POST.get('send_email', 'true').lower() == 'true'
        
        # Use transaction for data consistency
        with transaction.atomic():
            # Approve the plan
            if execution_plan.approve(request.user):
                # Create workflow history
                PlanWorkflowHistory.objects.create(
                    execution_plan=execution_plan,
                    from_status='pending_approval',
                    to_status='approved',
                    changed_by=request.user,
                    comments=comments or 'Plan approved'
                )
                
                # Add comment if provided
                if comments:
                    PlanComment.objects.create(
                        execution_plan=execution_plan,
                        comment=comments,
                        commented_by=request.user,
                        is_internal=True
                    )
                
                # Send email notification if requested
                email_sent = False
                email_message = ""
                
                if send_email:
                    email_success, email_result = send_execution_plan_email(
                        execution_plan=execution_plan,
                        email_type='approved',
                        custom_message=comments,
                        include_excel=True,
                        send_to_rm=True,
                        send_to_client=True
                    )
                    
                    if email_success:
                        email_sent = True
                        email_message = email_result
                    else:
                        logger.warning(f"Email failed for approved plan {execution_plan.plan_id}: {email_result}")
                        email_message = f"Plan approved but email failed: {email_result}"
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Plan approved successfully',
                    'new_status': execution_plan.status,
                    'email_sent': email_sent,
                    'email_message': email_message
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'error': 'Failed to approve plan'
                }, status=400)
    
    except Exception as e:
        logger.error(f"Error approving plan with email: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': 'An error occurred while approving the plan'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def mark_client_approved_with_email(request, plan_id):
    """
    Mark plan as client approved and send notification
    """
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check permission - only RM who created the plan can mark as client approved
    if execution_plan.created_by != request.user:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if execution_plan.mark_client_approved():
        # Create workflow history
        PlanWorkflowHistory.objects.create(
            execution_plan=execution_plan,
            from_status='approved',
            to_status='client_approved',
            changed_by=request.user,
            comments='Client approval confirmed'
        )
        
        # Send email to operations team
        send_execution_plan_email(
            execution_plan=execution_plan,
            email_type='client_approved',
            include_excel=True,
            send_to_rm=True,
            send_to_client=False  # This goes to ops team
        )
        
        # Notify operations team
        notify_ops_team(execution_plan)
        
        return JsonResponse({
            'success': True, 
            'message': 'Plan marked as client approved and operations team notified'
        })
    else:
        return JsonResponse({'error': 'Failed to mark as client approved'}, status=400)

@login_required
@require_http_methods(["POST"])
def complete_execution(request, plan_id):
    """Mark execution plan as completed"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check permission
    if not can_execute_plan(request.user, execution_plan):
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    if execution_plan.status != 'in_progress':
        return JsonResponse({'success': False, 'error': 'Plan is not in progress'}, status=400)
    
    completion_notes = request.POST.get('completion_notes', '')
    notify_client = request.POST.get('notify_client') == 'true'
    generate_report = request.POST.get('generate_report') == 'true'
    
    try:
        with transaction.atomic():
            # Mark as completed
            execution_plan.status = 'completed'
            execution_plan.completed_at = timezone.now()
            execution_plan.save()
            
            # Create workflow history
            PlanWorkflowHistory.objects.create(
                execution_plan=execution_plan,
                from_status='in_progress',
                to_status='completed',
                changed_by=request.user,
                comments=completion_notes or 'Execution completed'
            )
            
            # Add completion notes if provided
            if completion_notes:
                PlanComment.objects.create(
                    execution_plan=execution_plan,
                    comment=f"Completion Notes: {completion_notes}",
                    commented_by=request.user,
                    is_internal=True
                )
            
            # Generate metrics
            if generate_report:
                metrics, created = ExecutionMetrics.objects.get_or_create(execution_plan=execution_plan)
                metrics.calculate_metrics()
        
        return JsonResponse({
            'success': True,
            'message': 'Execution plan marked as completed successfully',
            'new_status': execution_plan.status
        })
        
    except Exception as e:
        logger.error(f"Error completing execution plan: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to complete execution'}, status=500)
    
@login_required
@require_http_methods(["POST"])  
def complete_execution_with_email(request, plan_id):
    """
    Mark execution as completed and send completion email to client - FIXED VERSION
    """
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check permission
    if not can_execute_plan(request.user, execution_plan):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        with transaction.atomic():
            # Mark as completed
            execution_plan.status = 'completed'
            execution_plan.completed_at = timezone.now()
            execution_plan.save()
            
            # Create workflow history
            PlanWorkflowHistory.objects.create(
                execution_plan=execution_plan,
                from_status='in_execution',
                to_status='completed',
                changed_by=request.user,
                comments='Execution completed successfully'
            )
            
            # Generate fresh Excel with execution results
            try:
                excel_data = generate_execution_plan_excel_in_memory(execution_plan)
                if excel_data:
                    logger.info(f"Generated completion Excel for plan {execution_plan.plan_id}")
            except Exception as e:
                logger.warning(f"Could not generate completion Excel: {str(e)}")
            
            # FIXED: Send completion email without custom_message parameter
            try:
                # Get client email
                client_email = get_client_email(execution_plan)
                
                if client_email:
                    # Use the enhanced email system instead
                    from django.core.mail import EmailMultiAlternatives
                    from django.template.loader import render_to_string
                    
                    subject = f"✅ Investment Plan Completed - {execution_plan.plan_name}"
                    
                    # Get client name
                    client_name = execution_plan.client.name if hasattr(execution_plan.client, 'name') else 'Valued Client'
                    rm_name = execution_plan.created_by.get_full_name() or execution_plan.created_by.username
                    
                    # Create completion message
                    message = f"""Dear {client_name},

Excellent news! Your investment execution plan "{execution_plan.plan_name}" has been completed successfully.

All {execution_plan.actions.count()} action{'s' if execution_plan.actions.count() != 1 else ''} in your plan have been executed as planned.

Execution Summary:
- Plan ID: {execution_plan.plan_id}
- Completion Date: {execution_plan.completed_at.strftime('%B %d, %Y') if execution_plan.completed_at else 'Today'}
- Total Actions Executed: {execution_plan.actions.count()}
- Status: Completed Successfully

Your portfolio has been updated according to the executed plan. You can expect to see these changes reflected in your statements within the next 1-2 business days.

If you have any questions about the executed transactions or would like to discuss your portfolio further, please don't hesitate to contact me.

Thank you for your trust in our services.

Best regards,
{rm_name}
{execution_plan.created_by.email if execution_plan.created_by.email else ''}"""
                    
                    # Create HTML version
                    html_message = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <meta charset="utf-8">
                        <style>
                            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                            .header {{ background-color: #16a34a; color: white; padding: 20px; text-align: center; border-radius: 8px 8px 0 0; }}
                            .content {{ padding: 20px; background-color: #f9f9f9; }}
                            .summary {{ background-color: white; padding: 15px; margin: 15px 0; border-left: 4px solid #16a34a; border-radius: 4px; }}
                            .footer {{ background-color: #333; color: white; padding: 15px; text-align: center; font-size: 12px; border-radius: 0 0 8px 8px; }}
                        </style>
                    </head>
                    <body>
                        <div class="container">
                            <div class="header">
                                <h1>🎯 Plan Execution Completed!</h1>
                                <h2>{execution_plan.plan_name}</h2>
                            </div>
                            
                            <div class="content">
                                <p>Dear {client_name},</p>
                                
                                <p><strong>Excellent news!</strong> Your investment execution plan has been completed successfully.</p>
                                
                                <div class="summary">
                                    <h3>✅ Execution Summary</h3>
                                    <ul>
                                        <li><strong>Plan ID:</strong> {execution_plan.plan_id}</li>
                                        <li><strong>Completion Date:</strong> {execution_plan.completed_at.strftime('%B %d, %Y at %I:%M %p') if execution_plan.completed_at else 'Today'}</li>
                                        <li><strong>Total Actions:</strong> {execution_plan.actions.count()}</li>
                                        <li><strong>Status:</strong> ✅ Completed Successfully</li>
                                        <li><strong>Your RM:</strong> {rm_name}</li>
                                    </ul>
                                </div>
                                
                                <p>All actions in your plan have been executed as planned. Your portfolio has been updated accordingly.</p>
                                
                                <p>You can expect to see these changes reflected in your statements within the next 1-2 business days.</p>
                                
                                <p>If you have any questions about the executed transactions or would like to discuss your portfolio further, please don't hesitate to contact your relationship manager.</p>
                                
                                <p><strong>Thank you for your trust in our services.</strong></p>
                                
                                <p>Best regards,<br>{rm_name}</p>
                            </div>
                            
                            <div class="footer">
                                <p>Investment Management Services | Execution Completed Successfully</p>
                                <p>This is an automated message generated on {timezone.now().strftime('%B %d, %Y')}</p>
                            </div>
                        </div>
                    </body>
                    </html>
                    """
                    
                    # Prepare recipients
                    recipients = [client_email]
                    cc_recipients = []
                    
                    # Add RM to CC
                    if execution_plan.created_by and execution_plan.created_by.email:
                        if execution_plan.created_by.email not in recipients:
                            cc_recipients.append(execution_plan.created_by.email)
                    
                    # Create and send email
                    email = EmailMultiAlternatives(
                        subject=subject,
                        body=message,
                        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@yourcompany.com'),
                        to=recipients,
                        cc=cc_recipients if cc_recipients else None,
                        reply_to=[execution_plan.created_by.email] if execution_plan.created_by and execution_plan.created_by.email else None
                    )
                    
                    # Attach HTML version
                    email.attach_alternative(html_message, "text/html")
                    
                    # Attach Excel if available
                    excel_attached = False
                    if excel_data:
                        try:
                            filename = f"{execution_plan.plan_id}_execution_completed.xlsx"
                            email.attach(filename, excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                            excel_attached = True
                        except Exception as e:
                            logger.warning(f"Could not attach Excel: {str(e)}")
                    
                    # Send email
                    email.send(fail_silently=False)
                    
                    email_success = True
                    email_result = f'Completion email sent successfully to {client_email}'
                    if excel_attached:
                        email_result += ' with Excel attachment'
                    
                else:
                    email_success = False
                    email_result = 'No client email found'
                
            except Exception as e:
                logger.error(f"Error sending completion email: {str(e)}")
                email_success = False
                email_result = f'Failed to send email: {str(e)}'
            
            return JsonResponse({
                'success': True,
                'message': 'Execution completed successfully',
                'email_sent': email_success,
                'email_message': email_result
            })
            
    except Exception as e:
        logger.error(f"Error completing execution: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# Modified existing view to include automatic email
@login_required
@require_http_methods(["POST"])
def submit_for_approval_with_email(request, plan_id):
    """
    Submit execution plan for approval with email notification
    """
    try:
        execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
        
        # Check permission
        if execution_plan.created_by != request.user:
            return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
        
        if execution_plan.status != 'draft':
            return JsonResponse({
                'success': False, 
                'error': f'Plan cannot be submitted for approval. Current status: {execution_plan.get_status_display()}'
            }, status=400)
        
        # Validate that plan has actions
        if not execution_plan.actions.exists():
            return JsonResponse({
                'success': False, 
                'error': 'Cannot submit plan without any actions'
            }, status=400)
        
        # Use transaction to ensure data consistency
        with transaction.atomic():
            # Update plan status
            execution_plan.status = 'pending_approval'
            execution_plan.submitted_at = timezone.now()
            execution_plan.save()
            
            # Create workflow history
            PlanWorkflowHistory.objects.create(
                execution_plan=execution_plan,
                from_status='draft',
                to_status='pending_approval',
                changed_by=request.user,
                comments='Plan submitted for approval'
            )
            
            # Send email notification to approval managers
            try:
                # Find approval managers
                approval_managers = User.objects.filter(
                    role__in=['rm_head', 'business_head', 'business_head_ops', 'top_management']
                )
                
                for manager in approval_managers:
                    if manager.email:
                        send_approval_notification_email(execution_plan, manager)
                        
            except Exception as e:
                logger.warning(f"Failed to send approval notifications: {str(e)}")
        
        return JsonResponse({
            'success': True, 
            'message': 'Plan submitted for approval successfully',
            'new_status': execution_plan.status
        })
        
    except Exception as e:
        logger.error(f"Error submitting plan for approval: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': 'Failed to submit plan for approval'
        }, status=500)


def send_approval_notification_email(execution_plan, manager):
    """
    Send email to manager for approval request
    """
    try:
        subject = f"📋 Execution Plan Approval Required - {execution_plan.plan_name}"
        
        html_message = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background-color: #1C64FF; color: white; padding: 20px; text-align: center;">
                <h2>Execution Plan Approval Required</h2>
            </div>
            
            <div style="padding: 20px; background-color: #f9f9f9;">
                <p>Dear {manager.get_full_name() or manager.username},</p>
                
                <p>A new execution plan requires your approval:</p>
                
                <div style="background-color: white; padding: 15px; margin: 15px 0; border-left: 4px solid #1C64FF;">
                    <h3>Plan Details:</h3>
                    <ul>
                        <li><strong>Plan Name:</strong> {execution_plan.plan_name}</li>
                        <li><strong>Plan ID:</strong> {execution_plan.plan_id}</li>
                        <li><strong>Client:</strong> {execution_plan.client.name if hasattr(execution_plan.client, 'name') else 'N/A'}</li>
                        <li><strong>Created By:</strong> {execution_plan.created_by.get_full_name() or execution_plan.created_by.username}</li>
                        <li><strong>Submitted:</strong> {execution_plan.submitted_at.strftime('%B %d, %Y at %I:%M %p') if execution_plan.submitted_at else 'N/A'}</li>
                        <li><strong>Total Actions:</strong> {execution_plan.actions.count()}</li>
                    </ul>
                </div>
                
                <p>Please log into the system to review and approve this execution plan.</p>
                
                <div style="text-align: center; margin: 20px 0;">
                    <a href="{getattr(settings, 'SITE_URL', 'https://your-domain.com')}/execution-plans/{execution_plan.id}/" 
                       style="background-color: #1C64FF; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                        Review Plan
                    </a>
                </div>
            </div>
            
            <div style="background-color: #333; color: white; padding: 15px; text-align: center; font-size: 12px;">
                <p>This is an automated notification from the Investment Management System</p>
            </div>
        </div>
        """
        
        text_message = f"""
Execution Plan Approval Required

Dear {manager.get_full_name() or manager.username},

A new execution plan requires your approval:

Plan Details:
- Plan Name: {execution_plan.plan_name}
- Plan ID: {execution_plan.plan_id}  
- Client: {execution_plan.client.name if hasattr(execution_plan.client, 'name') else 'N/A'}
- Created By: {execution_plan.created_by.get_full_name() or execution_plan.created_by.username}
- Submitted: {execution_plan.submitted_at.strftime('%B %d, %Y at %I:%M %p') if execution_plan.submitted_at else 'N/A'}
- Total Actions: {execution_plan.actions.count()}

Please log into the system to review and approve this execution plan.

Best regards,
Investment Management System
        """
        
        send_mail(
            subject=subject,
            message=text_message,
            html_message=html_message,
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@yourcompany.com'),
            recipient_list=[manager.email],
            fail_silently=False
        )
        
        logger.info(f"Approval notification sent to {manager.email} for plan {execution_plan.plan_id}")
        
    except Exception as e:
        logger.error(f"Error sending approval notification to {manager.email}: {str(e)}")


# Utility function to send bulk emails
def send_bulk_execution_plan_emails(execution_plans, email_type='updated'):
    """
    Send emails for multiple execution plans
    """
    results = []
    
    for plan in execution_plans:
        try:
            success, message = send_execution_plan_email(
                execution_plan=plan,
                email_type=email_type,
                include_excel=True,
                send_to_client=True,
                send_to_rm=True
            )
            
            results.append({
                'plan_id': plan.plan_id,
                'success': success,
                'message': message
            })
            
        except Exception as e:
            results.append({
                'plan_id': plan.plan_id,
                'success': False,
                'message': str(e)
            })
    
    return results

@login_required
@require_http_methods(["POST"])
def send_to_client_enhanced(request, plan_id):
    """Enhanced send to client with comprehensive email handling and automatic client email detection"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    # Check access permission
    if not can_access_plan(request.user, execution_plan):
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        # Get form data
        email_to = request.POST.get('email_to', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        include_excel = request.POST.get('include_excel', 'false').lower() == 'true'
        generate_fresh = request.POST.get('generate_fresh', 'false').lower() == 'true'
        copy_to_rm = request.POST.get('copy_to_rm', 'false').lower() == 'true'
        copy_to_ops = request.POST.get('copy_to_ops', 'false').lower() == 'true'
        
        # Auto-detect client email if not provided
        if not email_to:
            email_to = get_client_email(execution_plan)
            if not email_to:
                return JsonResponse({
                    'success': False, 
                    'error': 'No client email provided or found in database. Please enter email manually.'
                })
        
        # Validate email format
        from django.core.validators import validate_email
        try:
            validate_email(email_to)
        except ValidationError:
            return JsonResponse({
                'success': False, 
                'error': f'Invalid email address format: {email_to}'
            })
        
        # Set default subject if not provided
        if not subject:
            if execution_plan.status == 'approved':
                subject = f"✅ Your Investment Plan is Ready for Review - {execution_plan.plan_name}"
            elif execution_plan.status == 'completed':
                subject = f"🎯 Investment Plan Executed Successfully - {execution_plan.plan_name}"
            else:
                subject = f"📋 Investment Plan Update - {execution_plan.plan_name}"
        
        # Set default message if not provided
        if not message:
            client_name = execution_plan.client.name if hasattr(execution_plan.client, 'name') else 'Valued Client'
            rm_name = execution_plan.created_by.get_full_name() or execution_plan.created_by.username
            
            if execution_plan.status == 'approved':
                message = f"""Dear {client_name},

I hope this email finds you well.

I'm pleased to inform you that your investment execution plan "{execution_plan.plan_name}" has been approved and is ready for your review.

The plan includes {execution_plan.actions.count()} carefully selected action{'s' if execution_plan.actions.count() != 1 else ''} designed to optimize your investment portfolio based on our recent discussions and market analysis.

Please review the attached execution plan details. Once you're satisfied with the plan, please confirm your approval so we can proceed with the execution.

Key highlights of your plan:
- Plan ID: {execution_plan.plan_id}
- Total Actions: {execution_plan.actions.count()}
- Created Date: {execution_plan.created_at.strftime('%B %d, %Y') if execution_plan.created_at else 'N/A'}

If you have any questions or would like to discuss any aspect of the plan, please don't hesitate to reach out to me directly.

Thank you for your continued trust in our services.

Best regards,
{rm_name}
{execution_plan.created_by.email if execution_plan.created_by.email else ''}"""
            
            elif execution_plan.status == 'completed':
                message = f"""Dear {client_name},

Excellent news! Your investment execution plan "{execution_plan.plan_name}" has been completed successfully.

All {execution_plan.actions.count()} action{'s' if execution_plan.actions.count() != 1 else ''} in your plan have been executed as planned. Please find attached the detailed execution summary with all transaction confirmations.

Execution Summary:
- Plan ID: {execution_plan.plan_id}
- Completion Date: {execution_plan.completed_at.strftime('%B %d, %Y') if execution_plan.completed_at else 'Today'}
- Total Actions Executed: {execution_plan.actions.count()}
- Success Rate: 100%

Your portfolio has been updated according to the executed plan. You can expect to see these changes reflected in your statements within the next 1-2 business days.

If you have any questions about the executed transactions or would like to discuss your portfolio further, please don't hesitate to contact me.

Thank you for your trust in our services.

Best regards,
{rm_name}"""
            
            else:
                message = f"""Dear {client_name},

I hope this email finds you well.

Your investment execution plan "{execution_plan.plan_name}" has been updated with new information.

Plan Details:
- Plan ID: {execution_plan.plan_id}
- Current Status: {execution_plan.get_status_display() if hasattr(execution_plan, 'get_status_display') else execution_plan.status}
- Total Actions: {execution_plan.actions.count()}

Please review the attached plan details. If you have any questions or concerns, please don't hesitate to contact me.

Best regards,
{rm_name}"""
        
        # Prepare email recipients
        recipients = [email_to]
        cc_recipients = []
        bcc_recipients = []
        
        # Add RM to CC if requested
        if copy_to_rm and execution_plan.created_by and execution_plan.created_by.email:
            if execution_plan.created_by.email not in recipients:
                cc_recipients.append(execution_plan.created_by.email)
        
        # Add operations team to BCC if requested
        if copy_to_ops:
            ops_emails = User.objects.filter(
                role__in=['ops_team_lead', 'ops_exec', 'business_head_ops'],
                email__isnull=False
            ).exclude(email='').values_list('email', flat=True)
            
            for ops_email in ops_emails:
                if ops_email not in recipients and ops_email not in cc_recipients:
                    bcc_recipients.append(ops_email)
        
        # Generate fresh Excel if requested and include_excel is True
        excel_file_path = None
        if include_excel:
            if generate_fresh or not execution_plan.excel_file:
                try:
                    excel_file_path = generate_execution_plan_excel(execution_plan)
                    if excel_file_path:
                        execution_plan.excel_file = excel_file_path
                        execution_plan.save()
                        logger.info(f"Generated fresh Excel file for plan {execution_plan.plan_id}")
                except Exception as e:
                    logger.warning(f"Could not generate fresh Excel: {str(e)}")
                    # Continue with existing file if available
            
            # Use existing file if fresh generation failed or wasn't requested
            if execution_plan.excel_file:
                excel_file_path = execution_plan.excel_file.path if hasattr(execution_plan.excel_file, 'path') else os.path.join(settings.MEDIA_ROOT, str(execution_plan.excel_file))
        
        # Create email message
        email = EmailMultiAlternatives(
            subject=subject,
            body=message,  # Plain text version
            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@yourcompany.com'),
            to=recipients,
            cc=cc_recipients if cc_recipients else None,
            bcc=bcc_recipients if bcc_recipients else None,
            reply_to=[execution_plan.created_by.email] if execution_plan.created_by and execution_plan.created_by.email else None
        )
        
        # Create HTML version of the email
        company_name = getattr(settings, 'COMPANY_NAME', 'Investment Management')
        current_date = timezone.now().strftime('%B %d, %Y')
        
        html_message = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{subject}</title>
            <style>
                body {{ 
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
                    line-height: 1.6; 
                    color: #333; 
                    margin: 0; 
                    padding: 0; 
                    background-color: #f5f5f5;
                }}
                .container {{ 
                    max-width: 600px; 
                    margin: 20px auto; 
                    background-color: #ffffff;
                    border-radius: 10px;
                    overflow: hidden;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                }}
                .header {{ 
                    background: linear-gradient(135deg, #1C64FF 0%, #0052CC 100%);
                    color: white; 
                    padding: 30px 20px; 
                    text-align: center; 
                }}
                .header h1 {{
                    margin: 0 0 10px 0;
                    font-size: 24px;
                    font-weight: 600;
                }}
                .header h2 {{
                    margin: 0;
                    font-size: 18px;
                    font-weight: 400;
                    opacity: 0.9;
                }}
                .content {{ 
                    padding: 30px 20px; 
                    white-space: pre-line; 
                    font-size: 16px;
                }}
                .plan-details {{ 
                    background-color: #f8f9fa; 
                    padding: 20px; 
                    margin: 20px 0; 
                    border-left: 4px solid #1C64FF; 
                    border-radius: 5px;
                }}
                .plan-details h3 {{
                    margin-top: 0;
                    color: #1C64FF;
                    font-size: 18px;
                }}
                .plan-details ul {{
                    margin: 0;
                    padding-left: 20px;
                }}
                .plan-details li {{
                    margin-bottom: 8px;
                }}
                .status-badge {{
                    display: inline-block;
                    padding: 4px 12px;
                    border-radius: 20px;
                    font-size: 12px;
                    font-weight: 600;
                    text-transform: uppercase;
                    letter-spacing: 0.5px;
                }}
                .status-approved {{ background-color: #d4edda; color: #155724; }}
                .status-completed {{ background-color: #d1ecf1; color: #0c5460; }}
                .status-pending {{ background-color: #fff3cd; color: #856404; }}
                .footer {{ 
                    background-color: #343a40; 
                    color: white; 
                    padding: 20px; 
                    text-align: center; 
                    font-size: 14px;
                }}
                .footer p {{
                    margin: 5px 0;
                }}
                .disclaimer {{
                    background-color: #e9ecef;
                    padding: 15px;
                    margin: 20px 0;
                    border-radius: 5px;
                    font-size: 12px;
                    color: #6c757d;
                    text-align: center;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>{company_name}</h1>
                    <h2>Investment Execution Plan</h2>
                </div>
                
                <div class="content">
                    {message}
                    
                    <div class="plan-details">
                        <h3>📋 Plan Information</h3>
                        <ul>
                            <li><strong>Plan Name:</strong> {execution_plan.plan_name}</li>
                            <li><strong>Plan ID:</strong> {execution_plan.plan_id}</li>
                            <li><strong>Status:</strong> 
                                <span class="status-badge status-{execution_plan.status}">
                                    {execution_plan.get_status_display() if hasattr(execution_plan, 'get_status_display') else execution_plan.status}
                                </span>
                            </li>
                            <li><strong>Total Actions:</strong> {execution_plan.actions.count()}</li>
                            <li><strong>Created Date:</strong> {execution_plan.created_at.strftime('%B %d, %Y') if execution_plan.created_at else 'N/A'}</li>
                            <li><strong>Your Relationship Manager:</strong> {execution_plan.created_by.get_full_name() or execution_plan.created_by.username}</li>
                        </ul>
                    </div>
                    
                    {"<p><strong>📎 Attachment:</strong> Detailed execution plan (Excel format)</p>" if include_excel else ""}
                </div>
                
                <div class="disclaimer">
                    This email contains confidential information intended only for the specified recipient(s). 
                    If you have received this email in error, please notify the sender immediately.
                </div>
                
                <div class="footer">
                    <p><strong>{company_name}</strong> | Investment Management Services</p>
                    <p>Email sent on {current_date}</p>
                    <p>This is an automated message from our Investment Management System</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Attach HTML version
        email.attach_alternative(html_message, "text/html")
        
        # Attach Excel file if requested and available
        excel_attached = False
        if include_excel and excel_file_path and os.path.exists(excel_file_path):
            try:
                filename = f"{execution_plan.plan_id}_execution_plan.xlsx"
                with open(excel_file_path, 'rb') as f:
                    email.attach(
                        filename, 
                        f.read(), 
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                excel_attached = True
                logger.info(f"Excel file attached: {filename}")
            except Exception as e:
                logger.error(f"Error attaching Excel file: {str(e)}")
                excel_attached = False
        
        # Send email
        try:
            email.send(fail_silently=False)
            logger.info(f"Enhanced client email sent successfully for plan {execution_plan.plan_id} to {email_to}")
            
            # Create comment record for audit trail
            attachment_note = " with Excel attachment" if excel_attached else ""
            cc_note = f" (CC: {', '.join(cc_recipients)})" if cc_recipients else ""
            bcc_note = f" (BCC: Operations team)" if bcc_recipients else ""
            
            PlanComment.objects.create(
                execution_plan=execution_plan,
                comment=f"Plan sent to client via enhanced email: {email_to}{attachment_note}{cc_note}{bcc_note}",
                commented_by=request.user,
                is_internal=True
            )
            
            # Update plan metadata if needed
            if not hasattr(execution_plan, 'client_communication_sent') or not execution_plan.client_communication_sent:
                try:
                    execution_plan.client_communication_sent = True
                    execution_plan.save(update_fields=['client_communication_sent'])
                except:
                    pass  # Field might not exist in model
            
            return JsonResponse({
                'success': True,
                'message': f'Email sent successfully to {email_to}',
                'details': {
                    'recipient': email_to,
                    'cc_recipients': cc_recipients,
                    'excel_attached': excel_attached,
                    'subject': subject,
                    'timestamp': timezone.now().isoformat()
                }
            })
            
        except Exception as e:
            logger.error(f"Error sending enhanced client email: {str(e)}")
            return JsonResponse({
                'success': False, 
                'error': f'Failed to send email: {str(e)}'
            })
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request data'})
    except Exception as e:
        logger.error(f"Unexpected error in send_to_client_enhanced: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': f'An unexpected error occurred: {str(e)}'
        })


# Helper function to send completion email (used by completion notification)
@login_required
@require_http_methods(["POST"])
def send_completion_email(request, plan_id):
    """Send completion email directly (used by auto-completion detection)"""
    execution_plan = get_object_or_404(ExecutionPlan, id=plan_id)
    
    if not can_access_plan(request.user, execution_plan):
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        # Auto-detect client email
        client_email = get_client_email(execution_plan)
        
        if not client_email:
            return JsonResponse({
                'success': False, 
                'error': 'No client email found for completion notification'
            })
        
        # Send completion email using the enhanced email system
        email_success, email_result = send_execution_plan_email(
            execution_plan=execution_plan,
            email_type='completed',
            custom_message="All transactions in your execution plan have been completed successfully.",
            include_excel=True,
            send_to_rm=True,
            send_to_client=True
        )
        
        if email_success:
            return JsonResponse({
                'success': True,
                'message': email_result,
                'recipient': client_email
            })
        else:
            return JsonResponse({
                'success': False,
                'error': email_result
            })
            
    except Exception as e:
        logger.error(f"Error sending completion email: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_http_methods(["GET", "POST"])
def change_password(request):
    """
    Simple and secure password change view
    """
    if request.method == 'POST':
        # Handle AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                data = json.loads(request.body)
                old_password = data.get('old_password')
                new_password1 = data.get('new_password1')
                new_password2 = data.get('new_password2')
                
                # Validate old password
                if not request.user.check_password(old_password):
                    return JsonResponse({
                        'success': False,
                        'error': 'Current password is incorrect.'
                    })
                
                # Validate new passwords
                if not new_password1 or not new_password2:
                    return JsonResponse({
                        'success': False,
                        'error': 'Both new password fields are required.'
                    })
                
                if new_password1 != new_password2:
                    return JsonResponse({
                        'success': False,
                        'error': 'New passwords do not match.'
                    })
                
                if len(new_password1) < 8:
                    return JsonResponse({
                        'success': False,
                        'error': 'Password must be at least 8 characters long.'
                    })
                
                # Check if new password is same as old password
                if old_password == new_password1:
                    return JsonResponse({
                        'success': False,
                        'error': 'New password must be different from current password.'
                    })
                
                # Change password
                request.user.set_password(new_password1)
                request.user.save()
                
                # Keep user logged in after password change
                update_session_auth_hash(request, request.user)
                
                return JsonResponse({
                    'success': True,
                    'message': 'Password changed successfully!'
                })
                
            except json.JSONDecodeError:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid request format.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': 'An error occurred while changing password.'
                })
        
        # Handle regular form submission
        else:
            form = PasswordChangeForm(request.user, request.POST)
            if form.is_valid():
                user = form.save()
                # Keep user logged in after password change
                update_session_auth_hash(request, user)
                messages.success(request, 'Your password was successfully updated!')
                return redirect('change_password')
            else:
                # Extract error messages
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, error)
    else:
        form = PasswordChangeForm(request.user)
    
    context = {
        'form': form,
        'user': request.user,
    }
    
    return render(request, 'registration/change_password.html', context)